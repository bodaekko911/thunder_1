import asyncio
import io
import json
from types import SimpleNamespace

import openpyxl

from app.core.log import ActivityLog
from app.models.product import Product
from app.routers import import_data
from app.services import receive_import_service


def _run(coro):
    return asyncio.run(coro)


async def _read_streaming_response(response) -> bytes:
    chunks = []
    async for chunk in response.body_iterator:
        chunks.append(chunk)
    return b"".join(chunks)


def _make_xlsx(rows: list[list[object]], headers: list[str] | None = None) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers or ["SKU", "Product", "QTY", "Unit Price", "Product Type", "Date"])
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class _FakeResult:
    def __init__(self, values):
        self._values = list(values) if isinstance(values, list) else ([values] if values is not None else [])

    def scalar_one_or_none(self):
        return self._values[0] if self._values else None

    def scalars(self):
        return self

    def all(self):
        return list(self._values)


class FakeReceiveImportSession:
    def __init__(self, *, products=None, logs=None):
        self.products = list(products or [])
        self.logs = list(logs or [])
        self.added = []
        self.commits = 0

    async def execute(self, stmt):
        entity = None
        for desc in getattr(stmt, "column_descriptions", []):
            if desc.get("entity") is not None:
                entity = desc["entity"]
                break
        if entity is Product:
            return _FakeResult(self.products)
        if entity is ActivityLog:
            items = list(self.logs)
            where_text = " ".join(str(expr) for expr in getattr(stmt, "_where_criteria", []))
            params = stmt.compile().params
            if "activity_logs.module" in where_text:
                items = [log for log in items if log.module == params.get("module_1")]
            if "activity_logs.action =" in where_text:
                action = params.get("action_1")
                if action is not None:
                    items = [log for log in items if log.action == action]
            if "activity_logs.action IN" in where_text:
                allowed = set()
                for key, value in params.items():
                    if not key.startswith("action_"):
                        continue
                    if isinstance(value, (list, tuple, set)):
                        allowed.update(value)
                    else:
                        allowed.add(value)
                items = [log for log in items if log.action in allowed]
            if "activity_logs.ref_id =" in where_text:
                ref_id = params.get("ref_id_1")
                if ref_id is not None:
                    items = [log for log in items if log.ref_id == ref_id]
            if "activity_logs.ref_type =" in where_text:
                ref_type = params.get("ref_type_1")
                if ref_type is not None:
                    items = [log for log in items if log.ref_type == ref_type]
            return _FakeResult(items)
        return _FakeResult([])

    def add(self, obj):
        self.added.append(obj)
        if isinstance(obj, ActivityLog):
            obj.id = obj.id or len(self.logs) + 1
            self.logs.append(obj)

    async def commit(self):
        self.commits += 1


def _make_product(product_id: int, sku: str, name: str) -> Product:
    product = Product(id=product_id, sku=sku, name=name, is_active=True, stock=5, cost=2, unit="pcs")
    return product


def test_receive_import_dry_run_validates_and_maps_product_types() -> None:
    db = FakeReceiveImportSession(
        products=[
            _make_product(1, "SKU-001", "Olive Oil"),
            _make_product(2, "PKG-010", "Glass Jar Lid"),
        ]
    )
    workbook = _make_xlsx(
        [
            ["SKU-001", "Olive Oil", 10, 3.5, "Products", "2026-04-20"],
            ["PKG-010", "Glass Jar Lid", 50, 1.2, "packaging materials", "20/04/2026"],
        ]
    )

    result = _run(
        receive_import_service.import_receive_products(
            db=db,
            workbook_bytes=workbook,
            filename="receive.xlsx",
            current_user=SimpleNamespace(id=1, name="Admin", role="admin"),
            dry_run=True,
        )
    )

    assert result["summary"]["rows_read"] == 2
    assert result["summary"]["receive_records_created"] == 2
    assert result["summary"]["products_resolved"] == 2
    assert result["summary"]["products_rows_count"] == 1
    assert result["summary"]["packaging_rows_count"] == 1
    assert result["errors"] == []


def test_receive_import_real_run_uses_existing_receipt_flow(monkeypatch) -> None:
    db = FakeReceiveImportSession(products=[_make_product(1, "SKU-001", "Olive Oil")])
    created_payloads = []

    async def fake_create_receipt(_db, data, current_user):
        created_payloads.append((data, current_user))
        return {"id": len(created_payloads), "ref_number": f"RCV-{len(created_payloads):05d}"}

    monkeypatch.setattr(receive_import_service, "create_receipt", fake_create_receipt)

    workbook = _make_xlsx(
        [["SKU-001", "Olive Oil", 10, 3.5, "Products", "2026-04-20"]]
    )
    result = _run(
        receive_import_service.import_receive_products(
            db=db,
            workbook_bytes=workbook,
            filename="receive.xlsx",
            current_user=SimpleNamespace(id=2, name="Receiver", role="admin"),
            dry_run=False,
        )
    )

    assert result["summary"]["receive_records_created"] == 1
    assert result["summary"]["stock_moves_created"] == 1
    assert result["batch_id"]
    assert result["revert_available"] is True
    assert len(created_payloads) == 1
    assert created_payloads[0][0].product_id == 1
    assert created_payloads[0][0].product_type == "products"
    assert any(log.action == "receive_import_batch" for log in db.logs)
    assert any(log.action == "receive_import_item" for log in db.logs)


def test_receive_import_reports_missing_sku_and_invalid_type() -> None:
    db = FakeReceiveImportSession(products=[_make_product(1, "SKU-001", "Olive Oil")])
    workbook = _make_xlsx(
        [
            ["", "Olive Oil", 10, 3.5, "Products", "2026-04-20"],
            ["MISSING", "Unknown", 10, 3.5, "Bad Type", "not-a-date"],
        ]
    )

    result = _run(
        receive_import_service.import_receive_products(
            db=db,
            workbook_bytes=workbook,
            filename="receive.xlsx",
            current_user=SimpleNamespace(id=1, name="Admin", role="admin"),
            dry_run=True,
        )
    )

    assert result["summary"]["rows_skipped"] == 2
    assert len(result["errors"]) == 2
    assert "SKU is required" in result["errors"][0]["reason"]
    assert "was not found" in result["errors"][1]["reason"]
    assert "Product Type is required" in result["errors"][1]["reason"]
    assert "valid date" in result["errors"][1]["reason"]


def test_download_receive_products_template_contains_expected_headers() -> None:
    response = _run(import_data.download_receive_products_template(SimpleNamespace(id=1)))
    payload = _run(_read_streaming_response(response))
    workbook = openpyxl.load_workbook(io.BytesIO(payload), data_only=True)
    sheet = workbook.active

    assert [sheet.cell(1, col).value for col in range(1, 7)] == ["SKU", "Product", "QTY", "Unit Price", "Product Type", "Date"]
    assert [sheet.cell(2, col).value for col in range(1, 7)] == ["SKU-001", "Olive Oil 500ml", 24, 12.5, "Products", "2026-04-20"]
    assert [sheet.cell(3, col).value for col in range(1, 7)] == ["PKG-010", "Glass Jar Lid", 200, 1.15, "Packaging Materials", "2026-04-20"]


def test_list_and_revert_receive_import_batches(monkeypatch) -> None:
    batch_id = "receive-batch-1"
    db = FakeReceiveImportSession(
        logs=[
            ActivityLog(
                id=1,
                module="Import",
                action="receive_import_batch",
                description=json.dumps(
                    {
                        "filename": "receive.xlsx",
                        "rows_read": 2,
                        "rows_imported": 2,
                        "rows_skipped": 0,
                        "receive_records_created": 2,
                        "stock_moves_created": 2,
                        "products_rows_count": 1,
                        "packaging_rows_count": 1,
                    }
                ),
                ref_type="receive_import_batch",
                ref_id=batch_id,
            ),
            ActivityLog(
                id=2,
                module="Import",
                action="receive_import_item",
                description=json.dumps({"batch_id": batch_id}),
                ref_type=batch_id,
                ref_id="51",
            ),
            ActivityLog(
                id=3,
                module="Import",
                action="receive_import_item",
                description=json.dumps({"batch_id": batch_id}),
                ref_type=batch_id,
                ref_id="52",
            ),
        ]
    )

    listed = _run(receive_import_service.list_receive_import_batches(db))
    assert listed["batches"][0]["batch_id"] == batch_id
    assert listed["batches"][0]["stock_moves_created"] == 2
    assert listed["batches"][0]["reverted"] is False

    deleted = []

    async def fake_delete_receipt(_db, receipt_id, current_user):
        deleted.append((receipt_id, current_user.id))
        return {"ok": True}

    monkeypatch.setattr(receive_import_service, "delete_receipt", fake_delete_receipt)

    reverted = _run(
        receive_import_service.revert_receive_import_batch(
            db,
            batch_id,
            SimpleNamespace(id=9, name="Admin", role="admin"),
        )
    )

    assert reverted["ok"] is True
    assert reverted["deleted_receipts"] == 2
    assert sorted(deleted) == [(51, 9), (52, 9)]
    assert any(log.action == "receive_import_revert" for log in db.logs)
