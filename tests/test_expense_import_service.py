import asyncio
import io
import json
from types import SimpleNamespace

import openpyxl

from app.core.log import ActivityLog
from app.models.expense import ExpenseCategory
from app.models.farm import Farm
from app.routers import import_data
from app.services import expense_import_service


def _run(coro):
    return asyncio.run(coro)


async def _read_streaming_response(response) -> bytes:
    chunks = []
    async for chunk in response.body_iterator:
        chunks.append(chunk)
    return b"".join(chunks)


def _make_xlsx(rows: list[list[object]], headers: list[str] | None = None) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers or ["Category", "Amount", "Farm", "Date"])
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class _FakeResult:
    def __init__(self, values):
        self._values = values

    def scalar_one_or_none(self):
        return self._values[0] if self._values else None

    def scalars(self):
        return self

    def all(self):
        return list(self._values)


class FakeExpenseImportSession:
    def __init__(self, *, categories=None, farms=None, logs=None):
        self.categories = list(categories or [])
        self.farms = list(farms or [])
        self.logs = list(logs or [])
        self.added = []
        self.commits = 0

    async def execute(self, stmt):
        entity = None
        for desc in getattr(stmt, "column_descriptions", []):
            if desc.get("entity") is not None:
                entity = desc["entity"]
                break
        if entity is ExpenseCategory:
            return _FakeResult(self.categories)
        if entity is Farm:
            return _FakeResult(self.farms)
        if entity is ActivityLog:
            items = list(self.logs)
            where_text = " ".join(str(expr) for expr in getattr(stmt, "_where_criteria", []))
            compiled = stmt.compile()
            params = compiled.params
            if "activity_logs.module" in where_text:
                items = [log for log in items if log.module == params.get("module_1")]
            if "activity_logs.action =" in where_text:
                action = params.get("action_1")
                if action is not None:
                    items = [log for log in items if log.action == action]
            if "activity_logs.action IN" in where_text:
                allowed = set()
                for key, value in params.items():
                    if not key.startswith("action_"):
                        continue
                    if isinstance(value, (list, tuple, set)):
                        allowed.update(value)
                    else:
                        allowed.add(value)
                items = [log for log in items if log.action in allowed]
            if "activity_logs.ref_id =" in where_text:
                ref_id = params.get("ref_id_1")
                if ref_id is not None:
                    items = [log for log in items if log.ref_id == ref_id]
            if "activity_logs.ref_type =" in where_text:
                ref_type = params.get("ref_type_1")
                if ref_type is not None:
                    items = [log for log in items if log.ref_type == ref_type]
            return _FakeResult(items)
        return _FakeResult([])

    def add(self, obj):
        self.added.append(obj)
        if isinstance(obj, ActivityLog):
            obj.id = obj.id or len(self.logs) + 1
            self.logs.append(obj)

    async def commit(self):
        self.commits += 1


def _make_category(category_id: int, name: str, account_code: str = "5001") -> ExpenseCategory:
    category = ExpenseCategory(id=category_id, name=name, account_code=account_code, is_active="1")
    return category


def _make_farm(farm_id: int, name: str) -> Farm:
    farm = Farm(id=farm_id, name=name, is_active=1)
    return farm


def test_expense_import_dry_run_matches_categories_and_general_expense_rows() -> None:
    db = FakeExpenseImportSession(
        categories=[_make_category(1, "Fuel")],
        farms=[_make_farm(10, "North Farm")],
    )
    workbook = _make_xlsx(
        [
            [" fuel ", "150.25", " north farm ", "2026-04-10"],
            ["Office Supplies", 42, "", "11/04/2026"],
        ]
    )

    result = _run(
        expense_import_service.import_expenses(
            db=db,
            workbook_bytes=workbook,
            filename="expenses.xlsx",
            current_user=SimpleNamespace(id=1, name="Admin"),
            dry_run=True,
        )
    )

    assert result["dry_run"] is True
    assert result["summary"]["rows_read"] == 2
    assert result["summary"]["expenses_would_create"] == 2
    assert result["summary"]["categories_auto_created"] == 1
    assert result["summary"]["farms_resolved"] == 1
    assert result["summary"]["general_expense_rows"] == 1
    assert result["summary"]["notes_imported"] == 0
    assert result["errors"] == []
    assert result["auto_created_categories"] == [{"name": "Office Supplies", "account_code": None}]


def test_expense_import_real_run_auto_creates_category_and_uses_null_farm(monkeypatch) -> None:
    db = FakeExpenseImportSession(
        categories=[_make_category(1, "Fuel")],
        farms=[_make_farm(10, "North Farm")],
    )
    created_payloads = []

    async def fake_create_category(_db, data):
        category = _make_category(2, data.name, "5002")
        db.categories.append(category)
        return {"id": category.id, "name": category.name, "account_code": category.account_code}

    async def fake_create_expense_entry(_db, data, current_user):
        created_payloads.append((data, current_user))
        return {"id": len(created_payloads), "ref_number": f"EXP-{len(created_payloads):05d}"}

    monkeypatch.setattr(expense_import_service, "create_category", fake_create_category)
    monkeypatch.setattr(expense_import_service, "create_expense_entry", fake_create_expense_entry)

    workbook = _make_xlsx(
        [
            ["Fuel", 120, "North Farm", "2026-04-10", "Fuel refill"],
            ["Office Supplies", 75, "", "2026-04-11", ""],
        ],
        headers=["Category", "Amount", "Farm", "Date", "Notes"],
    )

    result = _run(
        expense_import_service.import_expenses(
                db=db,
                workbook_bytes=workbook,
                filename="expenses.xlsx",
                current_user=SimpleNamespace(id=7, name="Importer", role="admin"),
                dry_run=False,
            )
        )

    assert result["summary"]["expenses_created"] == 2
    assert result["summary"]["categories_auto_created"] == 1
    assert result["summary"]["notes_imported"] == 1
    assert result["revert_available"] is True
    assert result["batch_id"]
    assert len(created_payloads) == 2
    assert created_payloads[0][0].category_id == 1
    assert created_payloads[0][0].farm_id == 10
    assert created_payloads[0][0].description == "Fuel refill"
    assert created_payloads[1][0].category_id == 2
    assert created_payloads[1][0].farm_id is None
    assert any(isinstance(item, ActivityLog) and item.action == "expense_import_batch" for item in db.logs)
    assert any(isinstance(item, ActivityLog) and item.action == "expense_import_item" for item in db.logs)


def test_expense_import_reports_invalid_rows() -> None:
    db = FakeExpenseImportSession(
        categories=[_make_category(1, "Fuel")],
        farms=[_make_farm(10, "North Farm")],
    )
    workbook = _make_xlsx(
        [
            ["Fuel", "", "North Farm", "2026-04-10"],
            ["Fuel", 20, "Missing Farm", "not-a-date"],
        ]
    )

    result = _run(
        expense_import_service.import_expenses(
            db=db,
            workbook_bytes=workbook,
            filename="expenses.xlsx",
            current_user=SimpleNamespace(id=1, name="Admin"),
            dry_run=True,
        )
    )

    assert result["summary"]["rows_read"] == 2
    assert result["summary"]["rows_skipped"] == 2
    assert len(result["errors"]) == 2
    assert "Amount is required" in result["errors"][0]["reason"]
    assert "Farm 'Missing Farm' was not found" in result["errors"][1]["reason"]
    assert "valid date" in result["errors"][1]["reason"]


def test_download_expenses_template_contains_expected_headers_and_examples() -> None:
    response = _run(import_data.download_expenses_template(SimpleNamespace(id=1)))
    payload = _run(_read_streaming_response(response))
    workbook = openpyxl.load_workbook(io.BytesIO(payload), data_only=True)
    sheet = workbook.active

    assert [sheet.cell(1, col).value for col in range(1, 6)] == ["Category", "Amount", "Farm", "Date", "Notes"]
    assert [sheet.cell(2, col).value for col in range(1, 6)] == ["Fuel", 850.5, "North Farm", "2026-04-10", "Diesel for irrigation pump"]
    assert [sheet.cell(3, col).value for col in range(1, 6)] == ["Office Supplies", 120, None, "2026-04-11", "Admin stationery - should import as General Expense"]


def test_list_and_revert_expense_import_batches(monkeypatch) -> None:
    batch_id = "batch-123"
    db = FakeExpenseImportSession(
        logs=[
            ActivityLog(
                id=1,
                module="Import",
                action="expense_import_batch",
                description=json.dumps(
                    {
                        "filename": "expenses.xlsx",
                        "rows_read": 3,
                        "rows_imported": 2,
                        "rows_skipped": 1,
                        "expense_records_created": 2,
                        "notes_imported": 1,
                        "general_expense_rows": 1,
                    }
                ),
                ref_type="expense_import_batch",
                ref_id=batch_id,
            ),
            ActivityLog(
                id=2,
                module="Import",
                action="expense_import_item",
                description=json.dumps({"batch_id": batch_id}),
                ref_type=batch_id,
                ref_id="41",
            ),
            ActivityLog(
                id=3,
                module="Import",
                action="expense_import_item",
                description=json.dumps({"batch_id": batch_id}),
                ref_type=batch_id,
                ref_id="42",
            ),
        ]
    )

    listed = _run(expense_import_service.list_expense_import_batches(db))
    assert listed["batches"][0]["batch_id"] == batch_id
    assert listed["batches"][0]["notes_imported"] == 1
    assert listed["batches"][0]["reverted"] is False

    deleted = []

    async def fake_delete_expense_entry(_db, expense_id, current_user):
        deleted.append((expense_id, current_user.id))
        return {"ok": True}

    monkeypatch.setattr(expense_import_service, "delete_expense_entry", fake_delete_expense_entry)

    reverted = _run(
        expense_import_service.revert_expense_import_batch(
            db,
            batch_id,
            SimpleNamespace(id=9, name="Admin", role="admin"),
        )
    )
    assert reverted["ok"] is True
    assert reverted["deleted_expenses"] == 2
    assert sorted(deleted) == [(41, 9), (42, 9)]
    assert any(log.action == "expense_import_revert" for log in db.logs)
