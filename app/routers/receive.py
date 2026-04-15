from typing import Optional

from datetime import date
import io

from fastapi import APIRouter, Depends
from fastapi.responses import HTMLResponse, StreamingResponse
from sqlalchemy import or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.permissions import require_action, require_permission
from app.database import get_async_session
from app.models.product import Product
from app.models.user import User
from app.services.receive_service import (
    BatchReceiptCreate,
    ReceiptCreate,
    ReceiptUpdate,
    create_receipt,
    create_receipt_batch,
    delete_receipt,
    list_receipts,
    update_receipt,
)

router = APIRouter(
    prefix="/receive",
    tags=["Receive Products"],
    dependencies=[Depends(require_permission("page_receive_products"))],
)


def to_xlsx(headers, rows, sheet_name="Receive Products"):
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    hfill = PatternFill("solid", fgColor="2a7a2a")
    hfont = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F5FAF5")
    for column in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in column), default=10)
        ws.column_dimensions[column[0].column_letter].width = min(max_len + 4, 40)
    ws.row_dimensions[1].height = 20
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── API ───────────────────────────────────────────────────────────────────────

@router.get("/api/products")
async def get_products(db: AsyncSession = Depends(get_async_session)):
    result = await db.execute(
        select(Product)
        .where(or_(Product.is_active.is_(True), Product.is_active.is_(None)))
        .order_by(Product.name)
    )
    return [
        {
            "id":    p.id,
            "sku":   p.sku,
            "name":  p.name,
            "unit":  p.unit,
            "cost":  float(p.cost)  if p.cost  is not None else 0.0,
            "stock": float(p.stock) if p.stock is not None else 0.0,
        }
        for p in result.scalars().all()
    ]


@router.post("/api/receive-batch", status_code=201)
async def receive_products_batch(
    data: BatchReceiptCreate,
    db: AsyncSession = Depends(get_async_session),
    current_user: User = Depends(require_action("receive_products", "receipts", "create")),
):
    return await create_receipt_batch(db, data, current_user)


# Single-product endpoint kept for backward compatibility / API clients.
@router.post("/api/receive", status_code=201)
async def receive_products(
    data: ReceiptCreate,
    db: AsyncSession = Depends(get_async_session),
    current_user: User = Depends(require_action("receive_products", "receipts", "create")),
):
    return await create_receipt(db, data, current_user)


@router.get("/api/history")
async def get_receipt_history(
    skip:       int           = 0,
    limit:      int           = 50,
    product_id: Optional[int] = None,
    db: AsyncSession = Depends(get_async_session),
):
    return await list_receipts(db, skip=skip, limit=limit, product_id=product_id)


@router.put("/api/receipt/{receipt_id}")
async def edit_receipt(
    receipt_id: int,
    data: ReceiptUpdate,
    db: AsyncSession = Depends(get_async_session),
    current_user: User = Depends(require_action("receive_products", "receipts", "update")),
):
    return await update_receipt(db, receipt_id, data, current_user)


@router.delete("/api/receipt/{receipt_id}")
async def remove_receipt(
    receipt_id: int,
    db: AsyncSession = Depends(get_async_session),
    current_user: User = Depends(require_action("receive_products", "receipts", "delete")),
):
    return await delete_receipt(db, receipt_id, current_user)


@router.get("/api/export.xlsx")
async def export_receipts_excel(
    product_id: Optional[int] = None,
    db: AsyncSession = Depends(get_async_session),
    _: User = Depends(require_action("receive_products", "receipts", "export")),
):
    data = await list_receipts(db, skip=0, limit=10000, product_id=product_id)
    headers = [
        "Receipt #",
        "Receive Date",
        "Product",
        "SKU",
        "Qty",
        "Unit Cost",
        "Total Cost",
        "Expense Ref",
        "Supplier Ref",
        "Notes",
        "Received By",
        "Created At",
    ]
    rows = [
        [
            item["ref_number"],
            item["receive_date"],
            item["product_name"],
            item["product_sku"],
            item["qty"],
            item["unit_cost"],
            item["total_cost"],
            item["expense_ref"],
            item["supplier_ref"],
            item["notes"],
            item["received_by"],
            item["created_at"],
        ]
        for item in data["items"]
    ]
    buf = to_xlsx(headers, rows, "Receive Products")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=receive_products_{date.today()}.xlsx"},
    )


# ── UI ────────────────────────────────────────────────────────────────────────

@router.get("/", response_class=HTMLResponse)
def receive_ui():
    return """<!DOCTYPE html>
<html>
<head>
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Receive Products</title>
<link href="https://fonts.googleapis.com/css2?family=Outfit:wght@400;500;600;700;800;900&family=JetBrains+Mono:wght@400;500;700&display=swap" rel="stylesheet">
<style>
:root{
  --bg:#060810;--surface:#0a0d18;--card:#0f1424;--card2:#151c30;
  --border:rgba(255,255,255,0.06);--border2:rgba(255,255,255,0.11);
  --green:#00ff9d;--blue:#4d9fff;--amber:#ffb547;--danger:#ff4d6d;--purple:#a855f7;
  --text:#f0f4ff;--sub:#8899bb;--muted:#445066;
  --sans:'Outfit',sans-serif;--mono:'JetBrains Mono',monospace;
}
body.light{
  --bg:#f4f5ef;--surface:#f1f3eb;--card:#eceee6;--card2:#e4e6de;
  --border:rgba(0,0,0,0.08);--border2:rgba(0,0,0,0.14);
  --green:#0f8a43;
  --text:#1a1e14;--sub:#4a5040;--muted:#7b816f;
}
body.light nav{background:rgba(244,245,239,.92);}
body.light .picker-list{background:#fff;border-color:rgba(0,0,0,.12);}
body.light .picker-item:hover,.picker-item.active{background:rgba(77,159,255,.1);}
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
body{font-family:var(--sans);background:var(--bg);color:var(--text);min-height:100vh;font-size:14px}

/* ── nav ── */
nav{position:sticky;top:0;z-index:200;display:flex;align-items:center;
  justify-content:space-between;gap:12px;padding:0 24px;height:58px;
  background:rgba(6,8,16,.92);backdrop-filter:blur(20px);border-bottom:1px solid var(--border)}
.nav-left{display:flex;align-items:center;gap:16px}
.nav-logo{font-size:17px;font-weight:900;text-decoration:none;
  background:linear-gradient(135deg,var(--green),var(--blue));
  -webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text}
.nav-title{font-size:14px;font-weight:600;color:var(--sub)}
.nav-right{display:flex;align-items:center;gap:10px}
.mode-btn{display:flex;align-items:center;justify-content:center;width:36px;height:36px;
  border-radius:10px;border:1px solid var(--border);background:var(--card);
  color:var(--sub);font-size:16px;cursor:pointer;transition:all .2s;font-family:var(--sans)}
.mode-btn:hover{border-color:var(--border2);transform:scale(1.06)}
.account-menu{position:relative}
.user-pill{display:flex;align-items:center;gap:10px;background:var(--card);
  border:1px solid var(--border);border-radius:40px;padding:7px 16px 7px 10px;
  cursor:pointer;transition:all .2s}
.user-pill:hover,.user-pill.open{border-color:var(--border2)}
.user-avatar{width:28px;height:28px;background:linear-gradient(135deg,#7ecb6f,#d4a256);
  border-radius:50%;display:flex;align-items:center;justify-content:center;
  font-size:12px;font-weight:700;color:#0a0c08}
.user-name{font-size:13px;font-weight:500;color:var(--sub)}
.menu-caret{font-size:11px;color:var(--muted)}
.account-dropdown{position:absolute;right:0;top:calc(100% + 10px);min-width:220px;
  background:var(--card);border:1px solid var(--border2);border-radius:14px;padding:8px;
  box-shadow:0 24px 50px rgba(0,0,0,.35);display:none;z-index:500}
.account-dropdown.open{display:block}
.account-head{padding:10px 12px 8px;border-bottom:1px solid var(--border);margin-bottom:6px}
.account-label{font-size:11px;color:var(--muted);text-transform:uppercase;letter-spacing:1px}
.account-email{font-size:12px;color:var(--sub);margin-top:4px;word-break:break-word}
.account-item{width:100%;display:flex;align-items:center;gap:10px;padding:10px 12px;border:none;
  background:transparent;border-radius:10px;color:var(--sub);font-family:var(--sans);
  font-size:13px;text-decoration:none;cursor:pointer;text-align:left}
.account-item:hover{background:var(--card2);color:var(--text)}
.account-item.danger:hover{color:var(--danger)}
.logout-btn{background:transparent;border:1px solid var(--border);color:var(--muted);
  font-family:var(--sans);font-size:12px;padding:8px 16px;border-radius:8px;
  cursor:pointer;transition:all .2s}
.logout-btn:hover{border-color:var(--danger);color:var(--danger)}

/* ── layout ── */
.page{max-width:1100px;margin:0 auto;padding:32px 24px 60px}
.page-header{margin-bottom:28px}
.page-header h1{font-size:24px;font-weight:700;margin-bottom:4px}
.page-header p{font-size:13px;color:var(--muted)}
.card{background:var(--card);border:1px solid var(--border);border-radius:16px;padding:24px;margin-bottom:28px}
.section-title{font-size:11px;font-weight:600;letter-spacing:2px;text-transform:uppercase;
  color:var(--muted);margin-bottom:20px;display:flex;align-items:center;gap:12px}
.section-title::after{content:'';flex:1;height:1px;background:linear-gradient(90deg,var(--border2),transparent)}
.section-head{display:flex;align-items:center;justify-content:space-between;gap:12px;margin-bottom:20px}
.section-head .section-title{margin-bottom:0;flex:1}

/* ── meta fields (date / supplier / notes) ── */
.meta-grid{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:20px}
@media(max-width:600px){.meta-grid{grid-template-columns:1fr}}
.field{display:flex;flex-direction:column;gap:6px}
.field.full{grid-column:1/-1}
label{font-size:11px;font-weight:600;color:var(--sub);letter-spacing:.5px;text-transform:uppercase}
input[type=text],input[type=number],input[type=date],textarea{
  background:var(--surface);border:1px solid var(--border);border-radius:10px;
  color:var(--text);font-family:var(--sans);font-size:14px;padding:9px 13px;
  transition:border-color .2s;outline:none;width:100%}
input:focus,textarea:focus{border-color:var(--blue)}
textarea{resize:vertical;min-height:60px}

/* ── product rows table ── */
.rows-wrap{overflow:visible}
.rows-table{width:100%;border-collapse:collapse;margin-bottom:12px}
.rows-table th{
  text-align:left;padding:9px 12px;color:var(--muted);font-size:11px;
  font-weight:600;letter-spacing:1px;text-transform:uppercase;
  border-bottom:1px solid var(--border);white-space:nowrap}
.rows-table td{padding:8px 8px;vertical-align:middle}
.rows-table tr.data-row:last-child td{border-bottom:none}
.rows-table tr.data-row td{border-bottom:1px solid var(--border)}

/* ── searchable picker ── */
.picker{position:relative}
.picker-input{width:100%;min-width:280px}
.picker-list{
  position:absolute;top:calc(100% + 4px);left:0;z-index:300;
  min-width:380px;width:max-content;max-width:520px;
  background:var(--card2);border:1px solid var(--border2);border-radius:10px;
  max-height:280px;overflow-y:auto;display:none;box-shadow:0 8px 32px rgba(0,0,0,.5)}
.picker-list.open{display:block}
.picker-item{padding:10px 16px;cursor:pointer;font-size:13px;border-radius:6px;display:flex;align-items:center;gap:8px}
.picker-item:hover,.picker-item.highlighted{background:rgba(77,159,255,.12);color:var(--blue)}
.picker-item .sku{font-family:var(--mono);font-size:11px;color:var(--muted)}
.picker-item .stock{margin-left:auto;font-family:var(--mono);font-size:11px;color:var(--muted);white-space:nowrap}
.picker-empty{padding:12px 16px;color:var(--muted);font-size:13px}

/* ── row inputs ── */
.qty-input,.cost-input{width:100px;text-align:right;font-family:var(--mono)}
.unit-cell{font-size:12px;color:var(--sub);white-space:nowrap;padding:0 8px}
.row-total{font-family:var(--mono);font-size:13px;color:var(--amber);
  text-align:right;white-space:nowrap;padding:0 8px;min-width:80px}
.remove-btn{background:transparent;border:1px solid var(--border);color:var(--muted);
  width:28px;height:28px;border-radius:8px;cursor:pointer;font-size:15px;
  display:flex;align-items:center;justify-content:center;transition:all .2s}
.remove-btn:hover{border-color:var(--danger);color:var(--danger)}
.remove-btn:disabled{opacity:.3;cursor:default}

/* ── footer row ── */
.add-row-btn{
  display:inline-flex;align-items:center;gap:6px;
  background:transparent;border:1px dashed var(--border2);color:var(--sub);
  font-family:var(--sans);font-size:13px;font-weight:500;
  padding:8px 16px;border-radius:10px;cursor:pointer;transition:all .2s}
.add-row-btn:hover{border-color:var(--blue);color:var(--blue)}
.grand-total{
  display:flex;align-items:center;justify-content:flex-end;gap:12px;
  margin-top:16px;padding-top:14px;border-top:1px solid var(--border)}
.grand-total-label{font-size:12px;font-weight:600;color:var(--muted);text-transform:uppercase;letter-spacing:1px}
.grand-total-value{font-family:var(--mono);font-size:20px;font-weight:700;color:var(--amber)}
.submit-btn{
  width:100%;padding:13px;border-radius:12px;border:none;cursor:pointer;
  background:var(--green);color:#0a0c08;font-family:var(--sans);
  font-size:14px;font-weight:700;letter-spacing:.3px;transition:all .2s;margin-top:16px}
.submit-btn:hover:not(:disabled){filter:brightness(1.1);transform:translateY(-1px)}
.submit-btn:active:not(:disabled){transform:translateY(0)}
.submit-btn:disabled{opacity:.4;cursor:not-allowed}

/* ── history table ── */
.table-wrap{overflow-x:auto}
table.hist{width:100%;border-collapse:collapse;font-size:13px}
table.hist th{text-align:left;padding:9px 13px;color:var(--muted);font-size:11px;
  font-weight:600;letter-spacing:1px;text-transform:uppercase;
  border-bottom:1px solid var(--border);white-space:nowrap}
table.hist td{padding:10px 13px;border-bottom:1px solid var(--border);vertical-align:middle}
table.hist tr:last-child td{border-bottom:none}
table.hist tr:hover td{background:rgba(255,255,255,.025)}
body.light table.hist tr:hover td{background:rgba(0,0,0,.03)}
.badge{display:inline-block;padding:3px 9px;border-radius:6px;
  font-family:var(--mono);font-size:11px;font-weight:600;
  background:rgba(77,159,255,.12);color:var(--blue);border:1px solid rgba(77,159,255,.2)}
.badge-exp{background:rgba(0,255,157,.1);color:var(--green);border-color:rgba(0,255,157,.2)}
.badge-none{background:rgba(68,80,102,.2);color:var(--muted);border-color:transparent}
.empty-row{text-align:center;padding:40px;color:var(--muted)}
.history-actions{display:flex;gap:8px;flex-wrap:wrap}
.action-btn{background:transparent;border:1px solid var(--border2);color:var(--sub);padding:6px 10px;border-radius:8px;cursor:pointer;font-size:12px;font-family:var(--sans);transition:all .2s}
.action-btn:hover{border-color:var(--blue);color:var(--blue)}
.action-btn.danger:hover{border-color:var(--danger);color:var(--danger)}
.action-btn.export{border-color:rgba(0,255,157,.25);color:var(--green)}
.action-btn.export:hover{border-color:var(--green)}

/* modal */
.modal-wrap{position:fixed;inset:0;background:rgba(4,7,14,.72);backdrop-filter:blur(8px);display:none;align-items:center;justify-content:center;padding:24px;z-index:999}
.modal-wrap.open{display:flex}
.modal-card{width:min(560px,100%);background:var(--card);border:1px solid var(--border2);border-radius:18px;padding:24px;box-shadow:0 24px 50px rgba(0,0,0,.45)}
.modal-title{font-size:18px;font-weight:800;margin-bottom:6px}
.modal-sub{font-size:13px;color:var(--muted);margin-bottom:18px}
.modal-actions{display:flex;justify-content:flex-end;gap:10px;margin-top:18px}
.modal-btn{border:none;border-radius:10px;padding:10px 16px;font-family:var(--sans);font-size:13px;font-weight:700;cursor:pointer}
.modal-btn.secondary{background:var(--card2);color:var(--sub);border:1px solid var(--border2)}
.modal-btn.primary{background:var(--green);color:#0a0c08}

/* ── toast ── */
.toast{position:fixed;bottom:24px;right:24px;z-index:9999;
  padding:12px 20px;border-radius:12px;font-size:13px;font-weight:500;
  box-shadow:0 8px 32px rgba(0,0,0,.4);transition:opacity .3s;opacity:0;pointer-events:none}
.toast.show{opacity:1;pointer-events:auto}
.toast.ok{background:#0f2918;border:1px solid var(--green);color:var(--green)}
.toast.err{background:#240f14;border:1px solid var(--danger);color:var(--danger)}
</style>
    <script src="/static/auth-guard.js"></script>
</head>
<body>

<nav>
  <div class="nav-left">
    <a href="/home" class="nav-logo">&#9889; Thunder ERP</a>
    <span class="nav-title">Receive Products</span>
  </div>
  <div class="nav-right">
    <button class="mode-btn" id="mode-btn" onclick="toggleMode()">&#127769;</button>
    <div class="account-menu">
      <button class="user-pill" id="account-trigger" onclick="toggleAccountMenu(event)" aria-haspopup="menu" aria-expanded="false">
        <div class="user-avatar" id="user-avatar">A</div>
        <span class="user-name" id="user-name">...</span>
        <span class="menu-caret">&#9662;</span>
      </button>
      <div class="account-dropdown" id="account-dropdown" role="menu">
        <div class="account-head">
          <div class="account-label">Signed in as</div>
          <div class="account-email" id="user-email">&mdash;</div>
        </div>
        <a href="/users/password" class="account-item" role="menuitem">Change Password</a>
        <button class="account-item danger" onclick="logout()" role="menuitem">Sign out</button>
      </div>
    </div>
  </div>
</nav>

<div class="page">

  <div class="page-header">
    <h1>&#128507; Receive Products</h1>
    <p>Add one or more products, enter quantities and costs, then submit. Costs are posted automatically as expenses.</p>
  </div>

  <!-- ── Receive form ── -->
  <div class="card" id="receive-form-card">
    <div class="section-title">New Receipt</div>

    <form id="receive-form" onsubmit="submitBatch(event)">

      <!-- batch-level fields -->
      <div class="meta-grid">
        <div class="field">
          <label>Receive Date *</label>
          <input type="date" id="receive-date" required>
        </div>
        <div class="field">
          <label>Supplier / Reference <span style="color:var(--muted);font-weight:400">(optional)</span></label>
          <input type="text" id="supplier-ref" maxlength="150" placeholder="e.g. Acme Supplies / INV-2026-001">
        </div>
        <div class="field full">
          <label>Notes <span style="color:var(--muted);font-weight:400">(optional)</span></label>
          <textarea id="notes" maxlength="500" placeholder="Any additional details…" rows="2"></textarea>
        </div>
      </div>

      <!-- product rows -->
      <div class="rows-wrap">
        <table class="rows-table">
          <thead>
            <tr>
              <th style="min-width:320px">Product</th>
              <th>Qty</th>
              <th></th><!-- unit label -->
              <th>Unit Cost</th>
              <th style="text-align:right">Total</th>
              <th></th><!-- remove -->
            </tr>
          </thead>
          <tbody id="rows-body"></tbody>
        </table>
      </div>

      <button type="button" class="add-row-btn" id="add-row-btn" onclick="addRow()">&#43; Add Product</button>

      <div class="grand-total">
        <span class="grand-total-label">Grand Total</span>
        <span class="grand-total-value" id="grand-total">—</span>
      </div>

      <button type="submit" class="submit-btn" id="submit-btn" disabled>
        &#10003; Receive Stock
      </button>

    </form>
  </div>

  <!-- ── History ── -->
  <div class="card">
    <div class="section-head">
      <div class="section-title">Receipt History</div>
      <button type="button" class="action-btn export" id="export-btn" onclick="exportReceipts()" style="display:none">Export Excel</button>
    </div>
    <div class="table-wrap">
      <table class="hist">
        <thead>
          <tr>
            <th>Receipt #</th><th>Date</th><th>Product</th>
            <th>Qty</th><th>Unit Cost</th><th>Total</th>
            <th>Expense</th><th>Supplier Ref</th><th>Notes</th><th>By</th><th>Actions</th>
          </tr>
        </thead>
        <tbody id="history-body">
          <tr><td colspan="10" class="empty-row">Loading…</td></tr>
        </tbody>
      </table>
    </div>
  </div>

</div>

<div class="modal-wrap" id="edit-modal">
  <div class="modal-card">
    <div class="modal-title">Edit Receipt</div>
    <div class="modal-sub" id="edit-modal-sub">Update this received stock entry.</div>
    <div class="meta-grid">
      <div class="field">
        <label>Product</label>
        <input type="text" id="edit-product" readonly>
      </div>
      <div class="field">
        <label>Receive Date *</label>
        <input type="date" id="edit-date" required>
      </div>
      <div class="field">
        <label>Quantity *</label>
        <input type="number" id="edit-qty" min="0.001" step="0.001" required>
      </div>
      <div class="field">
        <label>Unit Cost</label>
        <input type="number" id="edit-cost" min="0" step="0.01">
      </div>
      <div class="field full">
        <label>Supplier / Reference</label>
        <input type="text" id="edit-supplier" maxlength="150">
      </div>
      <div class="field full">
        <label>Notes</label>
        <textarea id="edit-notes" maxlength="500" rows="3"></textarea>
      </div>
    </div>
    <div class="modal-actions">
      <button type="button" class="modal-btn secondary" onclick="closeEditModal()">Cancel</button>
      <button type="button" class="modal-btn primary" id="edit-save-btn" onclick="saveEditReceipt()">Save Changes</button>
    </div>
  </div>
</div>

<div class="toast" id="toast"></div>

<script>
// ── Auth guard ──────────────────────────────────────────────────────────────
if (!document.cookie.split(';').some(c => c.trim().startsWith('logged_in='))) {
  _redirectToLogin();
}

// ── State ───────────────────────────────────────────────────────────────────
let _products  = [];   // [{id, sku, name, unit, cost, stock}, …]
let _rowSeq    = 0;    // monotonic counter for unique row IDs
let _currentUserRole = '';
let _currentUserPermissions = new Set();
let _historyItems = [];
let _editingReceipt = null;

function hasPermission(permission) {
  return _currentUserRole === 'admin' || _currentUserPermissions.has(permission);
}

function applyReceivePermissions() {
  const canCreate = hasPermission('action_receive_products_create');
  const canExport = hasPermission('action_receive_products_export');
  const formCard = document.getElementById('receive-form-card');
  const addRowBtn = document.getElementById('add-row-btn');
  const submitBtn = document.getElementById('submit-btn');
  const exportBtn = document.getElementById('export-btn');
  if (formCard) formCard.style.display = canCreate ? '' : 'none';
  if (addRowBtn) addRowBtn.style.display = canCreate ? '' : 'none';
  if (submitBtn) submitBtn.style.display = canCreate ? '' : 'none';
  if (exportBtn) exportBtn.style.display = canExport ? '' : 'none';
}

// ── Bootstrap ───────────────────────────────────────────────────────────────
async function init() {
  if (localStorage.getItem('colorMode') === 'light') {
    document.body.classList.add('light');
    document.getElementById('mode-btn').innerHTML = '&#9728;&#65039;';
  }
  await Promise.all([initUser(), loadProducts()]);
  document.getElementById('receive-date').value = todayIso();
  addRow();          // start with one empty row
  await loadHistory();
}

async function initUser() {
  try {
    const r = await fetch('/auth/me');
    if (!r.ok) { _redirectToLogin(); return; }
    const u = await r.json();
    _currentUserRole = u.role || '';
    _currentUserPermissions = new Set(
      (typeof u.permissions === 'string' ? u.permissions.split(',') : (u.permissions || []))
        .map(v => v.trim())
        .filter(Boolean)
    );
    document.getElementById('user-name').innerText   = u.name;
    document.getElementById('user-avatar').innerText = u.name.charAt(0).toUpperCase();
    const emailEl = document.getElementById('user-email');
    if (emailEl) emailEl.innerText = u.email;
    applyReceivePermissions();
  } catch { _redirectToLogin(); }
}

function toggleAccountMenu(event){
  event.stopPropagation();
  const trigger = document.getElementById('account-trigger');
  const dropdown = document.getElementById('account-dropdown');
  const open = dropdown.classList.toggle('open');
  trigger.classList.toggle('open', open);
  trigger.setAttribute('aria-expanded', open ? 'true' : 'false');
}

document.addEventListener('click', e => {
  const menu = document.getElementById('account-dropdown');
  const trigger = document.getElementById('account-trigger');
  if(!menu || !trigger) return;
  if(menu.contains(e.target) || trigger.contains(e.target)) return;
  menu.classList.remove('open');
  trigger.classList.remove('open');
  trigger.setAttribute('aria-expanded', 'false');
});

async function loadProducts() {
  const r = await fetch('/receive/api/products');
  if (!r.ok) return;
  _products = await r.json();
}

// ── Row management ──────────────────────────────────────────────────────────
function addRow() {
  if (!hasPermission('action_receive_products_create')) return;
  const id  = ++_rowSeq;
  const tr  = document.createElement('tr');
  tr.className = 'data-row';
  tr.dataset.row = id;
  tr.innerHTML = `
    <td>
      <div class="picker" id="picker-${id}">
        <input type="text" class="picker-input" id="psearch-${id}"
               autocomplete="off" placeholder="Search product…"
               oninput="onPickerInput(${id})"
               onfocus="onPickerFocus(${id})"
               onkeydown="onPickerKey(event,${id})">
        <input type="hidden" id="pid-${id}">
        <div class="picker-list" id="plist-${id}"></div>
      </div>
    </td>
    <td><input type="number" class="qty-input" id="qty-${id}"
               min="0.001" step="0.001" placeholder="0.000"
               oninput="recalcRow(${id})"></td>
    <td class="unit-cell" id="unit-${id}">—</td>
    <td><input type="number" class="cost-input" id="cost-${id}"
               min="0" step="0.01" placeholder="0.00"
               oninput="recalcRow(${id})"></td>
    <td class="row-total" id="total-${id}">—</td>
    <td>
      <button type="button" class="remove-btn" id="rem-${id}"
              onclick="removeRow(${id})" title="Remove row">&#215;</button>
    </td>`;
  document.getElementById('rows-body').appendChild(tr);
  refreshRemoveButtons();
  document.getElementById(`psearch-${id}`).focus();
}

function removeRow(id) {
  const tr = document.querySelector(`tr[data-row="${id}"]`);
  if (tr) tr.remove();
  refreshRemoveButtons();
  updateGrandTotal();
  validateSubmit();
}

function refreshRemoveButtons() {
  const rows = document.querySelectorAll('#rows-body tr.data-row');
  rows.forEach(r => {
    const btn = r.querySelector('.remove-btn');
    if (btn) btn.disabled = rows.length === 1;
  });
}

// ── Picker logic ────────────────────────────────────────────────────────────
function onPickerInput(id) {
  const query = document.getElementById(`psearch-${id}`).value.trim().toLowerCase();
  renderPickerList(id, query);
  // Clear selection if text was changed
  document.getElementById(`pid-${id}`).value = '';
  recalcRow(id);
  validateSubmit();
}

function onPickerFocus(id) {
  const query = document.getElementById(`psearch-${id}`).value.trim().toLowerCase();
  renderPickerList(id, query);
}

function renderPickerList(id, query) {
  const list = document.getElementById(`plist-${id}`);
  const hits  = query
    ? _products.filter(p =>
        p.name.toLowerCase().includes(query) ||
        p.sku.toLowerCase().includes(query))
    : _products;

  if (hits.length === 0) {
    list.innerHTML = `<div class="picker-empty">No products found</div>`;
  } else {
    list.innerHTML = hits.map(p => `
      <div class="picker-item" data-id="${p.id}"
           onmousedown="selectProduct(event,${id},${p.id},'${esc(p.name)}','${esc(p.sku)}','${esc(p.unit)}',${p.cost},${p.stock})">
        <span>${esc(p.name)}</span>
        <span class="sku">${esc(p.sku)}</span>
        <span class="stock">${p.stock.toFixed(3)}&thinsp;${esc(p.unit)}</span>
      </div>`).join('');
  }
  list.classList.add('open');
}

function selectProduct(e, rowId, productId, name, sku, unit, cost, stock) {
  e.preventDefault();
  document.getElementById(`psearch-${rowId}`).value  = `${name}`;
  document.getElementById(`pid-${rowId}`).value      = productId;
  document.getElementById(`unit-${rowId}`).textContent = unit;
  document.getElementById(`plist-${rowId}`).classList.remove('open');

  // Pre-fill cost with last known cost if field is empty or zero
  const costInput = document.getElementById(`cost-${rowId}`);
  if ((!costInput.value || parseFloat(costInput.value) === 0) && cost > 0) {
    costInput.value = cost.toFixed(2);
  }

  recalcRow(rowId);
  validateSubmit();
  // Move focus to qty
  const qtyInput = document.getElementById(`qty-${rowId}`);
  if (!qtyInput.value) qtyInput.focus();
}

// Close picker when clicking outside
document.addEventListener('click', e => {
  document.querySelectorAll('.picker-list.open').forEach(list => {
    if (!list.closest('.picker').contains(e.target)) {
      list.classList.remove('open');
    }
  });
});

// Basic keyboard navigation in picker
function onPickerKey(e, id) {
  const list  = document.getElementById(`plist-${id}`);
  const items = list.querySelectorAll('.picker-item');
  if (!items.length) return;

  let cur = Array.from(items).findIndex(i => i.classList.contains('highlighted'));

  if (e.key === 'ArrowDown') {
    e.preventDefault();
    cur = Math.min(cur + 1, items.length - 1);
    items.forEach((item, i) => item.classList.toggle('highlighted', i === cur));
    items[cur]?.scrollIntoView({block:'nearest'});
  } else if (e.key === 'ArrowUp') {
    e.preventDefault();
    cur = Math.max(cur - 1, 0);
    items.forEach((item, i) => item.classList.toggle('highlighted', i === cur));
    items[cur]?.scrollIntoView({block:'nearest'});
  } else if (e.key === 'Enter') {
    e.preventDefault();
    if (cur >= 0) items[cur]?.dispatchEvent(new MouseEvent('mousedown', {bubbles:true}));
  } else if (e.key === 'Escape') {
    list.classList.remove('open');
  }
}

// ── Totals ──────────────────────────────────────────────────────────────────
function recalcRow(id) {
  const qty  = parseFloat(document.getElementById(`qty-${id}`)?.value)  || 0;
  const cost = parseFloat(document.getElementById(`cost-${id}`)?.value) || 0;
  const el   = document.getElementById(`total-${id}`);
  if (!el) return;
  if (qty > 0 && cost > 0) {
    el.textContent = (qty * cost).toFixed(2);
    el.style.color = 'var(--amber)';
  } else {
    el.textContent = '—';
    el.style.color = 'var(--muted)';
  }
  updateGrandTotal();
  validateSubmit();
}

function updateGrandTotal() {
  let sum = 0;
  document.querySelectorAll('#rows-body tr.data-row').forEach(tr => {
    const id    = tr.dataset.row;
    const qty   = parseFloat(document.getElementById(`qty-${id}`)?.value)  || 0;
    const cost  = parseFloat(document.getElementById(`cost-${id}`)?.value) || 0;
    if (qty > 0 && cost > 0) sum += qty * cost;
  });
  const el = document.getElementById('grand-total');
  if (sum > 0) { el.textContent = sum.toFixed(2); el.style.color = 'var(--amber)'; }
  else         { el.textContent = '—';             el.style.color = 'var(--muted)'; }
}

function validateSubmit() {
  const rows = document.querySelectorAll('#rows-body tr.data-row');
  const valid = Array.from(rows).some(tr => {
    const id  = tr.dataset.row;
    const pid = document.getElementById(`pid-${id}`)?.value;
    const qty = parseFloat(document.getElementById(`qty-${id}`)?.value) || 0;
    return pid && qty > 0;
  });
  document.getElementById('submit-btn').disabled = !valid;
}

// ── Submit ──────────────────────────────────────────────────────────────────
async function submitBatch(e) {
  e.preventDefault();
  if (!hasPermission('action_receive_products_create')) {
    showToast('Permission denied: action_receive_products_create', 'err');
    return;
  }
  const btn = document.getElementById('submit-btn');
  btn.disabled = true;
  btn.textContent = 'Receiving…';

  const rows    = document.querySelectorAll('#rows-body tr.data-row');
  const items   = [];
  const missing = [];

  rows.forEach(tr => {
    const id   = tr.dataset.row;
    const pid  = document.getElementById(`pid-${id}`)?.value;
    const qty  = parseFloat(document.getElementById(`qty-${id}`)?.value);
    const cost = document.getElementById(`cost-${id}`)?.value.trim();
    if (!pid || !qty || qty <= 0) return;   // skip incomplete rows
    const item = { product_id: parseInt(pid), qty };
    if (cost) item.unit_cost = parseFloat(cost);
    items.push(item);
    missing.push(!pid);
  });

  if (items.length === 0) {
    showToast('Add at least one product with a quantity.', 'err');
    btn.disabled = false; btn.textContent = '✓ Receive Stock';
    return;
  }

  const payload = {
    receive_date: document.getElementById('receive-date').value,
    supplier_ref: document.getElementById('supplier-ref').value.trim() || null,
    notes:        document.getElementById('notes').value.trim() || null,
    items,
  };

  try {
    const r = await fetch('/receive/api/receive-batch', {
      method:  'POST',
      headers: {'Content-Type': 'application/json'},
      body:    JSON.stringify(payload),
    });
    if (!r.ok) {
      const err = await r.json().catch(() => ({}));
      showToast(err.detail || 'Receive failed', 'err');
    } else {
      const data = await r.json();
      const expCount = data.receipts.filter(r => r.expense_ref).length;
      const msg = `${data.count} product${data.count > 1 ? 's' : ''} received`
        + (data.total_cost ? ` · Total ${data.total_cost.toFixed(2)}` : '')
        + (expCount        ? ` · ${expCount} expense${expCount > 1 ? 's' : ''} posted` : '');
      showToast(msg, 'ok');
      resetForm();
      await loadProducts();
      await loadHistory();
    }
  } catch { showToast('Network error', 'err'); }

  btn.disabled = false; btn.textContent = '✓ Receive Stock';
}

function resetForm() {
  document.getElementById('supplier-ref').value = '';
  document.getElementById('notes').value = '';
  document.getElementById('receive-date').value = todayIso();
  document.getElementById('rows-body').innerHTML = '';
  _rowSeq = 0;
  addRow();
  updateGrandTotal();
}

// ── History ─────────────────────────────────────────────────────────────────
async function loadHistory() {
  const r     = await fetch('/receive/api/history?limit=100');
  const tbody = document.getElementById('history-body');
  if (!r.ok) { tbody.innerHTML = `<tr><td colspan="11" class="empty-row">Could not load.</td></tr>`; return; }
  const data  = await r.json();
  _historyItems = data.items || [];
  const canUpdate = hasPermission('action_receive_products_update');
  const canDelete = hasPermission('action_receive_products_delete');
  const canManage = canUpdate || canDelete;
  if (!_historyItems.length) {
    tbody.innerHTML = `<tr><td colspan="11" class="empty-row">No receipts yet.</td></tr>`;
    return;
  }
  tbody.innerHTML = _historyItems.map(row => `<tr>
    <td><span class="badge">${esc(row.ref_number)}</span></td>
    <td>${esc(row.receive_date||'')}</td>
    <td>
      <div style="font-weight:600">${esc(row.product_name||'')}</div>
      <div style="font-family:var(--mono);font-size:11px;color:var(--muted)">${esc(row.product_sku||'')}</div>
    </td>
    <td style="font-family:var(--mono)">${parseFloat(row.qty).toFixed(3)}</td>
    <td style="font-family:var(--mono)">${row.unit_cost!=null ? parseFloat(row.unit_cost).toFixed(2) : '<span style="color:var(--muted)">—</span>'}</td>
    <td style="font-family:var(--mono);color:var(--amber)">${row.total_cost!=null ? parseFloat(row.total_cost).toFixed(2) : '<span style="color:var(--muted)">—</span>'}</td>
    <td>${row.expense_ref ? `<span class="badge badge-exp">${esc(row.expense_ref)}</span>` : '<span class="badge badge-none">—</span>'}</td>
    <td style="color:var(--sub)">${esc(row.supplier_ref||'—')}</td>
    <td style="color:var(--sub);max-width:160px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap"
        title="${esc(row.notes||'')}">${esc(row.notes||'—')}</td>
    <td style="color:var(--muted)">${esc(row.received_by||'—')}</td>
    <td>${canManage ? `<div class="history-actions">
      ${canUpdate ? `<button type="button" class="action-btn" onclick="openEditModal(${row.id})">Edit</button>` : ''}
      ${canDelete ? `<button type="button" class="action-btn danger" onclick="deleteReceipt(${row.id})">Delete</button>` : ''}
    </div>` : '<span style="color:var(--muted)">-</span>'}</td>
  </tr>`).join('');
}

// ── Utils ───────────────────────────────────────────────────────────────────
function openEditModal(receiptId) {
  if (!hasPermission('action_receive_products_update')) return;
  const receipt = _historyItems.find(item => item.id === receiptId);
  if (!receipt) {
    showToast('Receipt not found', 'err');
    return;
  }
  _editingReceipt = receipt;
  document.getElementById('edit-modal-sub').textContent = `Editing ${receipt.ref_number}`;
  document.getElementById('edit-product').value = receipt.product_name || '';
  document.getElementById('edit-date').value = receipt.receive_date || todayIso();
  document.getElementById('edit-qty').value = parseFloat(receipt.qty || 0).toFixed(3);
  document.getElementById('edit-cost').value = receipt.unit_cost != null ? parseFloat(receipt.unit_cost).toFixed(2) : '';
  document.getElementById('edit-supplier').value = receipt.supplier_ref || '';
  document.getElementById('edit-notes').value = receipt.notes || '';
  document.getElementById('edit-modal').classList.add('open');
}

function closeEditModal() {
  _editingReceipt = null;
  document.getElementById('edit-modal').classList.remove('open');
}

async function saveEditReceipt() {
  if (!_editingReceipt || !hasPermission('action_receive_products_update')) return;
  const btn = document.getElementById('edit-save-btn');
  btn.disabled = true;
  btn.textContent = 'Saving...';
  const payload = {
    receive_date: document.getElementById('edit-date').value,
    qty: parseFloat(document.getElementById('edit-qty').value),
    supplier_ref: document.getElementById('edit-supplier').value.trim() || null,
    notes: document.getElementById('edit-notes').value.trim() || null,
  };
  const costValue = document.getElementById('edit-cost').value.trim();
  payload.unit_cost = costValue === '' ? null : parseFloat(costValue);

  try {
    const r = await fetch(`/receive/api/receipt/${_editingReceipt.id}`, {
      method: 'PUT',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify(payload),
    });
    const data = await r.json().catch(() => ({}));
    if (!r.ok) {
      showToast(data.detail || 'Could not update receipt', 'err');
      return;
    }
    closeEditModal();
    showToast(`Receipt ${data.ref_number} updated`, 'ok');
    await loadProducts();
    await loadHistory();
  } catch {
    showToast('Network error', 'err');
  } finally {
    btn.disabled = false;
    btn.textContent = 'Save Changes';
  }
}

async function deleteReceipt(receiptId) {
  if (!hasPermission('action_receive_products_delete')) return;
  const receipt = _historyItems.find(item => item.id === receiptId);
  if (!receipt) {
    showToast('Receipt not found', 'err');
    return;
  }
  if (!confirm(`Delete receipt ${receipt.ref_number}? This will reverse its stock receipt.`)) return;
  try {
    const r = await fetch(`/receive/api/receipt/${receiptId}`, {method: 'DELETE'});
    const data = await r.json().catch(() => ({}));
    if (!r.ok) {
      showToast(data.detail || 'Could not delete receipt', 'err');
      return;
    }
    if (_editingReceipt && _editingReceipt.id === receiptId) closeEditModal();
    showToast(`Receipt ${receipt.ref_number} deleted`, 'ok');
    await loadProducts();
    await loadHistory();
  } catch {
    showToast('Network error', 'err');
  }
}

function exportReceipts() {
  if (!hasPermission('action_receive_products_export')) return;
  showToast('Preparing Excel export...', 'ok');
  window.location.href = '/receive/api/export.xlsx';
}

function todayIso() { return new Date().toISOString().slice(0, 10); }

function esc(s) {
  return String(s)
    .replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;')
    .replace(/"/g,'&quot;').replace(/'/g,'&#039;');
}

function showToast(msg, type) {
  const t = document.getElementById('toast');
  t.textContent = msg; t.className = `toast ${type} show`;
  setTimeout(() => t.className = 'toast', 3800);
}

function toggleMode() {
  const light = document.body.classList.toggle('light');
  document.getElementById('mode-btn').innerHTML = light ? '&#9728;&#65039;' : '&#127769;';
  localStorage.setItem('colorMode', light ? 'light' : 'dark');
}

async function logout() {
  await fetch('/auth/logout', {method:'POST'});
  window.location.href = '/';
}

init();
</script>
</body>
</html>"""
