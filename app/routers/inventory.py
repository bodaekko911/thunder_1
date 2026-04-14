from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import HTMLResponse, StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import func, or_, select
from sqlalchemy.orm import selectinload
from typing import Optional
from pydantic import BaseModel
from datetime import datetime, date
import io

from app.database import get_async_session
from app.core.permissions import get_current_user, require_permission
from app.core.log import record
from app.models.product import Product
from app.models.user import User
from app.models.inventory import LocationStock, StockLocation, StockMove, StockTransfer
from app.services.location_inventory_service import (
    create_stock_transfer,
    get_or_create_location_stock,
    quantize_qty,
    serialize_location,
    serialize_product_location_stock,
    serialize_transfer,
)
from app.services.replenishment_service import (
    create_or_reuse_draft_purchases,
    is_low_stock,
    serialize_low_stock_product,
    suggested_reorder_qty,
)

router = APIRouter(
    prefix="/inventory",
    tags=["Inventory"],
    dependencies=[Depends(require_permission("page_inventory"))],
)


def to_xlsx(headers, rows, sheet_name="Report"):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        hfill  = PatternFill("solid", fgColor="2a7a2a")
        hfont  = Font(bold=True, color="FFFFFF", size=11)
        thin   = Side(style="thin", color="DDDDDD")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = hfill; c.font = hfont
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border
        for ri, row in enumerate(rows, 2):
            for ci, val in enumerate(row, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.border = border
                c.alignment = Alignment(vertical="center")
                if ri % 2 == 0:
                    c.fill = PatternFill("solid", fgColor="F5FAF5")
        for col in ws.columns:
            mx = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(mx + 4, 40)
        ws.row_dimensions[1].height = 20
        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        return buf
    except ImportError:
        raise Exception("Run: pip install openpyxl --break-system-packages")


# ── Schemas ────────────────────────────────────────────
class StockAdjustment(BaseModel):
    product_id: int
    qty:        float
    location_id: Optional[int] = None
    note:       Optional[str] = None


class LowStockDraftRequest(BaseModel):
    product_ids: list[int]


class LocationCreate(BaseModel):
    name: str
    code: Optional[str] = None
    location_type: str = "warehouse"
    is_active: bool = True


class StockTransferCreate(BaseModel):
    source_location_id: int
    destination_location_id: int
    product_id: int
    qty: float
    note: Optional[str] = None


# ── API ────────────────────────────────────────────────
@router.get("/api/stock")
async def get_stock(
    q:         str  = "",
    low_stock: bool = False,
    skip:      int  = 0,
    limit:     int  = 50,
    db: AsyncSession = Depends(get_async_session),
):
    low_stock_threshold = func.coalesce(Product.reorder_level, Product.min_stock)
    stmt = select(Product).where(or_(Product.is_active.is_(True), Product.is_active.is_(None)))
    if q:
        stmt = stmt.where(
            Product.name.ilike(f"%{q}%") | Product.sku.ilike(f"%{q}%")
        )
    if low_stock:
        stmt = stmt.where(Product.stock <= low_stock_threshold)
    cnt_result = await db.execute(select(func.count()).select_from(stmt.subquery()))
    total = cnt_result.scalar()
    result = await db.execute(stmt.order_by(Product.name).offset(skip).limit(limit))
    items = result.scalars().all()
    return {
        "total": total,
        "items": [
            {
                "id":        p.id,
                "sku":       p.sku,
                "name":      p.name,
                "stock":     float(p.stock or 0),
                "min_stock": float(p.min_stock or 0),
                "reorder_level": float(p.reorder_level) if p.reorder_level is not None else None,
                "reorder_qty": float(p.reorder_qty) if p.reorder_qty is not None else None,
                "preferred_supplier_id": p.preferred_supplier_id,
                "suggested_reorder_qty": float(suggested_reorder_qty(p)),
                "unit":      p.unit,
                "low":       is_low_stock(p),
                "alert_state": "low_stock" if is_low_stock(p) else "ok",
            }
            for p in items
        ],
    }


@router.get("/api/low-stock")
async def get_low_stock_products(
    q: str = "",
    supplier_id: int | None = None,
    skip: int = 0,
    limit: int = 50,
    db: AsyncSession = Depends(get_async_session),
):
    low_stock_threshold = func.coalesce(Product.reorder_level, Product.min_stock)
    stmt = (
        select(Product)
        .options(selectinload(Product.preferred_supplier))
        .where(or_(Product.is_active.is_(True), Product.is_active.is_(None)), Product.stock <= low_stock_threshold)
    )
    if q:
        stmt = stmt.where(Product.name.ilike(f"%{q}%") | Product.sku.ilike(f"%{q}%"))
    if supplier_id is not None:
        stmt = stmt.where(Product.preferred_supplier_id == supplier_id)

    count_result = await db.execute(select(func.count()).select_from(stmt.subquery()))
    total = count_result.scalar() or 0
    result = await db.execute(stmt.order_by(Product.name).offset(skip).limit(limit))
    products = result.scalars().all()
    return {
        "total": total,
        "items": [serialize_low_stock_product(product) for product in products],
    }


@router.post("/api/low-stock/draft-purchases")
async def create_low_stock_draft_purchases(
    data: LowStockDraftRequest,
    db: AsyncSession = Depends(get_async_session),
    current_user: User = Depends(get_current_user),
):
    product_ids = sorted(set(data.product_ids))
    if not product_ids:
        raise HTTPException(status_code=400, detail="Select at least one low-stock product")

    result = await db.execute(
        select(Product)
        .options(selectinload(Product.preferred_supplier))
        .where(or_(Product.is_active.is_(True), Product.is_active.is_(None)), Product.id.in_(product_ids))
    )
    products = result.scalars().all()
    found_ids = {product.id for product in products}
    missing_ids = [product_id for product_id in product_ids if product_id not in found_ids]
    if missing_ids:
        raise HTTPException(status_code=404, detail=f"Product IDs not found: {', '.join(str(product_id) for product_id in missing_ids)}")

    not_low_stock = [product.name for product in products if not is_low_stock(product)]
    if not_low_stock:
        raise HTTPException(status_code=400, detail=f"Selected products are not low stock: {', '.join(not_low_stock)}")

    missing_supplier = [product.name for product in products if product.preferred_supplier_id is None]
    if missing_supplier:
        raise HTTPException(status_code=400, detail=f"Preferred supplier is required for: {', '.join(missing_supplier)}")

    zero_suggestion = [product.name for product in products if suggested_reorder_qty(product) <= 0]
    if zero_suggestion:
        raise HTTPException(status_code=400, detail=f"Reorder suggestion is zero for: {', '.join(zero_suggestion)}")

    purchases = await create_or_reuse_draft_purchases(
        db,
        products=products,
        user_id=current_user.id,
    )
    await db.commit()
    return {
        "ok": True,
        "created": sum(1 for purchase in purchases if not purchase["reused"]),
        "reused": sum(1 for purchase in purchases if purchase["reused"]),
        "purchases": purchases,
    }


@router.get("/api/locations")
async def get_locations(
    q: str = "",
    include_inactive: bool = False,
    db: AsyncSession = Depends(get_async_session),
):
    stmt = select(StockLocation)
    if not include_inactive:
        stmt = stmt.where(StockLocation.is_active == True)
    if q:
        stmt = stmt.where(
            StockLocation.name.ilike(f"%{q}%") |
            StockLocation.code.ilike(f"%{q}%") |
            StockLocation.location_type.ilike(f"%{q}%")
        )

    result = await db.execute(stmt.order_by(StockLocation.name))
    locations = result.scalars().all()

    summary_result = await db.execute(
        select(
            LocationStock.location_id,
            func.count(LocationStock.product_id),
            func.coalesce(func.sum(LocationStock.qty), 0),
        ).group_by(LocationStock.location_id)
    )
    summary_map = {
        location_id: {"product_count": product_count, "total_qty": total_qty}
        for location_id, product_count, total_qty in summary_result.all()
    }

    return {
        "total": len(locations),
        "items": [
            serialize_location(
                location,
                product_count=summary_map.get(location.id, {}).get("product_count", 0),
                total_qty=summary_map.get(location.id, {}).get("total_qty", 0),
            )
            for location in locations
        ],
    }


@router.post("/api/locations", dependencies=[Depends(require_permission("action_inventory_adjust"))])
async def create_location(
    data: LocationCreate,
    db: AsyncSession = Depends(get_async_session),
    current_user: User = Depends(get_current_user),
):
    clean_name = data.name.strip()
    if not clean_name:
        raise HTTPException(status_code=400, detail="Location name is required")

    existing_name = await db.execute(select(StockLocation).where(StockLocation.name == clean_name))
    if existing_name.scalar_one_or_none() is not None:
        raise HTTPException(status_code=400, detail="Location name already exists")

    clean_code = data.code.strip() if data.code else None
    if clean_code:
        existing_code = await db.execute(select(StockLocation).where(StockLocation.code == clean_code))
        if existing_code.scalar_one_or_none() is not None:
            raise HTTPException(status_code=400, detail="Location code already exists")

    location = StockLocation(
        name=clean_name,
        code=clean_code,
        location_type=data.location_type.strip() or "warehouse",
        is_active=data.is_active,
    )
    db.add(location)
    await db.flush()
    record(
        db,
        "Inventory",
        "create_location",
        f"Created stock location {location.name}",
        user=current_user,
        ref_type="stock_location",
        ref_id=location.id,
    )
    await db.commit()
    await db.refresh(location)
    return serialize_location(location)


@router.get("/api/location-stock")
async def get_location_stock(
    q: str = "",
    product_id: int | None = None,
    location_id: int | None = None,
    skip: int = 0,
    limit: int = 50,
    db: AsyncSession = Depends(get_async_session),
):
    stmt = select(Product).where(or_(Product.is_active.is_(True), Product.is_active.is_(None)))
    if q:
        stmt = stmt.where(Product.name.ilike(f"%{q}%") | Product.sku.ilike(f"%{q}%"))
    if product_id is not None:
        stmt = stmt.where(Product.id == product_id)
    if location_id is not None:
        stmt = stmt.where(
            Product.id.in_(
                select(LocationStock.product_id).where(LocationStock.location_id == location_id)
            )
        )

    count_result = await db.execute(select(func.count()).select_from(stmt.subquery()))
    total = count_result.scalar() or 0
    result = await db.execute(stmt.order_by(Product.name).offset(skip).limit(limit))
    products = result.scalars().all()

    location_stock_map: dict[int, list[LocationStock]] = {}
    if products:
        product_ids = [product.id for product in products]
        stock_stmt = (
            select(LocationStock)
            .options(selectinload(LocationStock.location))
            .where(LocationStock.product_id.in_(product_ids))
        )
        if location_id is not None:
            stock_stmt = stock_stmt.where(LocationStock.location_id == location_id)
        stock_result = await db.execute(stock_stmt.order_by(LocationStock.product_id, LocationStock.location_id))
        for location_stock in stock_result.scalars().all():
            location_stock_map.setdefault(location_stock.product_id, []).append(location_stock)

    return {
        "total": total,
        "items": [
            serialize_product_location_stock(product, location_stock_map.get(product.id, []))
            for product in products
        ],
    }


@router.post("/api/transfers", dependencies=[Depends(require_permission("action_inventory_adjust"))])
async def create_transfer(
    data: StockTransferCreate,
    db: AsyncSession = Depends(get_async_session),
    current_user: User = Depends(get_current_user),
):
    product_result = await db.execute(
        select(Product).where(
            Product.id == data.product_id,
            or_(Product.is_active.is_(True), Product.is_active.is_(None)),
        )
    )
    product = product_result.scalar_one_or_none()
    if product is None:
        raise HTTPException(status_code=404, detail="Product not found")

    locations_result = await db.execute(
        select(StockLocation).where(
            StockLocation.id.in_([data.source_location_id, data.destination_location_id])
        )
    )
    locations = {location.id: location for location in locations_result.scalars().all()}
    source_location = locations.get(data.source_location_id)
    destination_location = locations.get(data.destination_location_id)
    if source_location is None or destination_location is None:
        raise HTTPException(status_code=404, detail="Source or destination location not found")

    try:
        summary = await create_stock_transfer(
            db,
            product=product,
            source_location=source_location,
            destination_location=destination_location,
            qty=data.qty,
            user_id=current_user.id,
            note=data.note,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    summary["transfer"]["actor"] = current_user.name
    record(
        db,
        "Inventory",
        "create_transfer",
        (
            f"Transferred {float(quantize_qty(data.qty)):.3f} {product.unit} of {product.name} "
            f"from {source_location.name} to {destination_location.name}"
        ),
        user=current_user,
        ref_type="stock_transfer",
        ref_id=summary["transfer"]["id"],
    )
    await db.commit()
    return {"ok": True, **summary}


@router.get("/api/transfers")
async def get_transfers(
    q: str = "",
    product_id: int | None = None,
    source_location_id: int | None = None,
    destination_location_id: int | None = None,
    skip: int = 0,
    limit: int = 100,
    db: AsyncSession = Depends(get_async_session),
):
    stmt = (
        select(StockTransfer)
        .options(
            selectinload(StockTransfer.product),
            selectinload(StockTransfer.source_location),
            selectinload(StockTransfer.destination_location),
            selectinload(StockTransfer.user),
        )
        .join(Product, StockTransfer.product_id == Product.id)
    )
    if q:
        stmt = stmt.where(
            Product.name.ilike(f"%{q}%") |
            Product.sku.ilike(f"%{q}%") |
            StockTransfer.note.ilike(f"%{q}%")
        )
    if product_id is not None:
        stmt = stmt.where(StockTransfer.product_id == product_id)
    if source_location_id is not None:
        stmt = stmt.where(StockTransfer.source_location_id == source_location_id)
    if destination_location_id is not None:
        stmt = stmt.where(StockTransfer.destination_location_id == destination_location_id)

    count_result = await db.execute(select(func.count()).select_from(stmt.subquery()))
    total = count_result.scalar() or 0
    result = await db.execute(stmt.order_by(StockTransfer.created_at.desc()).offset(skip).limit(limit))
    transfers = result.scalars().all()
    return {
        "total": total,
        "items": [serialize_transfer(transfer) for transfer in transfers],
    }


@router.get("/api/moves")
async def get_moves(
    q:         str = "",
    date_from: str = None,
    date_to:   str = None,
    product_id: int = None,
    skip:       int = 0,
    limit:      int = 100,
    db: AsyncSession = Depends(get_async_session),
):
    count_stmt = select(StockMove).join(Product, StockMove.product_id == Product.id)
    stmt = (
        select(StockMove)
        .options(selectinload(StockMove.product))
        .join(Product, StockMove.product_id == Product.id)
    )
    if product_id:
        count_stmt = count_stmt.where(StockMove.product_id == product_id)
        stmt = stmt.where(StockMove.product_id == product_id)
    if q:
        count_stmt = count_stmt.where(
            Product.name.ilike(f"%{q}%") |
            Product.sku.ilike(f"%{q}%") |
            StockMove.ref_type.ilike(f"%{q}%") |
            StockMove.note.ilike(f"%{q}%")
        )
        stmt = stmt.where(
            Product.name.ilike(f"%{q}%") |
            Product.sku.ilike(f"%{q}%") |
            StockMove.ref_type.ilike(f"%{q}%") |
            StockMove.note.ilike(f"%{q}%")
        )
    if date_from:
        try:
            dt_from = datetime.fromisoformat(date_from)
            count_stmt = count_stmt.where(StockMove.created_at >= dt_from)
            stmt = stmt.where(StockMove.created_at >= dt_from)
        except ValueError:
            pass
    if date_to:
        try:
            dt_to = datetime.fromisoformat(date_to).replace(hour=23, minute=59, second=59)
            count_stmt = count_stmt.where(StockMove.created_at <= dt_to)
            stmt = stmt.where(StockMove.created_at <= dt_to)
        except ValueError:
            pass
    cnt_result = await db.execute(select(func.count()).select_from(count_stmt.subquery()))
    total = cnt_result.scalar()
    result = await db.execute(stmt.order_by(StockMove.created_at.desc()).offset(skip).limit(limit))
    moves = result.scalars().all()
    return {
        "total": total,
        "moves": [
            {
                "id":         m.id,
                "product":    m.product.name if m.product else "—",
                "sku":        m.product.sku  if m.product else "—",
                "type":       m.type,
                "qty":        float(m.qty),
                "qty_before": float(m.qty_before) if m.qty_before is not None else 0,
                "qty_after":  float(m.qty_after)  if m.qty_after  is not None else 0,
                "ref_type":   m.ref_type or "—",
                "ref_id":     m.ref_id,
                "note":       m.note or "—",
                "created_at": m.created_at.strftime("%Y-%m-%d %H:%M") if m.created_at else "—",
            }
            for m in moves
        ],
    }


@router.get("/export/moves")
async def export_moves(
    q:         str = "",
    date_from: str = None,
    date_to:   str = None,
    product_id: int = None,
    db: AsyncSession = Depends(get_async_session),
):
    data = await get_moves(q=q, date_from=date_from, date_to=date_to, product_id=product_id, skip=0, limit=100000, db=db)
    rows = [
        [m["created_at"], m["product"], m["sku"], m["type"], m["qty"], m["qty_before"], m["qty_after"], m["ref_type"], m["ref_id"], m["note"]]
        for m in data["moves"]
    ]
    buf = to_xlsx(
        ["Date", "Product", "SKU", "Type", "Qty", "Before", "After", "Reference", "Ref ID", "Note"],
        rows,
        "Stock Moves"
    )
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=inventory_moves_{date.today()}.xlsx"}
    )


@router.post("/api/adjust", dependencies=[Depends(require_permission("action_inventory_adjust"))])
async def adjust_stock(data: StockAdjustment, db: AsyncSession = Depends(get_async_session), current_user: User = Depends(get_current_user)):
    result = await db.execute(select(Product).where(Product.id == data.product_id))
    product = result.scalar_one_or_none()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")

    before = float(product.stock or 0)
    after  = before + data.qty

    if after < 0:
        raise HTTPException(status_code=400, detail=f"Stock cannot go below 0. Current stock: {before}")

    location = None
    location_after = None
    ref_type = "manual"
    ref_id = None
    note = data.note or "Manual adjustment"
    if data.location_id is not None:
        location_result = await db.execute(select(StockLocation).where(StockLocation.id == data.location_id))
        location = location_result.scalar_one_or_none()
        if location is None:
            raise HTTPException(status_code=404, detail="Location not found")

        location_stock = await get_or_create_location_stock(
            db,
            product_id=product.id,
            location_id=location.id,
        )
        location_before = float(location_stock.qty)
        location_after = location_before + data.qty
        if location_after < 0:
            raise HTTPException(
                status_code=400,
                detail=f"Location stock cannot go below 0. Current stock in {location.name}: {location_before}",
            )
        location_stock.qty = location_after
        ref_type = "location_adjust"
        ref_id = location.id
        note = data.note or f"Manual adjustment for {location.name}"

    product.stock = after

    move = StockMove(
        product_id=product.id,
        type="adjust",
        qty=data.qty,
        qty_before=before,
        qty_after=after,
        ref_type=ref_type,
        ref_id=ref_id,
        note=note,
        user_id=current_user.id,
    )
    db.add(move)
    description = (
        f"Stock adjusted: {product.name} - {before:+.3g} -> {after:.3g} (delta {data.qty:+.3g})"
        + (f" - {data.note}" if data.note else "")
    )
    if location is not None:
        description += f" at {location.name}"
    record(
        db,
        "Inventory",
        "adjust_stock",
        description,
        user=current_user,
        ref_type="stock_move",
        ref_id=move.id if hasattr(move, "id") else None,
    )
    if False:
        record(db, "Inventory", "adjust_stock",
           f"Stock adjusted: {product.name} — {before:+.3g} → {after:.3g} (Δ{data.qty:+.3g})"
           + (f" — {data.note}" if data.note else ""),
           ref_type="stock_move", ref_id=move.id if hasattr(move, 'id') else None)
    await db.commit()
    payload = {"ok": True, "new_stock": after}
    if location is not None:
        payload["location_id"] = location.id
        payload["new_location_stock"] = location_after
    return payload


@router.get("/api/summary")
async def get_summary(db: AsyncSession = Depends(get_async_session)):
    low_stock_threshold = func.coalesce(Product.reorder_level, Product.min_stock)
    active_filter = or_(Product.is_active.is_(True), Product.is_active.is_(None))
    r1 = await db.execute(select(func.count(Product.id)).where(active_filter))
    total_products = r1.scalar() or 0
    r2 = await db.execute(select(func.count(Product.id)).where(
        active_filter, Product.stock <= low_stock_threshold
    ))
    low_stock = r2.scalar() or 0
    r3 = await db.execute(select(func.count(Product.id)).where(
        active_filter, Product.stock <= 0
    ))
    out_of_stock = r3.scalar() or 0
    r4 = await db.execute(select(func.count(StockMove.id)))
    total_moves = r4.scalar() or 0
    return {
        "total_products": total_products,
        "low_stock":      low_stock,
        "out_of_stock":   out_of_stock,
        "total_moves":    total_moves,
    }


# ── UI ─────────────────────────────────────────────────
@router.get("/", response_class=HTMLResponse)
def inventory_ui():
    return """
<!DOCTYPE html>
<html>
<head>
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Inventory</title>
<link href="https://fonts.googleapis.com/css2?family=Outfit:wght@400;500;600;700;800;900&family=JetBrains+Mono:wght@400;500;700&display=swap" rel="stylesheet">
<style>
:root {
    --bg:      #060810;
    --surface: #0a0d18;
    --card:    #0f1424;
    --card2:   #151c30;
    --border:  rgba(255,255,255,0.06);
    --border2: rgba(255,255,255,0.11);
    --green:   #00ff9d;
    --blue:    #4d9fff;
    --purple:  #a855f7;
    --danger:  #ff4d6d;
    --warn:    #ffb547;
    --text:    #f0f4ff;
    --sub:     #8899bb;
    --muted:   #445066;
    --sans:    'Outfit', sans-serif;
    --mono:    'JetBrains Mono', monospace;
    --r:       12px;
}
body.light{
    --bg:#f4f5ef;--surface:#f1f3eb;--card:#eceee6;--card2:#e4e6de;
    --border:rgba(0,0,0,0.08);--border2:rgba(0,0,0,0.14);
    --green:#0f8a43;
    --text:#1a1e14;--sub:#4a5040;--muted:#7b816f;
}
body.light nav{background:rgba(244,245,239,.92);}
body.light .nav-link:hover{background:rgba(0,0,0,.05);}
body.light tr:hover td{background:rgba(0,0,0,.03);}
.mode-btn{display:flex;align-items:center;justify-content:center;width:36px;height:36px;border-radius:10px;border:1px solid var(--border);background:var(--card);color:var(--sub);font-size:16px;cursor:pointer;transition:all .2s;font-family:var(--sans);}
.mode-btn:hover{border-color:var(--border2);transform:scale(1.06);}
.topbar-right{display:flex;align-items:center;gap:12px;}
.user-pill{display:flex;align-items:center;gap:10px;background:var(--card);border:1px solid var(--border);border-radius:40px;padding:7px 16px 7px 10px;}
.user-avatar{width:28px;height:28px;background:linear-gradient(135deg,#7ecb6f,#d4a256);border-radius:50%;display:flex;align-items:center;justify-content:center;font-size:12px;font-weight:700;color:#0a0c08;}
.user-name{font-size:13px;font-weight:500;color:var(--sub);}
.logout-btn{background:transparent;border:1px solid var(--border);color:var(--muted);font-family:var(--sans);font-size:12px;font-weight:500;padding:8px 16px;border-radius:8px;cursor:pointer;transition:all .2s;letter-spacing:.3px;}
.logout-btn:hover{border-color:#c97a7a;color:#c97a7a;}
*, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
body { font-family: var(--sans); background: var(--bg); color: var(--text); min-height: 100vh; font-size: 14px; }

nav {
    position: sticky; top: 0; z-index: 100;
    display: flex; align-items: center; gap: 10px;
    padding: 0 24px; height: 58px;
    background: rgba(10,13,24,.92);
    backdrop-filter: blur(20px);
    border-bottom: 1px solid var(--border);
}
.logo {
    font-size: 18px; font-weight: 900;
    background: linear-gradient(135deg, var(--green), var(--blue));
    -webkit-background-clip: text; -webkit-text-fill-color: transparent;
    background-clip: text; margin-right: 12px;
}
.nav-link {
    padding: 7px 14px; border-radius: 8px;
    color: var(--sub); font-size: 13px; font-weight: 600;
    text-decoration: none; transition: all .2s;
}
.nav-link:hover { background: rgba(255,255,255,.05); color: var(--text); }
.nav-link.active { background: rgba(0,255,157,.1); color: var(--green); }
.nav-spacer { flex: 1; }

.content { max-width: 1300px; margin: 0 auto; padding: 28px 24px; display: flex; flex-direction: column; gap: 20px; }
.page-title { font-size: 24px; font-weight: 800; letter-spacing: -.5px; }
.page-sub   { color: var(--muted); font-size: 13px; margin-top: 3px; }

/* STAT CARDS */
.stats-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(180px,1fr)); gap: 14px; }
.stat-card {
    background: var(--card); border: 1px solid var(--border);
    border-radius: var(--r); padding: 18px 20px;
    display: flex; flex-direction: column; gap: 8px;
    position: relative; overflow: hidden;
}
.stat-card::before { content:''; position:absolute; top:0; left:0; right:0; height:2px; }
.stat-card.green::before  { background: linear-gradient(90deg, var(--green), transparent); }
.stat-card.warn::before   { background: linear-gradient(90deg, var(--warn), transparent); }
.stat-card.danger::before { background: linear-gradient(90deg, var(--danger), transparent); }
.stat-card.blue::before   { background: linear-gradient(90deg, var(--blue), transparent); }
.stat-label { font-size: 10px; font-weight: 700; letter-spacing: 1.5px; text-transform: uppercase; color: var(--muted); }
.stat-value { font-family: var(--mono); font-size: 28px; font-weight: 700; }
.stat-value.green  { color: var(--green); }
.stat-value.warn   { color: var(--warn); }
.stat-value.danger { color: var(--danger); }
.stat-value.blue   { color: var(--blue); }

/* TABS */
.tabs { display: flex; gap: 4px; background: var(--card); border: 1px solid var(--border); border-radius: var(--r); padding: 4px; width: fit-content; }
.tab {
    padding: 8px 20px; border-radius: 9px;
    font-size: 13px; font-weight: 700; cursor: pointer;
    border: none; background: transparent; color: var(--muted);
    transition: all .2s; font-family: var(--sans);
}
.tab.active { background: var(--card2); color: var(--text); }

/* TOOLBAR */
.toolbar { display: flex; align-items: center; gap: 10px; flex-wrap: wrap; }
.search-box {
    display: flex; align-items: center; gap: 9px;
    background: var(--card); border: 1px solid var(--border);
    border-radius: var(--r); padding: 0 14px; flex: 1; min-width: 200px;
    transition: border-color .2s;
}
.search-box:focus-within { border-color: rgba(0,255,157,.3); }
.search-box svg { color: var(--muted); flex-shrink: 0; }
.search-box input {
    background: transparent; border: none; outline: none;
    color: var(--text); font-family: var(--sans);
    font-size: 14px; padding: 11px 0; width: 100%;
}
.search-box input::placeholder { color: var(--muted); }
.btn {
    display: flex; align-items: center; gap: 7px;
    padding: 10px 16px; border-radius: var(--r);
    font-family: var(--sans); font-size: 13px; font-weight: 700;
    cursor: pointer; border: none; transition: all .2s; white-space: nowrap;
}
.btn-green  { background: linear-gradient(135deg, var(--green), #00d4ff); color: #021a10; }
.btn-green:hover { filter: brightness(1.1); transform: translateY(-1px); }
.btn-outline {
    background: transparent; border: 1px solid var(--border2); color: var(--sub);
}
.btn-outline:hover { border-color: var(--warn); color: var(--warn); }
.btn-outline.active { border-color: var(--warn); color: var(--warn); background: rgba(255,181,71,.08); }

/* TABLE */
.table-wrap { background: var(--card); border: 1px solid var(--border); border-radius: var(--r); overflow: hidden; }
table { width: 100%; border-collapse: collapse; }
thead { background: var(--card2); }
th { text-align: left; font-size: 10px; font-weight: 700; letter-spacing: 1px; text-transform: uppercase; color: var(--muted); padding: 12px 16px; }
td { padding: 12px 16px; border-top: 1px solid var(--border); color: var(--sub); font-size: 13px; }
tr:hover td { background: rgba(255,255,255,.02); }
td.name { color: var(--text); font-weight: 600; }
td.mono { font-family: var(--mono); }
.low-badge  { display:inline-flex;align-items:center;gap:4px;background:rgba(255,77,109,.1);border:1px solid rgba(255,77,109,.2);color:var(--danger);font-size:11px;font-weight:700;padding:2px 8px;border-radius:20px; }
.ok-badge   { font-family:var(--mono);font-size:13px;color:var(--green); }
.out-badge  { display:inline-flex;align-items:center;gap:4px;background:rgba(255,77,109,.15);border:1px solid rgba(255,77,109,.3);color:var(--danger);font-size:11px;font-weight:700;padding:2px 8px;border-radius:20px; }

.action-btn {
    background: transparent; border: 1px solid var(--border2);
    color: var(--sub); font-size: 12px; font-weight: 600;
    padding: 5px 10px; border-radius: 7px; cursor: pointer;
    transition: all .15s; font-family: var(--sans);
}
.action-btn:hover { border-color: var(--green); color: var(--green); }

/* MOVE TYPE BADGES */
.move-in     { color: var(--green); font-weight: 700; }
.move-out    { color: var(--danger); font-weight: 700; }
.move-adjust { color: var(--blue); font-weight: 700; }

/* PAGINATION */
.pagination {
    display: flex; align-items: center; justify-content: space-between;
    padding: 14px 16px; border-top: 1px solid var(--border); font-size: 13px; color: var(--muted);
}
.page-btns { display: flex; gap: 6px; }
.page-btn {
    background: var(--card2); border: 1px solid var(--border2);
    color: var(--sub); font-family: var(--sans); font-size: 12px;
    padding: 6px 12px; border-radius: 7px; cursor: pointer; transition: all .15s;
}
.page-btn:hover { border-color: var(--green); color: var(--green); }
.page-btn:disabled { opacity: .3; cursor: not-allowed; }

/* MODAL */
.modal-bg {
    position: fixed; inset: 0; z-index: 500;
    background: rgba(0,0,0,.7); backdrop-filter: blur(4px);
    display: none; align-items: center; justify-content: center;
}
.modal-bg.open { display: flex; }
.modal {
    background: var(--card); border: 1px solid var(--border2);
    border-radius: 16px; padding: 28px;
    width: 460px; max-width: 95vw;
    animation: modalIn .2s ease;
}
@keyframes modalIn { from{opacity:0;transform:scale(.95)} to{opacity:1;transform:scale(1)} }
.modal-title { font-size: 18px; font-weight: 800; margin-bottom: 6px; }
.modal-sub   { font-size: 13px; color: var(--muted); margin-bottom: 20px; }
.fld { display: flex; flex-direction: column; gap: 6px; margin-bottom: 14px; }
.fld label { font-size: 11px; font-weight: 700; letter-spacing: 1px; text-transform: uppercase; color: var(--muted); }
.fld input, .fld select {
    background: var(--card2); border: 1px solid var(--border2);
    border-radius: 10px; padding: 10px 12px;
    color: var(--text); font-family: var(--sans); font-size: 14px;
    outline: none; transition: border-color .2s; width: 100%;
}
.fld input:focus, .fld select:focus { border-color: rgba(0,255,157,.4); }
.current-stock-display {
    background: var(--card2); border: 1px solid var(--border2);
    border-radius: 10px; padding: 12px 14px;
    display: flex; justify-content: space-between; align-items: center;
    margin-bottom: 14px;
}
.modal-actions { display: flex; gap: 10px; margin-top: 6px; justify-content: flex-end; }
.btn-cancel {
    background: transparent; border: 1px solid var(--border2);
    color: var(--sub); padding: 10px 18px; border-radius: var(--r);
    font-family: var(--sans); font-size: 13px; font-weight: 700; cursor: pointer;
}
.btn-cancel:hover { border-color: var(--danger); color: var(--danger); }

.toast {
    position: fixed; bottom: 22px; left: 50%;
    transform: translateX(-50%) translateY(16px);
    background: var(--card2); border: 1px solid var(--border2);
    border-radius: var(--r); padding: 12px 20px;
    font-size: 13px; font-weight: 600; color: var(--text);
    box-shadow: 0 20px 50px rgba(0,0,0,.5);
    opacity: 0; pointer-events: none;
    transition: opacity .25s, transform .25s; z-index: 999;
}
.toast.show { opacity:1; transform: translateX(-50%) translateY(0); }
::-webkit-scrollbar { width: 4px; }
::-webkit-scrollbar-thumb { background: var(--border2); border-radius: 4px; }
</style>
</head>
<body>

<nav>
    <a href="/home" class="logo" style="text-decoration:none;display:flex;align-items:center;gap:8px;"><svg width="22" height="22" viewBox="0 0 24 24" fill="none"><polygon points="13,2 4,14 11,14 11,22 20,10 13,10" fill="#f59e0b" stroke="#fbbf24" stroke-width="0.5"/></svg>Thunder ERP</a>
    <a href="/dashboard"       class="nav-link">Dashboard</a>
    <a href="/pos"             class="nav-link">POS</a>
    <a href="/products/"       class="nav-link">Products</a>
    <a href="/customers-mgmt/" class="nav-link">Customers</a>
    <a href="/suppliers/"      class="nav-link">Suppliers</a>
    <a href="/inventory/"      class="nav-link active">Inventory</a>
    <a href="/import"          class="nav-link">Import</a>
    <span class="nav-spacer"></span>
    <div class="topbar-right">
        <button class="mode-btn" id="mode-btn" onclick="toggleMode()" title="Toggle color mode">??</button>
        <div class="user-pill">
            <div class="user-avatar" id="user-avatar">A</div>
            <span class="user-name" id="user-name">Admin</span>
        </div>
        <button class="logout-btn" onclick="logout()">Sign out</button>
    </div>
</nav>

<div class="content">
    <div>
        <div class="page-title">Inventory</div>
        <div class="page-sub">Track stock levels and movements</div>
    </div>

    <!-- STAT CARDS -->
    <div class="stats-grid">
        <div class="stat-card green">
            <div class="stat-label">Total Products</div>
            <div class="stat-value green" id="stat-total">—</div>
        </div>
        <div class="stat-card warn">
            <div class="stat-label">Low Stock</div>
            <div class="stat-value warn" id="stat-low">—</div>
        </div>
        <div class="stat-card danger">
            <div class="stat-label">Out of Stock</div>
            <div class="stat-value danger" id="stat-out">—</div>
        </div>
        <div class="stat-card blue">
            <div class="stat-label">Total Movements</div>
            <div class="stat-value blue" id="stat-moves">—</div>
        </div>
    </div>

    <!-- TABS -->
    <div style="display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:12px;">
        <div class="tabs">
            <button class="tab active" id="tab-stock" onclick="switchTab('stock')">Stock Levels</button>
            <button class="tab"        id="tab-moves" onclick="switchTab('moves')">Movements</button>
        </div>
    </div>

    <!-- STOCK LEVELS -->
    <div id="stock-section">
        <div class="toolbar">
            <div class="search-box">
                <svg width="15" height="15" fill="none" stroke="currentColor" stroke-width="2.5" viewBox="0 0 24 24">
                    <circle cx="11" cy="11" r="8"/><path d="m21 21-4.35-4.35"/>
                </svg>
                <input id="stock-search" placeholder="Search by name or SKU…" oninput="onStockSearch()">
            </div>
            <button class="btn btn-outline" id="low-stock-btn" onclick="toggleLowStock()">⚠ Low Stock Only</button>
        </div>
        <div class="table-wrap">
            <table>
                <thead>
                    <tr>
                        <th>SKU</th>
                        <th>Product</th>
                        <th>Current Stock</th>
                        <th>Min Stock</th>
                        <th>Unit</th>
                        <th>Status</th>
                        <th>Adjust</th>
                    </tr>
                </thead>
                <tbody id="stock-body">
                    <tr><td colspan="7" style="text-align:center;color:var(--muted);padding:40px">Loading…</td></tr>
                </tbody>
            </table>
            <div class="pagination">
                <span id="stock-page-info">—</span>
                <div class="page-btns">
                    <button class="page-btn" id="stock-prev" onclick="stockPrevPage()">← Prev</button>
                    <button class="page-btn" id="stock-next" onclick="stockNextPage()">Next →</button>
                </div>
            </div>
        </div>
    </div>
    <!-- MOVEMENTS -->
    <div id="moves-section" style="display:none">
        <div class="toolbar">
            <div class="search-box">
                <svg width="15" height="15" fill="none" stroke="currentColor" stroke-width="2.5" viewBox="0 0 24 24">
                    <circle cx="11" cy="11" r="8"/><path d="m21 21-4.35-4.35"/>
                </svg>
                <input id="moves-search" placeholder="Search by product, SKU, reference, or note..." oninput="onMovesSearch()">
            </div>
            <input type="date" id="moves-date-from" style="background:var(--card);border:1px solid var(--border);border-radius:var(--r);color:var(--text);padding:10px 12px;font-family:var(--sans);">
            <input type="date" id="moves-date-to" style="background:var(--card);border:1px solid var(--border);border-radius:var(--r);color:var(--text);padding:10px 12px;font-family:var(--sans);">
            <button class="btn btn-outline" onclick="applyMovesFilters()">Apply</button>
            <button class="btn btn-green" onclick="exportMoves()">Export Excel</button>
        </div>
        <div class="table-wrap">
            <table>
                <thead>
                    <tr>
                        <th>Date</th>
                        <th>Product</th>
                        <th>SKU</th>
                        <th>Type</th>
                        <th>Qty</th>
                        <th>Before</th>
                        <th>After</th>
                        <th>Reference</th>
                        <th>Note</th>
                    </tr>
                </thead>
                <tbody id="moves-body">
                    <tr><td colspan="9" style="text-align:center;color:var(--muted);padding:40px">Loading…</td></tr>
                </tbody>
            </table>
            <div class="pagination">
                <span id="moves-page-info">—</span>
                <div class="page-btns">
                    <button class="page-btn" id="moves-prev" onclick="movesPrevPage()">← Prev</button>
                    <button class="page-btn" id="moves-next" onclick="movesNextPage()">Next →</button>
                </div>
            </div>
        </div>
    </div>
</div>

<!-- ADJUST MODAL -->
<div class="modal-bg" id="adjust-modal">
    <div class="modal">
        <div class="modal-title">Adjust Stock</div>
        <div class="modal-sub" id="adjust-product-name">Product name</div>

        <div class="current-stock-display">
            <span style="color:var(--muted);font-size:13px;font-weight:600">Current Stock</span>
            <span style="font-family:var(--mono);font-size:22px;font-weight:700;color:var(--green)" id="adjust-current">0</span>
        </div>

        <div class="fld">
            <label>Adjustment Type</label>
            <select id="adj-type" onchange="updateAdjPreview()">
                <option value="add">➕ Add Stock (positive)</option>
                <option value="remove">➖ Remove Stock (negative)</option>
                <option value="set">🔄 Set Exact Amount</option>
            </select>
        </div>

        <div class="fld">
            <label>Quantity</label>
            <input id="adj-qty" type="number" placeholder="0" min="0" step="any" oninput="updateAdjPreview()">
        </div>

        <div class="current-stock-display" id="adj-preview" style="display:none">
            <span style="color:var(--muted);font-size:13px;font-weight:600">New Stock Will Be</span>
            <span style="font-family:var(--mono);font-size:22px;font-weight:700;color:var(--blue)" id="adj-preview-val">0</span>
        </div>

        <div class="fld">
            <label>Note (optional)</label>
            <input id="adj-note" placeholder="Reason for adjustment…">
        </div>

        <div class="modal-actions">
            <button class="btn-cancel" onclick="closeAdjustModal()">Cancel</button>
            <button class="btn btn-green" onclick="saveAdjustment()">Apply Adjustment</button>
        </div>
    </div>
</div>

<div class="toast" id="toast"></div>

<script>
  // Auth guard: redirect to login if the readable session cookie is absent
  function _hasAuthCookie() {
      return document.cookie.split(";").some(c => c.trim().startsWith("logged_in="));
  }
  if (!_hasAuthCookie()) { window.location.href = "/"; }

  function setModeButton(isLight){
    const btn = document.getElementById("mode-btn");
    if(btn) btn.innerText = isLight ? "☀️" : "🌙";
}
function toggleMode(){
    const isLight = document.body.classList.toggle("light");
    localStorage.setItem("colorMode", isLight ? "light" : "dark");
    setModeButton(isLight);
}
function initializeColorMode(){
    const isLight = localStorage.getItem("colorMode") === "light";
    document.body.classList.toggle("light", isLight);
    setModeButton(isLight);
}
async function initUser() {
    try {
        const r = await fetch("/auth/me");
        if (!r.ok) { window.location.href = "/"; return; }
        const u = await r.json();
        const nameEl = document.getElementById("user-name");
        const avatarEl = document.getElementById("user-avatar");
        if (nameEl) nameEl.innerText = u.name;
        if (avatarEl) avatarEl.innerText = u.name.charAt(0).toUpperCase();
        return u;
    } catch(e) { window.location.href = "/"; }
}
async function logout(){
    await fetch("/auth/logout", { method: "POST" });
    window.location.href = "/";
}
  function hasPermission(permission, u){
      const role = u ? (u.role || "") : "";
      const perms = new Set(u ? (u.permissions || []) : []);
      return role === "admin" || perms.has(permission);
  }
  function applyInventoryActionPermissions(u){
      if(hasPermission("action_inventory_adjust", u)) return;
      document.querySelectorAll("#stock-body tr td:last-child button").forEach(btn => btn.remove());
  }
  initializeColorMode();
  initUser().then(u => { if(u) applyInventoryActionPermissions(u); });
  let currentTab   = "stock";
let stockPage    = 0;
let movesPage    = 0;
let pageSize     = 50;
let stockTotal   = 0;
let movesTotal   = 0;
let lowStockOnly = false;
let adjustingProduct = null;
let searchTimer  = null;

/* ── INIT ── */
async function init(){
    await loadSummary();
    await loadStock();
}

/* ── SUMMARY ── */
async function loadSummary(){
    let d = await (await fetch("/inventory/api/summary")).json();
    document.getElementById("stat-total").innerText  = d.total_products;
    document.getElementById("stat-low").innerText    = d.low_stock;
    document.getElementById("stat-out").innerText    = d.out_of_stock;
    document.getElementById("stat-moves").innerText  = d.total_moves;
}

/* ── TABS ── */
function switchTab(tab){
    currentTab = tab;
    document.getElementById("tab-stock").classList.toggle("active", tab==="stock");
    document.getElementById("tab-moves").classList.toggle("active", tab==="moves");
    document.getElementById("stock-section").style.display = tab==="stock" ? "" : "none";
    document.getElementById("moves-section").style.display = tab==="moves" ? "" : "none";
    if(tab==="moves") loadMoves();
}

/* ── STOCK ── */
function onStockSearch(){
    clearTimeout(searchTimer);
    searchTimer = setTimeout(()=>{ stockPage=0; loadStock(); }, 300);
}

function onMovesSearch(){
    clearTimeout(searchTimer);
    searchTimer = setTimeout(()=>{ movesPage=0; loadMoves(); }, 300);
}

function applyMovesFilters(){
    movesPage = 0;
    loadMoves();
}

function toggleLowStock(){
    lowStockOnly = !lowStockOnly;
    stockPage = 0;
    document.getElementById("low-stock-btn").classList.toggle("active", lowStockOnly);
    loadStock();
}

async function loadStock(){
    let q   = document.getElementById("stock-search").value.trim();
    let url = `/inventory/api/stock?skip=${stockPage*pageSize}&limit=${pageSize}&low_stock=${lowStockOnly}`;
    if(q) url += `&q=${encodeURIComponent(q)}`;
    let data = await (await fetch(url)).json();
    stockTotal = data.total;

    document.getElementById("stock-page-info").innerText =
        `Showing ${Math.min(stockPage*pageSize+1,stockTotal)}–${Math.min((stockPage+1)*pageSize,stockTotal)} of ${stockTotal}`;
    document.getElementById("stock-prev").disabled = stockPage===0;
    document.getElementById("stock-next").disabled = (stockPage+1)*pageSize >= stockTotal;

    if(!data.items.length){
        document.getElementById("stock-body").innerHTML =
            `<tr><td colspan="7" style="text-align:center;color:var(--muted);padding:40px">No products found</td></tr>`;
        return;
    }

    document.getElementById("stock-body").innerHTML = data.items.map(p => `
        <tr>
            <td style="font-family:var(--mono);font-size:12px;color:var(--muted)">${p.sku}</td>
            <td class="name">${p.name}</td>
            <td>${p.stock <= 0
                ? `<span class="out-badge">✕ Out</span>`
                : p.low
                ? `<span class="low-badge">⚠ ${p.stock.toFixed(0)}</span>`
                : `<span class="ok-badge">${p.stock.toFixed(0)}</span>`}
            </td>
            <td style="font-family:var(--mono);color:var(--muted)">${p.min_stock.toFixed(0)}</td>
            <td style="color:var(--muted)">${p.unit}</td>
            <td>${p.stock <= 0
                ? `<span style="color:var(--danger);font-size:12px">Out of Stock</span>`
                : p.low
                ? `<span style="color:var(--warn);font-size:12px">⚠ Low</span>`
                : `<span style="color:var(--green);font-size:12px">● OK</span>`}
            </td>
            <td>
                <button class="action-btn" onclick="openAdjustModal(${p.id},'${p.name.replace(/'/g,"\\'")}',${p.stock})">
                    Adjust
                </button>
            </td>
        </tr>`).join("");
    applyInventoryActionPermissions();
}

function stockPrevPage(){ if(stockPage>0){ stockPage--; loadStock(); } }
function stockNextPage(){ if((stockPage+1)*pageSize<stockTotal){ stockPage++; loadStock(); } }

/* ── MOVEMENTS ── */
async function loadMoves(){
    let q    = document.getElementById("moves-search").value.trim();
    let from = document.getElementById("moves-date-from").value;
    let to   = document.getElementById("moves-date-to").value;
    let url  = `/inventory/api/moves?skip=${movesPage*pageSize}&limit=${pageSize}`;
    if(q) url += `&q=${encodeURIComponent(q)}`;
    if(from) url += `&date_from=${encodeURIComponent(from)}`;
    if(to) url += `&date_to=${encodeURIComponent(to)}`;
    let data = await (await fetch(url)).json();
    movesTotal = data.total;

    document.getElementById("moves-page-info").innerText =
        `Showing ${Math.min(movesPage*pageSize+1,movesTotal)}–${Math.min((movesPage+1)*pageSize,movesTotal)} of ${movesTotal}`;
    document.getElementById("moves-prev").disabled = movesPage===0;
    document.getElementById("moves-next").disabled = (movesPage+1)*pageSize >= movesTotal;

    if(!data.moves.length){
        document.getElementById("moves-body").innerHTML =
            `<tr><td colspan="9" style="text-align:center;color:var(--muted);padding:40px">No movements yet</td></tr>`;
        return;
    }

    document.getElementById("moves-body").innerHTML = data.moves.map(m => {
        let typeClass = m.type==="in"?"move-in":m.type==="out"?"move-out":"move-adjust";
        let typeLabel = m.type==="in"?"▲ IN":m.type==="out"?"▼ OUT":"⟳ ADJ";
        let qtySign   = m.qty >= 0 ? "+" : "";
        return `
        <tr>
            <td style="font-size:12px;color:var(--muted)">${m.created_at}</td>
            <td class="name" style="max-width:160px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">${m.product}</td>
            <td style="font-family:var(--mono);font-size:11px;color:var(--muted)">${m.sku}</td>
            <td><span class="${typeClass}">${typeLabel}</span></td>
            <td style="font-family:var(--mono);color:${m.qty>=0?'var(--green)':'var(--danger)'}">${qtySign}${m.qty.toFixed(0)}</td>
            <td style="font-family:var(--mono);color:var(--muted)">${m.qty_before.toFixed(0)}</td>
            <td style="font-family:var(--mono);color:var(--sub)">${m.qty_after.toFixed(0)}</td>
            <td style="font-size:12px;color:var(--blue);text-transform:capitalize">${m.ref_type}</td>
            <td style="font-size:12px;color:var(--muted);max-width:140px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">${m.note}</td>
        </tr>`;
    }).join("");
}

function movesPrevPage(){ if(movesPage>0){ movesPage--; loadMoves(); } }
function movesNextPage(){ if((movesPage+1)*pageSize<movesTotal){ movesPage++; loadMoves(); } }

function exportMoves(){
    let q    = document.getElementById("moves-search").value.trim();
    let from = document.getElementById("moves-date-from").value;
    let to   = document.getElementById("moves-date-to").value;
    let url  = `/inventory/export/moves?`;
    let params = [];
    if(q) params.push(`q=${encodeURIComponent(q)}`);
    if(from) params.push(`date_from=${encodeURIComponent(from)}`);
    if(to) params.push(`date_to=${encodeURIComponent(to)}`);
    window.location.href = url + params.join("&");
}

/* ── ADJUST MODAL ── */
function openAdjustModal(id, name, currentStock){
    adjustingProduct = {id, currentStock};
    document.getElementById("adjust-product-name").innerText = name;
    document.getElementById("adjust-current").innerText      = currentStock.toFixed(0);
    document.getElementById("adj-type").value  = "add";
    document.getElementById("adj-qty").value   = "";
    document.getElementById("adj-note").value  = "";
    document.getElementById("adj-preview").style.display = "none";
    document.getElementById("adjust-modal").classList.add("open");
}

function closeAdjustModal(){
    document.getElementById("adjust-modal").classList.remove("open");
}

function updateAdjPreview(){
    let type = document.getElementById("adj-type").value;
    let qty  = parseFloat(document.getElementById("adj-qty").value)||0;
    let curr = adjustingProduct ? adjustingProduct.currentStock : 0;
    let newVal;
    if(type==="add")    newVal = curr + qty;
    else if(type==="remove") newVal = curr - qty;
    else                newVal = qty;

    document.getElementById("adj-preview").style.display = qty > 0 ? "" : "none";
    document.getElementById("adj-preview-val").innerText = newVal.toFixed(0);
    document.getElementById("adj-preview-val").style.color = newVal < 0 ? "var(--danger)" : "var(--blue)";
}

async function saveAdjustment(){
    if(!adjustingProduct){ return; }
    let type = document.getElementById("adj-type").value;
    let qty  = parseFloat(document.getElementById("adj-qty").value)||0;
    let note = document.getElementById("adj-note").value.trim();
    let curr = adjustingProduct.currentStock;

    if(qty <= 0){ showToast("Enter a quantity greater than 0"); return; }

    let actualQty;
    if(type==="add")         actualQty = qty;
    else if(type==="remove") actualQty = -qty;
    else                     actualQty = qty - curr; // set exact

    let res  = await fetch("/inventory/api/adjust",{
        method:"POST",
        headers:{"Content-Type":"application/json"},
        body:JSON.stringify({
            product_id: adjustingProduct.id,
            qty:        actualQty,
            note:       note || `Manual ${type}`,
        }),
    });
    let data = await res.json();
    if(data.detail){ showToast("Error: "+data.detail); return; }

    closeAdjustModal();
    showToast(`Stock updated ✓ New stock: ${data.new_stock.toFixed(0)}`);
    loadStock();
    loadSummary();
}

document.getElementById("adjust-modal").addEventListener("click",function(e){
    if(e.target===this) closeAdjustModal();
});

let toastTimer=null;
function showToast(msg){
    let t=document.getElementById("toast");
    t.innerText=msg; t.classList.add("show");
    clearTimeout(toastTimer);
    toastTimer=setTimeout(()=>t.classList.remove("show"),3500);
}

init();
</script>
</body>
</html>
"""
