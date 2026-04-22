from decimal import Decimal
from typing import Optional
from fastapi import APIRouter, Form, UploadFile, File, Depends
from fastapi.responses import HTMLResponse, StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import delete, select
import openpyxl, io

from app.core.permissions import require_permission
from app.core.security import get_current_user
from app.database import get_async_session
from app.models.invoice import Invoice, InvoiceItem
from app.models.accounting import Journal
from app.models.inventory import StockMove
from app.models.product import Product
from app.models.customer import Customer
from app.models.refund import RetailRefund
from app.models.b2b import B2BInvoice as B2BInvoiceModel, B2BInvoiceItem, Consignment as ConsignmentModel
from app.models.supplier import PurchaseItem
from app.services.sales_import_service import import_sales
from app.services.b2b_sales_import_service import import_b2b_sales
from app.services.expense_import_service import import_expenses
from app.services.farm_intake_import_service import (
    import_farm_intake,
    list_farm_intake_import_batches,
    revert_farm_intake_import_batch,
)

router = APIRouter(
    prefix="/import",
    tags=["Import"],
    dependencies=[Depends(require_permission("page_import"))],
)

ITEM_TYPE_ALIASES = {
    "finished": "finished",
    "finished product": "finished",
    "product": "finished",
    "raw": "raw",
    "raw material": "raw",
    "fresh": "fresh",
    "packing": "packing",
    "packaging": "packing",
    "ingredient": "ingredient",
    "ingredients": "ingredient",
}


def find_col(raw_headers, names):
    for name in names:
        for i, h in enumerate(raw_headers):
            if h == name.lower().strip():
                return i + 1
    return None


def safe_str(v):
    return str(v).strip() if v is not None else None


def safe_float(v):
    try: return float(v)
    except (ValueError, TypeError): return None


def normalize_item_type(value):
    raw_value = (safe_str(value) or "").strip().lower()
    if not raw_value:
        return "finished"
    return ITEM_TYPE_ALIASES.get(raw_value, raw_value)


# ── PREVIEW ────────────────────────────────────────────
@router.post("/api/preview")
async def preview_file(file: UploadFile = File(...)):
    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    ws = wb.active
    max_col = min(ws.max_column, 10)
    headers = [str(ws.cell(1, c).value or "") for c in range(1, max_col + 1)]
    rows = []
    for row in range(2, min(ws.max_row + 1, 7)):
        rows.append([str(ws.cell(row, c).value or "") for c in range(1, max_col + 1)])
    return {"headers": headers, "rows": rows, "total_rows": max(ws.max_row - 1, 0)}


# ── PRODUCTS ───────────────────────────────────────────
@router.post("/api/products")
async def import_products(file: UploadFile = File(...), db: AsyncSession = Depends(get_async_session)):
    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    ws = wb.active
    hdrs = [str(ws.cell(1, c).value or "").strip().lower() for c in range(1, ws.max_column + 2)]

    col_sku  = find_col(hdrs, ["sku","code","item code"])
    col_name = find_col(hdrs, ["item","name","product","product name","description"])
    col_unit = find_col(hdrs, ["uom","unit","unit of measure"])
    col_cost = find_col(hdrs, ["unit cost","cost","cost price"])
    col_price= find_col(hdrs, ["sales price","price","sale price","selling price"])
    col_cat  = find_col(hdrs, ["group","category","category name"])
    col_type = find_col(hdrs, ["item type","type","product type"])

    if not col_name:
        return {"error": "Cannot find Item/Name column"}

    created = updated = 0
    errors  = []

    for row in range(2, ws.max_row + 1):
        def v(c):
            if not c: return None
            x = ws.cell(row, c).value
            return str(x).strip() if x is not None else None

        name = v(col_name)
        if not name or name.lower() == "none": continue

        raw_sku = ws.cell(row, col_sku).value if col_sku else None
        if raw_sku is not None:
            try:    sku = str(int(float(str(raw_sku))))
            except (ValueError, TypeError): sku = str(raw_sku).strip()
        else:
            sku = None

        if not sku:
            all_r = await db.execute(select(Product))
            nums = []
            for p in all_r.scalars().all():
                try: nums.append(int(p.sku))
                except (ValueError, TypeError): pass
            sku = str(max(nums) + 1) if nums else "10001"

        unit  = v(col_unit) or "gram"
        cost  = safe_float(ws.cell(row, col_cost).value  if col_cost  else None) or 0.0
        price = safe_float(ws.cell(row, col_price).value if col_price else None) or 0.0
        cat   = v(col_cat)

        item_type = normalize_item_type(v(col_type))

        ex_r = await db.execute(select(Product).where(Product.sku == sku))
        existing = ex_r.scalar_one_or_none()
        if not existing:
            ex_r = await db.execute(select(Product).where(Product.name == name, Product.is_active == True))
            existing = ex_r.scalar_one_or_none()

        if existing:
            existing.name = name
            existing.unit = unit
            if cost  > 0: existing.cost  = cost
            if price > 0: existing.price = price
            if cat and hasattr(existing, "category"):  existing.category  = cat
            if hasattr(existing, "item_type"):          existing.item_type = item_type
            if not existing.sku: existing.sku = sku
            updated += 1
        else:
            p = Product(sku=sku, name=name, unit=unit, cost=cost,
                        price=price, stock=0, min_stock=5)
            if cat and hasattr(p, "category"):  p.category  = cat
            if hasattr(p, "item_type"):          p.item_type = item_type
            db.add(p)
            created += 1

    try:
        await db.commit()
    except Exception as e:
        await db.rollback()
        return {"error": str(e)}

    return {"ok": True, "created": created, "updated": updated,
            "errors": errors, "message": f"Done: {created} created, {updated} updated"}


# ── STOCK ──────────────────────────────────────────────
@router.post("/api/stock")
async def import_stock(file: UploadFile = File(...), db: AsyncSession = Depends(get_async_session)):
    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    ws = wb.active
    hdrs = [str(ws.cell(1, c).value or "").strip().lower() for c in range(1, ws.max_column + 2)]

    col_sku   = find_col(hdrs, ["sku","code","item code"])
    col_name  = find_col(hdrs, ["item","name","product","description"])
    col_stock = find_col(hdrs, ["stock","qty","quantity","on hand","soh"])

    if not col_stock:
        return {"error": "Cannot find Stock/Qty column"}

    updated   = 0
    not_found = []

    for row in range(2, ws.max_row + 1):
        sku_raw   = ws.cell(row, col_sku).value   if col_sku  else None
        name_raw  = ws.cell(row, col_name).value  if col_name else None
        stock_raw = ws.cell(row, col_stock).value

        if stock_raw is None: continue
        new_stock = safe_float(stock_raw)
        if new_stock is None: continue

        if sku_raw is not None:
            try:    sku = str(int(float(str(sku_raw))))
            except (ValueError, TypeError): sku = str(sku_raw).strip()
        else:
            sku = None

        product = None
        if sku:
            _r = await db.execute(select(Product).where(Product.sku == sku, Product.is_active == True))
            product = _r.scalar_one_or_none()
        if not product and name_raw:
            _r = await db.execute(select(Product).where(Product.name == str(name_raw).strip(), Product.is_active == True))
            product = _r.scalar_one_or_none()

        if product:
            before = float(product.stock)
            product.stock = new_stock
            db.add(StockMove(
                product_id=product.id, type="adjust",
                qty=round(new_stock - before, 3),
                qty_before=before, qty_after=new_stock,
                ref_type="import", note="Stock import from Excel",
            ))
            updated += 1
        else:
            label = sku or (str(name_raw)[:30] if name_raw else f"row {row}")
            not_found.append(label)

    try:
        await db.commit()
    except Exception as e:
        await db.rollback()
        return {"error": str(e)}

    return {"ok": True, "updated": updated, "not_found": not_found[:30],
            "message": f"Done: {updated} updated" + (f", {len(not_found)} not found" if not_found else "")}


# ── CUSTOMERS ──────────────────────────────────────────
@router.post("/api/customers")
async def import_customers(file: UploadFile = File(...), db: AsyncSession = Depends(get_async_session)):
    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    ws = wb.active
    hdrs = [str(ws.cell(1, c).value or "").strip().lower() for c in range(1, ws.max_column + 2)]

    col_name = find_col(hdrs, ["name","customer name","client name","customer"])
    col_phone= find_col(hdrs, ["phone","mobile","tel","telephone"])
    col_email= find_col(hdrs, ["email","e-mail"])
    col_addr = find_col(hdrs, ["address","area","city","location"])

    if not col_name:
        return {"error": "Cannot find Name column"}

    created = skipped = 0

    for row in range(2, ws.max_row + 1):
        def v(c):
            if not c: return None
            x = ws.cell(row, c).value
            return str(x).strip() if x is not None else None

        name  = v(col_name)
        if not name or name.lower() == "none": continue
        phone = v(col_phone)
        email = v(col_email)
        addr  = v(col_addr)

        if phone:
            _r = await db.execute(select(Customer).where(Customer.phone == phone))
            if _r.scalar_one_or_none():
                skipped += 1; continue
        _r = await db.execute(select(Customer).where(Customer.name == name))
        if _r.scalar_one_or_none():
            skipped += 1; continue

        db.add(Customer(name=name, phone=phone, email=email, address=addr))
        created += 1

    try:
        await db.commit()
    except Exception as e:
        await db.rollback()
        return {"error": str(e)}

    return {"ok": True, "created": created, "skipped": skipped,
            "message": f"Done: {created} imported, {skipped} skipped"}


# ── SALES IMPORT ───────────────────────────────────────

@router.post("/api/sales")
async def import_sales_endpoint(
    file: UploadFile = File(...),
    dry_run: bool = Form(True),
    mode: str = Form("history_only"),
    force: bool = Form(False),
    default_cost_ratio: Optional[float] = Form(None),
    db: AsyncSession = Depends(get_async_session),
    current_user=Depends(get_current_user),
):
    """Import historical retail sales from an Excel file.

    dry_run=True (default): validate and preview without writing anything.
    mode: history_only | with_journals | with_stock_and_journals
    force=True: skip duplicate detection and re-import even if records exist.
    default_cost_ratio: if set, auto-created products get cost = price × ratio.
    """
    contents = await file.read()
    return await import_sales(
        db=db,
        workbook_bytes=contents,
        filename=file.filename or "upload.xlsx",
        current_user_id=current_user.id,
        dry_run=dry_run,
        mode=mode,
        force=force,
        default_cost_ratio=default_cost_ratio,
    )


@router.get("/api/sales/template")
async def download_sales_template(_=Depends(get_current_user)):
    """Return a pre-filled Excel template for the historical sales import."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales"

    headers = ["SKU", "Item", "QTY", "Price", "Customer", "Date"]
    hdr_font = Font(bold=True, color="00FF9D")
    hdr_fill = PatternFill("solid", fgColor="0F1424")
    for col, h in enumerate(headers, 1):
        c = ws.cell(1, col, h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center")
    for col, w in enumerate([14, 28, 8, 10, 22, 14], 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.append(["SKU-001", "Olive Oil 500ml",  3,  15.50, "Ahmed Al-Rashid", "2026-01-15"])
    ws.append(["SKU-002", "Tahini 250g",      10,  8.00, "Ahmed Al-Rashid", "2026-01-15"])
    ws.append(["SKU-001", "Olive Oil 500ml",  1,  15.50, "Sara Khalil",     "2026-01-20"])

    readme = wb.create_sheet("README")
    readme.column_dimensions["A"].width = 20
    readme.column_dimensions["B"].width = 70
    readme.append(["Column", "Rules"])
    readme["A1"].font = Font(bold=True)
    readme["B1"].font = Font(bold=True)
    rules = [
        ("SKU",      "Required. Must match a product SKU in the ERP (whitespace stripped). Numeric-looking SKUs are normalised (e.g. 12345.0 → 12345)."),
        ("Item",     "Optional. Product description — used only in error messages when the SKU is not found. Not stored on the invoice."),
        ("QTY",      "Required. Numeric, must be > 0. Decimals accepted."),
        ("Price",    "Required. Unit price at the time of sale. May differ from the current product price. Must be >= 0."),
        ("Customer", "Optional. Customer name. Leave blank for Walk-in Customer. If the name does not already exist, a new Customer record is created automatically."),
        ("Date",     "Required. Must be >= 2026-01-01. Accepted formats: YYYY-MM-DD, DD/MM/YYYY, MM/DD/YYYY. Excel date cells are also accepted."),
        ("", ""),
        ("Grouping", "Multiple rows with the same Customer + Date are combined into a single invoice. Leave Customer blank and they go to the Walk-in invoice for that date."),
        ("Dry run",  "Always preview with Dry run checked first. Uncheck Dry run only for the final confirmed import."),
    ]
    for key, val in rules:
        readme.append([key, val])
        if key:
            readme.cell(readme.max_row, 1).font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=sales_import_template.xlsx"},
    )


@router.get("/api/sales/batches")
async def list_import_batches(
    db: AsyncSession = Depends(get_async_session),
    _=Depends(get_current_user),
):
    """Return a summary of recent historical-import batches."""
    _r = await db.execute(
        select(
            Invoice.import_batch_id,
            Invoice.notes,
        )
        .where(Invoice.import_batch_id.isnot(None))
        .distinct()
    )
    batch_rows = _r.all()

    batches = []
    for (batch_id, notes) in batch_rows:
        _stats = await db.execute(
            select(Invoice).where(Invoice.import_batch_id == batch_id)
        )
        invs = _stats.scalars().all()
        total = sum(float(i.total) for i in invs)
        earliest = min((i.created_at for i in invs), default=None)
        batches.append({
            "batch_id":     batch_id,
            "filename":     _extract_filename(notes),
            "ran_on":       earliest.date().isoformat() if earliest else None,
            "invoice_count": len(invs),
            "total_value":  round(total, 2),
        })

    # Newest first
    batches.sort(key=lambda b: b["ran_on"] or "", reverse=True)
    return {"batches": batches}


def _extract_filename(notes: str | None) -> str:
    """Pull the filename out of 'Imported from <filename> on <date>'."""
    if not notes:
        return ""
    try:
        return notes.split("Imported from ")[1].split(" on ")[0]
    except (IndexError, AttributeError):
        return notes[:40] if notes else ""


@router.delete("/api/sales/batch/{batch_id}")
async def delete_import_batch(
    batch_id: str,
    db: AsyncSession = Depends(get_async_session),
    _=Depends(get_current_user),
):
    """Revert a historical-sales import batch.

    Deletes all invoices in the batch (cascades to invoice_items), removes any
    StockMoves and Journals that reference those invoice IDs, and (for
    with_stock_and_journals imports) restores product.stock.

    Also cascade-deletes auto-created products and customers that were stamped
    with this batch_id, but only if they have no references outside this batch.
    """
    _r = await db.execute(
        select(Invoice).where(Invoice.import_batch_id == batch_id)
    )
    invoices = _r.scalars().all()
    if not invoices:
        return {"ok": True, "deleted_invoices": 0}

    invoice_ids = [inv.id for inv in invoices]

    # 1. Restore stock for any stock moves that were written by this batch
    _sm = await db.execute(
        select(StockMove).where(
            StockMove.ref_type == "invoice",
            StockMove.ref_id.in_(invoice_ids),
        )
    )
    for move in _sm.scalars().all():
        _pr = await db.execute(
            select(Product).where(Product.id == move.product_id)
        )
        product = _pr.scalar_one_or_none()
        if product and move.qty is not None:
            product.stock = float(product.stock) - float(move.qty)

    # 2. Delete stock moves
    await db.execute(
        delete(StockMove).where(
            StockMove.ref_type == "invoice",
            StockMove.ref_id.in_(invoice_ids),
        )
    )

    # 3. Delete journals (cascade removes journal_entries)
    _jr = await db.execute(
        select(Journal).where(
            Journal.ref_type == "invoice",
            Journal.ref_id.in_(invoice_ids),
        )
    )
    for journal in _jr.scalars().all():
        await db.delete(journal)

    # 4. Delete invoices (cascade removes invoice_items)
    for inv in invoices:
        await db.delete(inv)

    # Flush so cascade deletes are visible to the reference-count queries below
    await db.flush()

    # 5. Cascade-delete auto-created products that now have zero references
    _ap = await db.execute(
        select(Product).where(Product.created_by_import_batch == batch_id)
    )
    deleted_products = 0
    for product in _ap.scalars().all():
        pid = product.id
        refs = []
        for model, col in [
            (InvoiceItem,    InvoiceItem.product_id),
            (B2BInvoiceItem, B2BInvoiceItem.product_id),
            (PurchaseItem,   PurchaseItem.product_id),
            (StockMove,      StockMove.product_id),
        ]:
            _ref = await db.execute(select(model).where(col == pid).limit(1))
            refs.append(_ref.scalar_one_or_none())
        if all(r is None for r in refs):
            await db.delete(product)
            deleted_products += 1

    # 6. Cascade-delete auto-created customers that now have zero references
    _ac = await db.execute(
        select(Customer).where(Customer.created_by_import_batch == batch_id)
    )
    deleted_customers = 0
    for customer in _ac.scalars().all():
        cid = customer.id
        _inv_ref = await db.execute(select(Invoice).where(Invoice.customer_id == cid).limit(1))
        _ref_ref = await db.execute(select(RetailRefund).where(RetailRefund.customer_id == cid).limit(1))
        if _inv_ref.scalar_one_or_none() is None and _ref_ref.scalar_one_or_none() is None:
            await db.delete(customer)
            deleted_customers += 1

    await db.commit()
    return {
        "ok": True,
        "deleted_invoices":  len(invoices),
        "deleted_products":  deleted_products,
        "deleted_customers": deleted_customers,
    }


# ── B2B SALES IMPORT ────────────────────────────────────────────────────────

@router.post("/api/b2b-sales")
async def import_b2b_sales_endpoint(
    file: UploadFile = File(...),
    dry_run: bool = Form(True),
    mode: str = Form("history_only"),
    force: bool = Form(False),
    db: AsyncSession = Depends(get_async_session),
    current_user=Depends(get_current_user),
):
    contents = await file.read()
    return await import_b2b_sales(
        db=db,
        workbook_bytes=contents,
        filename=file.filename or "upload.xlsx",
        current_user_id=current_user.id,
        dry_run=dry_run,
        mode=mode,
        force=force,
    )


@router.get("/api/b2b-sales/template")
async def download_b2b_sales_template(_=Depends(get_current_user)):
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "B2B Sales"

    headers = ["SKU", "Item", "QTY", "Price", "Discount", "Payment type", "Client name", "Date"]
    hdr_font = Font(bold=True, color="4D9FFF")
    hdr_fill = PatternFill("solid", fgColor="0F1424")
    for col, h in enumerate(headers, 1):
        c = ws.cell(1, col, h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center")
    for col, w in enumerate([14, 28, 8, 10, 10, 16, 26, 14], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.append(["SKU-001", "Olive Oil 500ml",  50,  14.00, 10, "cash",         "Nile Grocery",     "2026-01-10"])
    ws.append(["SKU-002", "Tahini 250g",      100,  7.00,  5, "full_payment", "Cairo Mart",       "2026-01-12"])
    ws.append(["SKU-001", "Olive Oil 500ml",  30,  14.00, 10, "full_payment", "Cairo Mart",       "2026-01-12"])
    ws.append(["SKU-003", "Sesame Oil 250ml",  20,  9.50,  0, "consignment",  "Delta Foods Co.",  "2026-01-15"])

    readme = wb.create_sheet("README")
    readme.column_dimensions["A"].width = 20
    readme.column_dimensions["B"].width = 75
    readme.append(["Column", "Rules"])
    readme["A1"].font = Font(bold=True)
    readme["B1"].font = Font(bold=True)
    rules = [
        ("SKU",          "Required. Matched by normalised SKU. If not found, the product is auto-created for this import."),
        ("Item",         "Optional. Used for auto-created product name when SKU is new; otherwise a placeholder Imported Product <SKU> is used."),
        ("QTY",          "Required. Numeric, must be > 0. Decimals accepted."),
        ("Price",        "Required. Unit price BEFORE discount. Must be >= 0."),
        ("Discount",     "Optional. Per-line discount PERCENTAGE (e.g. 10 = 10%). Blank/null = 0. Range 0–100."),
        ("Payment type", "Required. One of: cash, full_payment, consignment (or aliases). "
                         "cash aliases: paid, cod, immediate. "
                         "full_payment aliases: credit, net30, on account, invoiced. "
                         "consignment aliases: cons, sale or return, sor."),
        ("Client name",  "Required. B2B client name. Auto-created if not found."),
        ("Date",         "Required. Must be >= 2026-01-01. Accepted formats: YYYY-MM-DD, DD/MM/YYYY, MM/DD/YYYY. Excel date cells also accepted."),
        ("", ""),
        ("Grouping",     "Rows with same Client name + Date + Payment type become one invoice. "
                         "Same client can have cash + consignment rows on the same day — those create two separate invoices."),
        ("Dry run",      "Always preview with Dry run checked first. Uncheck only for the final confirmed import."),
    ]
    for key, val in rules:
        readme.append([key, val])
        if key:
            readme.cell(readme.max_row, 1).font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=b2b_sales_import_template.xlsx"},
    )


@router.get("/api/b2b-sales/batches")
async def list_b2b_import_batches(
    db: AsyncSession = Depends(get_async_session),
    _=Depends(get_current_user),
):
    _r = await db.execute(
        select(B2BInvoiceModel.import_batch_id, B2BInvoiceModel.notes)
        .where(B2BInvoiceModel.import_batch_id.isnot(None))
        .distinct(B2BInvoiceModel.import_batch_id)
    )
    batch_rows = _r.all()
    batches = []
    for (batch_id, notes) in batch_rows:
        _stats = await db.execute(
            select(B2BInvoiceModel).where(B2BInvoiceModel.import_batch_id == batch_id)
        )
        invs = _stats.scalars().all()
        total = sum(float(i.total) for i in invs)
        earliest = min((i.created_at for i in invs), default=None)
        batches.append({
            "batch_id":      batch_id,
            "filename":      _extract_filename(notes),
            "ran_on":        earliest.date().isoformat() if earliest else None,
            "invoice_count": len(invs),
            "total_value":   round(total, 2),
        })
    batches.sort(key=lambda b: b["ran_on"] or "", reverse=True)
    return {"batches": batches}


@router.delete("/api/b2b-sales/batch/{batch_id}")
async def delete_b2b_import_batch(
    batch_id: str,
    db: AsyncSession = Depends(get_async_session),
    _=Depends(get_current_user),
):
    """Revert a B2B historical-import batch.

    - Deletes b2b_invoices + b2b_invoice_items
    - Deletes consignments + consignment_items linked to those invoices
    - Reverses Journal account balances and deletes journals/entries
    - Reverses client.outstanding for full_payment/consignment invoices
    - Reverses stock if with_stock_adjustment mode was used (detected via StockMove rows)
    - Does NOT revert client.discount_pct or B2BClientPrice (durable settings)
    """
    from sqlalchemy.orm import selectinload
    from app.models.accounting import Journal, JournalEntry
    from app.models.b2b import B2BClient, B2BInvoiceItem, ConsignmentItem

    _r = await db.execute(
        select(B2BInvoiceModel)
        .where(B2BInvoiceModel.import_batch_id == batch_id)
        .options(selectinload(B2BInvoiceModel.items))
    )
    invoices = _r.scalars().all()
    if not invoices:
        return {"ok": True, "deleted_invoices": 0, "deleted_consignments": 0}

    invoice_ids = [inv.id for inv in invoices]

    # 1. Reverse client.outstanding for unpaid invoices
    client_adjustments: dict[int, float] = {}
    for inv in invoices:
        if inv.invoice_type in ("full_payment", "consignment"):
            unpaid = max(0.0, float(inv.total) - float(inv.amount_paid))
            client_adjustments[inv.client_id] = (
                client_adjustments.get(inv.client_id, 0.0) + unpaid
            )
    for client_id, delta in client_adjustments.items():
        _c = await db.execute(select(B2BClient).where(B2BClient.id == client_id))
        client = _c.scalar_one_or_none()
        if client:
            client.outstanding = Decimal(str(max(0.0, float(client.outstanding) - delta)))

    # 2. Reverse stock moves
    _sm = await db.execute(
        select(StockMove).where(
            StockMove.ref_type == "b2b",
            StockMove.ref_id.in_(invoice_ids),
        )
    )
    for move in _sm.scalars().all():
        _pr = await db.execute(select(Product).where(Product.id == move.product_id))
        product = _pr.scalar_one_or_none()
        if product and move.qty is not None:
            product.stock = float(product.stock) - float(move.qty)
    await db.execute(
        delete(StockMove).where(
            StockMove.ref_type == "b2b",
            StockMove.ref_id.in_(invoice_ids),
        )
    )

    # 3. Reverse journal account balances and delete journals
    _jr = await db.execute(
        select(Journal)
        .options(selectinload(Journal.entries).selectinload(JournalEntry.account))
        .where(Journal.ref_type == "b2b", Journal.ref_id.in_(invoice_ids))
    )
    for journal in _jr.scalars().all():
        for entry in journal.entries:
            if entry.account:
                entry.account.balance -= (
                    Decimal(str(entry.debit)) - Decimal(str(entry.credit))
                )
        await db.delete(journal)

    # 4. Delete consignments + items (by batch_id first, then by invoice linkage)
    _cons = await db.execute(
        select(ConsignmentModel)
        .options(selectinload(ConsignmentModel.items))
        .where(ConsignmentModel.import_batch_id == batch_id)
    )
    cons_list = _cons.scalars().all()
    cons_deleted = len(cons_list)
    for cons in cons_list:
        for ci in cons.items:
            await db.delete(ci)
        await db.delete(cons)

    # 5. Delete invoices (cascade removes invoice items via ORM relationship)
    for inv in invoices:
        await db.delete(inv)

    await db.commit()
    return {
        "ok": True,
        "deleted_invoices":     len(invoices),
        "deleted_consignments": cons_deleted,
        "note": "client.discount_pct and B2BClientPrice entries were NOT reverted (durable settings).",
    }


# ── UI ─────────────────────────────────────────────────
@router.post("/api/farm-intake")
async def import_farm_intake_endpoint(
    file: UploadFile = File(...),
    dry_run: bool = Form(True),
    record_stock_movement: bool = Form(True),
    db: AsyncSession = Depends(get_async_session),
    current_user=Depends(get_current_user),
):
    contents = await file.read()
    return await import_farm_intake(
        db=db,
        workbook_bytes=contents,
        filename=file.filename or "farm_intake.xlsx",
        current_user_id=current_user.id,
        dry_run=dry_run,
        record_stock_movement=record_stock_movement,
    )


@router.get("/api/farm-intake/template")
async def download_farm_intake_template(_=Depends(get_current_user)):
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Farm Intake"

    headers = ["SKU", "product", "QTY", "Farm", "Date"]
    hdr_font = Font(bold=True, color="00FF9D")
    hdr_fill = PatternFill("solid", fgColor="0F1424")
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")
    for col, width in enumerate([14, 28, 10, 22, 14], 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.append(["SKU-001", "Tomatoes", 120, "Organic Farm", "2026-04-10"])
    ws.append(["SKU-002", "Cucumbers", 75, "Organic Farm", "2026-04-10"])
    ws.append(["SKU-003", "Mint", 24.5, "New Valley Farm", "2026-04-11"])

    readme = wb.create_sheet("README")
    readme.column_dimensions["A"].width = 20
    readme.column_dimensions["B"].width = 78
    readme.append(["Column", "Rules"])
    readme["A1"].font = Font(bold=True)
    readme["B1"].font = Font(bold=True)
    rules = [
        ("SKU", "Required. Used as the main product match key. Must match an existing product SKU. Numeric-looking SKUs are normalized."),
        ("product", "Optional. Informational display name used only in preview/error messages."),
        ("QTY", "Required. Numeric and must be greater than 0. Decimals are accepted."),
        ("Farm", "Required. Existing farms are reused. Unknown farm names are auto-created during import."),
        ("Date", "Required. Accepted formats: YYYY-MM-DD, DD/MM/YYYY, MM/DD/YYYY. Excel date cells are also accepted."),
        ("", ""),
        ("Grouping", "Rows with the same Farm + Date are grouped into one Farm Intake delivery record."),
        ("Stock moves", "Keep 'Record stock movement / inventory history' checked to match normal Farm Intake behaviour. Uncheck it only when stock was already recorded elsewhere."),
        ("Dry run", "Always preview with Dry run checked first. Uncheck Dry run only when you are ready to save."),
    ]
    for key, value in rules:
        readme.append([key, value])
        if key:
            readme.cell(readme.max_row, 1).font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=farm_intake_import_template.xlsx"},
    )


@router.post("/api/expenses")
async def import_expenses_endpoint(
    file: UploadFile = File(...),
    dry_run: bool = Form(True),
    db: AsyncSession = Depends(get_async_session),
    current_user=Depends(get_current_user),
):
    contents = await file.read()
    return await import_expenses(
        db=db,
        workbook_bytes=contents,
        filename=file.filename or "expenses.xlsx",
        current_user=current_user,
        dry_run=dry_run,
    )


@router.get("/api/expenses/template")
async def download_expenses_template(_=Depends(get_current_user)):
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expenses"

    headers = ["Category", "Amount", "Farm", "Date"]
    hdr_font = Font(bold=True, color="FB923C")
    hdr_fill = PatternFill("solid", fgColor="0F1424")
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")
    for col, width in enumerate([24, 14, 24, 14], 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.append(["Fuel", 850.50, "North Farm", "2026-04-10"])
    ws.append(["Office Supplies", 120.00, "", "2026-04-11"])

    readme = wb.create_sheet("README")
    readme.column_dimensions["A"].width = 20
    readme.column_dimensions["B"].width = 78
    readme.append(["Column", "Rules"])
    readme["A1"].font = Font(bold=True)
    readme["B1"].font = Font(bold=True)
    rules = [
        ("Category", "Required. Matched case-insensitively after trimming whitespace. Missing categories are auto-created during import."),
        ("Amount", "Required. Must be numeric and greater than 0. Excel numeric cells and common formatted values are accepted."),
        ("Farm", "Optional. Existing farms are matched case-insensitively. Leave blank to record the row as General Expense."),
        ("Date", "Required. Accepted formats: YYYY-MM-DD, DD/MM/YYYY, MM/DD/YYYY. Excel date cells are also accepted."),
        ("", ""),
        ("General Expense", "A blank Farm keeps the expense unassigned to a farm by saving farm_id as empty/null."),
        ("Dry run", "Preview with Dry run checked first. Uncheck Dry run only when you are ready to save the expenses."),
    ]
    for key, value in rules:
        readme.append([key, value])
        if key:
            readme.cell(readme.max_row, 1).font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=expenses_import_template.xlsx"},
    )


@router.get("/api/farm-intake/batches")
async def list_farm_batches(
    db: AsyncSession = Depends(get_async_session),
    _=Depends(get_current_user),
):
    return await list_farm_intake_import_batches(db)


@router.delete("/api/farm-intake/batch/{batch_id}")
async def delete_farm_batch(
    batch_id: str,
    db: AsyncSession = Depends(get_async_session),
    _=Depends(get_current_user),
):
    return await revert_farm_intake_import_batch(db, batch_id)


@router.get("/", response_class=HTMLResponse)
def import_ui():
    return """<!DOCTYPE html>
<html>
<head>
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Import Data — Thunder ERP</title>
<link href="https://fonts.googleapis.com/css2?family=Outfit:wght@400;500;600;700;800;900&family=JetBrains+Mono:wght@400;500;700&display=swap" rel="stylesheet">
<style>
:root{--bg:#060810;--card:#0f1424;--card2:#151c30;--border:rgba(255,255,255,0.06);--border2:rgba(255,255,255,0.11);--green:#00ff9d;--blue:#4d9fff;--orange:#fb923c;--teal:#2dd4bf;--danger:#ff4d6d;--warn:#ffb547;--lime:#84cc16;--purple:#a855f7;--text:#f0f4ff;--sub:#8899bb;--muted:#445066;--sans:'Outfit',sans-serif;--mono:'JetBrains Mono',monospace;--r:12px;}
body.light{
    --bg:#f4f5ef;--surface:#f1f3eb;--card:#eceee6;--card2:#e4e6de;
    --border:rgba(0,0,0,0.08);--border2:rgba(0,0,0,0.14);
    --green:#0f8a43;
    --text:#1a1e14;--sub:#4a5040;--muted:#7b816f;
}
body.light nav{background:rgba(244,245,239,.92);}
body.light .nav-link:hover{background:rgba(0,0,0,.05);}
body.light tr:hover td{background:rgba(0,0,0,.03);}
.mode-btn{display:flex;align-items:center;justify-content:center;width:36px;height:36px;border-radius:10px;border:1px solid var(--border);background:var(--card);color:var(--sub);font-size:16px;cursor:pointer;transition:all .2s;font-family:var(--sans);}
.mode-btn:hover{border-color:var(--border2);transform:scale(1.06);}
.topbar-right{display:flex;align-items:center;gap:12px;}
.user-pill{display:flex;align-items:center;gap:10px;background:var(--card);border:1px solid var(--border);border-radius:40px;padding:7px 16px 7px 10px;}
.user-avatar{width:28px;height:28px;background:linear-gradient(135deg,#7ecb6f,#d4a256);border-radius:50%;display:flex;align-items:center;justify-content:center;font-size:12px;font-weight:700;color:#0a0c08;}
.user-name{font-size:13px;font-weight:500;color:var(--sub);}
.logout-btn{background:transparent;border:1px solid var(--border);color:var(--muted);font-family:var(--sans);font-size:12px;font-weight:500;padding:8px 16px;border-radius:8px;cursor:pointer;transition:all .2s;letter-spacing:.3px;}
.logout-btn:hover{border-color:#c97a7a;color:#c97a7a;}
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0;}
body{font-family:var(--sans);background:var(--bg);color:var(--text);min-height:100vh;font-size:14px;}
nav{position:sticky;top:0;z-index:100;display:flex;align-items:center;gap:8px;padding:0 24px;height:58px;background:rgba(10,13,24,.92);backdrop-filter:blur(20px);border-bottom:1px solid var(--border);}
.logo{font-size:17px;font-weight:900;background:linear-gradient(135deg,var(--green),var(--blue));-webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text;margin-right:10px;text-decoration:none;display:flex;align-items:center;gap:8px;}
.nav-link{padding:7px 12px;border-radius:8px;color:var(--sub);font-size:12px;font-weight:600;text-decoration:none;transition:all .2s;}
.nav-link:hover{background:rgba(255,255,255,.05);color:var(--text);}
.nav-link.active{background:rgba(77,159,255,.1);color:var(--blue);}
.nav-spacer{flex:1;}
.content{max-width:1100px;margin:0 auto;padding:32px 24px;display:flex;flex-direction:column;gap:24px;}
.page-title{font-size:24px;font-weight:800;letter-spacing:-.5px;}
.page-sub{color:var(--muted);font-size:13px;margin-top:3px;}
.import-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(320px,1fr));gap:20px;}
.import-card{background:var(--card);border:1px solid var(--border);border-radius:16px;overflow:hidden;display:flex;flex-direction:column;}
.import-card-header{padding:20px 22px 16px;border-bottom:1px solid var(--border);display:flex;align-items:center;gap:12px;}
.import-card-icon{width:42px;height:42px;border-radius:12px;display:flex;align-items:center;justify-content:center;font-size:20px;flex-shrink:0;}
.icon-products{background:rgba(132,204,22,.1);}
.icon-stock{background:rgba(45,212,191,.1);}
.icon-customers{background:rgba(77,159,255,.1);}
.import-card-title{font-size:15px;font-weight:800;}
.import-card-sub{font-size:12px;color:var(--muted);margin-top:2px;}
.import-card-body{padding:18px 22px;flex:1;display:flex;flex-direction:column;gap:14px;}
.col-map{background:var(--card2);border:1px solid var(--border);border-radius:10px;padding:12px 14px;}
.col-map-title{font-size:10px;font-weight:700;letter-spacing:1.5px;text-transform:uppercase;color:var(--muted);margin-bottom:10px;}
.col-row{display:flex;align-items:center;gap:8px;margin-bottom:6px;font-size:12px;}
.col-row:last-child{margin-bottom:0;}
.col-excel{font-family:var(--mono);color:var(--lime);font-size:11px;background:rgba(132,204,22,.08);padding:2px 7px;border-radius:4px;white-space:nowrap;}
.col-arrow{color:var(--muted);font-size:10px;}
.col-field{color:var(--sub);}
.col-opt{color:var(--muted);font-size:10px;font-style:italic;}
.drop-zone{border:2px dashed var(--border2);border-radius:12px;padding:28px 20px;text-align:center;cursor:pointer;transition:all .2s;position:relative;}
.drop-zone:hover,.drop-zone.drag-over{border-color:var(--blue);background:rgba(77,159,255,.04);}
.drop-zone input[type=file]{position:absolute;inset:0;opacity:0;cursor:pointer;width:100%;height:100%;}
.drop-icon{font-size:28px;margin-bottom:8px;}
.drop-text{font-size:13px;font-weight:600;color:var(--sub);}
.drop-hint{font-size:11px;color:var(--muted);margin-top:4px;}
.preview-wrap{overflow-x:auto;border:1px solid var(--border);border-radius:10px;}
.preview-info{font-size:12px;color:var(--muted);padding:8px 12px;border-bottom:1px solid var(--border);display:flex;justify-content:space-between;align-items:center;}
table{width:100%;border-collapse:collapse;font-size:12px;}
thead{background:var(--card2);}
th{text-align:left;font-size:10px;font-weight:700;letter-spacing:1px;text-transform:uppercase;color:var(--muted);padding:8px 12px;white-space:nowrap;}
td{padding:8px 12px;border-top:1px solid var(--border);color:var(--sub);white-space:nowrap;max-width:160px;overflow:hidden;text-overflow:ellipsis;}
.import-btn{width:100%;padding:12px;border-radius:var(--r);font-family:var(--sans);font-size:14px;font-weight:700;cursor:pointer;border:none;transition:all .2s;display:flex;align-items:center;justify-content:center;gap:8px;}
.import-btn:disabled{opacity:.4;cursor:not-allowed;}
.btn-lime{background:linear-gradient(135deg,var(--lime),var(--green));color:#0a1a00;}
.btn-teal{background:linear-gradient(135deg,var(--teal),var(--blue));color:#001a18;}
.btn-blue{background:linear-gradient(135deg,var(--blue),var(--purple));color:white;}
.btn-lime:not(:disabled):hover,.btn-teal:not(:disabled):hover,.btn-blue:not(:disabled):hover{filter:brightness(1.1);transform:translateY(-1px);}
.result-box{border-radius:10px;padding:12px 16px;font-size:13px;font-weight:600;display:none;}
.result-ok{background:rgba(0,255,157,.08);border:1px solid rgba(0,255,157,.2);color:var(--green);}
.result-err{background:rgba(255,77,109,.08);border:1px solid rgba(255,77,109,.2);color:var(--danger);}
.result-warn{background:rgba(255,181,71,.08);border:1px solid rgba(255,181,71,.2);color:var(--warn);}
.not-found-list{margin-top:6px;font-size:11px;font-weight:400;color:var(--muted);max-height:80px;overflow-y:auto;}
.progress-wrap{height:4px;background:var(--border2);border-radius:4px;overflow:hidden;display:none;}
.progress-fill{height:100%;border-radius:4px;transition:width .3s;background:linear-gradient(90deg,var(--green),var(--lime));}
.section-label{font-size:11px;font-weight:700;letter-spacing:2px;text-transform:uppercase;color:var(--muted);display:flex;align-items:center;gap:10px;}
.section-label::after{content:'';flex:1;height:1px;background:linear-gradient(90deg,var(--border2),transparent);}
::-webkit-scrollbar{width:4px;height:4px;}
::-webkit-scrollbar-thumb{background:var(--border2);border-radius:4px;}
</style>
    <script src="/static/auth-guard.js"></script>
</head>
<body>
<nav>
    <a href="/home" class="logo">
        <svg width="18" height="18" viewBox="0 0 24 24" fill="none"><polygon points="13,2 4,14 11,14 11,22 20,10 13,10" fill="#f59e0b"/></svg>
        Thunder ERP
    </a>
    <a href="/dashboard" class="nav-link">Dashboard</a>
    <a href="/products/"  class="nav-link">Products</a>
    <a href="/import/"    class="nav-link active">Import</a>
    <span class="nav-spacer"></span>
    <div class="topbar-right">
        <button class="mode-btn" id="mode-btn" onclick="toggleMode()" title="Toggle color mode">??</button>
        <div class="user-pill">
            <div class="user-avatar" id="user-avatar">A</div>
            <span class="user-name" id="user-name">Admin</span>
        </div>
        <button class="logout-btn" onclick="logout()">Sign out</button>
    </div>
</nav>

<div class="content">
    <div>
        <div class="page-title">Import Data</div>
        <div class="page-sub">Import or update products, stock, customers, and expenses from Excel (.xlsx)</div>
    </div>

    <div class="section-label">Products & Stock</div>
    <div class="import-grid">

        <!-- PRODUCTS -->
        <div class="import-card">
            <div class="import-card-header">
                <div class="import-card-icon icon-products">🌿</div>
                <div>
                    <div class="import-card-title">Import Products</div>
                    <div class="import-card-sub">Creates new products — updates existing by SKU</div>
                </div>
            </div>
            <div class="import-card-body">
                <div class="col-map">
                    <div class="col-map-title">Expected Excel Columns</div>
                    <div class="col-row"><span class="col-excel">SKU</span><span class="col-arrow">→</span><span class="col-field">Product SKU</span><span class="col-opt">(auto-generated if missing)</span></div>
                    <div class="col-row"><span class="col-excel">Item</span><span class="col-arrow">→</span><span class="col-field">Product Name</span><span style="color:var(--danger);font-size:10px;margin-left:4px">required</span></div>
                    <div class="col-row"><span class="col-excel">UOM</span><span class="col-arrow">→</span><span class="col-field">Unit</span><span class="col-opt">(gram / kg / pcs…)</span></div>
                    <div class="col-row"><span class="col-excel">Unit Cost</span><span class="col-arrow">→</span><span class="col-field">Cost Price</span></div>
                    <div class="col-row"><span class="col-excel">Sales price</span><span class="col-arrow">→</span><span class="col-field">Sale Price</span></div>
                    <div class="col-row"><span class="col-excel">Group</span><span class="col-arrow">→</span><span class="col-field">Category</span></div>
                    <div class="col-row"><span class="col-excel">Item Type</span><span class="col-arrow">→</span><span class="col-field">Raw / Finished / Fresh / Packing / Ingredient</span><span class="col-opt">(defaults to finished)</span></div>
                </div>
                <div class="drop-zone" id="drop-products" ondragover="onDrag(event,'products')" ondragleave="offDrag('products')" ondrop="onDrop(event,'products')">
                    <input type="file" accept=".xlsx,.xls" onchange="onFile(this,'products')">
                    <div class="drop-icon">📄</div>
                    <div class="drop-text">Click or drag products.xlsx here</div>
                    <div class="drop-hint" id="hint-products">Same SKU = update existing product</div>
                </div>
                <div class="progress-wrap" id="prog-products"><div class="progress-fill" id="progfill-products" style="width:0%"></div></div>
                <div id="preview-products"></div>
                <div class="result-box" id="res-products"></div>
                <button class="import-btn btn-lime" id="btn-products" onclick="doImport('products')" disabled>⬆ Import Products</button>
            </div>
        </div>

        <!-- STOCK -->
        <div class="import-card">
            <div class="import-card-header">
                <div class="import-card-icon icon-stock">📦</div>
                <div>
                    <div class="import-card-title">Import Stock on Hand</div>
                    <div class="import-card-sub">Updates current stock for existing products by SKU</div>
                </div>
            </div>
            <div class="import-card-body">
                <div class="col-map">
                    <div class="col-map-title">Expected Excel Columns</div>
                    <div class="col-row"><span class="col-excel">SKU</span><span class="col-arrow">→</span><span class="col-field">Must match existing product</span><span style="color:var(--danger);font-size:10px;margin-left:4px">required</span></div>
                    <div class="col-row"><span class="col-excel">Item</span><span class="col-arrow">→</span><span class="col-field">Fallback if no SKU match</span></div>
                    <div class="col-row"><span class="col-excel">Stock</span><span class="col-arrow">→</span><span class="col-field">New Stock Quantity</span><span style="color:var(--danger);font-size:10px;margin-left:4px">required</span></div>
                    <div style="margin-top:8px;font-size:11px;color:var(--warn);padding:8px 10px;background:rgba(255,181,71,.06);border-radius:6px;border:1px solid rgba(255,181,71,.15);">
                        ⚠ Sets stock to exact value. Import products first if they don't exist.
                    </div>
                </div>
                <div class="drop-zone" id="drop-stock" ondragover="onDrag(event,'stock')" ondragleave="offDrag('stock')" ondrop="onDrop(event,'stock')">
                    <input type="file" accept=".xlsx,.xls" onchange="onFile(this,'stock')">
                    <div class="drop-icon">📊</div>
                    <div class="drop-text">Click or drag SOH.xlsx here</div>
                    <div class="drop-hint" id="hint-stock">Overwrites current stock for matched SKUs</div>
                </div>
                <div class="progress-wrap" id="prog-stock"><div class="progress-fill" id="progfill-stock" style="width:0%"></div></div>
                <div id="preview-stock"></div>
                <div class="result-box" id="res-stock"></div>
                <button class="import-btn btn-teal" id="btn-stock" onclick="doImport('stock')" disabled>⬆ Import Stock</button>
            </div>
        </div>

    </div>

    <div class="section-label">Customers</div>
    <div class="import-grid" style="grid-template-columns:minmax(320px,500px)">

        <!-- CUSTOMERS -->
        <div class="import-card">
            <div class="import-card-header">
                <div class="import-card-icon icon-customers">👥</div>
                <div>
                    <div class="import-card-title">Import Customers</div>
                    <div class="import-card-sub">Skips duplicates by phone or name</div>
                </div>
            </div>
            <div class="import-card-body">
                <div class="col-map">
                    <div class="col-map-title">Expected Excel Columns</div>
                    <div class="col-row"><span class="col-excel">Name</span><span class="col-arrow">→</span><span class="col-field">Customer Name</span><span style="color:var(--danger);font-size:10px;margin-left:4px">required</span></div>
                    <div class="col-row"><span class="col-excel">Phone</span><span class="col-arrow">→</span><span class="col-field">Phone Number</span><span class="col-opt">(used for duplicate check)</span></div>
                    <div class="col-row"><span class="col-excel">Email</span><span class="col-arrow">→</span><span class="col-field">Email</span></div>
                    <div class="col-row"><span class="col-excel">Address</span><span class="col-arrow">→</span><span class="col-field">Address / Area</span></div>
                </div>
                <div class="drop-zone" id="drop-customers" ondragover="onDrag(event,'customers')" ondragleave="offDrag('customers')" ondrop="onDrop(event,'customers')">
                    <input type="file" accept=".xlsx,.xls" onchange="onFile(this,'customers')">
                    <div class="drop-icon">📋</div>
                    <div class="drop-text">Click or drag Customers.xlsx here</div>
                    <div class="drop-hint" id="hint-customers">Duplicates automatically skipped</div>
                </div>
                <div class="progress-wrap" id="prog-customers"><div class="progress-fill" id="progfill-customers" style="width:0%"></div></div>
                <div id="preview-customers"></div>
                <div class="result-box" id="res-customers"></div>
                <button class="import-btn btn-blue" id="btn-customers" onclick="doImport('customers')" disabled>⬆ Import Customers</button>
            </div>
        </div>

    </div>

    <div class="section-label">Expenses</div>
    <div class="import-grid" style="grid-template-columns:minmax(340px,680px)">
        <div class="import-card">
            <div class="import-card-header">
                <div class="import-card-icon" style="background:rgba(251,146,60,.12)">&#128184;</div>
                <div>
                    <div class="import-card-title">Expenses Import</div>
                    <div class="import-card-sub">Import expenses with preview, validation, and General Expense handling</div>
                </div>
            </div>
            <div class="import-card-body">
                <div class="col-map">
                    <div class="col-map-title">Expected Excel Columns</div>
                    <div class="col-row"><span class="col-excel">Category</span><span class="col-arrow">-></span><span class="col-field">Expense category</span><span style="color:var(--danger);font-size:10px;margin-left:4px">required</span></div>
                    <div class="col-row"><span class="col-excel">Amount</span><span class="col-arrow">-></span><span class="col-field">Expense amount</span><span style="color:var(--danger);font-size:10px;margin-left:4px">required</span></div>
                    <div class="col-row"><span class="col-excel">Farm</span><span class="col-arrow">-></span><span class="col-field">Existing farm name</span><span class="col-opt">blank becomes General Expense</span></div>
                    <div class="col-row"><span class="col-excel">Date</span><span class="col-arrow">-></span><span class="col-field">Expense date</span><span style="color:var(--danger);font-size:10px;margin-left:4px">required</span></div>
                    <div style="margin-top:8px;font-size:11px;color:var(--sub);padding:8px 10px;background:rgba(251,146,60,.08);border-radius:6px;border:1px solid rgba(251,146,60,.18);">
                        Existing categories are reused case-insensitively. Missing categories are auto-created. Blank Farm rows are imported as General Expense.
                    </div>
                    <div style="margin-top:6px">
                        <a href="/import/api/expenses/template" download style="font-size:11px;color:var(--blue);text-decoration:none">Download Expenses template</a>
                    </div>
                </div>

                <div style="display:flex;flex-direction:column;gap:8px;">
                    <label style="display:flex;align-items:center;gap:8px;cursor:pointer;font-size:13px;color:var(--sub)">
                        <input type="checkbox" id="chk-expenses-dryrun" checked style="accent-color:var(--orange)">
                        <span><b style="color:var(--text)">Dry run</b> - preview without saving (recommended first step)</span>
                    </label>
                </div>

                <div class="drop-zone" id="drop-expenses" ondragover="onDrag(event,'expenses')" ondragleave="offDrag('expenses')" ondrop="onDrop(event,'expenses')">
                    <input type="file" accept=".xlsx,.xls" onchange="onFile(this,'expenses')">
                    <div class="drop-icon">&#128184;</div>
                    <div class="drop-text">Click or drag expenses.xlsx here</div>
                    <div class="drop-hint" id="hint-expenses">Blank Farm rows are recorded as General Expense</div>
                </div>
                <div class="progress-wrap" id="prog-expenses"><div class="progress-fill" id="progfill-expenses" style="width:0%"></div></div>
                <div id="preview-expenses"></div>
                <div class="result-box" id="res-expenses"></div>
                <button class="import-btn" style="background:linear-gradient(135deg,var(--orange),#ffd166);color:#2e1300" id="btn-expenses" onclick="doImportExpenses()" disabled>Import Expenses</button>
            </div>
        </div>
    </div>

    <div class="section-label">Historical Sales</div>
    <div class="import-grid" style="grid-template-columns:minmax(340px,680px)">

        <!-- HISTORICAL SALES -->
        <div class="import-card">
            <div class="import-card-header">
                <div class="import-card-icon" style="background:rgba(251,146,60,.1)">🧾</div>
                <div>
                    <div class="import-card-title">Historical Sales</div>
                    <div class="import-card-sub">Import past retail sales from Excel — one-time backfill</div>
                </div>
            </div>
            <div class="import-card-body">
                <div class="col-map">
                    <div class="col-map-title">Expected Excel Columns</div>
                    <div class="col-row"><span class="col-excel">SKU</span><span class="col-arrow">→</span><span class="col-field">Product SKU — auto-created if unknown</span><span style="color:var(--danger);font-size:10px;margin-left:4px">required</span></div>
                    <div class="col-row"><span class="col-excel">Item</span><span class="col-arrow">→</span><span class="col-field">Product name (used when auto-creating)</span><span class="col-opt">optional</span></div>
                    <div class="col-row"><span class="col-excel">QTY</span><span class="col-arrow">→</span><span class="col-field">Quantity sold (&gt; 0)</span><span style="color:var(--danger);font-size:10px;margin-left:4px">required</span></div>
                    <div class="col-row"><span class="col-excel">Price</span><span class="col-arrow">→</span><span class="col-field">Unit price at time of sale</span><span style="color:var(--danger);font-size:10px;margin-left:4px">required</span></div>
                    <div class="col-row"><span class="col-excel">Customer</span><span class="col-arrow">→</span><span class="col-field">Customer name — auto-created if unknown (blank = Walk-in)</span><span class="col-opt">optional</span></div>
                    <div class="col-row"><span class="col-excel">Date</span><span class="col-arrow">→</span><span class="col-field">Sale date ≥ 2026-01-01</span><span style="color:var(--danger);font-size:10px;margin-left:4px">required</span></div>
                    <div style="margin-top:8px;font-size:11px;color:var(--sub);padding:8px 10px;background:rgba(77,159,255,.06);border-radius:6px;border:1px solid rgba(77,159,255,.15);">
                        Unknown SKUs and unknown customers will be automatically created from the sheet data.
                        Auto-created products get <code style="font-family:var(--mono)">cost = 0</code> by default — set this via the 'Default cost ratio' option below, or adjust after import in Products → "Imported - Historical".
                    </div>
                    <div style="margin-top:6px">
                        <a href="/import/api/sales/template" download style="font-size:11px;color:var(--blue);text-decoration:none">⬇ Download template</a>
                    </div>
                </div>

                <!-- Mode radios -->
                <div style="background:var(--card2);border:1px solid var(--border);border-radius:10px;padding:12px 14px;">
                    <div class="col-map-title" style="margin-bottom:10px">Import Mode</div>
                    <label style="display:flex;align-items:flex-start;gap:8px;margin-bottom:10px;cursor:pointer;font-size:13px;color:var(--sub)">
                        <input type="radio" name="sales-mode" value="history_only" checked style="margin-top:2px;accent-color:var(--green)">
                        <span><b style="color:var(--text)">Reports only</b> — recommended<br><span style="font-size:11px;color:var(--muted)">Creates Invoice records only. No stock adjustment. No accounting entries.</span></span>
                    </label>
                    <label style="display:flex;align-items:flex-start;gap:8px;margin-bottom:10px;cursor:pointer;font-size:13px;color:var(--sub)">
                        <input type="radio" name="sales-mode" value="with_journals" style="margin-top:2px;accent-color:var(--blue)">
                        <span><b style="color:var(--text)">Also post to accounting</b><br><span style="font-size:11px;color:var(--muted)">Posts 1000/4000 journal entries dated to the historical sale date.</span></span>
                    </label>
                    <label style="display:flex;align-items:flex-start;gap:8px;cursor:pointer;font-size:13px;color:var(--sub)">
                        <input type="radio" name="sales-mode" value="with_stock_and_journals" style="margin-top:2px;accent-color:var(--danger)">
                        <span><b style="color:var(--danger)">Also adjust stock — use with extreme caution</b><br><span style="font-size:11px;color:var(--muted)">Full live-sale behaviour: decrements stock, writes StockMoves, posts journals. Almost never appropriate for backfills.</span></span>
                    </label>
                </div>

                <!-- Options -->
                <div style="display:flex;flex-direction:column;gap:8px;">
                    <label style="display:flex;align-items:center;gap:8px;cursor:pointer;font-size:13px;color:var(--sub)">
                        <input type="checkbox" id="chk-dryrun" checked style="accent-color:var(--green)">
                        <span><b style="color:var(--text)">Dry run</b> — preview without saving (recommended first step)</span>
                    </label>
                    <label style="display:flex;align-items:center;gap:8px;cursor:pointer;font-size:13px;color:var(--sub)">
                        <input type="checkbox" id="chk-force" style="accent-color:var(--warn)">
                        <span>Force import even if duplicates detected</span>
                    </label>
                    <div style="display:flex;flex-direction:column;gap:4px;">
                        <label style="font-size:12px;font-weight:600;color:var(--sub)">
                            Default cost ratio for auto-created products
                            <span style="font-weight:400;color:var(--muted)">(optional)</span>
                        </label>
                        <input type="number" id="inp-cost-ratio" step="0.01" min="0" max="1"
                            placeholder="e.g. 0.65 — leave blank for cost = 0"
                            style="background:var(--card2);border:1px solid var(--border2);border-radius:8px;
                            padding:8px 12px;color:var(--text);font-family:var(--sans);font-size:13px;width:100%;">
                        <span style="font-size:11px;color:var(--muted)">
                            If set, auto-created products get cost = price × this ratio. Leave blank for cost = 0.
                        </span>
                    </div>
                </div>

                <div class="drop-zone" id="drop-sales" ondragover="onDrag(event,'sales')" ondragleave="offDrag('sales')" ondrop="onDrop(event,'sales')">
                    <input type="file" accept=".xlsx,.xls" onchange="onFile(this,'sales')">
                    <div class="drop-icon">🧾</div>
                    <div class="drop-text">Click or drag sales.xlsx here</div>
                    <div class="drop-hint" id="hint-sales">Rows with same Customer + Date become one invoice</div>
                </div>
                <div class="progress-wrap" id="prog-sales"><div class="progress-fill" id="progfill-sales" style="width:0%"></div></div>
                <div id="preview-sales"></div>
                <div class="result-box" id="res-sales"></div>
                <button class="import-btn" style="background:linear-gradient(135deg,var(--orange),var(--warn));color:#1a0a00" id="btn-sales" onclick="doImportSales()" disabled>⬆ Import Sales</button>
            </div>
        </div>
    </div>

    <!-- Recent Retail Batches -->
    <div class="section-label">Recent Sales Import Batches</div>
    <div id="batches-panel" style="background:var(--card);border:1px solid var(--border);border-radius:12px;padding:18px 20px;">
        <div style="color:var(--muted);font-size:13px">Loading…</div>
    </div>

<div class="section-label" style="margin-top:32px">Farm Intake</div>
    <div class="import-grid" style="grid-template-columns:minmax(340px,760px)">
        <div class="import-card">
            <div class="import-card-header">
                <div class="import-card-icon" style="background:rgba(45,212,191,.12)">&#127806;</div>
                <div>
                    <div class="import-card-title">Farm Intake Import</div>
                    <div class="import-card-sub">Import farm delivery history from Excel and optionally record stock movement</div>
                </div>
            </div>
            <div class="import-card-body">
                <div class="col-map">
                    <div class="col-map-title">Expected Excel Columns</div>
                    <div class="col-row"><span class="col-excel">SKU</span><span class="col-arrow">-></span><span class="col-field">Existing product SKU</span><span style="color:var(--danger);font-size:10px;margin-left:4px">required</span></div>
                    <div class="col-row"><span class="col-excel">product</span><span class="col-arrow">-></span><span class="col-field">Product name / display hint</span><span class="col-opt">optional</span></div>
                    <div class="col-row"><span class="col-excel">QTY</span><span class="col-arrow">-></span><span class="col-field">Delivered quantity (&gt; 0)</span><span style="color:var(--danger);font-size:10px;margin-left:4px">required</span></div>
                    <div class="col-row"><span class="col-excel">Farm</span><span class="col-arrow">-></span><span class="col-field">Farm name - auto-created if new</span><span style="color:var(--danger);font-size:10px;margin-left:4px">required</span></div>
                    <div class="col-row"><span class="col-excel">Date</span><span class="col-arrow">-></span><span class="col-field">Farm intake date</span><span style="color:var(--danger);font-size:10px;margin-left:4px">required</span></div>
                    <div style="margin-top:8px;font-size:11px;color:var(--sub);padding:8px 10px;background:rgba(45,212,191,.06);border-radius:6px;border:1px solid rgba(45,212,191,.15);">
                        Rows with the same Farm + Date are grouped into one farm delivery record so they appear like manual Farm Intake entries.
                    </div>
                    <div style="margin-top:6px">
                        <a href="/import/api/farm-intake/template" download style="font-size:11px;color:var(--blue);text-decoration:none">Download Farm Intake template</a>
                    </div>
                </div>

                <div style="display:flex;flex-direction:column;gap:8px;">
                    <label style="display:flex;align-items:center;gap:8px;cursor:pointer;font-size:13px;color:var(--sub)">
                        <input type="checkbox" id="chk-farm-dryrun" checked style="accent-color:var(--teal)">
                        <span><b style="color:var(--text)">Dry run</b> - preview without saving (recommended first step)</span>
                    </label>
                    <label style="display:flex;align-items:flex-start;gap:8px;cursor:pointer;font-size:13px;color:var(--sub)">
                        <input type="checkbox" id="chk-farm-stock" checked style="margin-top:2px;accent-color:var(--green)">
                        <span><b style="color:var(--text)">Record stock movement / inventory history</b><br><span style="font-size:11px;color:var(--muted)">Checked by default to match normal Farm Intake behaviour. Uncheck only if stock was already recorded elsewhere.</span></span>
                    </label>
                </div>

                <div class="drop-zone" id="drop-farm-intake" ondragover="onDrag(event,'farm-intake')" ondragleave="offDrag('farm-intake')" ondrop="onDrop(event,'farm-intake')">
                    <input type="file" accept=".xlsx,.xls" onchange="onFile(this,'farm-intake')">
                    <div class="drop-icon">&#127806;</div>
                    <div class="drop-text">Click or drag farm_intake.xlsx here</div>
                    <div class="drop-hint" id="hint-farm-intake">Grouped by Farm + Date into delivery records</div>
                </div>
                <div class="progress-wrap" id="prog-farm-intake"><div class="progress-fill" id="progfill-farm-intake" style="width:0%"></div></div>
                <div id="preview-farm-intake"></div>
                <div class="result-box" id="res-farm-intake"></div>
                <button class="import-btn" style="background:linear-gradient(135deg,var(--teal),var(--green));color:#042118" id="btn-farm-intake" onclick="doImportFarmIntake()" disabled>Import Farm Intake</button>
            </div>
        </div>
    </div>

    <div class="section-label">Recent Farm Intake Import Batches</div>
    <div id="farm-batches-panel" style="background:var(--card);border:1px solid var(--border);border-radius:12px;padding:18px 20px;">
        <div style="color:var(--muted);font-size:13px">Loading...</div>
    </div>

    <div class="section-label" style="margin-top:32px">Historical B2B Sales</div>
    <div class="import-grid" style="grid-template-columns:minmax(340px,780px)">

        <!-- HISTORICAL B2B SALES -->
        <div class="import-card">
            <div class="import-card-header">
                <div class="import-card-icon" style="background:rgba(77,159,255,.10)">🏭</div>
                <div>
                    <div class="import-card-title">Historical B2B Sales</div>
                    <div class="import-card-sub">Import wholesale/B2B sales from Excel with per-line discounts and payment types</div>
                </div>
            </div>
            <div class="import-card-body">
                <div class="col-map">
                    <div class="col-map-title">Expected Excel Columns</div>
                    <div class="col-row"><span class="col-excel">SKU</span><span class="col-arrow">→</span><span class="col-field">Must match existing product</span><span style="color:var(--danger);font-size:10px;margin-left:4px">required</span></div>
                    <div class="col-row"><span class="col-excel">Item</span><span class="col-arrow">→</span><span class="col-field">Product name (informational)</span><span class="col-opt">optional</span></div>
                    <div class="col-row"><span class="col-excel">QTY</span><span class="col-arrow">→</span><span class="col-field">Quantity (&gt; 0)</span><span style="color:var(--danger);font-size:10px;margin-left:4px">required</span></div>
                    <div class="col-row"><span class="col-excel">Price</span><span class="col-arrow">→</span><span class="col-field">Unit price before discount</span><span style="color:var(--danger);font-size:10px;margin-left:4px">required</span></div>
                    <div class="col-row"><span class="col-excel">Discount</span><span class="col-arrow">→</span><span class="col-field">Per-line discount % (blank = 0)</span><span class="col-opt">optional</span></div>
                    <div class="col-row"><span class="col-excel">Payment type</span><span class="col-arrow">→</span><span class="col-field">cash · full_payment · consignment</span><span style="color:var(--danger);font-size:10px;margin-left:4px">required</span></div>
                    <div class="col-row"><span class="col-excel">Client name</span><span class="col-arrow">→</span><span class="col-field">B2B client (auto-created if new)</span><span style="color:var(--danger);font-size:10px;margin-left:4px">required</span></div>
                    <div class="col-row"><span class="col-excel">Date</span><span class="col-arrow">→</span><span class="col-field">Sale date ≥ 2026-01-01</span><span style="color:var(--danger);font-size:10px;margin-left:4px">required</span></div>
                    <div style="margin-top:8px">
                        <a href="/import/api/b2b-sales/template" download style="font-size:11px;color:var(--blue);text-decoration:none">⬇ Download B2B template</a>
                    </div>
                </div>

                <!-- Mode radios -->
                <div style="background:var(--card2);border:1px solid var(--border);border-radius:10px;padding:12px 14px;">
                    <div class="col-map-title" style="margin-bottom:10px">Import Mode</div>
                    <label style="display:flex;align-items:flex-start;gap:8px;margin-bottom:10px;cursor:pointer;font-size:13px;color:var(--sub)">
                        <input type="radio" name="b2b-mode" value="history_only" checked style="margin-top:2px;accent-color:var(--blue)">
                        <span><b style="color:var(--text)">History only</b> — recommended<br><span style="font-size:11px;color:var(--muted)">Creates B2B invoices + journal entries (AR/cash). No stock adjustment.</span></span>
                    </label>
                    <label style="display:flex;align-items:flex-start;gap:8px;cursor:pointer;font-size:13px;color:var(--sub)">
                        <input type="radio" name="b2b-mode" value="with_stock_adjustment" style="margin-top:2px;accent-color:var(--danger)">
                        <span><b style="color:var(--danger)">Also adjust stock — use with extreme caution</b><br><span style="font-size:11px;color:var(--muted)">Decrements products.stock and writes StockMoves. Only use if stock movement was NOT already recorded.</span></span>
                    </label>
                </div>

                <!-- Options -->
                <div style="display:flex;flex-direction:column;gap:8px;">
                    <label style="display:flex;align-items:center;gap:8px;cursor:pointer;font-size:13px;color:var(--sub)">
                        <input type="checkbox" id="chk-b2b-dryrun" checked style="accent-color:var(--blue)">
                        <span><b style="color:var(--text)">Dry run</b> — preview without saving (recommended first step)</span>
                    </label>
                    <label style="display:flex;align-items:center;gap:8px;cursor:pointer;font-size:13px;color:var(--sub)">
                        <input type="checkbox" id="chk-b2b-force" style="accent-color:var(--warn)">
                        <span>Force import even if duplicates detected</span>
                    </label>
                </div>

                <div class="drop-zone" id="drop-b2b-sales" ondragover="onDrag(event,'b2b-sales')" ondragleave="offDrag('b2b-sales')" ondrop="onDrop(event,'b2b-sales')">
                    <input type="file" accept=".xlsx,.xls" onchange="onFile(this,'b2b-sales')">
                    <div class="drop-icon">🏭</div>
                    <div class="drop-text">Click or drag b2b_sales.xlsx here</div>
                    <div class="drop-hint" id="hint-b2b-sales">Rows grouped by Client + Date + Payment type</div>
                </div>
                <div class="progress-wrap" id="prog-b2b-sales"><div class="progress-fill" id="progfill-b2b-sales" style="width:0%"></div></div>
                <div id="preview-b2b-sales"></div>
                <div class="result-box" id="res-b2b-sales"></div>
                <button class="import-btn" style="background:linear-gradient(135deg,var(--blue),#4d6fff);color:#fff" id="btn-b2b-sales" onclick="doImportB2B()" disabled>⬆ Import B2B Sales</button>
            </div>
        </div>
    </div>

    <!-- Recent B2B Batches -->
    <div class="section-label">Recent B2B Import Batches</div>
    <div id="b2b-batches-panel" style="background:var(--card);border:1px solid var(--border);border-radius:12px;padding:18px 20px;">
        <div style="color:var(--muted);font-size:13px">Loading…</div>
    </div>

</div>

<script>
  // Auth guard: redirect to login if the readable session cookie is absent
  function _hasAuthCookie() {
      return document.cookie.split(";").some(c => c.trim().startsWith("logged_in="));
  }
  if (!_hasAuthCookie()) { _redirectToLogin(); }

  function setModeButton(isLight){
    const btn = document.getElementById("mode-btn");
    if(btn) btn.innerText = isLight ? "☀️" : "🌙";
}
function toggleMode(){
    const isLight = document.body.classList.toggle("light");
    localStorage.setItem("colorMode", isLight ? "light" : "dark");
    setModeButton(isLight);
}
function initializeColorMode(){
    const isLight = localStorage.getItem("colorMode") === "light";
    document.body.classList.toggle("light", isLight);
    setModeButton(isLight);
}
async function initUser() {
    try {
        const r = await fetch("/auth/me");
        if (!r.ok) { _redirectToLogin(); return; }
        const u = await r.json();
        const nameEl = document.getElementById("user-name");
        const avatarEl = document.getElementById("user-avatar");
        if (nameEl) nameEl.innerText = u.name;
        if (avatarEl) avatarEl.innerText = u.name.charAt(0).toUpperCase();
        return u;
    } catch(e) { _redirectToLogin(); }
}
async function logout(){
    await fetch("/auth/logout", { method: "POST" });
    window.location.href = "/";
}
  initializeColorMode();
  initUser();
  const files = {products:null, stock:null, customers:null, expenses:null, sales:null, 'farm-intake':null, 'b2b-sales':null};

function onDrag(e,t){ e.preventDefault(); document.getElementById('drop-'+t).classList.add('drag-over'); }
function offDrag(t){ document.getElementById('drop-'+t).classList.remove('drag-over'); }
function onDrop(e,t){ e.preventDefault(); offDrag(t); let f=e.dataTransfer.files[0]; if(f) loadFile(f,t); }
function onFile(inp,t){ let f=inp.files[0]; if(f) loadFile(f,t); }

async function loadFile(f, type){
    files[type] = f;
    document.getElementById('drop-'+type).querySelector('.drop-text').innerText = f.name;
    document.getElementById('btn-'+type).disabled = false;
    showResult(type, '', '');

    let fd = new FormData(); fd.append('file', f);
    let prev = await (await fetch('/import/api/preview', {method:'POST', body:fd})).json();
    if(prev.headers){
        document.getElementById('hint-'+type).innerText = prev.total_rows + ' rows detected';
        document.getElementById('preview-'+type).innerHTML = `
            <div class="preview-wrap">
                <div class="preview-info">
                    <span>Preview — first 5 rows</span>
                    <span style="color:var(--lime);font-family:var(--mono)">${prev.total_rows} total rows</span>
                </div>
                <table><thead><tr>${prev.headers.map(h=>`<th>${h||'—'}</th>`).join('')}</tr></thead>
                <tbody>${prev.rows.map(r=>`<tr>${r.map(c=>`<td>${c}</td>`).join('')}</tr>`).join('')}</tbody>
                </table>
            </div>`;
    }
}

function showResult(type, msg, kind){
    let el = document.getElementById('res-'+type);
    if(!msg){ el.style.display='none'; return; }
    el.className = 'result-box result-'+kind;
    el.innerHTML = msg;
    el.style.display = 'block';
}

function showProg(type, pct){
    let w=document.getElementById('prog-'+type);
    let f=document.getElementById('progfill-'+type);
    w.style.display='block'; f.style.width=pct+'%';
    if(pct>=100) setTimeout(()=>{w.style.display='none';f.style.width='0%';},800);
}

async function doImport(type){
    let f = files[type];
    if(!f){ showResult(type,'Please select a file first','err'); return; }
    let btn = document.getElementById('btn-'+type);
    btn.disabled=true; btn.innerHTML='⏳ Importing…';
    showProg(type, 40); showResult(type,'','');

    let fd = new FormData(); fd.append('file', f);
    let res  = await fetch('/import/api/'+type, {method:'POST', body:fd});
    let data = await res.json();
    showProg(type, 100);

    let cap = type.charAt(0).toUpperCase()+type.slice(1);
    btn.disabled=false; btn.innerHTML='⬆ Import '+cap;

    if(data.error){ showResult(type, '✗ '+data.error, 'err'); return; }

    let msg='', kind='ok';
    if(type==='products'){
        msg = `✓ <b>${data.created}</b> products created &nbsp;·&nbsp; <b>${data.updated}</b> updated`;
        if(data.errors&&data.errors.length){ msg+=`<br><span style="font-size:11px;font-weight:400">${data.errors.join(', ')}</span>`; kind='warn'; }
    } else if(type==='stock'){
        msg = `✓ <b>${data.updated}</b> products stock updated`;
        if(data.not_found&&data.not_found.length){
            msg += `<br><b>${data.not_found.length} SKUs not found:</b>`;
            msg += `<div class="not-found-list">${data.not_found.join(', ')}</div>`;
            kind='warn';
        }
    } else {
        msg = `✓ <b>${data.created}</b> customers imported &nbsp;·&nbsp; <b>${data.skipped}</b> skipped`;
    }
    showResult(type, msg, kind);
}

// ── Historical Sales ────────────────────────────────────────────────────────

async function doImportExpenses() {
    const f = files['expenses'];
    if (!f) { showResult('expenses', 'Please select a file first', 'err'); return; }

    const btn = document.getElementById('btn-expenses');
    btn.disabled = true;
    btn.innerHTML = 'Processing...';
    showProg('expenses', 40);
    showResult('expenses', '', '');

    const dryRun = document.getElementById('chk-expenses-dryrun').checked;
    const fd = new FormData();
    fd.append('file', f);
    fd.append('dry_run', dryRun ? 'true' : 'false');

    let res, data;
    try {
        res = await fetch('/import/api/expenses', { method: 'POST', body: fd });
        data = await res.json();
    } catch (e) {
        showProg('expenses', 100);
        btn.disabled = false;
        btn.innerHTML = 'Import Expenses';
        showResult('expenses', 'Error: Network error: ' + e.message, 'err');
        return;
    }

    showProg('expenses', 100);
    btn.disabled = false;
    btn.innerHTML = 'Import Expenses';

    if (!res.ok) {
        const detail = data?.detail;
        const msg = Array.isArray(detail)
            ? detail.map(e => e.msg || JSON.stringify(e)).join('; ')
            : (detail || data?.error || `HTTP ${res.status}`);
        showResult('expenses', 'Error: ' + msg, 'err');
        return;
    }

    if (data.error) {
        showResult('expenses', 'Error: ' + data.error, 'err');
        return;
    }

    renderExpensesResult(data);
}

function renderExpensesResult(data) {
    if (!data || !data.summary) {
        showResult('expenses', 'Error: Unexpected response from server - no summary returned.', 'err');
        return;
    }
    const s = data.summary || {};
    const isDry = !!data.dry_run;
    const n = value => {
        const parsed = Number(value);
        return Number.isFinite(parsed) ? parsed : 0;
    };
    const txt = value => (value === undefined || value === null || value === '') ? '-' : String(value);

    let html = `<div style="font-size:13px">
        ${isDry ? '<span style="color:var(--warn)">Dry run - nothing was saved</span><br>' : ''}
        <b>${n(s.rows_read)}</b> rows read &nbsp;|&nbsp;
        <b>${n(isDry ? s.expenses_would_create : s.expenses_created)}</b> expense records ${isDry ? 'valid for import' : 'created'} &nbsp;|&nbsp;
        <b>${n(s.rows_skipped)}</b> skipped<br>
        <span style="color:var(--sub);font-size:12px">
            Categories ${isDry ? 'to auto-create' : 'auto-created'}: <b>${n(s.categories_auto_created)}</b> &nbsp;|&nbsp;
            Farm-linked rows: <b>${n(s.farms_resolved)}</b> &nbsp;|&nbsp;
            General Expense rows: <b>${n(s.general_expense_rows)}</b>
        </span><br>
        <span style="color:var(--sub);font-size:12px">
            Date range: ${txt(s.earliest_date)} -> ${txt(s.latest_date)} &nbsp;|&nbsp;
            Total amount: <b>${n(s.total_amount).toFixed(2)}</b>
        </span>`;

    if (data.auto_created_categories && data.auto_created_categories.length) {
        html += `<br><details style="margin-top:6px" ${isDry ? 'open' : ''}><summary style="font-size:12px;cursor:pointer;color:var(--sub)">
            ${isDry ? 'Categories to auto-create' : 'Auto-created categories'} (${data.auto_created_categories.length})
        </summary><div style="margin-top:4px;font-size:11px;color:var(--muted)">
            ${data.auto_created_categories.map(c => c.account_code ? `${c.name} (${c.account_code})` : c.name).join(', ')}
        </div></details>`;
    }

    if (data.warnings && data.warnings.length) {
        html += `<br><div style="margin-top:6px;padding:8px 12px;background:rgba(255,181,71,.08);
            border:1px solid rgba(255,181,71,.2);border-radius:8px;font-size:12px;color:var(--warn)">
            ${data.warnings.map(w => `Warning: ${w}`).join('<br>')}
        </div>`;
    }

    if (data.errors && data.errors.length) {
        html += `<br><br><b style="color:var(--danger)">${data.errors.length} error(s):</b>
        <div style="max-height:180px;overflow-y:auto;margin-top:6px">
        <table style="font-size:11px;width:100%">
            <thead><tr><th>Row</th><th>Category</th><th>Amount</th><th>Farm</th><th>Date</th><th>Reason</th></tr></thead>
            <tbody>${data.errors.map(e => `<tr>
                <td>${e.row}</td>
                <td>${e.category || ''}</td>
                <td>${e.amount || ''}</td>
                <td>${e.farm || 'General Expense'}</td>
                <td>${e.date || ''}</td>
                <td style="color:var(--danger)">${e.reason}</td>
            </tr>`).join('')}</tbody>
        </table></div>`;
    }

    if (isDry && (!data.errors || data.errors.length === 0)) {
        html += `<br><br>
        <button onclick="runRealExpensesImport()" style="width:100%;padding:10px;border-radius:8px;
            background:linear-gradient(135deg,var(--orange),#ffd166);color:#2e1300;
            font-weight:800;font-size:13px;border:none;cursor:pointer;">
            Run real expenses import
        </button>`;
    }

    html += '</div>';
    const kind = data.errors && data.errors.length ? 'warn' : 'ok';
    showResult('expenses', html, kind);
}

async function runRealExpensesImport() {
    document.getElementById('chk-expenses-dryrun').checked = false;
    await doImportExpenses();
}

async function doImportSales() {
    const f = files['sales'];
    if (!f) { showResult('sales', 'Please select a file first', 'err'); return; }
    const btn = document.getElementById('btn-sales');
    btn.disabled = true; btn.innerHTML = '⏳ Processing…';
    showProg('sales', 40); showResult('sales', '', '');

    const mode      = document.querySelector('input[name="sales-mode"]:checked').value;
    const dryRun    = document.getElementById('chk-dryrun').checked;
    const force     = document.getElementById('chk-force').checked;
    const costRatio = document.getElementById('inp-cost-ratio').value.trim();

    const fd = new FormData();
    fd.append('file',    f);
    fd.append('dry_run', dryRun ? 'true' : 'false');
    fd.append('mode',    mode);
    fd.append('force',   force ? 'true' : 'false');
    if (costRatio !== '') fd.append('default_cost_ratio', costRatio);

    let res, data;
    try {
        res  = await fetch('/import/api/sales', { method: 'POST', body: fd });
        data = await res.json();
    } catch (e) {
        showProg('sales', 100);
        btn.disabled = false; btn.innerHTML = '⬆ Import Sales';
        showResult('sales', '✗ Network error: ' + e.message, 'err');
        return;
    }
    showProg('sales', 100);
    btn.disabled = false;
    btn.innerHTML = '⬆ Import Sales';

    if (!res.ok) {
        // FastAPI error format: {"detail": "..."} or {"detail": [{...}]}
        const detail = data.detail;
        const msg = Array.isArray(detail)
            ? detail.map(e => e.msg || JSON.stringify(e)).join('; ')
            : (detail || data.error || `HTTP ${res.status}`);
        showResult('sales', '✗ ' + msg, 'err');
        return;
    }

    if (data.error) { showResult('sales', '✗ ' + data.error, 'err'); return; }

    renderSalesResult(data, dryRun);
    loadBatches();
}

function renderSalesResult(data, wasDryRun) {
    if (!data || !data.summary) {
        showResult('sales', '✗ Unexpected response from server — no summary returned.', 'err');
        return;
    }
    const s = data.summary;
    const isDry = !!data.dry_run;
    // Guard helper: treat null/undefined as 0 for numeric display
    const n  = v => (v != null && !isNaN(+v)) ? +v : 0;
    const fmt  = v => n(v).toLocaleString();
    const fmt2 = v => n(v).toFixed(2);

    const created = isDry ? n(s.invoices_would_create) : n(s.invoices_created);
    const label   = isDry ? 'would create' : 'created';

    let html = `<div style="font-size:13px">
        ${isDry ? '<span style="color:var(--warn)">⚠ DRY RUN — nothing was saved</span><br>' : ''}
        <b>${fmt(s.rows_read)}</b> rows read &nbsp;·&nbsp;
        <b>${fmt(created)}</b> invoices ${label} &nbsp;·&nbsp;
        <b>${fmt(s.line_items)}</b> line items &nbsp;·&nbsp;
        <b>${fmt(s.rows_skipped)}</b> skipped<br>
        <span style="color:var(--sub);font-size:12px">
            Customers ${isDry?'would create':'created'}: <b>${fmt(s.customers_auto_created)}</b> &nbsp;·&nbsp;
            Products ${isDry?'would create':'created'}: <b>${fmt(s.products_auto_created)}</b> &nbsp;·&nbsp;
            Date range: ${s.earliest_date||'–'} → ${s.latest_date||'–'} &nbsp;·&nbsp;
            Total value: <b>${fmt2(s.total_value)}</b>
        </span>`;

    if (data.batch_id) {
        html += `<br><span style="color:var(--muted);font-size:11px">Batch ID: ${data.batch_id}</span>`;
    }

    // Auto-created customers
    if (data.auto_created_customers && data.auto_created_customers.length) {
        html += `<br><details style="margin-top:6px"><summary style="font-size:12px;cursor:pointer;color:var(--sub)">
            Auto-created customers (${data.auto_created_customers.length})
        </summary><div style="margin-top:4px;font-size:11px;color:var(--muted)">
            ${data.auto_created_customers.slice(0,10).join(', ')}
            ${data.auto_created_customers.length > 10 ? ` <i>and ${data.auto_created_customers.length - 10} more…</i>` : ''}
        </div></details>`;
    }

    // Auto-created products
    if (data.auto_created_products && data.auto_created_products.length) {
        const nProd = data.auto_created_products.length;
        const costRatioSet = document.getElementById('inp-cost-ratio').value.trim() !== '';
        const warnAmber = isDry && !costRatioSet && nProd > 5;
        html += `<br><details style="margin-top:4px" ${isDry ? 'open' : ''}><summary style="font-size:12px;cursor:pointer;color:${warnAmber?'var(--warn)':'var(--sub)'}">
            ${warnAmber ? '⚠ ' : ''}Auto-created products (${nProd})${warnAmber ? ' — cost = 0, consider setting Default cost ratio' : ''}
        </summary><div style="margin-top:4px">
        <table style="font-size:11px;width:100%">
            <thead><tr><th>SKU</th><th>Name</th><th>Price</th><th>Cost</th></tr></thead>
            <tbody>${data.auto_created_products.slice(0,10).map(p=>`<tr>
                <td style="font-family:var(--mono)">${p.sku}</td>
                <td>${p.name}</td>
                <td>${p.price.toFixed(2)}</td>
                <td>${p.cost === 0 ? '<span style="color:var(--warn)">0 — not set</span>' : p.cost.toFixed(2)}</td>
            </tr>`).join('')}</tbody>
        </table>
        ${nProd > 10 ? `<div style="font-size:11px;color:var(--muted);padding-top:4px">…and ${nProd - 10} more</div>` : ''}
        </div></details>`;
    }

    // Warnings
    if (data.warnings && data.warnings.length) {
        html += `<br><div style="margin-top:6px;padding:8px 12px;background:rgba(255,181,71,.08);
            border:1px solid rgba(255,181,71,.2);border-radius:8px;font-size:12px;color:var(--warn)">
            ${data.warnings.map(w=>`⚠ ${w}`).join('<br>')}
        </div>`;
    }

    if (data.errors && data.errors.length) {
        html += `<br><br><b style="color:var(--danger)">${data.errors.length} error(s):</b>
        <div style="max-height:180px;overflow-y:auto;margin-top:6px">
        <table style="font-size:11px;width:100%">
            <thead><tr><th>Row</th><th>SKU</th><th>Customer</th><th>Date</th><th>Reason</th></tr></thead>
            <tbody>${data.errors.map(e=>`<tr>
                <td>${e.row}</td><td style="font-family:var(--mono)">${e.sku}</td>
                <td>${e.customer}</td><td>${e.date}</td>
                <td style="color:var(--danger)">${e.reason}</td>
            </tr>`).join('')}</tbody>
        </table></div>`;
    }

    // If dry_run + no errors → show "Run real import" button
    if (isDry && (!data.errors || data.errors.length === 0)) {
        html += `<br><br>
        <button onclick="runRealImport()" style="width:100%;padding:10px;border-radius:8px;
            background:linear-gradient(135deg,var(--green),var(--lime));color:#0a1a00;
            font-weight:800;font-size:13px;border:none;cursor:pointer;">
            ✓ Run real import
        </button>`;
    }

    html += '</div>';
    const kind = data.errors && data.errors.length ? 'warn' : 'ok';
    showResult('sales', html, kind);
}

async function runRealImport() {
    const f = files['sales'];
    if (!f) { return; }
    document.getElementById('chk-dryrun').checked = false;
    await doImportSales();
}

// ── Recent batches ──────────────────────────────────────────────────────────

async function loadBatches() {
    const panel = document.getElementById('batches-panel');
    try {
        const r = await fetch('/import/api/sales/batches');
        const d = await r.json();
        const batches = d.batches || [];
        if (!batches.length) {
            panel.innerHTML = '<div style="color:var(--muted);font-size:12px">No import batches found.</div>';
            return;
        }
        panel.innerHTML = `<table style="width:100%;font-size:12px">
            <thead><tr>
                <th>Batch ID</th><th>File</th><th>Date</th>
                <th>Invoices</th><th>Total Value</th><th></th>
            </tr></thead>
            <tbody>${batches.map(b=>`<tr>
                <td style="font-family:var(--mono);font-size:10px">${b.batch_id.slice(0,12)}…</td>
                <td>${b.filename}</td>
                <td>${b.ran_on||'–'}</td>
                <td>${b.invoice_count}</td>
                <td>${b.total_value.toFixed(2)}</td>
                <td><button onclick="revertBatch('${b.batch_id}')"
                    style="padding:4px 10px;border-radius:6px;border:1px solid var(--danger);
                    background:rgba(255,77,109,.08);color:var(--danger);
                    font-size:11px;cursor:pointer;font-family:var(--sans)">
                    Revert
                </button></td>
            </tr>`).join('')}
            </tbody></table>`;
    } catch(e) {
        panel.innerHTML = '<div style="color:var(--muted);font-size:12px">Could not load batches.</div>';
    }
}

async function revertBatch(batchId) {
    if (!confirm('Delete all invoices in batch ' + batchId.slice(0,8) + '…? This cannot be undone.')) return;
    const r = await fetch('/import/api/sales/batch/' + batchId, { method: 'DELETE' });
    const d = await r.json();
    if (d.ok) {
        let msg = `✓ Batch reverted — ${d.deleted_invoices} invoices deleted`;
        if (d.deleted_products)  msg += `, ${d.deleted_products} auto-created products removed`;
        if (d.deleted_customers) msg += `, ${d.deleted_customers} auto-created customers removed`;
        showResult('sales', msg + '.', 'ok');
    } else {
        showResult('sales', '✗ Revert failed', 'err');
    }
    loadBatches();
}

// Load batches on page load
loadBatches();

async function doImportFarmIntake() {
    const f = files['farm-intake'];
    if (!f) { showResult('farm-intake', 'Please select a file first', 'err'); return; }

    const btn = document.getElementById('btn-farm-intake');
    btn.disabled = true;
    btn.innerHTML = 'Processing...';
    showProg('farm-intake', 40);
    showResult('farm-intake', '', '');

    const dryRun = document.getElementById('chk-farm-dryrun').checked;
    const recordStock = document.getElementById('chk-farm-stock').checked;

    const fd = new FormData();
    fd.append('file', f);
    fd.append('dry_run', dryRun ? 'true' : 'false');
    fd.append('record_stock_movement', recordStock ? 'true' : 'false');

    let res, data;
    try {
        res = await fetch('/import/api/farm-intake', { method: 'POST', body: fd });
        data = await res.json();
    } catch (e) {
        showProg('farm-intake', 100);
        btn.disabled = false;
        btn.innerHTML = 'Import Farm Intake';
        showResult('farm-intake', 'Error: Network error: ' + e.message, 'err');
        return;
    }

    showProg('farm-intake', 100);
    btn.disabled = false;
    btn.innerHTML = 'Import Farm Intake';

    if (!res.ok) {
        const detail = data?.detail;
        const msg = Array.isArray(detail)
            ? detail.map(e => e.msg || JSON.stringify(e)).join('; ')
            : (detail || data?.error || `HTTP ${res.status}`);
        showResult('farm-intake', 'Error: ' + msg, 'err');
        return;
    }

    if (data.error) {
        showResult('farm-intake', 'Error: ' + data.error, 'err');
        return;
    }

    renderFarmIntakeResult(data);
    loadFarmBatches();
}

function renderFarmIntakeResult(data) {
    if (!data || !data.summary) {
        showResult('farm-intake', 'Error: Unexpected response from server - no summary returned.', 'err');
        return;
    }
    const s = data.summary || {};
    const isDry = !!data.dry_run;
    const n = v => (v != null && !isNaN(+v)) ? +v : 0;
    const txt = v => (v == null || v === '') ? '-' : String(v);

    let html = `<div style="font-size:13px">
        ${isDry ? '<span style="color:var(--warn)">Dry run - nothing was saved</span><br>' : ''}
        <b>${n(s.rows_read)}</b> rows read &nbsp;|&nbsp;
        <b>${n(s.rows_imported)}</b> rows ${isDry ? 'valid for import' : 'imported'} &nbsp;|&nbsp;
        <b>${n(s.rows_skipped)}</b> skipped<br>
        <span style="color:var(--sub);font-size:12px">
            Farm intake records ${isDry ? 'would create' : 'created'}: <b>${n(s.farm_deliveries_created)}</b> &nbsp;|&nbsp;
            Farms ${isDry ? 'would create' : 'auto-created'}: <b>${n(s.farms_auto_created)}</b> &nbsp;|&nbsp;
            Products auto-created: <b>${n(s.products_auto_created)}</b>
        </span><br>
        <span style="color:var(--sub);font-size:12px">
            Stock movement recorded: <b>${s.stock_movement_recorded ? 'Yes' : 'No'}</b> &nbsp;|&nbsp;
            Stock move records ${isDry ? 'would create' : 'created'}: <b>${n(s.stock_moves_created)}</b> &nbsp;|&nbsp;
            Date range: ${txt(s.earliest_date)} -> ${txt(s.latest_date)}
        </span>`;

    if (data.batch_id) {
        html += `<br><span style="color:var(--muted);font-size:11px">Batch ID: ${data.batch_id}</span>`;
    }

    if (data.auto_created_farms && data.auto_created_farms.length) {
        html += `<br><details style="margin-top:6px"><summary style="font-size:12px;cursor:pointer;color:var(--sub)">
            Auto-created farms (${data.auto_created_farms.length})
        </summary><div style="margin-top:4px;font-size:11px;color:var(--muted)">
            ${data.auto_created_farms.map(f=>f.name).join(', ')}
        </div></details>`;
    }

    if (data.warnings && data.warnings.length) {
        html += `<br><div style="margin-top:6px;padding:8px 12px;background:rgba(255,181,71,.08);
            border:1px solid rgba(255,181,71,.2);border-radius:8px;font-size:12px;color:var(--warn)">
            ${data.warnings.map(w=>`Warning: ${w}`).join('<br>')}
        </div>`;
    }

    if (data.errors && data.errors.length) {
        html += `<br><br><b style="color:var(--danger)">${data.errors.length} error(s):</b>
        <div style="max-height:180px;overflow-y:auto;margin-top:6px">
        <table style="font-size:11px;width:100%">
            <thead><tr><th>Row</th><th>SKU</th><th>Product</th><th>Farm</th><th>Date</th><th>Reason</th></tr></thead>
            <tbody>${data.errors.map(e=>`<tr>
                <td>${e.row}</td>
                <td style="font-family:var(--mono)">${e.sku || ''}</td>
                <td>${e.product || ''}</td>
                <td>${e.farm || ''}</td>
                <td>${e.date || ''}</td>
                <td style="color:var(--danger)">${e.reason}</td>
            </tr>`).join('')}</tbody>
        </table></div>`;
    }

    if (isDry && (!data.errors || data.errors.length === 0)) {
        html += `<br><br>
        <button onclick="runRealFarmIntakeImport()" style="width:100%;padding:10px;border-radius:8px;
            background:linear-gradient(135deg,var(--teal),var(--green));color:#042118;
            font-weight:800;font-size:13px;border:none;cursor:pointer;">
            Run real farm intake import
        </button>`;
    }

    html += '</div>';
    const kind = data.errors && data.errors.length ? 'warn' : 'ok';
    showResult('farm-intake', html, kind);
}

async function runRealFarmIntakeImport() {
    document.getElementById('chk-farm-dryrun').checked = false;
    await doImportFarmIntake();
}

async function loadFarmBatches() {
    const panel = document.getElementById('farm-batches-panel');
    try {
        const r = await fetch('/import/api/farm-intake/batches');
        const d = await r.json();
        const batches = d.batches || [];
        if (!batches.length) {
            panel.innerHTML = '<div style="color:var(--muted);font-size:12px">No Farm Intake import batches found.</div>';
            return;
        }
        panel.innerHTML = `<table style="width:100%;font-size:12px">
            <thead><tr>
                <th>Batch ID</th><th>File</th><th>Date</th><th>Deliveries</th><th>Rows</th><th>Stock</th><th></th>
            </tr></thead>
            <tbody>${batches.map(b=>`<tr>
                <td style="font-family:var(--mono);font-size:10px">${b.batch_id.slice(0,12)}...</td>
                <td>${b.filename || '-'}</td>
                <td>${b.ran_on || '-'}</td>
                <td>${b.delivery_count}</td>
                <td>${b.row_count}</td>
                <td>${b.stock_recorded ? 'Yes' : 'No'}</td>
                <td><button onclick="revertFarmBatch('${b.batch_id}')"
                    style="padding:4px 10px;border-radius:6px;border:1px solid var(--danger);
                    background:rgba(255,77,109,.08);color:var(--danger);
                    font-size:11px;cursor:pointer;font-family:var(--sans)">
                    Revert
                </button></td>
            </tr>`).join('')}
            </tbody></table>`;
    } catch (e) {
        panel.innerHTML = '<div style="color:var(--muted);font-size:12px">Could not load Farm Intake batches.</div>';
    }
}

async function revertFarmBatch(batchId) {
    if (!confirm('Delete all farm intake records in batch ' + batchId.slice(0,8) + '...? This cannot be undone.')) return;
    const r = await fetch('/import/api/farm-intake/batch/' + batchId, { method: 'DELETE' });
    const d = await r.json();
    if (d.ok) {
        showResult('farm-intake', `Batch reverted - ${d.deleted_deliveries} deliveries deleted, ${d.deleted_stock_moves} stock moves removed.`, 'ok');
    } else {
        showResult('farm-intake', 'Error: Revert failed', 'err');
    }
    loadFarmBatches();
}

loadFarmBatches();

// ── Historical B2B Sales ────────────────────────────────────────────────────

async function doImportB2B() {
    const f = files['b2b-sales'];
    if (!f) { showResult('b2b-sales', 'Please select a file first', 'err'); return; }
    const btn = document.getElementById('btn-b2b-sales');
    btn.disabled = true; btn.innerHTML = '⏳ Processing…';
    showProg('b2b-sales', 40); showResult('b2b-sales', '', '');

    const mode   = document.querySelector('input[name="b2b-mode"]:checked').value;
    const dryRun = document.getElementById('chk-b2b-dryrun').checked;
    const force  = document.getElementById('chk-b2b-force').checked;

    const fd = new FormData();
    fd.append('file',    f);
    fd.append('dry_run', dryRun ? 'true' : 'false');
    fd.append('mode',    mode);
    fd.append('force',   force ? 'true' : 'false');

    let res, data;
    try {
        res = await fetch('/import/api/b2b-sales', { method: 'POST', body: fd });
        data = await res.json();
    } catch (e) {
        showProg('b2b-sales', 100);
        btn.disabled = false;
        btn.innerHTML = '⬆ Import B2B Sales';
        showResult('b2b-sales', '✗ Network error: ' + e.message, 'err');
        return;
    }
    showProg('b2b-sales', 100);
    btn.disabled = false;
    btn.innerHTML = '⬆ Import B2B Sales';

    if (!res.ok) {
        const detail = data?.detail;
        const msg = Array.isArray(detail)
            ? detail.map(e => e.msg || JSON.stringify(e)).join('; ')
            : (detail || data?.error || `HTTP ${res.status}`);
        showResult('b2b-sales', '✗ ' + msg, 'err');
        return;
    }

    if (data.error) { showResult('b2b-sales', '✗ ' + data.error, 'err'); return; }

    renderB2BResult(data);
    loadB2BBatches();
}

function renderB2BResult(data) {
    const isDry = !!data?.dry_run;
    const s = data?.summary || {};
    const pick = (...values) => values.find(v => v !== undefined && v !== null);
    const n = (...values) => {
        const value = pick(...values);
        const parsed = Number(value);
        return Number.isFinite(parsed) ? parsed : 0;
    };
    const text = (...values) => {
        const value = pick(...values);
        return value === undefined || value === null || value === '' ? '–' : String(value);
    };

    const rowsRead = n(s.rows_read, data.rows_read, data.rowsRead);
    const invoicesCreated = n(s.invoices_created, s.invoices_would_create, data.invoices_created, data.invoices_would_create);
    const lineItems = n(s.line_items, data.line_items, data.lineItems);
    const rowsSkipped = n(s.rows_skipped, data.rows_skipped, data.skipped);
    const clientsAutoCreated = n(s.clients_auto_created, data.clients_auto_created);
    const productsAutoCreated = n(s.products_auto_created, data.products_auto_created);
    const consignmentsCount = isDry
        ? n(s.consignments_would_create, data.consignments_would_create)
        : n(s.consignments_created, data.consignments_created);
    const earliestDate = text(s.earliest_date, data.earliest_date);
    const latestDate = text(s.latest_date, data.latest_date);
    const totalSubtotal = n(s.total_subtotal, data.total_subtotal);
    const totalDiscount = n(s.total_discount, data.total_discount);
    const totalInvoiced = n(s.total_invoiced, data.total_invoiced);

    if (!data || (!data.summary && !data.error && !data.errors)) {
        showResult('b2b-sales', '✗ Unexpected response from server — no summary returned.', 'err');
        return;
    }

    const created = invoicesCreated;
    const label   = isDry ? 'would create' : 'created';

    let html = `<div style="font-size:13px">
        ${isDry ? '<span style="color:var(--warn)">⚠ DRY RUN — nothing was saved</span><br>' : ''}
        <b>${rowsRead}</b> rows read &nbsp;·&nbsp;
        <b>${created}</b> invoices ${label} &nbsp;·&nbsp;
        <b>${lineItems}</b> line items &nbsp;·&nbsp;
        <b>${rowsSkipped}</b> skipped<br>
        <span style="color:var(--sub);font-size:12px">
            Clients auto-created: <b>${clientsAutoCreated}</b> &nbsp;·&nbsp;
            Products ${isDry ? 'would create' : 'created'}: <b>${productsAutoCreated}</b> &nbsp;·&nbsp;
            Consignments: <b>${consignmentsCount}</b> &nbsp;·&nbsp;
            Date range: ${earliestDate} → ${latestDate}
        </span><br>
        <span style="color:var(--sub);font-size:12px">
            Subtotal: <b>${totalSubtotal.toFixed(2)}</b> &nbsp;·&nbsp;
            Discount: <b>${totalDiscount.toFixed(2)}</b> &nbsp;·&nbsp;
            Invoiced: <b>${totalInvoiced.toFixed(2)}</b>
        </span>`;

    if (data.batch_id) {
        html += `<br><span style="color:var(--muted);font-size:11px">Batch ID: ${data.batch_id}</span>`;
    }

    // by_payment_type breakdown
    const bpt = s.by_payment_type || {};
    if (Object.keys(bpt).length) {
        html += `<br><br><b style="font-size:12px">By payment type:</b>
        <table style="font-size:11px;margin-top:4px">
            <thead><tr><th>Type</th><th>Invoices</th><th>Total</th></tr></thead>
            <tbody>${Object.entries(bpt).map(([t,v])=>`<tr>
                <td>${t}</td><td>${v.invoices}</td><td>${(v.total||0).toFixed(2)}</td>
            </tr>`).join('')}</tbody>
        </table>`;
    }

    // discount_pct_suggestions
    if (data.discount_pct_suggestions && data.discount_pct_suggestions.length) {
        html += `<br><details style="margin-top:8px"><summary style="font-size:12px;cursor:pointer;color:var(--blue)">
            Discount % suggestions (${data.discount_pct_suggestions.length})
        </summary><div style="margin-top:6px">`;
        data.discount_pct_suggestions.forEach(s => {
            const badge = s.applied
                ? `<span style="color:var(--green);font-size:10px">✓ auto-applied</span>`
                : `<span style="color:var(--warn);font-size:10px">⚠ manual review needed</span>`;
            html += `<div style="font-size:11px;padding:4px 0;border-bottom:1px solid var(--border)">
                <b>${s.client}</b>: current ${s.current}% → suggested <b>${s.suggested}%</b> ${badge}`;
            if (!s.applied && !isDry) {
                html += ` <button onclick="applyClientDiscount('${s.client}', ${s.suggested})"
                    style="margin-left:6px;padding:2px 8px;border-radius:4px;border:1px solid var(--blue);
                    background:transparent;color:var(--blue);font-size:10px;cursor:pointer">Apply</button>`;
            }
            html += '</div>';
        });
        html += '</div></details>';
    }

    // auto_created_clients
    if (data.auto_created_clients && data.auto_created_clients.length) {
        html += `<br><details style="margin-top:4px"><summary style="font-size:12px;cursor:pointer;color:var(--sub)">
            Auto-created clients (${data.auto_created_clients.length})
        </summary><div style="margin-top:4px;font-size:11px;color:var(--muted)">
        ${data.auto_created_clients.map(c=>
            `${c.name} · ${c.payment_terms} · discount ${c.discount_pct}%`
        ).join('<br>')}
        </div></details>`;
    }

    if (data.auto_created_products && data.auto_created_products.length) {
        html += `<br><details style="margin-top:4px" ${isDry ? 'open' : ''}><summary style="font-size:12px;cursor:pointer;color:var(--sub)">
            Auto-created products (${data.auto_created_products.length})
        </summary><div style="margin-top:4px">
        <table style="font-size:11px;width:100%">
            <thead><tr><th>SKU</th><th>Name</th></tr></thead>
            <tbody>${data.auto_created_products.slice(0,10).map(p=>`<tr>
                <td style="font-family:var(--mono)">${p.sku}</td>
                <td>${p.name}</td>
            </tr>`).join('')}</tbody>
        </table>
        ${data.auto_created_products.length > 10 ? `<div style="font-size:11px;color:var(--muted);padding-top:4px">…and ${data.auto_created_products.length - 10} more</div>` : ''}
        </div></details>`;
    }

    if (data.warnings && data.warnings.length) {
        html += `<br><div style="margin-top:6px;padding:8px 12px;background:rgba(255,181,71,.08);
            border:1px solid rgba(255,181,71,.2);border-radius:8px;font-size:12px;color:var(--warn)">
            ${data.warnings.map(w=>`⚠ ${w}`).join('<br>')}
        </div>`;
    }

    if (data.errors && data.errors.length) {
        html += `<br><br><b style="color:var(--danger)">${data.errors.length} error(s):</b>
        <div style="max-height:180px;overflow-y:auto;margin-top:6px">
        <table style="font-size:11px;width:100%">
            <thead><tr><th>Row</th><th>SKU</th><th>Client</th><th>Date</th><th>Reason</th></tr></thead>
            <tbody>${data.errors.map(e=>`<tr>
                <td>${e.row}</td><td style="font-family:var(--mono)">${e.sku||''}</td>
                <td>${e.client||''}</td><td>${e.date||''}</td>
                <td style="color:var(--danger)">${e.reason}</td>
            </tr>`).join('')}</tbody>
        </table></div>`;
    }

    if (isDry && (!data.errors || data.errors.length === 0)) {
        html += `<br><br>
        <button onclick="runRealB2BImport()" style="width:100%;padding:10px;border-radius:8px;
            background:linear-gradient(135deg,var(--blue),#4d6fff);color:#fff;
            font-weight:800;font-size:13px;border:none;cursor:pointer;">
            ✓ Run real B2B import
        </button>`;
    }

    html += '</div>';
    const kind = data.errors && data.errors.length ? 'warn' : 'ok';
    showResult('b2b-sales', html, kind);
}

async function runRealB2BImport() {
    document.getElementById('chk-b2b-dryrun').checked = false;
    await doImportB2B();
}

async function applyClientDiscount(clientName, discountPct) {
    // Placeholder — future: call PUT /b2b/api/clients/{id} with new discount_pct
    alert('To apply: go to B2B → Clients → ' + clientName + ' → edit discount to ' + discountPct + '%');
}

async function loadB2BBatches() {
    const panel = document.getElementById('b2b-batches-panel');
    try {
        const r = await fetch('/import/api/b2b-sales/batches');
        const d = await r.json();
        const batches = d.batches || [];
        if (!batches.length) {
            panel.innerHTML = '<div style="color:var(--muted);font-size:12px">No B2B import batches found.</div>';
            return;
        }
        panel.innerHTML = `<table style="width:100%;font-size:12px">
            <thead><tr>
                <th>Batch ID</th><th>File</th><th>Date</th>
                <th>Invoices</th><th>Total Value</th><th></th>
            </tr></thead>
            <tbody>${batches.map(b=>`<tr>
                <td style="font-family:var(--mono);font-size:10px">${b.batch_id.slice(0,12)}…</td>
                <td>${b.filename}</td><td>${b.ran_on||'–'}</td>
                <td>${b.invoice_count}</td><td>${b.total_value.toFixed(2)}</td>
                <td><button onclick="revertB2BBatch('${b.batch_id}')"
                    style="padding:4px 10px;border-radius:6px;border:1px solid var(--danger);
                    background:rgba(255,77,109,.08);color:var(--danger);
                    font-size:11px;cursor:pointer;font-family:var(--sans)">
                    Revert
                </button></td>
            </tr>`).join('')}
            </tbody></table>`;
    } catch(e) {
        panel.innerHTML = '<div style="color:var(--muted);font-size:12px">Could not load batches.</div>';
    }
}

async function revertB2BBatch(batchId) {
    if (!confirm('Delete all B2B invoices in batch ' + batchId.slice(0,8) + '…? This cannot be undone.')) return;
    const r = await fetch('/import/api/b2b-sales/batch/' + batchId, { method: 'DELETE' });
    const d = await r.json();
    if (d.ok) {
        showResult('b2b-sales', `✓ Batch reverted — ${d.deleted_invoices} invoices, ${d.deleted_consignments} consignments deleted.`, 'ok');
    } else {
        showResult('b2b-sales', '✗ Revert failed', 'err');
    }
    loadB2BBatches();
}

loadB2BBatches();
</script>
</body>
</html>"""
