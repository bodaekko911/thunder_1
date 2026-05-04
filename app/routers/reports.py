from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import HTMLResponse, StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
from sqlalchemy import func, or_, select
from datetime import datetime, date, timedelta, timezone
from collections import defaultdict
from typing import Optional
import io
import re

from app.core.permissions import require_permission
from app.database import get_async_session
from app.core.navigation import render_app_header
from app.models.product import Product
from app.models.invoice import Invoice
from app.models.b2b import B2BClient, B2BInvoice, B2BInvoiceItem, B2BRefund
from app.models.inventory import StockMove
from app.models.farm import Farm, FarmDelivery, FarmDeliveryItem
from app.models.spoilage import SpoilageRecord
from app.models.refund import RetailRefund
from app.models.production import ProductionBatch, BatchInput, BatchOutput
from app.models.accounting import Account, Journal, JournalEntry
from app.models.receipt import ProductReceipt
from app.models.expense import Expense
from app.models.user import User

router = APIRouter(
    prefix="/reports",
    tags=["Reports"],
    dependencies=[Depends(require_permission("page_reports"))],
)


# ── EXCEL HELPER ───────────────────────────────────────
def _excel_dependencies():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        return openpyxl, Font, PatternFill, Alignment, Border, Side, get_column_letter
    except ImportError:
        raise Exception("Run: pip install openpyxl --break-system-packages")


def _coerce_excel_value(value, fmt):
    if value in (None, ""):
        return value
    if fmt == "date" and isinstance(value, str):
        try:
            return datetime.fromisoformat(value[:10]).date()
        except ValueError:
            return value
    if fmt == "datetime" and isinstance(value, str):
        normalized = value.replace("T", " ")
        for parser in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
            try:
                return datetime.strptime(normalized, parser)
            except ValueError:
                continue
    return value


def _apply_excel_number_format(cell, fmt):
    if fmt == "money":
        cell.number_format = '#,##0.00'
    elif fmt == "qty":
        cell.number_format = '#,##0.00'
    elif fmt == "int":
        cell.number_format = '#,##0'
    elif fmt == "percent":
        cell.number_format = '0.00%'
    elif fmt == "percent_value":
        cell.number_format = '0.00"%"'
    elif fmt == "date":
        cell.number_format = 'yyyy-mm-dd'
    elif fmt == "datetime":
        cell.number_format = 'yyyy-mm-dd hh:mm'


def _autosize_report_sheet(ws, get_column_letter, min_width=10, max_width=42):
    for col_idx in range(1, ws.max_column + 1):
        values = []
        for row_idx in range(1, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is not None:
                values.append(str(value))
        max_len = max((len(v) for v in values), default=min_width)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(min(max_len + 3, max_width), min_width)


def add_report_sheet(
    wb,
    *,
    sheet_name,
    report_title,
    headers,
    rows,
    metadata=None,
    column_formats=None,
    wrap_columns=None,
    total_row_indices=None,
    tab_color="1F4E78",
):
    openpyxl, Font, PatternFill, Alignment, Border, Side, get_column_letter = _excel_dependencies()
    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90
    ws.sheet_properties.tabColor = tab_color

    title_fill = PatternFill("solid", fgColor="1F4E78")
    meta_fill = PatternFill("solid", fgColor="EAF1FB")
    header_fill = PatternFill("solid", fgColor="2F6F4F")
    alt_fill = PatternFill("solid", fgColor="F7FAFC")
    total_fill = PatternFill("solid", fgColor="E3F2E8")
    thin = Side(style="thin", color="D9E2EC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    current_row = 1
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=max(len(headers), 2))
    title_cell = ws.cell(row=current_row, column=1, value=report_title)
    title_cell.fill = title_fill
    title_cell.font = Font(bold=True, color="FFFFFF", size=15)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.border = border
    ws.row_dimensions[current_row].height = 24
    current_row += 1

    generated_cell = ws.cell(row=current_row, column=1, value=f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    generated_cell.font = Font(italic=True, color="5B6B7A", size=10)
    current_row += 1

    for label, value in (metadata or []):
        label_cell = ws.cell(row=current_row, column=1, value=label)
        value_cell = ws.cell(row=current_row, column=2, value=value)
        label_cell.font = Font(bold=True, color="334E68")
        label_cell.fill = meta_fill
        value_cell.fill = meta_fill
        label_cell.border = border
        value_cell.border = border
        current_row += 1

    current_row += 1
    header_row = current_row
    for col_no, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_no, value=header)
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[header_row].height = 20

    column_formats = column_formats or {}
    wrap_columns = set(wrap_columns or [])
    total_row_indices = set(total_row_indices or [])
    for row_idx, row in enumerate(rows, start=1):
        excel_row = header_row + row_idx
        is_total_row = row_idx in total_row_indices
        for col_idx, value in enumerate(row, 1):
            header = headers[col_idx - 1]
            fmt = column_formats.get(header)
            cell = ws.cell(row=excel_row, column=col_idx, value=_coerce_excel_value(value, fmt))
            cell.border = border
            if is_total_row:
                cell.fill = total_fill
                cell.font = Font(bold=True)
            elif row_idx % 2 == 1:
                cell.fill = alt_fill
            horizontal = "left"
            if fmt in {"money", "qty", "int", "percent", "percent_value"}:
                horizontal = "right"
            elif fmt in {"date", "datetime"}:
                horizontal = "center"
            cell.alignment = Alignment(horizontal=horizontal, vertical="top", wrap_text=(header in wrap_columns))
            _apply_excel_number_format(cell, fmt)

    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{header_row}"
    _autosize_report_sheet(ws, get_column_letter)
    return ws


def build_report_workbook(sheet_specs):
    openpyxl, *_ = _excel_dependencies()
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for spec in sheet_specs:
        add_report_sheet(wb, **spec)
    return wb


def workbook_to_buffer(wb):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def to_xlsx(headers, rows, sheet_name="Report", report_title=None, metadata=None, column_formats=None, wrap_columns=None, total_row_indices=None):
    wb = build_report_workbook([
        {
            "sheet_name": sheet_name,
            "report_title": report_title or sheet_name,
            "headers": headers,
            "rows": rows,
            "metadata": metadata or [],
            "column_formats": column_formats or {},
            "wrap_columns": wrap_columns or set(),
            "total_row_indices": total_row_indices or set(),
        }
    ])
    return workbook_to_buffer(wb)


def parse_dates(date_from, date_to):
    now = datetime.now(timezone.utc)
    if date_from and date_to:
        try:
            d_from = datetime.fromisoformat(date_from).replace(tzinfo=timezone.utc)
            d_to   = datetime.fromisoformat(date_to).replace(hour=23, minute=59, second=59, tzinfo=timezone.utc)
        except Exception:
            d_from = now.replace(day=1, hour=0, minute=0, second=0)
            d_to   = now
    else:
        d_from = now.replace(day=1, hour=0, minute=0, second=0)
        d_to   = now
    return d_from, d_to


def _resolve_pagination(skip, limit, default_limit=100):
    skip_value = getattr(skip, "default", skip)
    limit_value = getattr(limit, "default", limit)
    try:
        skip_value = max(int(skip_value or 0), 0)
    except (TypeError, ValueError):
        skip_value = 0
    try:
        limit_value = max(int(limit_value or default_limit), 0)
    except (TypeError, ValueError):
        limit_value = default_limit
    return skip_value, limit_value


def _num(value) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _paginate_rows(rows, skip, limit, include_all=False):
    if include_all:
        return rows
    if limit == 0:
        return []
    return rows[skip : skip + limit]


def _channel_totals():
    return {
        "gross_sales": 0.0,
        "cash_collected": 0.0,
        "outstanding": 0.0,
        "count": 0,
    }


async def _load_b2b_client_payment_records(
    db: AsyncSession,
    *,
    d_from: datetime,
    d_to: datetime,
):
    payment_ref_types = ("consignment_client_payment", "consignment_payment", "b2b_payment", "b2b_collection")
    payment_result = await db.execute(
        select(Journal)
        .where(
            Journal.ref_type.in_(payment_ref_types),
            Journal.created_at >= d_from,
            Journal.created_at <= d_to,
        )
        .options(selectinload(Journal.entries).selectinload(JournalEntry.account), selectinload(Journal.user))
        .order_by(Journal.created_at.desc(), Journal.id.desc())
    )
    journals = payment_result.scalars().all()
    client_ids = {journal.ref_id for journal in journals if journal.ref_id}
    invoice_ids = set()
    invoice_numbers = set()
    invoice_pattern = re.compile(r"(B2B-\d{5,})", re.IGNORECASE)
    for journal in journals:
        if journal.ref_type == "consignment_client_payment":
            continue
        if journal.ref_id:
            invoice_ids.add(journal.ref_id)
        match = invoice_pattern.search(journal.description or "")
        if match:
            invoice_numbers.add(match.group(1).upper())

    invoice_map_by_id = {}
    invoice_map_by_number = {}
    if invoice_ids or invoice_numbers:
        conditions = []
        if invoice_ids:
            conditions.append(B2BInvoice.id.in_(invoice_ids))
        if invoice_numbers:
            conditions.append(func.upper(B2BInvoice.invoice_number).in_(invoice_numbers))
        invoice_result = await db.execute(
            select(B2BInvoice)
            .where(or_(*conditions))
            .options(selectinload(B2BInvoice.client))
        )
        invoices = invoice_result.scalars().all()
        invoice_map_by_id = {invoice.id: invoice for invoice in invoices}
        invoice_map_by_number = {str(invoice.invoice_number or "").upper(): invoice for invoice in invoices}
        client_ids.update(invoice.client_id for invoice in invoices if invoice.client_id)

    client_map = {}
    if client_ids:
        client_result = await db.execute(select(B2BClient).where(B2BClient.id.in_(client_ids)))
        client_map = {client.id: client for client in client_result.scalars().all()}

    payment_records = []
    for journal in journals:
        amount = 0.0
        for entry in journal.entries:
            if entry.account and entry.account.code == "1000" and _num(entry.debit) > 0:
                amount = _num(entry.debit)
                break
        if amount <= 0:
            amount = max((_num(entry.debit) for entry in journal.entries), default=0.0)
        client = client_map.get(journal.ref_id)
        reference = f"BCP-{journal.id}"
        if journal.ref_type != "consignment_client_payment":
            invoice = invoice_map_by_id.get(journal.ref_id) if journal.ref_id else None
            if not invoice:
                match = invoice_pattern.search(journal.description or "")
                if match:
                    reference = match.group(1).upper()
                    invoice = invoice_map_by_number.get(reference)
            if invoice:
                reference = invoice.invoice_number or reference
                client = invoice.client or client_map.get(invoice.client_id)
        payment_records.append({
            "journal_id": journal.id,
            "reference": reference,
            "client_id": client.id if client else journal.ref_id,
            "client": client.name if client else "—",
            "datetime": journal.created_at.strftime("%Y-%m-%d %H:%M") if journal.created_at else "—",
            "date": journal.created_at.strftime("%Y-%m-%d") if journal.created_at else "",
            "user_name": journal.user.name if journal.user else "—",
            "amount": round(amount, 2),
            "notes": journal.description or "",
            "payment_method": "cash",
            "status": "posted",
            "journal_ref_type": journal.ref_type or "—",
        })
    return payment_records


async def _build_sales_report(
    db: AsyncSession,
    *,
    d_from: datetime,
    d_to: datetime,
    skip: int = 0,
    limit: int = 100,
    include_all: bool = False,
):
    b2b_payment_records = await _load_b2b_client_payment_records(db, d_from=d_from, d_to=d_to)
    result = await db.execute(
        select(Invoice)
        .where(Invoice.created_at >= d_from, Invoice.created_at <= d_to)
        .options(selectinload(Invoice.items), selectinload(Invoice.user), selectinload(Invoice.customer))
    )
    pos_invoices = result.scalars().all()

    result = await db.execute(
        select(B2BInvoice)
        .where(B2BInvoice.created_at >= d_from, B2BInvoice.created_at <= d_to)
        .options(
            selectinload(B2BInvoice.items).selectinload(B2BInvoiceItem.product),
            selectinload(B2BInvoice.client),
            selectinload(B2BInvoice.user),
        )
    )
    b2b_invoices = result.scalars().all()

    result = await db.execute(
        select(RetailRefund)
        .where(RetailRefund.created_at >= d_from, RetailRefund.created_at <= d_to)
        .options(selectinload(RetailRefund.customer), selectinload(RetailRefund.user))
    )
    retail_refunds = result.scalars().all()

    result = await db.execute(
        select(B2BRefund)
        .where(B2BRefund.created_at >= d_from, B2BRefund.created_at <= d_to)
        .options(selectinload(B2BRefund.client), selectinload(B2BRefund.user))
    )
    b2b_refunds = result.scalars().all()

    channels = {"pos": _channel_totals(), "b2b": _channel_totals()}
    daily = defaultdict(lambda: {"gross_sales": 0.0, "refunds": 0.0, "cash_collected": 0.0})
    product_sales = defaultdict(lambda: {"qty": 0.0, "revenue": 0.0})

    pos_records = []
    for inv in sorted(pos_invoices, key=lambda x: x.created_at or datetime.min.replace(tzinfo=timezone.utc), reverse=True):
        total = _num(inv.total)
        collected = total if (inv.status or "").lower() == "paid" else 0.0
        outstanding = max(total - collected, 0.0)
        day_key = inv.created_at.strftime("%Y-%m-%d") if inv.created_at else ""
        channels["pos"]["gross_sales"] += total
        channels["pos"]["cash_collected"] += collected
        channels["pos"]["outstanding"] += outstanding
        channels["pos"]["count"] += 1
        daily[day_key]["gross_sales"] += total
        daily[day_key]["cash_collected"] += collected
        for item in inv.items:
            item_name = item.name or "â€”"
            product_sales[item_name]["qty"] += _num(item.qty)
            product_sales[item_name]["revenue"] += _num(item.total)
        pos_records.append({
            "invoice_number": inv.invoice_number or f"POS-{inv.id}",
            "datetime": inv.created_at.strftime("%Y-%m-%d %H:%M") if inv.created_at else "â€”",
            "customer": inv.customer.name if inv.customer else "Walk-in",
            "user_name": inv.user.name if inv.user else "â€”",
            "payment": inv.payment_method or "â€”",
            "status": inv.status or "â€”",
            "items": [{"name": it.name, "qty": _num(it.qty), "unit_price": _num(it.unit_price), "total": _num(it.total)} for it in inv.items],
            "total": total,
            "cash_collected": collected,
            "outstanding": outstanding,
        })

    b2b_records = []
    for inv in sorted(b2b_invoices, key=lambda x: x.created_at or datetime.min.replace(tzinfo=timezone.utc), reverse=True):
        total = _num(inv.total)
        amount_paid = _num(inv.amount_paid)
        collected = 0.0 if (inv.invoice_type or "").lower() == "consignment" else amount_paid
        outstanding = max(total - amount_paid, 0.0)
        day_key = inv.created_at.strftime("%Y-%m-%d") if inv.created_at else ""
        channels["b2b"]["gross_sales"] += total
        channels["b2b"]["cash_collected"] += collected
        channels["b2b"]["outstanding"] += outstanding
        channels["b2b"]["count"] += 1
        daily[day_key]["gross_sales"] += total
        daily[day_key]["cash_collected"] += collected
        items_data = []
        for it in inv.items:
            product_name = it.product.name if it.product else "â€”"
            item_qty = _num(it.qty)
            item_total = _num(it.total)
            items_data.append({"name": product_name, "qty": item_qty, "unit_price": _num(it.unit_price), "total": item_total})
            product_sales[product_name]["qty"] += item_qty
            product_sales[product_name]["revenue"] += item_total
        b2b_records.append({
            "invoice_number": inv.invoice_number,
            "client": inv.client.name if inv.client else "â€”",
            "datetime": inv.created_at.strftime("%Y-%m-%d %H:%M") if inv.created_at else "â€”",
            "user_name": inv.user.name if inv.user else "â€”",
            "invoice_type": inv.invoice_type,
            "status": inv.status or "â€”",
            "items": items_data,
            "total": total,
            "amount_paid": amount_paid,
            "balance_due": outstanding,
        })

    for payment in b2b_payment_records:
        channels["b2b"]["cash_collected"] += payment["amount"]
        if payment["date"]:
            daily[payment["date"]]["cash_collected"] += payment["amount"]

    refund_records = []
    retail_cash_refunds = 0.0
    b2b_cash_refunds = 0.0
    total_refunds = 0.0
    for refund in sorted(retail_refunds, key=lambda x: x.created_at or datetime.min.replace(tzinfo=timezone.utc), reverse=True):
        refund_total = _num(refund.total)
        day_key = refund.created_at.strftime("%Y-%m-%d") if refund.created_at else ""
        daily[day_key]["refunds"] += refund_total
        total_refunds += refund_total
        if (refund.refund_method or "").lower() == "cash":
            retail_cash_refunds += refund_total
        refund_records.append({
            "refund_number": refund.refund_number,
            "source": "Retail",
            "counterparty": refund.customer.name if refund.customer else "â€”",
            "datetime": refund.created_at.strftime("%Y-%m-%d %H:%M") if refund.created_at else "â€”",
            "processed_by": refund.user.name if refund.user else "â€”",
            "reason": refund.reason or "â€”",
            "refund_method": refund.refund_method or "â€”",
            "total": refund_total,
        })
    for refund in sorted(b2b_refunds, key=lambda x: x.created_at or datetime.min.replace(tzinfo=timezone.utc), reverse=True):
        refund_total = _num(refund.total)
        day_key = refund.created_at.strftime("%Y-%m-%d") if refund.created_at else ""
        daily[day_key]["refunds"] += refund_total
        total_refunds += refund_total
        b2b_cash_refunds += refund_total
        refund_records.append({
            "refund_number": refund.refund_number,
            "source": "B2B",
            "counterparty": refund.client.name if refund.client else "â€”",
            "datetime": refund.created_at.strftime("%Y-%m-%d %H:%M") if refund.created_at else "â€”",
            "processed_by": refund.user.name if refund.user else "â€”",
            "reason": refund.notes or "â€”",
            "refund_method": "cash",
            "total": refund_total,
        })

    gross_sales = channels["pos"]["gross_sales"] + channels["b2b"]["gross_sales"]
    cash_collected = channels["pos"]["cash_collected"] + channels["b2b"]["cash_collected"] - retail_cash_refunds - b2b_cash_refunds
    outstanding = channels["pos"]["outstanding"] + channels["b2b"]["outstanding"]
    net_sales = gross_sales - total_refunds

    daily_rows = []
    for day_key, bucket in sorted(daily.items()):
        daily_rows.append({
            "date": day_key,
            "gross_sales": round(bucket["gross_sales"], 2),
            "refunds": round(bucket["refunds"], 2),
            "net_sales": round(bucket["gross_sales"] - bucket["refunds"], 2),
            "cash_collected": round(bucket["cash_collected"], 2),
        })

    top_products = sorted(product_sales.items(), key=lambda x: x[1]["revenue"], reverse=True)[:10]
    return {
        "gross_sales": round(gross_sales, 2),
        "refunds": round(total_refunds, 2),
        "net_sales": round(net_sales, 2),
        "cash_collected": round(cash_collected, 2),
        "outstanding": round(outstanding, 2),
        "channels": {
            "pos": {
                "gross_sales": round(channels["pos"]["gross_sales"], 2),
                "cash_collected": round(channels["pos"]["cash_collected"] - retail_cash_refunds, 2),
                "outstanding": round(channels["pos"]["outstanding"], 2),
                "count": channels["pos"]["count"],
            },
            "b2b": {
                "gross_sales": round(channels["b2b"]["gross_sales"], 2),
                "cash_collected": round(channels["b2b"]["cash_collected"] - b2b_cash_refunds, 2),
                "outstanding": round(channels["b2b"]["outstanding"], 2),
                "count": channels["b2b"]["count"],
            },
        },
        "refund_breakdown": {
            "retail": round(sum(_num(r.total) for r in retail_refunds), 2),
            "b2b": round(sum(_num(r.total) for r in b2b_refunds), 2),
        },
        "daily": daily_rows,
        "top_products": [{"name": name, "qty": round(values["qty"], 2), "revenue": round(values["revenue"], 2)} for name, values in top_products],
        "pos_records": _paginate_rows(pos_records, skip, limit, include_all=include_all),
        "b2b_records": _paginate_rows(b2b_records, skip, limit, include_all=include_all),
        "b2b_payment_records": _paginate_rows(b2b_payment_records, skip, limit, include_all=include_all),
        "refund_records": _paginate_rows(refund_records, skip, limit, include_all=include_all),
        "pos_count": len(pos_invoices),
        "b2b_count": len(b2b_invoices),
        "b2b_payment_count": len(b2b_payment_records),
        "refund_count": len(refund_records),
        "date_from": d_from.strftime("%Y-%m-%d"),
        "date_to": d_to.strftime("%Y-%m-%d"),
    }


# ── SALES ──────────────────────────────────────────────
@router.get("/api/sales")
async def sales_report(date_from: Optional[str] = None, date_to: Optional[str] = None, skip: int = 0, limit: int = Query(default=100, le=500), db: AsyncSession = Depends(get_async_session), _=Depends(require_permission("tab_reports_sales"))):
    d_from, d_to = parse_dates(date_from, date_to)
    if (d_to - d_from).days > 366:
        raise HTTPException(status_code=400, detail="Date range cannot exceed 1 year")
    skip, limit = _resolve_pagination(skip, limit)
    return await _build_sales_report(db, d_from=d_from, d_to=d_to, skip=skip, limit=limit)
    result = await db.execute(
        select(Invoice)
        .where(Invoice.created_at >= d_from, Invoice.created_at <= d_to, Invoice.status == "paid")
        .options(selectinload(Invoice.items), selectinload(Invoice.user), selectinload(Invoice.customer))
    )
    pos_invoices = result.scalars().all()
    result = await db.execute(
        select(B2BInvoice)
        .where(B2BInvoice.created_at >= d_from, B2BInvoice.created_at <= d_to)
        .options(selectinload(B2BInvoice.items).selectinload(B2BInvoiceItem.product),
                 selectinload(B2BInvoice.client), selectinload(B2BInvoice.user))
    )
    b2b_invoices = result.scalars().all()
    result = await db.execute(
        select(RetailRefund)
        .where(RetailRefund.created_at >= d_from, RetailRefund.created_at <= d_to)
        .options(selectinload(RetailRefund.customer), selectinload(RetailRefund.user))
    )
    refunds = result.scalars().all()

    pos_total    = sum(float(i.total) for i in pos_invoices)
    b2b_total    = sum(float(i.amount_paid) for i in b2b_invoices)
    refund_total = sum(float(r.total) for r in refunds)
    pos_total    = max(0, pos_total - refund_total)

    daily = {}
    for i in pos_invoices:
        d = i.created_at.strftime("%Y-%m-%d")
        daily.setdefault(d, {"pos": 0, "b2b": 0, "refunds": 0})
        daily[d]["pos"] += float(i.total)
    for i in b2b_invoices:
        d = i.created_at.strftime("%Y-%m-%d")
        daily.setdefault(d, {"pos": 0, "b2b": 0, "refunds": 0})
        daily[d]["b2b"] += float(i.amount_paid)
    for r in refunds:
        d = r.created_at.strftime("%Y-%m-%d")
        daily.setdefault(d, {"pos": 0, "b2b": 0, "refunds": 0})
        daily[d]["refunds"] += float(r.total)
    daily_list = [{"date": k, "pos": round(max(0, v["pos"] - v["refunds"]), 2), "b2b": round(v["b2b"], 2), "refunds": round(v["refunds"], 2), "total": round(max(0, v["pos"] - v["refunds"]) + v["b2b"], 2)} for k, v in sorted(daily.items())]

    product_sales = {}
    for inv in pos_invoices:
        for item in inv.items:
            product_sales.setdefault(item.name, {"qty": 0, "revenue": 0})
            product_sales[item.name]["qty"]     += float(item.qty)
            product_sales[item.name]["revenue"] += float(item.total)
    top = sorted(product_sales.items(), key=lambda x: x[1]["revenue"], reverse=True)[:10]

    # Detailed POS records
    pos_records = []
    for inv in sorted(pos_invoices, key=lambda x: x.created_at, reverse=True):
        items = inv.items
        pos_records.append({
            "invoice_number": inv.invoice_number or f"POS-{inv.id}",
            "datetime": inv.created_at.strftime("%Y-%m-%d %H:%M") if inv.created_at else "—",
            "user_name": inv.user.name if inv.user else "—",
            "payment": inv.payment_method or "—",
            "items": [{"name": it.name, "qty": float(it.qty), "unit_price": float(it.unit_price), "total": float(it.total)} for it in items],
            "total": float(inv.total),
        })

    # Detailed B2B records
    b2b_records = []
    for inv in sorted(b2b_invoices, key=lambda x: x.created_at, reverse=True):
        items_data = []
        for it in inv.items:
            items_data.append({"name": it.product.name if it.product else "—", "qty": float(it.qty), "unit_price": float(it.unit_price), "total": float(it.total)})
        b2b_records.append({
            "invoice_number": inv.invoice_number,
            "client": inv.client.name if inv.client else "—",
            "datetime": inv.created_at.strftime("%Y-%m-%d %H:%M") if inv.created_at else "—",
            "user_name": inv.user.name if inv.user else "—",
            "invoice_type": inv.invoice_type,
            "status": inv.status,
            "items": items_data,
            "total": float(inv.total),
            "amount_paid": float(inv.amount_paid),
            "balance_due": float(inv.total) - float(inv.amount_paid),
        })

    # Detailed refund records
    refund_records = []
    for r in sorted(refunds, key=lambda x: x.created_at, reverse=True):
        refund_records.append({
            "refund_number":  r.refund_number,
            "customer":       r.customer.name if r.customer else "—",
            "datetime":       r.created_at.strftime("%Y-%m-%d %H:%M") if r.created_at else "—",
            "processed_by":   r.user.name if r.user else "—",
            "reason":         r.reason or "—",
            "refund_method":  r.refund_method,
            "total":          float(r.total),
        })

    pos_records = pos_records[skip : skip + limit]
    b2b_records = b2b_records[skip : skip + limit]
    refund_records = refund_records[skip : skip + limit]

    return {"pos_total": round(pos_total, 2), "b2b_total": round(b2b_total, 2),
            "refund_total": round(refund_total, 2),
            "grand_total": round(pos_total + b2b_total, 2),
            "pos_count": len(pos_invoices), "b2b_count": len(b2b_invoices), "refund_count": len(refunds),
            "daily": daily_list, "top_products": [{"name": k, "qty": round(v["qty"], 2), "revenue": round(v["revenue"], 2)} for k, v in top],
            "pos_records": pos_records, "b2b_records": b2b_records, "refund_records": refund_records,
            "date_from": d_from.strftime("%Y-%m-%d"), "date_to": d_to.strftime("%Y-%m-%d")}

@router.get("/export/sales", dependencies=[Depends(require_permission("action_export_excel"))])
async def export_sales(date_from: str = None, date_to: str = None, db: AsyncSession = Depends(get_async_session)):
    d_from, d_to = parse_dates(date_from, date_to)
    data = await _build_sales_report(db, d_from=d_from, d_to=d_to, include_all=True)
    wb = build_report_workbook([
        {
            "sheet_name": "Summary",
            "report_title": "Sales Report Summary",
            "headers": ["Metric", "Value"],
            "rows": [
                ["Gross Sales", data["gross_sales"]],
                ["Refunds", data["refunds"]],
                ["Net Sales", data["net_sales"]],
                ["Cash Collected", data["cash_collected"]],
                ["Outstanding", data["outstanding"]],
                ["POS Gross Sales", data["channels"]["pos"]["gross_sales"]],
                ["POS Cash Collected", data["channels"]["pos"]["cash_collected"]],
                ["POS Outstanding", data["channels"]["pos"]["outstanding"]],
                ["B2B Gross Sales", data["channels"]["b2b"]["gross_sales"]],
                ["B2B Cash Collected", data["channels"]["b2b"]["cash_collected"]],
                ["B2B Outstanding", data["channels"]["b2b"]["outstanding"]],
                ["B2B Client Payments", data["b2b_payment_count"]],
                ["Retail Refunds", data["refund_breakdown"]["retail"]],
                ["B2B Refunds", data["refund_breakdown"]["b2b"]],
            ],
            "metadata": [
                ("Date Range", f"{data['date_from']} to {data['date_to']}"),
                ("POS Invoices", data["pos_count"]),
                ("B2B Invoices", data["b2b_count"]),
                ("B2B Payment Records", data["b2b_payment_count"]),
                ("Refund Records", data["refund_count"]),
            ],
            "column_formats": {"Value": "money"},
            "total_row_indices": {1, 2, 3, 4, 5},
            "tab_color": "1F4E78",
        },
        {
            "sheet_name": "Daily",
            "report_title": "Sales Daily Breakdown",
            "headers": ["Date", "Gross Sales", "Refunds", "Net Sales", "Cash Collected"],
            "rows": [[row["date"], row["gross_sales"], row["refunds"], row["net_sales"], row["cash_collected"]] for row in data["daily"]],
            "metadata": [("Date Range", f"{data['date_from']} to {data['date_to']}")],
            "column_formats": {"Date": "date", "Gross Sales": "money", "Refunds": "money", "Net Sales": "money", "Cash Collected": "money"},
            "tab_color": "2F6F4F",
        },
        {
            "sheet_name": "POS Invoices",
            "report_title": "POS Invoice Detail",
            "headers": ["Invoice #", "Date / Time", "Customer", "User", "Payment", "Status", "Invoice Total", "Cash Collected", "Outstanding"],
            "rows": [[row["invoice_number"], row["datetime"], row["customer"], row["user_name"], row["payment"], row["status"], row["total"], row["cash_collected"], row["outstanding"]] for row in data["pos_records"]],
            "metadata": [("Date Range", f"{data['date_from']} to {data['date_to']}"), ("Records", len(data["pos_records"]))],
            "column_formats": {"Date / Time": "datetime", "Invoice Total": "money", "Cash Collected": "money", "Outstanding": "money"},
            "tab_color": "4F81BD",
        },
        {
            "sheet_name": "B2B Invoices",
            "report_title": "B2B Invoice Detail",
            "headers": ["Invoice #", "Client", "Date / Time", "User", "Type", "Status", "Total Invoiced", "Amount Paid", "Outstanding"],
            "rows": [[row["invoice_number"], row["client"], row["datetime"], row["user_name"], row["invoice_type"], row["status"], row["total"], row["amount_paid"], row["balance_due"]] for row in data["b2b_records"]],
            "metadata": [("Date Range", f"{data['date_from']} to {data['date_to']}"), ("Records", len(data["b2b_records"]))],
            "column_formats": {"Date / Time": "datetime", "Total Invoiced": "money", "Amount Paid": "money", "Outstanding": "money"},
            "tab_color": "C55A11",
        },
        {
            "sheet_name": "B2B Collections",
            "report_title": "B2B Client Payment Detail",
            "headers": ["Reference", "Client", "Date / Time", "User", "Amount", "Notes"],
            "rows": [[row["reference"], row["client"], row["datetime"], row["user_name"], row["amount"], row["notes"]] for row in data["b2b_payment_records"]],
            "metadata": [("Date Range", f"{data['date_from']} to {data['date_to']}"), ("Records", len(data["b2b_payment_records"]))],
            "column_formats": {"Date / Time": "datetime", "Amount": "money"},
            "wrap_columns": {"Notes"},
            "tab_color": "2F6F4F",
        },
        {
            "sheet_name": "Refunds",
            "report_title": "Refund Detail",
            "headers": ["Refund #", "Source", "Counterparty", "Date / Time", "Processed By", "Method", "Reason", "Amount"],
            "rows": [[row["refund_number"], row["source"], row["counterparty"], row["datetime"], row["processed_by"], row["refund_method"], row["reason"], row["total"]] for row in data["refund_records"]],
            "metadata": [("Date Range", f"{data['date_from']} to {data['date_to']}"), ("Records", len(data["refund_records"]))],
            "column_formats": {"Date / Time": "datetime", "Amount": "money"},
            "wrap_columns": {"Reason"},
            "tab_color": "C00000",
        },
        {
            "sheet_name": "Top Products",
            "report_title": "Top Products",
            "headers": ["Product", "Qty Sold", "Gross Sales"],
            "rows": [[row["name"], row["qty"], row["revenue"]] for row in data["top_products"]],
            "metadata": [("Date Range", f"{data['date_from']} to {data['date_to']}")],
            "column_formats": {"Qty Sold": "qty", "Gross Sales": "money"},
            "tab_color": "70AD47",
        },
    ])
    buf = workbook_to_buffer(wb)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=sales_report_{date.today()}.xlsx"},
    )


@router.get("/api/b2b-statement")
async def b2b_statement(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    skip: int = 0,
    limit: int = Query(default=100, le=500),
    db: AsyncSession = Depends(get_async_session),
    _=Depends(require_permission("tab_reports_b2b")),
):
    d_from, d_to = parse_dates(date_from, date_to)
    if (d_to - d_from).days > 366:
        raise HTTPException(status_code=400, detail="Date range cannot exceed 1 year")
    skip, limit = _resolve_pagination(skip, limit)
    res = await db.execute(select(B2BClient).where(B2BClient.is_active == True).order_by(B2BClient.name))
    clients = res.scalars().all()
    result = []
    for client in clients:
        agg_res = await db.execute(
            select(
                func.count(B2BInvoice.id),
                func.sum(B2BInvoice.total),
                func.sum(B2BInvoice.amount_paid),
            ).where(
                B2BInvoice.client_id == client.id,
                B2BInvoice.created_at >= d_from,
                B2BInvoice.created_at <= d_to,
            )
        )
        invoice_count, total_invoiced, total_paid = agg_res.one()
        if not invoice_count:
            continue
        total_invoiced = _num(total_invoiced)
        total_paid = _num(total_paid)
        result.append(
            {
                "id": client.id,
                "name": client.name,
                "phone": client.phone or "-",
                "payment_terms": client.payment_terms or "-",
                "total_invoiced": round(total_invoiced, 2),
                "total_paid": round(total_paid, 2),
                "outstanding": round(total_invoiced - total_paid, 2),
                "invoice_count": int(invoice_count or 0),
            }
        )
    return result[skip : skip + limit]


@router.get("/export/b2b-statement", dependencies=[Depends(require_permission("action_export_excel"))])
async def export_b2b(date_from: str = None, date_to: str = None, db: AsyncSession = Depends(get_async_session)):
    d_from, d_to = parse_dates(date_from, date_to)
    data = await b2b_statement(date_from=date_from, date_to=date_to, skip=0, limit=100000, db=db)
    rows = [[d["name"],d["phone"],d["payment_terms"],d["total_invoiced"],d["total_paid"],d["outstanding"],d["invoice_count"]] for d in data]
    buf = to_xlsx(
        ["Client","Phone","Payment Terms","Total Invoiced","Total Paid","Outstanding","Invoices"],
        rows,
        "B2B Statement",
        report_title="B2B Statement",
        metadata=[("Date Range", f"{d_from.strftime('%Y-%m-%d')} to {d_to.strftime('%Y-%m-%d')}"), ("Rows Exported", len(rows))],
        column_formats={"Total Invoiced": "money", "Total Paid": "money", "Outstanding": "money", "Invoices": "int"},
    )
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=b2b_statement_{date.today()}.xlsx"})


# ── INVENTORY ──────────────────────────────────────────
async def _build_inventory_report(
    db: AsyncSession,
    *,
    mode: str,
    d_from: Optional[datetime] = None,
    d_to: Optional[datetime] = None,
    skip: int = 0,
    limit: int = 100,
    include_all: bool = False,
):
    if mode == "movement":
        move_res = await db.execute(
            select(StockMove)
            .where(StockMove.created_at >= d_from, StockMove.created_at <= d_to)
            .options(selectinload(StockMove.product))
            .order_by(StockMove.created_at.desc(), StockMove.id.desc())
        )
        moves = move_res.scalars().all()
        grouped = {}
        detail_rows = []
        for move in moves:
            product = move.product
            if product is None:
                continue
            if product.id not in grouped:
                grouped[product.id] = {
                    "sku": product.sku,
                    "name": product.name,
                    "category": product.category or "—",
                    "unit": product.unit,
                    "stock_in": 0.0,
                    "stock_out": 0.0,
                    "receipts": 0.0,
                    "sales_usage": 0.0,
                    "spoilage": 0.0,
                    "transfers_in": 0.0,
                    "transfers_out": 0.0,
                    "adjustments_net": 0.0,
                    "net_movement": 0.0,
                }
            row = grouped[product.id]
            qty = abs(_num(move.qty))
            is_in = (move.type or "").lower() == "in"
            signed_qty = qty if is_in else -qty
            row["net_movement"] += signed_qty
            if is_in:
                row["stock_in"] += qty
            else:
                row["stock_out"] += qty
            ref_type = (move.ref_type or "").lower()
            if ref_type == "receipt":
                row["receipts"] += qty
            elif ref_type in {"invoice", "b2b", "consignment"}:
                row["sales_usage"] += qty
            elif ref_type == "spoilage":
                row["spoilage"] += qty
            elif ref_type == "transfer":
                if is_in:
                    row["transfers_in"] += qty
                else:
                    row["transfers_out"] += qty
            else:
                row["adjustments_net"] += signed_qty
            detail_rows.append({
                "date": move.created_at.strftime("%Y-%m-%d %H:%M") if move.created_at else "—",
                "sku": product.sku,
                "product": product.name,
                "transaction_type": ref_type or (move.type or "move"),
                "direction": "in" if is_in else "out",
                "qty": qty,
                "unit": product.unit,
                "reference": f"{move.ref_type or 'move'}:{move.ref_id or ''}".strip(":"),
                "note": move.note or "",
            })
        rows = sorted(grouped.values(), key=lambda x: x["name"])
        for row in rows:
            for key in ("stock_in", "stock_out", "receipts", "sales_usage", "spoilage", "transfers_in", "transfers_out", "adjustments_net", "net_movement"):
                row[key] = round(row[key], 2)
        return {
            "mode": "movement",
            "date_from": d_from.strftime("%Y-%m-%d"),
            "date_to": d_to.strftime("%Y-%m-%d"),
            "products": _paginate_rows(rows, skip, limit, include_all=include_all),
            "detail_rows": detail_rows if include_all else _paginate_rows(detail_rows, skip, limit, include_all=False),
            "total_products": len(rows),
            "summary": {
                "stock_in": round(sum(r["stock_in"] for r in rows), 2),
                "stock_out": round(sum(r["stock_out"] for r in rows), 2),
                "receipts": round(sum(r["receipts"] for r in rows), 2),
                "spoilage": round(sum(r["spoilage"] for r in rows), 2),
            },
        }

    prod_res = await db.execute(select(Product).where(Product.is_active == True).order_by(Product.name))
    products = prod_res.scalars().all()
    rows = []
    dead_stock_cutoff = datetime.now(timezone.utc) - timedelta(days=90)
    dead_stock_count = 0
    for product in products:
        threshold = _num(product.reorder_level if product.reorder_level is not None else product.min_stock if product.min_stock is not None else 5)
        last_move_res = await db.execute(select(func.max(StockMove.created_at)).where(StockMove.product_id == product.id))
        last_move_at = last_move_res.scalar()
        is_dead_stock = _num(product.stock) > 0 and (last_move_at is None or last_move_at < dead_stock_cutoff)
        low_stock = _num(product.stock) <= threshold
        if is_dead_stock:
            dead_stock_count += 1
        rows.append({
            "sku": product.sku,
            "name": product.name,
            "category": product.category or "—",
            "stock": _num(product.stock),
            "unit": product.unit,
            "price": _num(product.price),
            "value": round(_num(product.stock) * _num(product.price), 2),
            "threshold": round(threshold, 2),
            "reorder_qty": round(_num(product.reorder_qty), 2),
            "low_stock": low_stock,
            "dead_stock": is_dead_stock,
            "last_move_at": last_move_at.strftime("%Y-%m-%d") if last_move_at else "—",
        })
    return {
        "mode": "snapshot",
        "date_from": None,
        "date_to": None,
        "products": _paginate_rows(rows, skip, limit, include_all=include_all),
        "total_value": round(sum(r["value"] for r in rows), 2),
        "low_count": sum(1 for r in rows if r["low_stock"]),
        "dead_stock_count": dead_stock_count,
        "total_products": len(rows),
    }


@router.get("/api/inventory")
async def inventory_report(mode: str = "snapshot", date_from: Optional[str] = None, date_to: Optional[str] = None, skip: int = 0, limit: int = Query(default=100, le=500), db: AsyncSession = Depends(get_async_session), _=Depends(require_permission("tab_reports_inventory"))):
    skip, limit = _resolve_pagination(skip, limit)
    if mode == "movement":
        d_from, d_to = parse_dates(date_from, date_to)
        if (d_to - d_from).days > 366:
            raise HTTPException(status_code=400, detail="Date range cannot exceed 1 year")
        return await _build_inventory_report(db, mode="movement", d_from=d_from, d_to=d_to, skip=skip, limit=limit)
    return await _build_inventory_report(db, mode="snapshot", skip=skip, limit=limit)
    prod_res = await db.execute(select(Product).where(Product.is_active == True).order_by(Product.name))
    products = prod_res.scalars().all()
    rows = []
    for p in products:
        in_res  = await db.execute(select(func.sum(StockMove.qty)).where(StockMove.product_id==p.id, StockMove.type=="in"))
        out_res = await db.execute(select(func.sum(StockMove.qty)).where(StockMove.product_id==p.id, StockMove.type=="out"))
        total_in  = float(in_res.scalar() or 0)
        total_out = abs(float(out_res.scalar() or 0))
        rows.append({"sku":p.sku,"name":p.name,"stock":float(p.stock),"unit":p.unit,"price":float(p.price),
            "value":round(float(p.stock)*float(p.price),2),"total_in":round(total_in,2),"total_out":round(total_out,2),"low_stock":float(p.stock)<=5})
    total_value   = round(sum(r["value"] for r in rows), 2)
    low_count     = sum(1 for r in rows if r["low_stock"])
    total_products = len(rows)
    rows = rows[skip : skip + limit]
    return {"products":rows,"total_value":total_value,"low_count":low_count,"total_products":total_products}

@router.get("/export/inventory", dependencies=[Depends(require_permission("action_export_excel"))])
async def export_inventory(mode: str = "snapshot", date_from: Optional[str] = None, date_to: Optional[str] = None, db: AsyncSession = Depends(get_async_session)):
    if mode == "movement":
        d_from, d_to = parse_dates(date_from, date_to)
        data = await _build_inventory_report(db, mode="movement", d_from=d_from, d_to=d_to, include_all=True)
        rows = [[p["sku"], p["name"], p["category"], p["unit"], p["stock_in"], p["stock_out"], p["receipts"], p["sales_usage"], p["spoilage"], p["transfers_in"], p["transfers_out"], p["adjustments_net"], p["net_movement"]] for p in data["products"]]
        buf = to_xlsx(
            ["SKU","Product","Category","Unit","Stock In","Stock Out","Receipts","Sales/Usage","Spoilage","Transfers In","Transfers Out","Adjustments Net","Net Movement"],
            rows,
            "Inventory Movement",
            report_title="Inventory Movement Report",
            metadata=[
                ("Date Range", f"{data['date_from']} to {data['date_to']}"),
                ("Products", data["total_products"]),
            ],
            column_formats={"Stock In": "qty", "Stock Out": "qty", "Receipts": "qty", "Sales/Usage": "qty", "Spoilage": "qty", "Transfers In": "qty", "Transfers Out": "qty", "Adjustments Net": "qty", "Net Movement": "qty"},
        )
    else:
        data = await _build_inventory_report(db, mode="snapshot", include_all=True)
        rows = [[p["sku"], p["name"], p["category"], p["stock"], p["unit"], p["price"], p["value"], p["threshold"], p["reorder_qty"], p["last_move_at"], "YES" if p["low_stock"] else "", "YES" if p["dead_stock"] else ""] for p in data["products"]]
        buf = to_xlsx(
            ["SKU","Product","Category","Stock","Unit","Price (EGP)","Stock Value","Threshold","Reorder Qty","Last Move","Low Stock","Dead Stock"],
            rows,
            "Inventory Snapshot",
            report_title="Inventory Snapshot Report",
            metadata=[
                ("Products", data["total_products"]),
                ("Low Stock Items", data["low_count"]),
                ("Dead Stock Items", data["dead_stock_count"]),
                ("Total Stock Value", f"{data['total_value']:.2f}"),
            ],
            column_formats={"Stock": "qty", "Price (EGP)": "money", "Stock Value": "money", "Threshold": "qty", "Reorder Qty": "qty", "Last Move": "date"},
        )
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=inventory_{date.today()}.xlsx"})


# ── FARM INTAKE ────────────────────────────────────────
async def _build_farm_intake_report(
    db: AsyncSession,
    *,
    d_from: datetime,
    d_to: datetime,
    skip: int = 0,
    limit: int = 100,
    include_all: bool = False,
):
    delivery_res = await db.execute(
        select(FarmDelivery)
        .where(FarmDelivery.delivery_date >= d_from.date(), FarmDelivery.delivery_date <= d_to.date())
        .options(
            selectinload(FarmDelivery.farm),
            selectinload(FarmDelivery.items).selectinload(FarmDeliveryItem.product),
            selectinload(FarmDelivery.user),
        )
        .order_by(FarmDelivery.delivery_date.desc(), FarmDelivery.id.desc())
    )
    deliveries = delivery_res.scalars().all()
    farm_summary = {}
    detail_rows = []
    for delivery in deliveries:
        farm_name = delivery.farm.name if delivery.farm and delivery.farm.name else f"Farm {delivery.farm_id}"
        summary = farm_summary.setdefault(farm_name, {"farm": farm_name, "delivery_count": 0, "line_count": 0, "total_qty": 0.0, "products": defaultdict(float)})
        summary["delivery_count"] += 1
        for item in delivery.items:
            product = item.product
            sku = product.sku if product else "—"
            name = product.name if product else "—"
            qty = _num(item.qty)
            unit = item.unit or (product.unit if product else "")
            summary["line_count"] += 1
            summary["total_qty"] += qty
            summary["products"][f"{sku}|{name}|{unit}"] += qty
            detail_rows.append({
                "farm": farm_name,
                "date": str(delivery.delivery_date),
                "delivery_number": delivery.delivery_number,
                "sku": sku,
                "product": name,
                "qty": round(qty, 2),
                "unit": unit,
                "received_by": delivery.received_by or "—",
                "user_name": delivery.user.name if delivery.user else "—",
                "notes": item.notes or delivery.notes or "",
            })
    summary_rows = []
    for farm_name, summary in sorted(farm_summary.items()):
        top_product = ""
        if summary["products"]:
            top_key, top_qty = max(summary["products"].items(), key=lambda x: x[1])
            _, top_name, top_unit = top_key.split("|")
            top_product = f"{top_name} ({round(top_qty, 2)} {top_unit})"
        summary_rows.append({
            "farm": farm_name,
            "delivery_count": summary["delivery_count"],
            "line_count": summary["line_count"],
            "total_qty": round(summary["total_qty"], 2),
            "top_product": top_product or "—",
        })
    return {
        "date_from": d_from.strftime("%Y-%m-%d"),
        "date_to": d_to.strftime("%Y-%m-%d"),
        "summary": summary_rows,
        "detail": _paginate_rows(detail_rows, skip, limit, include_all=include_all),
        "totals": {
            "delivery_count": len(deliveries),
            "line_count": len(detail_rows),
            "total_qty": round(sum(r["qty"] for r in detail_rows), 2),
            "farm_count": len(summary_rows),
        },
    }


@router.get("/api/farm-intake")
async def farm_intake_report(date_from: Optional[str] = None, date_to: Optional[str] = None, skip: int = 0, limit: int = Query(default=100, le=500), db: AsyncSession = Depends(get_async_session), _=Depends(require_permission("tab_reports_farm"))):
    d_from, d_to = parse_dates(date_from, date_to)
    if (d_to - d_from).days > 366:
        raise HTTPException(status_code=400, detail="Date range cannot exceed 1 year")
    skip, limit = _resolve_pagination(skip, limit)
    return await _build_farm_intake_report(db, d_from=d_from, d_to=d_to, skip=skip, limit=limit)
    farm_res = await db.execute(select(Farm).where(Farm.is_active == 1))
    farms = farm_res.scalars().all()
    # Auto-fix unnamed farms
    default_names = ["Organic Farm", "Regenerative Farm"]
    for i, farm in enumerate(farms):
        if not farm.name or str(farm.name).strip().lower() in ("none", ""):
            farm.name = default_names[i] if i < len(default_names) else f"Farm {farm.id}"
    try: await db.commit()
    except Exception: await db.rollback()
    result = []
    delivery_rows = []
    for farm in farms:
        del_res = await db.execute(
            select(FarmDelivery)
            .where(FarmDelivery.farm_id==farm.id, FarmDelivery.delivery_date>=d_from.date(), FarmDelivery.delivery_date<=d_to.date())
            .options(selectinload(FarmDelivery.items).selectinload(FarmDeliveryItem.product),
                     selectinload(FarmDelivery.user))
        )
        deliveries = del_res.scalars().all()
        product_totals = {}
        for d in deliveries:
            delivery_rows.append({
                "delivery_number": d.delivery_number,
                "farm": farm.name or f"Farm {farm.id}",
                "delivery_date": str(d.delivery_date),
                "received_by": d.received_by or "—",
                "user_name": d.user.name if d.user else "—",
                "total_items": len(d.items),
                "total_qty": round(sum(float(item.qty) for item in d.items), 2),
                "notes": d.notes or "",
            })
            for item in d.items:
                name = item.product.name if item.product else "—"
                unit = item.unit or ""
                key  = f"{name}|{unit}"
                product_totals[key] = product_totals.get(key, 0) + float(item.qty)
        products = [{"name":k.split("|")[0],"unit":k.split("|")[1],"total_qty":round(v,2)} for k,v in sorted(product_totals.items(), key=lambda x: x[1], reverse=True)]
        result.append({"name": farm.name or f"Farm {farm.id}", "delivery_count":len(deliveries), "products":products, "total_qty":round(sum(p["total_qty"] for p in products),2)})
    delivery_rows.sort(key=lambda row: (row["delivery_date"], row["delivery_number"]), reverse=True)
    delivery_rows = delivery_rows[skip : skip + limit]
    return {"farms": result, "deliveries": delivery_rows}

@router.get("/export/farm-intake", dependencies=[Depends(require_permission("action_export_excel"))])
async def export_farm(date_from: str = None, date_to: str = None, db: AsyncSession = Depends(get_async_session)):
    d_from, d_to = parse_dates(date_from, date_to)
    data = await _build_farm_intake_report(db, d_from=d_from, d_to=d_to, include_all=True)
    wb = build_report_workbook([
        {
            "sheet_name": "Farm Intake Summary",
            "report_title": "Farm Intake Summary",
            "headers": ["Farm", "Deliveries", "Line Items", "Total Qty", "Top Product"],
            "rows": [[row["farm"], row["delivery_count"], row["line_count"], row["total_qty"], row["top_product"]] for row in data["summary"]],
            "metadata": [
                ("Date Range", f"{data['date_from']} to {data['date_to']}"),
                ("Farms", data["totals"]["farm_count"]),
                ("Deliveries", data["totals"]["delivery_count"]),
                ("Line Items", data["totals"]["line_count"]),
                ("Total Qty", data["totals"]["total_qty"]),
            ],
            "column_formats": {"Deliveries": "int", "Line Items": "int", "Total Qty": "qty"},
            "tab_color": "70AD47",
        },
        {
            "sheet_name": "Farm Intake Detail",
            "report_title": "Farm Intake Detail",
            "headers": ["Farm", "Date", "Delivery #", "SKU", "Product", "Qty", "Unit", "Received By", "Performed By", "Notes"],
            "rows": [[row["farm"], row["date"], row["delivery_number"], row["sku"], row["product"], row["qty"], row["unit"], row["received_by"], row["user_name"], row["notes"]] for row in data["detail"]],
            "metadata": [("Date Range", f"{data['date_from']} to {data['date_to']}"), ("Rows Exported", len(data["detail"]))],
            "column_formats": {"Date": "date", "Qty": "qty"},
            "wrap_columns": {"Notes"},
            "tab_color": "548235",
        },
    ])
    buf = workbook_to_buffer(wb)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=farm_intake_{date.today()}.xlsx"})


# ── SPOILAGE ───────────────────────────────────────────
@router.get("/api/spoilage")
async def spoilage_report(date_from: Optional[str] = None, date_to: Optional[str] = None, skip: int = 0, limit: int = Query(default=100, le=500), db: AsyncSession = Depends(get_async_session), _=Depends(require_permission("tab_reports_spoilage"))):
    d_from, d_to = parse_dates(date_from, date_to)
    if (d_to - d_from).days > 366:
        raise HTTPException(status_code=400, detail="Date range cannot exceed 1 year")
    skip, limit = _resolve_pagination(skip, limit)
    sp_res = await db.execute(
        select(SpoilageRecord)
        .where(SpoilageRecord.spoilage_date>=d_from.date(), SpoilageRecord.spoilage_date<=d_to.date())
        .order_by(SpoilageRecord.spoilage_date.desc())
        .options(selectinload(SpoilageRecord.product), selectinload(SpoilageRecord.farm), selectinload(SpoilageRecord.user))
    )
    records = sp_res.scalars().all()
    by_product, by_reason, rows = {}, {}, []
    for r in records:
        name = r.product.name if r.product else "—"; unit = r.product.unit if r.product else ""; reason = r.reason or "—"
        by_product[name]  = by_product.get(name, 0)  + float(r.qty)
        by_reason[reason] = by_reason.get(reason, 0) + float(r.qty)
        rows.append({"ref":r.ref_number,"product":name,"qty":float(r.qty),"unit":unit,"reason":reason,"farm":r.farm.name if r.farm else "—","date":str(r.spoilage_date),"user_name":r.user.name if r.user else "—","notes":r.notes or ""})
    total_qty   = round(sum(float(r.qty) for r in records), 2)
    total_count = len(records)
    rows = rows[skip : skip + limit]
    return {"records":rows,"total_qty":total_qty,"total_count":total_count,
            "by_product":[{"name":k,"qty":round(v,2)} for k,v in sorted(by_product.items(),key=lambda x:x[1],reverse=True)[:8]],
            "by_reason": [{"reason":k,"qty":round(v,2)} for k,v in sorted(by_reason.items(), key=lambda x:x[1],reverse=True)]}

@router.get("/export/spoilage", dependencies=[Depends(require_permission("action_export_excel"))])
async def export_spoilage(date_from: str = None, date_to: str = None, db: AsyncSession = Depends(get_async_session)):
    d_from, d_to = parse_dates(date_from, date_to)
    data = await spoilage_report(date_from=date_from, date_to=date_to, skip=0, limit=100000, db=db)
    rows = [[r["ref"],r["product"],r["qty"],r["unit"],r["reason"],r["farm"],r["date"],r["user_name"],r["notes"]] for r in data["records"]]
    buf = to_xlsx(
        ["Ref #","Product","Qty","Unit","Reason","Farm","Date","Performed By","Notes"],
        rows,
        "Spoilage",
        report_title="Spoilage Report",
        metadata=[("Date Range", f"{d_from.strftime('%Y-%m-%d')} to {d_to.strftime('%Y-%m-%d')}"), ("Records", data["total_count"]), ("Total Qty", data["total_qty"])],
        column_formats={"Qty": "qty", "Date": "date"},
        wrap_columns={"Notes", "Reason"},
    )
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=spoilage_{date.today()}.xlsx"})


# ── PRODUCTION ─────────────────────────────────────────
@router.get("/api/production")
async def production_report(date_from: Optional[str] = None, date_to: Optional[str] = None, skip: int = 0, limit: int = Query(default=100, le=500), db: AsyncSession = Depends(get_async_session), _=Depends(require_permission("tab_reports_production"))):
    d_from, d_to = parse_dates(date_from, date_to)
    if (d_to - d_from).days > 366:
        raise HTTPException(status_code=400, detail="Date range cannot exceed 1 year")
    skip, limit = _resolve_pagination(skip, limit)
    batch_res = await db.execute(
        select(ProductionBatch)
        .where(ProductionBatch.created_at>=d_from, ProductionBatch.created_at<=d_to)
        .order_by(ProductionBatch.created_at.desc())
        .options(selectinload(ProductionBatch.inputs).selectinload(BatchInput.product),
                 selectinload(ProductionBatch.outputs).selectinload(BatchOutput.product),
                 selectinload(ProductionBatch.recipe), selectinload(ProductionBatch.user))
    )
    batches = batch_res.scalars().all()
    rows, losses, total_proc, total_pkg = [], [], 0, 0
    for b in batches:
        is_pkg = b.batch_number.startswith("PKG")
        inputs_str  = ", ".join(f"{float(i.qty):.0f}{i.product.unit if i.product else ''} {i.product.name if i.product else '—'}" for i in b.inputs)
        outputs_str = ", ".join(f"{float(o.qty):.0f}{o.product.unit if o.product else ''} {o.product.name if o.product else '—'}" for o in b.outputs)
        rows.append({"batch_number":b.batch_number,"type":"Packaging" if is_pkg else "Processing",
            "recipe":b.recipe.name if b.recipe else "Custom","waste_pct":float(b.waste_pct),
            "notes":b.notes or "","date":b.created_at.strftime("%Y-%m-%d") if b.created_at else "—",
            "inputs_str":inputs_str,"outputs_str":outputs_str,"user_name":b.user.name if b.user else "—"})
        if is_pkg: total_pkg  += 1
        else:      total_proc += 1; losses.append(float(b.waste_pct))
    total_batches = len(rows)
    rows = rows[skip : skip + limit]
    return {"batches":rows,"total_processing":total_proc,"total_packaging":total_pkg,
            "avg_loss_pct":round(sum(losses)/len(losses),2) if losses else 0,"total_batches":total_batches}

@router.get("/export/production", dependencies=[Depends(require_permission("action_export_excel"))])
async def export_production(date_from: str = None, date_to: str = None, db: AsyncSession = Depends(get_async_session)):
    d_from, d_to = parse_dates(date_from, date_to)
    data = await production_report(date_from=date_from, date_to=date_to, skip=0, limit=100000, db=db)
    rows = [[b["batch_number"],b["type"],b["recipe"],b["inputs_str"],b["outputs_str"],b["waste_pct"],b["date"],b["user_name"],b["notes"]] for b in data["batches"]]
    buf = to_xlsx(
        ["Batch #","Type","Recipe","Inputs","Outputs","Loss %","Date","Performed By","Notes"],
        rows,
        "Production",
        report_title="Production Report",
        metadata=[
            ("Date Range", f"{d_from.strftime('%Y-%m-%d')} to {d_to.strftime('%Y-%m-%d')}"),
            ("Total Batches", data["total_batches"]),
            ("Processing Batches", data["total_processing"]),
            ("Packaging Batches", data["total_packaging"]),
            ("Average Loss %", f"{data['avg_loss_pct']:.2f}%"),
        ],
        column_formats={"Loss %": "percent_value", "Date": "date"},
        wrap_columns={"Inputs", "Outputs", "Notes"},
    )
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=production_{date.today()}.xlsx"})


# ── P&L ────────────────────────────────────────────────
@router.get("/api/pl")
async def pl_report(date_from: Optional[str] = None, date_to: Optional[str] = None, db: AsyncSession = Depends(get_async_session), _=Depends(require_permission("tab_reports_pl"))):
    d_from, d_to = parse_dates(date_from, date_to)
    if (d_to - d_from).days > 366:
        raise HTTPException(status_code=400, detail="Date range cannot exceed 1 year")
    acc_res = await db.execute(
        select(Account)
        .options(selectinload(Account.entries).selectinload(JournalEntry.journal))
    )
    accounts = acc_res.scalars().all()

    def acc_entries(acc):
        return [e for e in acc.entries
                if e.journal and d_from <= e.journal.created_at <= d_to]

    def acc_movement(acc):
        entries = acc_entries(acc)
        return sum(float(e.credit) - float(e.debit) for e in entries)

    def entry_details(acc):
        entries = sorted(acc_entries(acc), key=lambda e: e.journal.created_at, reverse=True)
        result = []
        for e in entries:
            j = e.journal
            result.append({
                "date": j.created_at.strftime("%Y-%m-%d %H:%M") if j.created_at else "—",
                "ref_type": j.ref_type or "manual",
                "description": j.description or "—",
                "debit": float(e.debit),
                "credit": float(e.credit),
                "amount": abs(float(e.credit) - float(e.debit)),
            })
        return result

    revenue_lines = []
    for a in accounts:
        if a.type == "revenue":
            mv = round(acc_movement(a), 2)
            if mv != 0:
                revenue_lines.append({"name": a.name, "code": a.code, "amount": round(abs(mv), 2), "entries": entry_details(a)})

    expense_lines = []
    for a in accounts:
        if a.type == "expense":
            mv = round(acc_movement(a), 2)
            if mv != 0:
                expense_lines.append({"name": a.name, "code": a.code, "amount": round(abs(mv), 2), "entries": entry_details(a)})

    total_revenue = sum(r["amount"] for r in revenue_lines)
    total_expense = sum(e["amount"] for e in expense_lines)
    return {"revenue_lines": revenue_lines, "expense_lines": expense_lines,
            "total_revenue": round(total_revenue, 2), "total_expense": round(total_expense, 2),
            "net_profit": round(total_revenue - total_expense, 2),
            "date_from": d_from.strftime("%Y-%m-%d"), "date_to": d_to.strftime("%Y-%m-%d"),
            "used_balance_fallback": False,
            "warning": None if (revenue_lines or expense_lines) else "No journal-backed revenue or expense movement was found in this period. Current account balances were not used as a substitute."}

@router.get("/export/pl", dependencies=[Depends(require_permission("action_export_excel"))])
async def export_pl(date_from: str = None, date_to: str = None, db: AsyncSession = Depends(get_async_session)):
    data = await pl_report(date_from=date_from, date_to=date_to, db=db)
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()

        green_fill  = PatternFill("solid", fgColor="2a7a2a")
        red_fill    = PatternFill("solid", fgColor="8a1a2a")
        blue_fill   = PatternFill("solid", fgColor="1a3a7a")
        white_font  = Font(bold=True, color="FFFFFF", size=11)
        thin  = Side(style="thin", color="CCCCCC")
        bord  = Border(left=thin, right=thin, top=thin, bottom=thin)
        alt   = PatternFill("solid", fgColor="F5FAF5")
        alt2  = PatternFill("solid", fgColor="FFF5F5")
        total_font  = Font(bold=True, size=11)
        total_green = PatternFill("solid", fgColor="D0F0D0")
        total_red   = PatternFill("solid", fgColor="F0D0D0")
        section_font = Font(bold=True, size=12, color="FFFFFF")

        def add_cell(ws, row, col, value, fill=None, font=None, align="left"):
            c = ws.cell(row=row, column=col, value=value)
            c.border = bord
            c.alignment = Alignment(horizontal=align, vertical="center")
            if fill: c.fill = fill
            if font: c.font = font
            return c

        def auto_width(ws):
            for ci, col in enumerate(ws.columns, 1):
                mx = max((len(str(c.value or "")) for c in col), default=10)
                ws.column_dimensions[get_column_letter(ci)].width = min(mx + 4, 50)

        # ── Sheet 1: P&L Summary ──
        ws1 = wb.active
        ws1.title = "P&L Summary"

        # Title row
        ws1.merge_cells("A1:D1")
        tc = ws1["A1"]
        tc.value = f"Profit & Loss Statement  |  {data['date_from']}  →  {data['date_to']}"
        tc.font = Font(bold=True, size=13, color="FFFFFF")
        tc.fill = green_fill
        tc.alignment = Alignment(horizontal="center", vertical="center")
        ws1.row_dimensions[1].height = 28

        # Headers
        ri = 2
        for ci, h in enumerate(["Category", "Code", "Account", "Amount (EGP)"], 1):
            add_cell(ws1, ri, ci, h, fill=PatternFill("solid", fgColor="4a4a4a"), font=Font(bold=True, color="FFFFFF", size=10), align="center")
        ri += 1

        # Revenue section header
        ws1.merge_cells(f"A{ri}:D{ri}")
        c = ws1.cell(row=ri, column=1, value="REVENUE")
        c.font = section_font; c.fill = green_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        ri += 1

        for line in data["revenue_lines"]:
            fill = alt if ri % 2 == 0 else None
            add_cell(ws1, ri, 1, "Revenue", fill=fill)
            add_cell(ws1, ri, 2, line["code"], fill=fill)
            add_cell(ws1, ri, 3, line["name"], fill=fill)
            add_cell(ws1, ri, 4, line["amount"], fill=fill, align="right")
            ri += 1

        add_cell(ws1, ri, 1, "", fill=total_green)
        add_cell(ws1, ri, 2, "", fill=total_green)
        add_cell(ws1, ri, 3, "TOTAL REVENUE", fill=total_green, font=total_font)
        add_cell(ws1, ri, 4, data["total_revenue"], fill=total_green, font=Font(bold=True, size=12), align="right")
        ri += 2

        # Expenses section header
        ws1.merge_cells(f"A{ri}:D{ri}")
        c = ws1.cell(row=ri, column=1, value="EXPENSES")
        c.font = section_font; c.fill = red_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        ri += 1

        for line in data["expense_lines"]:
            fill = alt2 if ri % 2 == 0 else None
            add_cell(ws1, ri, 1, "Expense", fill=fill)
            add_cell(ws1, ri, 2, line["code"], fill=fill)
            add_cell(ws1, ri, 3, line["name"], fill=fill)
            add_cell(ws1, ri, 4, line["amount"], fill=fill, align="right")
            ri += 1

        add_cell(ws1, ri, 1, "", fill=total_red)
        add_cell(ws1, ri, 2, "", fill=total_red)
        add_cell(ws1, ri, 3, "TOTAL EXPENSES", fill=total_red, font=total_font)
        add_cell(ws1, ri, 4, data["total_expense"], fill=total_red, font=Font(bold=True, size=12), align="right")
        ri += 2

        # Net profit/loss
        is_profit = data["net_profit"] >= 0
        net_fill = PatternFill("solid", fgColor="B0E0B0") if is_profit else PatternFill("solid", fgColor="E0B0B0")
        net_font = Font(bold=True, size=13, color="1a5a1a" if is_profit else "8a0000")
        ws1.merge_cells(f"A{ri}:C{ri}")
        c = ws1.cell(row=ri, column=1, value="NET PROFIT" if is_profit else "NET LOSS")
        c.font = net_font; c.fill = net_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = bord
        ws1.cell(row=ri, column=2).border = bord
        ws1.cell(row=ri, column=3).border = bord
        add_cell(ws1, ri, 4, abs(data["net_profit"]), fill=net_fill, font=net_font, align="right")
        ws1.row_dimensions[ri].height = 24

        auto_width(ws1)

        # ── Sheet 2: Revenue Entries ──
        ws2 = wb.create_sheet("Revenue Entries")
        for ci, h in enumerate(["Account Code", "Account Name", "Date", "Type", "Description", "Amount (EGP)"], 1):
            c = ws2.cell(row=1, column=ci, value=h)
            c.fill = green_fill; c.font = white_font; c.border = bord
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[1].height = 22
        ri = 2
        for line in data["revenue_lines"]:
            for entry in line["entries"]:
                fill = alt if ri % 2 == 0 else None
                for ci, val in enumerate([line["code"], line["name"], entry["date"], entry["ref_type"], entry["description"], entry["amount"]], 1):
                    c = ws2.cell(row=ri, column=ci, value=val)
                    c.border = bord
                    c.alignment = Alignment(vertical="center")
                    if fill: c.fill = fill
                ri += 1
        auto_width(ws2)

        # ── Sheet 3: Expense Entries ──
        ws3 = wb.create_sheet("Expense Entries")
        for ci, h in enumerate(["Account Code", "Account Name", "Date", "Type", "Description", "Amount (EGP)"], 1):
            c = ws3.cell(row=1, column=ci, value=h)
            c.fill = red_fill; c.font = white_font; c.border = bord
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws3.row_dimensions[1].height = 22
        ri = 2
        for line in data["expense_lines"]:
            for entry in line["entries"]:
                fill = alt2 if ri % 2 == 0 else None
                for ci, val in enumerate([line["code"], line["name"], entry["date"], entry["ref_type"], entry["description"], entry["amount"]], 1):
                    c = ws3.cell(row=ri, column=ci, value=val)
                    c.border = bord
                    c.alignment = Alignment(vertical="center")
                    if fill: c.fill = fill
                ri += 1
        auto_width(ws3)

        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=pl_report_{date.today()}.xlsx"})
    except ImportError:
        raise Exception("Run: pip install openpyxl --break-system-packages")


# ── TRANSACTIONS ────────────────────────────────────────
async def _build_transactions_report(
    db: AsyncSession,
    *,
    d_from: datetime,
    d_to: datetime,
    source: Optional[str] = None,
):
    from app.models.refund import RetailRefundItem

    b2b_payment_records = await _load_b2b_client_payment_records(db, d_from=d_from, d_to=d_to)
    rows = []

    if not source or source == "pos":
        pos_res = await db.execute(
            select(Invoice)
            .where(Invoice.created_at >= d_from, Invoice.created_at <= d_to)
            .options(selectinload(Invoice.items), selectinload(Invoice.user), selectinload(Invoice.customer))
        )
        for inv in pos_res.scalars().all():
            for item in inv.items:
                rows.append({
                    "date": inv.created_at.strftime("%Y-%m-%d %H:%M") if inv.created_at else "—",
                    "reference": inv.invoice_number,
                    "transaction_type": "POS Sale",
                    "source": "POS",
                    "counterparty_type": "Customer",
                    "counterparty_name": inv.customer.name if inv.customer else "Walk-in",
                    "sku": item.sku or "—",
                    "product": item.name or "—",
                    "qty": _num(item.qty),
                    "unit_price": _num(item.unit_price),
                    "money_effect": _num(item.total),
                    "stock_effect": -_num(item.qty),
                    "direction": "out",
                    "payment_method": inv.payment_method or "cash",
                    "status": inv.status or "—",
                    "user_name": inv.user.name if inv.user else "—",
                    "notes": inv.notes or "",
                })

    if not source or source == "b2b":
        b2b_res = await db.execute(
            select(B2BInvoice)
            .where(B2BInvoice.created_at >= d_from, B2BInvoice.created_at <= d_to)
            .options(selectinload(B2BInvoice.items).selectinload(B2BInvoiceItem.product), selectinload(B2BInvoice.client), selectinload(B2BInvoice.user))
        )
        for inv in b2b_res.scalars().all():
            for item in inv.items:
                product = item.product
                rows.append({
                    "date": inv.created_at.strftime("%Y-%m-%d %H:%M") if inv.created_at else "—",
                    "reference": inv.invoice_number,
                    "transaction_type": "B2B Invoice",
                    "source": "B2B",
                    "counterparty_type": "Client",
                    "counterparty_name": inv.client.name if inv.client else "—",
                    "sku": product.sku if product else "—",
                    "product": product.name if product else "—",
                    "qty": _num(item.qty),
                    "unit_price": _num(item.unit_price),
                    "money_effect": _num(item.total),
                    "stock_effect": -_num(item.qty),
                    "direction": "out",
                    "payment_method": inv.payment_method or inv.invoice_type,
                    "status": inv.status or "—",
                    "user_name": inv.user.name if inv.user else "—",
                    "notes": inv.notes or "",
                })
        for payment in b2b_payment_records:
            rows.append({
                "date": payment["datetime"],
                "reference": payment["reference"],
                "transaction_type": "B2B Client Payment",
                "source": "B2B Collection",
                "counterparty_type": "Client",
                "counterparty_name": payment["client"],
                "sku": "—",
                "product": "Consignment client payment",
                "qty": 0.0,
                "unit_price": payment["amount"],
                "money_effect": payment["amount"],
                "stock_effect": 0.0,
                "direction": "in",
                "payment_method": payment["payment_method"],
                "status": payment["status"],
                "user_name": payment["user_name"],
                "notes": payment["notes"],
            })

    if not source or source == "refund":
        refund_res = await db.execute(
            select(RetailRefund)
            .where(RetailRefund.created_at >= d_from, RetailRefund.created_at <= d_to)
            .options(selectinload(RetailRefund.customer), selectinload(RetailRefund.user), selectinload(RetailRefund.items).selectinload(RetailRefundItem.product))
        )
        for refund in refund_res.scalars().all():
            for item in refund.items:
                product = item.product
                rows.append({
                    "date": refund.created_at.strftime("%Y-%m-%d %H:%M") if refund.created_at else "—",
                    "reference": refund.refund_number,
                    "transaction_type": "Retail Refund",
                    "source": "Refund",
                    "counterparty_type": "Customer",
                    "counterparty_name": refund.customer.name if refund.customer else "—",
                    "sku": product.sku if product else "—",
                    "product": product.name if product else "—",
                    "qty": _num(item.qty),
                    "unit_price": _num(item.unit_price),
                    "money_effect": -_num(item.total),
                    "stock_effect": _num(item.qty),
                    "direction": "in",
                    "payment_method": refund.refund_method or "—",
                    "status": "refunded",
                    "user_name": refund.user.name if refund.user else "—",
                    "notes": refund.reason or "",
                })

    if not source or source == "receive":
        rec_res = await db.execute(
            select(ProductReceipt)
            .where(ProductReceipt.receive_date >= d_from.date(), ProductReceipt.receive_date <= d_to.date())
            .options(selectinload(ProductReceipt.product), selectinload(ProductReceipt.user), selectinload(ProductReceipt.expense))
        )
        for rec in rec_res.scalars().all():
            product = rec.product
            rows.append({
                "date": rec.created_at.strftime("%Y-%m-%d %H:%M") if rec.created_at else (rec.receive_date.isoformat() if rec.receive_date else "—"),
                "reference": rec.ref_number,
                "transaction_type": "Stock Receipt",
                "source": "Receive",
                "counterparty_type": "Supplier",
                "counterparty_name": rec.supplier_ref or "—",
                "sku": product.sku if product else "—",
                "product": product.name if product else "—",
                "qty": _num(rec.qty),
                "unit_price": _num(rec.unit_cost),
                "money_effect": -_num(rec.total_cost),
                "stock_effect": _num(rec.qty),
                "direction": "in",
                "payment_method": rec.expense.payment_method if rec.expense and rec.expense.payment_method else "cash",
                "status": "received",
                "user_name": rec.user.name if rec.user else "—",
                "notes": rec.notes or "",
            })

    if not source or source == "expense":
        receipt_expense_ids_res = await db.execute(select(ProductReceipt.expense_id).where(ProductReceipt.expense_id.is_not(None)))
        receipt_expense_ids = {row[0] for row in receipt_expense_ids_res.all() if row[0] is not None}
        exp_res = await db.execute(
            select(Expense)
            .where(Expense.expense_date >= d_from.date(), Expense.expense_date <= d_to.date())
            .options(selectinload(Expense.category), selectinload(Expense.user))
        )
        for exp in exp_res.scalars().all():
            if exp.id in receipt_expense_ids:
                continue
            rows.append({
                "date": exp.created_at.strftime("%Y-%m-%d %H:%M") if exp.created_at else (exp.expense_date.isoformat() if exp.expense_date else "—"),
                "reference": exp.ref_number,
                "transaction_type": "Expense",
                "source": "Expense",
                "counterparty_type": "Vendor",
                "counterparty_name": exp.vendor or "—",
                "sku": "—",
                "product": exp.category.name if exp.category else "Expense",
                "qty": 0.0,
                "unit_price": _num(exp.amount),
                "money_effect": -_num(exp.amount),
                "stock_effect": 0.0,
                "direction": "out",
                "payment_method": exp.payment_method or "cash",
                "status": "posted",
                "user_name": exp.user.name if exp.user else "—",
                "notes": exp.description or "",
            })

    rows.sort(key=lambda x: x["date"], reverse=True)
    return {
        "rows": rows,
        "total_rows": len(rows),
        "money_in": round(sum(r["money_effect"] for r in rows if r["money_effect"] > 0), 2),
        "money_out": round(abs(sum(r["money_effect"] for r in rows if r["money_effect"] < 0)), 2),
        "net_money": round(sum(r["money_effect"] for r in rows), 2),
        "stock_in": round(sum(r["stock_effect"] for r in rows if r["stock_effect"] > 0), 2),
        "stock_out": round(abs(sum(r["stock_effect"] for r in rows if r["stock_effect"] < 0)), 2),
    }


@router.get("/api/transactions")
async def transactions_report(
    date_from: Optional[str] = None,
    date_to:   Optional[str] = None,
    source:    Optional[str] = None,
    db: AsyncSession = Depends(get_async_session),
    _=Depends(require_permission("tab_reports_transactions")),
):
    d_from, d_to = parse_dates(date_from, date_to)
    return await _build_transactions_report(db, d_from=d_from, d_to=d_to, source=source)
    from app.models.refund import RetailRefundItem
    receipt_expense_ids_res = await db.execute(
        select(ProductReceipt.expense_id).where(ProductReceipt.expense_id.is_not(None))
    )
    receipt_expense_ids = {row[0] for row in receipt_expense_ids_res.all() if row[0] is not None}
    d_from, d_to = parse_dates(date_from, date_to)
    rows   = []

    # POS
    if not source or source == "pos":
        pos_res = await db.execute(
            select(Invoice)
            .where(Invoice.created_at >= d_from, Invoice.created_at <= d_to)
            .order_by(Invoice.created_at.desc())
            .options(selectinload(Invoice.items), selectinload(Invoice.user), selectinload(Invoice.customer))
        )
        for inv in pos_res.scalars().all():
            cname    = inv.customer.name if inv.customer else "Walk-in"
            items    = inv.items
            disc_per = round(float(inv.discount)/len(items), 2) if items else 0
            disc_pct = round(float(inv.discount)/float(inv.subtotal)*100, 1) if float(inv.subtotal) > 0 else 0
            for item in items:
                rows.append({
                    "date":           inv.created_at.strftime("%Y-%m-%d %H:%M") if inv.created_at else "—",
                    "invoice_number": inv.invoice_number,
                    "source":         "POS",
                    "customer":       cname,
                    "sku":            item.sku or "—",
                    "user_name":      inv.user.name if inv.user else "—",
                    "product":        item.name or "—",
                    "qty":            float(item.qty),
                    "unit_price":     float(item.unit_price),
                    "line_total":     float(item.total),
                    "discount":       disc_per,
                    "discount_pct":   disc_pct,
                    "payment_method": inv.payment_method or "cash",
                    "invoice_total":  float(inv.total),
                    "status":         inv.status,
                    "row_type":       "sale",
                })

    # B2B
    if not source or source == "b2b":
        b2b_res = await db.execute(
            select(B2BInvoice)
            .where(B2BInvoice.created_at >= d_from, B2BInvoice.created_at <= d_to)
            .order_by(B2BInvoice.created_at.desc())
            .options(selectinload(B2BInvoice.items).selectinload(B2BInvoiceItem.product),
                     selectinload(B2BInvoice.client), selectinload(B2BInvoice.user))
        )
        for inv in b2b_res.scalars().all():
            cname    = inv.client.name if inv.client else "—"
            items    = inv.items
            disc_per = round(float(inv.discount)/len(items), 2) if items else 0
            disc_pct = round(float(inv.discount)/float(inv.subtotal)*100, 1) if float(inv.subtotal) > 0 else 0
            for item in items:
                product = item.product
                rows.append({
                    "date":           inv.created_at.strftime("%Y-%m-%d %H:%M") if inv.created_at else "—",
                    "invoice_number": inv.invoice_number,
                    "user_name":      inv.user.name if inv.user else "—",
                    "source":         f"B2B ({inv.invoice_type.replace('_', ' ').title()})",
                    "customer":       cname,
                    "sku":            product.sku if product else "—",
                    "product":        product.name if product else "—",
                    "qty":            float(item.qty),
                    "unit_price":     float(item.unit_price),
                    "line_total":     float(item.total),
                    "discount":       disc_per,
                    "discount_pct":   disc_pct,
                    "payment_method": inv.payment_method or inv.invoice_type,
                    "invoice_total":  float(inv.total),
                    "status":         inv.status,
                    "row_type":       "sale",
                })

    # Refunds
    if not source or source == "refund":
        ref_res = await db.execute(
            select(RetailRefund)
            .where(RetailRefund.created_at >= d_from, RetailRefund.created_at <= d_to)
            .order_by(RetailRefund.created_at.desc())
            .options(selectinload(RetailRefund.customer), selectinload(RetailRefund.user),
                     selectinload(RetailRefund.items).selectinload(RetailRefundItem.product))
        )
        for ref in ref_res.scalars().all():
            cname = ref.customer.name if ref.customer else "—"
            for item in ref.items:
                product = item.product
                rows.append({
                    "date":           ref.created_at.strftime("%Y-%m-%d %H:%M") if ref.created_at else "—",
                    "invoice_number": ref.refund_number,
                    "user_name":      ref.user.name if ref.user else "—",
                    "source":         "Refund",
                    "customer":       cname,
                    "sku":            product.sku if product else "—",
                    "product":        product.name if product else "—",
                    "qty":            -float(item.qty),
                    "unit_price":     float(item.unit_price),
                    "line_total":     -float(item.total),
                    "discount":       0,
                    "discount_pct":   0,
                    "payment_method": ref.refund_method,
                    "invoice_total":  -float(ref.total),
                    "status":         "refunded",
                    "row_type":       "refund",
                    "reason":         ref.reason or "—",
                })

    # Receive
    if not source or source == "receive":
        rec_res = await db.execute(
            select(ProductReceipt)
            .where(
                ProductReceipt.receive_date >= d_from.date(),
                ProductReceipt.receive_date <= d_to.date(),
            )
            .order_by(ProductReceipt.receive_date.desc(), ProductReceipt.id.desc())
            .options(
                selectinload(ProductReceipt.product),
                selectinload(ProductReceipt.user),
                selectinload(ProductReceipt.expense),
            )
        )
        for rec in rec_res.scalars().all():
            product = rec.product
            expense = rec.expense
            total_cost = float(rec.total_cost or 0)
            rows.append({
                "date":           rec.created_at.strftime("%Y-%m-%d %H:%M") if rec.created_at else (rec.receive_date.isoformat() if rec.receive_date else "—"),
                "invoice_number": rec.ref_number,
                "user_name":      rec.user.name if rec.user else "—",
                "source":         "Receive",
                "customer":       rec.supplier_ref or "—",
                "sku":            product.sku if product else "—",
                "product":        product.name if product else "—",
                "qty":            float(rec.qty),
                "unit_price":     float(rec.unit_cost or 0),
                "line_total":     -total_cost,
                "discount":       0,
                "discount_pct":   0,
                "payment_method": expense.payment_method if expense and expense.payment_method else "cash",
                "invoice_total":  -total_cost,
                "status":         "received",
                "row_type":       "receipt",
                "reason":         rec.notes or "—",
            })

    # Expenses
    if not source or source == "expense":
        exp_res = await db.execute(
            select(Expense)
            .where(
                Expense.expense_date >= d_from.date(),
                Expense.expense_date <= d_to.date(),
            )
            .order_by(Expense.expense_date.desc(), Expense.id.desc())
            .options(selectinload(Expense.category), selectinload(Expense.user))
        )
        for exp in exp_res.scalars().all():
            if exp.id in receipt_expense_ids:
                continue
            rows.append({
                "date":           exp.created_at.strftime("%Y-%m-%d %H:%M") if exp.created_at else (exp.expense_date.isoformat() if exp.expense_date else "—"),
                "invoice_number": exp.ref_number,
                "user_name":      exp.user.name if exp.user else "—",
                "source":         "Expense",
                "customer":       exp.vendor or "—",
                "sku":            "—",
                "product":        exp.category.name if exp.category else "Expense",
                "qty":            1.0,
                "unit_price":     float(exp.amount),
                "line_total":     -float(exp.amount),
                "discount":       0,
                "discount_pct":   0,
                "payment_method": exp.payment_method or "cash",
                "invoice_total":  -float(exp.amount),
                "status":         "posted",
                "row_type":       "expense",
                "reason":         exp.description or "—",
            })

    rows.sort(key=lambda x: x["date"], reverse=True)
    total_revenue  = sum(r["line_total"] for r in rows)
    total_qty      = sum(r["qty"]        for r in rows)
    total_discount = sum(r["discount"]   for r in rows)
    return {
        "rows":           rows,
        "total_rows":     len(rows),
        "total_revenue":  round(total_revenue,  2),
        "total_qty":      round(total_qty,       2),
        "total_discount": round(total_discount,  2),
    }

@router.get("/export/transactions", dependencies=[Depends(require_permission("action_export_excel"))])
async def export_transactions(date_from: str = None, date_to: str = None, source: str = None, db: AsyncSession = Depends(get_async_session)):
    d_from, d_to = parse_dates(date_from, date_to)
    data = await _build_transactions_report(db, d_from=d_from, d_to=d_to, source=source)
    wb = build_report_workbook([
        {
            "sheet_name": "Summary",
            "report_title": "Transactions Summary",
            "headers": ["Metric", "Value"],
            "rows": [
                ["Money In", data["money_in"]],
                ["Money Out", data["money_out"]],
                ["Net Money", data["net_money"]],
            ],
            "metadata": [("Date Range", f"{d_from.strftime('%Y-%m-%d')} to {d_to.strftime('%Y-%m-%d')}"), ("Source Filter", source or "All"), ("Rows", data["total_rows"])],
            "column_formats": {"Value": "money"},
            "tab_color": "1F4E78",
        },
        {
            "sheet_name": "Transactions",
            "report_title": "Transaction Detail",
            "headers": ["Date","Reference","Transaction Type","Source","Counterparty Type","Counterparty","Performed By","SKU","Product","Qty","Unit Price","Money Effect","Stock Effect","Direction","Payment Method","Status","Notes"],
            "rows": [[r["date"], r["reference"], r["transaction_type"], r["source"], r["counterparty_type"], r["counterparty_name"], r["user_name"], r["sku"], r["product"], r["qty"], r["unit_price"], r["money_effect"], r["stock_effect"], r["direction"], r["payment_method"], r["status"], r["notes"]] for r in data["rows"]],
            "metadata": [("Date Range", f"{d_from.strftime('%Y-%m-%d')} to {d_to.strftime('%Y-%m-%d')}"), ("Source Filter", source or "All"), ("Rows Exported", data["total_rows"])],
            "column_formats": {"Date": "datetime", "Qty": "qty", "Unit Price": "money", "Money Effect": "money", "Stock Effect": "qty"},
            "wrap_columns": {"Notes"},
            "tab_color": "2F6F4F",
        },
    ])
    buf = workbook_to_buffer(wb)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=transactions_{date.today()}.xlsx"})


# ── UI ─────────────────────────────────────────────────
@router.get("/", response_class=HTMLResponse)
def reports_ui(current_user: User = Depends(require_permission("page_reports"))):
    return """<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Reports — Thunder ERP</title>
<script src="/static/theme-init.js"></script>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Outfit:wght@400;500;600;700;800;900&family=JetBrains+Mono:wght@400;500;700&display=swap" rel="stylesheet">
<script src="/static/theme.js"></script>
<style>
:root{
    --bg:#060810;--card:#0f1424;--card2:#151c30;
    --border:rgba(255,255,255,0.06);--border2:rgba(255,255,255,0.11);
    --green:#00ff9d;--blue:#4d9fff;--orange:#fb923c;--teal:#2dd4bf;
    --danger:#ff4d6d;--warn:#ffb547;--lime:#84cc16;--purple:#a855f7;
    --text:#f0f4ff;--sub:#8899bb;--muted:#445066;
    --sans:'Outfit',sans-serif;--mono:'JetBrains Mono',monospace;--r:12px;
}
body.light{
    --bg:#f4f5ef;--surface:#f1f3eb;--card:#eceee6;--card2:#e4e6de;
    --border:rgba(0,0,0,0.08);--border2:rgba(0,0,0,0.14);
    --green:#0f8a43;
    --text:#1a1e14;--sub:#4a5040;--muted:#7b816f;
}
body.light nav{background:rgba(244,245,239,.92);}
body.light .nav-link:hover{background:rgba(0,0,0,.05);}
body.light tr:hover td{background:rgba(0,0,0,.03);}
.topbar-right{display:flex;align-items:center;gap:12px;}
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0;}
body{font-family:var(--sans);background:var(--bg);color:var(--text);min-height:100vh;font-size:14px;}
nav{position:sticky;top:0;z-index:100;display:flex;align-items:center;gap:8px;padding:0 24px;height:58px;background:rgba(10,13,24,.92);backdrop-filter:blur(20px);border-bottom:1px solid var(--border);}
.logo{font-size:17px;font-weight:900;background:linear-gradient(135deg,var(--green),var(--blue));-webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text;text-decoration:none;display:flex;align-items:center;gap:8px;margin-right:10px;}
.nav-link{padding:7px 12px;border-radius:8px;color:var(--sub);font-size:12px;font-weight:600;text-decoration:none;transition:all .2s;}
.nav-link:hover{background:rgba(255,255,255,.05);color:var(--text);}
.nav-link.active{background:rgba(132,204,22,.1);color:var(--lime);}
.nav-spacer{flex:1;}
.content{max-width:1300px;margin:0 auto;padding:28px 24px;display:flex;flex-direction:column;gap:18px;}
.page-title{font-size:24px;font-weight:800;letter-spacing:-.5px;}
.page-sub{color:var(--muted);font-size:13px;margin-top:3px;}
.tabs{display:flex;gap:4px;background:var(--card);border:1px solid var(--border);border-radius:var(--r);padding:4px;flex-wrap:wrap;}
.tab{padding:7px 13px;border-radius:9px;font-size:12px;font-weight:700;cursor:pointer;border:none;background:transparent;color:var(--muted);transition:all .2s;font-family:var(--sans);white-space:nowrap;}
.tab.active{background:var(--card2);color:var(--text);}
.section{display:none;flex-direction:column;gap:16px;}
.section.active{display:flex;}
/* FILTER BAR */
.filter-bar{display:flex;align-items:center;gap:8px;flex-wrap:wrap;background:var(--card);border:1px solid var(--border);border-radius:var(--r);padding:12px 16px;}
.filter-bar label{font-size:11px;font-weight:700;letter-spacing:1px;text-transform:uppercase;color:var(--muted);white-space:nowrap;}
.filter-bar input[type=date]{background:var(--card2);border:1px solid var(--border2);border-radius:8px;padding:7px 11px;color:var(--text);font-family:var(--sans);font-size:13px;outline:none;}
.filter-bar input[type=date]:focus{border-color:var(--lime);}
.filter-sep{width:1px;height:24px;background:var(--border2);margin:0 4px;}
.btn{display:inline-flex;align-items:center;gap:6px;padding:8px 14px;border-radius:var(--r);font-family:var(--sans);font-size:12px;font-weight:700;cursor:pointer;border:none;transition:all .2s;white-space:nowrap;}
.btn-lime {background:linear-gradient(135deg,var(--lime),var(--green));color:#0a1a00;}
.btn-lime:hover{filter:brightness(1.1);}
.btn-excel{background:linear-gradient(135deg,#217346,#1e6b3f);color:white;}
.btn-excel:hover{filter:brightness(1.1);}
.btn-print{background:var(--card2);border:1px solid var(--border2);color:var(--sub);}
.btn-print:hover{border-color:var(--blue);color:var(--blue);}
/* STATS */
.stats-row{display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));gap:12px;}
.stat-card{background:var(--card);border:1px solid var(--border);border-radius:var(--r);padding:14px 16px;position:relative;overflow:hidden;}
.stat-card::before{content:'';position:absolute;top:0;left:0;right:0;height:2px;}
.sc-green::before {background:linear-gradient(90deg,var(--green),transparent);}
.sc-blue::before  {background:linear-gradient(90deg,var(--blue),transparent);}
.sc-orange::before{background:linear-gradient(90deg,var(--orange),transparent);}
.sc-danger::before{background:linear-gradient(90deg,var(--danger),transparent);}
.sc-teal::before  {background:linear-gradient(90deg,var(--teal),transparent);}
.stat-label{font-size:10px;font-weight:700;letter-spacing:1.5px;text-transform:uppercase;color:var(--muted);margin-bottom:6px;}
.stat-value{font-family:var(--mono);font-size:22px;font-weight:700;}
.sv-green {color:var(--green);}
.sv-blue  {color:var(--blue);}
.sv-orange{color:var(--orange);}
.sv-danger{color:var(--danger);}
.sv-teal  {color:var(--teal);}
/* TABLE */
.table-wrap{background:var(--card);border:1px solid var(--border);border-radius:var(--r);overflow:hidden;}
.table-title{padding:12px 16px;font-size:11px;font-weight:700;letter-spacing:.5px;text-transform:uppercase;border-bottom:1px solid var(--border);color:var(--muted);display:flex;justify-content:space-between;align-items:center;}
table{width:100%;border-collapse:collapse;}
thead{background:var(--card2);}
th{text-align:left;font-size:10px;font-weight:700;letter-spacing:1px;text-transform:uppercase;color:var(--muted);padding:10px 14px;}
td{padding:10px 14px;border-top:1px solid var(--border);color:var(--sub);font-size:13px;}
tr:hover td{background:rgba(255,255,255,.02);}
td.name{color:var(--text);font-weight:600;}
td.mono{font-family:var(--mono);}
/* CHARTS */
.two-col{display:grid;grid-template-columns:1fr 1fr;gap:14px;}
.chart-card{background:var(--card);border:1px solid var(--border);border-radius:var(--r);padding:16px;}
.chart-title{font-size:10px;font-weight:700;letter-spacing:1.5px;text-transform:uppercase;color:var(--muted);margin-bottom:12px;}
.bar-row{display:flex;align-items:center;gap:10px;margin-bottom:8px;}
.bar-label{font-size:12px;color:var(--sub);width:180px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;flex-shrink:0;}
.bar-track{flex:1;background:var(--card2);border-radius:4px;height:8px;overflow:hidden;}
.bar-fill{height:100%;border-radius:4px;transition:width .5s ease;}
.bar-val{font-family:var(--mono);font-size:11px;width:70px;text-align:right;flex-shrink:0;}
/* BADGES */
.badge{display:inline-flex;padding:2px 8px;border-radius:20px;font-size:11px;font-weight:700;}
.badge-low{background:rgba(255,77,109,.1);color:var(--danger);}
.badge-ok {background:rgba(0,255,157,.1);color:var(--green);}
/* P&L */
.pl-section{background:var(--card);border:1px solid var(--border);border-radius:var(--r);overflow:hidden;margin-bottom:8px;}
.pl-header{background:var(--card2);padding:10px 16px;font-size:10px;font-weight:700;letter-spacing:1px;text-transform:uppercase;color:var(--muted);}
.pl-row{display:flex;justify-content:space-between;padding:9px 16px;border-top:1px solid var(--border);font-size:13px;}
.pl-row:hover{background:rgba(255,255,255,.02);}
.pl-total{font-weight:700;font-size:14px;background:rgba(0,0,0,.15);}
.pl-net{font-size:16px;font-weight:800;padding:12px 16px;}
/* TOAST */
.toast{position:fixed;bottom:22px;left:50%;transform:translateX(-50%) translateY(16px);background:var(--card2);border:1px solid var(--border2);border-radius:var(--r);padding:12px 20px;font-size:13px;font-weight:600;color:var(--text);opacity:0;pointer-events:none;transition:opacity .25s,transform .25s;z-index:999;}
.toast.show{opacity:1;transform:translateX(-50%) translateY(0);}
::-webkit-scrollbar{width:4px;}
::-webkit-scrollbar-thumb{background:var(--border2);border-radius:4px;}
@media(max-width:700px){.two-col{grid-template-columns:1fr;}}

/* ══════════════ PRINT STYLES ══════════════ */
.print-header{display:none;}
@media print{
    nav,.filter-bar,.tabs,.no-print{display:none!important;}
    body{background:white;color:#111;font-family:Arial,sans-serif;padding:0;}
    .content{padding:10px;max-width:100%;}
    .section{display:flex!important;}
    .section:not(.active){display:none!important;}
    .print-header{display:flex;align-items:center;justify-content:space-between;padding-bottom:14px;margin-bottom:20px;border-bottom:3px solid #2a7a2a;}
    .stat-card{border:1px solid #ddd!important;background:white!important;}
    .stat-label{color:#555!important;}
    .stat-value{color:#111!important;}
    .table-wrap{border:1px solid #ddd!important;background:white!important;}
    table thead{background:#2a7a2a!important;-webkit-print-color-adjust:exact;print-color-adjust:exact;}
    th{color:white!important;}
    td{color:#333!important;border-top:1px solid #eee!important;}
    td.name{color:#111!important;}
    .table-title{color:#555!important;background:white!important;}
    .chart-card{border:1px solid #ddd!important;background:white!important;}
    .chart-title{color:#555!important;}
    .bar-label,.bar-val{color:#333!important;}
    .bar-track{background:#eee!important;}
    .bar-fill{-webkit-print-color-adjust:exact;print-color-adjust:exact;}
    .pl-section{border:1px solid #ddd!important;background:white!important;}
    .pl-header{background:#f0f0f0!important;color:#555!important;-webkit-print-color-adjust:exact;print-color-adjust:exact;}
    .pl-row{color:#333!important;}
    .pl-total{background:#f5f5f5!important;-webkit-print-color-adjust:exact;print-color-adjust:exact;}
    .badge-low{background:#fee!important;color:#c00!important;-webkit-print-color-adjust:exact;print-color-adjust:exact;}
    .badge-ok {background:#efe!important;color:#2a7a2a!important;-webkit-print-color-adjust:exact;print-color-adjust:exact;}
    .two-col{grid-template-columns:1fr 1fr;}
}
</style>
    <script src="/static/auth-guard.js"></script>
</head>
<body>
""" + render_app_header(current_user, "page_reports") + """

<div class="content">
    <div class="no-print">
        <div class="page-title">📊 Reports</div>
        <div class="page-sub">Filter by date period · Export to Excel · Print with logo</div>
    </div>

    <div class="tabs no-print">
        <button class="tab active" onclick="switchTab('sales')">📈 Sales</button>
        <button class="tab"        onclick="switchTab('transactions')">🧾 Transactions</button>
        <button class="tab"        onclick="switchTab('b2b')">🤝 B2B Statement</button>
        <button class="tab"        onclick="switchTab('inventory')">📦 Inventory</button>
        <button class="tab"        onclick="switchTab('farm')">🌾 Farm Intake</button>
        <button class="tab"        onclick="switchTab('spoilage')">🗑 Spoilage</button>
        <button class="tab"        onclick="switchTab('production')">⚙️ Production</button>
        <button class="tab"        onclick="switchTab('pl')">💰 P&amp;L</button>
    </div>

    <!-- ──────────── SALES ──────────── -->
    <div id="section-sales" class="section active">
        <div class="print-header">
            <div style="display:flex;align-items:center;gap:14px">
                <img src="/static/Logo.png" style="height:120px;object-fit:contain">
                <div>
                    <div style="font-size:16px;font-weight:900;color:#2a7a2a">Habiba Organic Farm</div>
                    <div style="font-size:11px;color:#666;margin-top:2px">Commercial registry: 126278 &nbsp;|&nbsp; Tax ID: 560042604</div>
                </div>
            </div>
            <div style="text-align:right">
                <div style="font-size:18px;font-weight:800;color:#2a7a2a">Sales Report</div>
                <div style="font-size:12px;color:#666;margin-top:4px" id="ph-sales-dates"></div>
            </div>
        </div>
        <div class="filter-bar no-print">
            <label>From</label><input type="date" id="sales-from">
            <label>To</label>  <input type="date" id="sales-to">
            <div class="filter-sep"></div>
            <button class="btn btn-lime"  onclick="loadSales()">Apply</button>
            <button class="btn btn-excel" onclick="exportSection('sales')">⬇ Excel</button>
            <button class="btn btn-print" onclick="window.print()">🖨 Print</button>
        </div>
        <div class="stats-row">
            <div class="stat-card sc-green"><div class="stat-label">Net Revenue</div><div class="stat-value sv-green" id="s-total">—</div></div>
            <div class="stat-card sc-blue" ><div class="stat-label">POS Revenue</div> <div class="stat-value sv-blue"  id="s-pos">—</div></div>
            <div class="stat-card sc-orange"><div class="stat-label">B2B Revenue</div><div class="stat-value sv-orange" id="s-b2b">—</div></div>
            <div class="stat-card" style="border-color:rgba(255,77,109,.3);background:rgba(255,77,109,.04);position:relative;overflow:hidden;">
                <div style="position:absolute;top:0;left:0;right:0;height:2px;background:linear-gradient(90deg,#ff4d6d,transparent)"></div>
                <div class="stat-label" style="color:#ff4d6d">↩ Refunds</div>
                <div class="stat-value" style="color:#ff4d6d;font-family:var(--mono)" id="s-refunds">—</div>
                <div style="font-size:11px;color:rgba(255,77,109,.6);margin-top:4px" id="s-refund-count">— refunds</div>
            </div>
        </div>
        <div class="two-col">
            <div class="table-wrap">
                <div class="table-title">Daily Breakdown</div>
                <table><thead><tr><th>Date</th><th>POS</th><th>B2B</th><th style="color:#ff4d6d">Refunds</th><th>Net Total</th></tr></thead>
                <tbody id="sales-daily"></tbody></table>
            </div>
            <div class="chart-card">
                <div class="chart-title">Top Products by Revenue</div>
                <div id="sales-top"></div>
            </div>
        </div>
        <div id="sales-records"></div>
    </div>

    <!-- ──────────── B2B ──────────── -->
    <!-- ── TRANSACTIONS ── -->
    <div id="section-transactions" class="section">
        <div class="print-header">
            <div style="display:flex;align-items:center;gap:14px">
                <img src="/static/Logo.png" style="height:120px;object-fit:contain">
                <div>
                    <div style="font-size:20px;font-weight:900;color:#2a7a2a">Habiba Organic Farm</div>
                    <div style="font-size:13px;color:#555;margin-top:2px">Transactions Report</div>
                    <div style="font-size:12px;color:#555" id="ph-tx-dates"></div>
                </div>
            </div>
        </div>
        <div class="filter-bar no-print">
            <label>From</label><input type="date" id="tx-from">
            <label>To</label><input type="date" id="tx-to">
            <select id="tx-source">
                <option value="">All Sources</option>
                <option value="pos">POS Only</option>
                <option value="b2b">B2B Only</option>
                <option value="refund">Refunds Only</option>
                <option value="receive">Receive Only</option>
                <option value="expense">Expenses Only</option>
            </select>
            <button class="btn btn-lime" onclick="loadTransactions()">Apply</button>
            <div class="filter-sep"></div>
            <button class="btn btn-excel no-print" onclick="exportSection('transactions')">Export Excel</button>
            <button class="btn-print no-print" onclick="printSection()">🖨 Print</button>
        </div>
        <div class="stats-row">
            <div class="stat-card lime" ><div class="stat-label">Total Lines</div><div class="stat-value lime"   id="tx-count">—</div></div>
            <div class="stat-card green"><div class="stat-label">Total Revenue</div><div class="stat-value green" id="tx-revenue">—</div></div>
            <div class="stat-card blue" ><div class="stat-label">Total Qty Sold</div><div class="stat-value blue"  id="tx-qty">—</div></div>
            <div class="stat-card warn" ><div class="stat-label">Total Discount</div><div class="stat-value warn"  id="tx-discount">—</div></div>
        </div>
        <div class="table-wrap">
            <div class="table-title">All Transactions</div>
            <div style="overflow-x:auto">
            <table>
                <thead><tr>
                    <th>Date</th><th>Invoice #</th><th>Source</th><th>Customer</th><th>By</th>
                    <th>SKU</th><th>Product</th><th>QTY</th><th>Unit Price</th>
                    <th>Line Total</th><th>Discount</th><th>Disc %</th>
                    <th>Payment</th><th>Inv. Total</th><th>Status</th>
                </tr></thead>
                <tbody id="tx-body"></tbody>
            </table>
            </div>
        </div>
    </div>

    <div id="section-b2b" class="section">
        <div class="print-header">
            <div style="display:flex;align-items:center;gap:14px">
                <img src="/static/Logo.png" style="height:120px;object-fit:contain">
                <div>
                    <div style="font-size:16px;font-weight:900;color:#2a7a2a">Habiba Organic Farm</div>
                    <div style="font-size:11px;color:#666;margin-top:2px">Commercial registry: 126278 &nbsp;|&nbsp; Tax ID: 560042604</div>
                </div>
            </div>
            <div style="text-align:right">
                <div style="font-size:18px;font-weight:800;color:#2a7a2a">B2B Client Statement</div>
                <div style="font-size:12px;color:#666;margin-top:4px" id="ph-b2b-dates"></div>
            </div>
        </div>
        <div class="filter-bar no-print">
            <label>From</label><input type="date" id="b2b-from">
            <label>To</label>  <input type="date" id="b2b-to">
            <div class="filter-sep"></div>
            <button class="btn btn-lime"  onclick="loadB2B()">Apply</button>
            <button class="btn btn-excel" onclick="exportSection('b2b')">⬇ Excel</button>
            <button class="btn btn-print" onclick="window.print()">🖨 Print</button>
        </div>
        <div class="stats-row">
            <div class="stat-card sc-blue"  ><div class="stat-label">Clients</div>      <div class="stat-value sv-blue"   id="b-clients">—</div></div>
            <div class="stat-card sc-green" ><div class="stat-label">Total Invoiced</div><div class="stat-value sv-green"  id="b-invoiced">—</div></div>
            <div class="stat-card sc-danger"><div class="stat-label">Outstanding</div>   <div class="stat-value sv-danger" id="b-outstanding">—</div></div>
        </div>
        <div class="table-wrap">
            <div class="table-title">Client Statements</div>
            <table><thead><tr><th>Client</th><th>Phone</th><th>Terms</th><th>Invoiced</th><th>Paid</th><th>Outstanding</th><th>Invoices</th></tr></thead>
            <tbody id="b2b-body"></tbody></table>
        </div>
    </div>

    <!-- ──────────── INVENTORY ──────────── -->
    <div id="section-inventory" class="section">
        <div class="print-header">
            <div style="display:flex;align-items:center;gap:14px">
                <img src="/static/Logo.png" style="height:120px;object-fit:contain">
                <div>
                    <div style="font-size:16px;font-weight:900;color:#2a7a2a">Habiba Organic Farm</div>
                    <div style="font-size:11px;color:#666;margin-top:2px">Commercial registry: 126278 &nbsp;|&nbsp; Tax ID: 560042604</div>
                </div>
            </div>
            <div style="text-align:right">
                <div style="font-size:18px;font-weight:800;color:#2a7a2a">Inventory Report</div>
                <div style="font-size:12px;color:#666;margin-top:4px" id="ph-inv-dates"></div>
            </div>
        </div>
        <div class="filter-bar no-print">
            <span style="font-size:12px;color:var(--muted)">Current stock snapshot</span>
            <div class="filter-sep"></div>
            <button class="btn btn-lime"  onclick="loadInventory()">Refresh</button>
            <button class="btn btn-excel" onclick="exportSection('inventory')">⬇ Excel</button>
            <button class="btn btn-print" onclick="window.print()">🖨 Print</button>
        </div>
        <div class="stats-row">
            <div class="stat-card sc-blue"  ><div class="stat-label">Products</div>   <div class="stat-value sv-blue"   id="inv-count">—</div></div>
            <div class="stat-card sc-green" ><div class="stat-label">Stock Value</div><div class="stat-value sv-green"  id="inv-value">—</div></div>
            <div class="stat-card sc-danger"><div class="stat-label">Low Stock</div>   <div class="stat-value sv-danger" id="inv-low">—</div></div>
        </div>
        <div class="table-wrap">
            <div class="table-title">Stock Levels</div>
            <table><thead><tr><th>SKU</th><th>Product</th><th>Stock</th><th>Unit</th><th>Price</th><th>Stock Value</th><th>Status</th></tr></thead>
            <tbody id="inv-body"></tbody></table>
        </div>
    </div>

    <!-- ──────────── FARM ──────────── -->
    <div id="section-farm" class="section">
        <div class="print-header">
            <div style="display:flex;align-items:center;gap:14px">
                <img src="/static/Logo.png" style="height:120px;object-fit:contain">
                <div>
                    <div style="font-size:16px;font-weight:900;color:#2a7a2a">Habiba Organic Farm</div>
                    <div style="font-size:11px;color:#666;margin-top:2px">Commercial registry: 126278 &nbsp;|&nbsp; Tax ID: 560042604</div>
                </div>
            </div>
            <div style="text-align:right">
                <div style="font-size:18px;font-weight:800;color:#2a7a2a">Farm Intake Report</div>
                <div style="font-size:12px;color:#666;margin-top:4px" id="ph-farm-dates"></div>
            </div>
        </div>
        <div class="filter-bar no-print">
            <label>From</label><input type="date" id="farm-from">
            <label>To</label>  <input type="date" id="farm-to">
            <div class="filter-sep"></div>
            <button class="btn btn-lime"  onclick="loadFarm()">Apply</button>
            <button class="btn btn-excel" onclick="exportSection('farm')">⬇ Excel</button>
            <button class="btn btn-print" onclick="window.print()">🖨 Print</button>
        </div>
        <div id="farm-content"></div>
    </div>

    <!-- ──────────── SPOILAGE ──────────── -->
    <div id="section-spoilage" class="section">
        <div class="print-header">
            <div style="display:flex;align-items:center;gap:14px">
                <img src="/static/Logo.png" style="height:120px;object-fit:contain">
                <div>
                    <div style="font-size:16px;font-weight:900;color:#2a7a2a">Habiba Organic Farm</div>
                    <div style="font-size:11px;color:#666;margin-top:2px">Commercial registry: 126278 &nbsp;|&nbsp; Tax ID: 560042604</div>
                </div>
            </div>
            <div style="text-align:right">
                <div style="font-size:18px;font-weight:800;color:#2a7a2a">Spoilage Report</div>
                <div style="font-size:12px;color:#666;margin-top:4px" id="ph-spl-dates"></div>
            </div>
        </div>
        <div class="filter-bar no-print">
            <label>From</label><input type="date" id="spl-from">
            <label>To</label>  <input type="date" id="spl-to">
            <div class="filter-sep"></div>
            <button class="btn btn-lime"  onclick="loadSpoilage()">Apply</button>
            <button class="btn btn-excel" onclick="exportSection('spoilage')">⬇ Excel</button>
            <button class="btn btn-print" onclick="window.print()">🖨 Print</button>
        </div>
        <div class="stats-row">
            <div class="stat-card sc-danger"><div class="stat-label">Total Records</div><div class="stat-value sv-danger" id="spl-count">—</div></div>
            <div class="stat-card sc-orange"><div class="stat-label">Total Qty Lost</div><div class="stat-value sv-orange" id="spl-qty">—</div></div>
        </div>
        <div class="two-col">
            <div class="chart-card"><div class="chart-title">By Product</div><div id="spl-by-product"></div></div>
            <div class="chart-card"><div class="chart-title">By Reason</div> <div id="spl-by-reason"></div></div>
        </div>
        <div class="table-wrap">
            <div class="table-title">All Records</div>
            <table><thead><tr><th>Ref #</th><th>Product</th><th>Qty</th><th>Reason</th><th>Farm</th><th>Date</th><th>By</th><th>Notes</th></tr></thead>
            <tbody id="spl-body"></tbody></table>
        </div>
    </div>

    <!-- ──────────── PRODUCTION ──────────── -->
    <div id="section-production" class="section">
        <div class="print-header">
            <div style="display:flex;align-items:center;gap:14px">
                <img src="/static/Logo.png" style="height:120px;object-fit:contain">
                <div>
                    <div style="font-size:16px;font-weight:900;color:#2a7a2a">Habiba Organic Farm</div>
                    <div style="font-size:11px;color:#666;margin-top:2px">Commercial registry: 126278 &nbsp;|&nbsp; Tax ID: 560042604</div>
                </div>
            </div>
            <div style="text-align:right">
                <div style="font-size:18px;font-weight:800;color:#2a7a2a">Production Report</div>
                <div style="font-size:12px;color:#666;margin-top:4px" id="ph-prod-dates"></div>
            </div>
        </div>
        <div class="filter-bar no-print">
            <label>From</label><input type="date" id="prod-from">
            <label>To</label>  <input type="date" id="prod-to">
            <div class="filter-sep"></div>
            <button class="btn btn-lime"  onclick="loadProduction()">Apply</button>
            <button class="btn btn-excel" onclick="exportSection('production')">⬇ Excel</button>
            <button class="btn btn-print" onclick="window.print()">🖨 Print</button>
        </div>
        <div class="stats-row">
            <div class="stat-card sc-orange"><div class="stat-label">Processing Batches</div><div class="stat-value sv-orange" id="prod-proc">—</div></div>
            <div class="stat-card sc-teal"  ><div class="stat-label">Packaging Runs</div>   <div class="stat-value sv-teal"   id="prod-pkg">—</div></div>
            <div class="stat-card sc-danger"><div class="stat-label">Avg Loss %</div>        <div class="stat-value sv-danger" id="prod-loss">—</div></div>
        </div>
        <div class="table-wrap">
            <div class="table-title">All Batches</div>
            <table><thead><tr><th>Batch #</th><th>Type</th><th>Recipe</th><th>Inputs</th><th>Outputs</th><th>Loss %</th><th>Date</th><th>By</th></tr></thead>
            <tbody id="prod-body"></tbody></table>
        </div>
    </div>

    <!-- ──────────── P&L ──────────── -->
    <div id="section-pl" class="section">
        <div class="print-header">
            <div style="display:flex;align-items:center;gap:14px">
                <img src="/static/Logo.png" style="height:120px;object-fit:contain">
                <div>
                    <div style="font-size:16px;font-weight:900;color:#2a7a2a">Habiba Organic Farm</div>
                    <div style="font-size:11px;color:#666;margin-top:2px">Commercial registry: 126278 &nbsp;|&nbsp; Tax ID: 560042604</div>
                </div>
            </div>
            <div style="text-align:right">
                <div style="font-size:18px;font-weight:800;color:#2a7a2a">Profit &amp; Loss Statement</div>
                <div style="font-size:12px;color:#666;margin-top:4px" id="ph-pl-dates"></div>
            </div>
        </div>
        <div class="filter-bar no-print">
            <label>From</label><input type="date" id="pl-from">
            <label>To</label>  <input type="date" id="pl-to">
            <div class="filter-sep"></div>
            <button class="btn btn-lime"  onclick="loadPL()">Apply</button>
            <button class="btn btn-excel" onclick="exportSection('pl')">⬇ Excel</button>
            <button class="btn btn-print" onclick="window.print()">🖨 Print</button>
        </div>
        <div id="pl-content"></div>
    </div>

</div><!-- end .content -->
<div class="toast" id="toast"></div>

<script>
  let __currentUser = null;

  // Theme is handled by window.__appTheme (theme.js / theme-init.js).
  // toggleMode is kept as a named alias so window.__appNav.toggleTheme() can find it.
  function toggleMode(){
    if(window.__appTheme) window.__appTheme.toggle();
  }

async function initUser() {
    try {
        const r = await fetch("/auth/me");
        if (!r.ok) { _redirectToLogin(); return; }
        const u = await r.json();
        __currentUser = u;
        return u;
    } catch(e) { _redirectToLogin(); }
}

  function hasPermission(permission){
      const role = __currentUser ? (__currentUser.role || "") : "";
      let permsArray = [];
      if (__currentUser && __currentUser.permissions) {
          permsArray = typeof __currentUser.permissions === "string" 
              ? __currentUser.permissions.split(",").map(v => String(v).trim()).filter(Boolean) 
              : __currentUser.permissions;
      }
      const perms = new Set(permsArray);
      return role === "admin" || perms.has(permission);
  }
  function configureReportsPermissions(){
      ensureTabMetadata();
      let firstAllowed = null;
      document.querySelectorAll(".tabs .tab").forEach((btn) => {
          const tab = btn.dataset.tab;
          if(!tab) return;
          if(!isTabAllowed(tab)){
              btn.style.display = "none";
          } else if(!firstAllowed) {
              firstAllowed = tab;
          }
      });
      if(!isTabAllowed(currentTab) && firstAllowed){
          currentTab = firstAllowed;
      }
      if(!hasPermission("action_export_excel")){
          document.querySelectorAll(".btn-excel").forEach(btn => btn.style.display = "none");
      }
  }
  let currentTab = "sales";
let toastTimer = null;
initUser().then(u => {
    if(!u) return;
    configureReportsPermissions();
    switchTab(currentTab);
});

function switchTab(tab){
    ensureTabMetadata();
    if(!isTabAllowed(tab)){
        const fallback = REPORT_TAB_ORDER.find(isTabAllowed);
        if(!fallback){
            showToast("No report tabs available for this account.");
            return;
        }
        tab = fallback;
    }
    currentTab = tab;
    document.querySelectorAll(".tab").forEach(btn => btn.classList.toggle("active", btn.dataset.tab===tab));
    document.querySelectorAll(".section").forEach(s => s.classList.remove("active"));
    const section = document.getElementById("section-"+tab);
    if(!section) return;
    section.classList.add("active");
    const loaders = {sales:loadSales, transactions:loadTransactions, b2b:loadB2B, inventory:loadInventory, farm:loadFarm, spoilage:loadSpoilage, production:loadProduction, pl:loadPL};
    if(loaders[tab]){
        loaders[tab]();
    } else {
        setSectionStatus(tab, "error", "This report tab is not wired correctly.");
    }
}

function today(){ return new Date().toISOString().split("T")[0]; }
function monthStart(){ let d=new Date(); d.setDate(1); return d.toISOString().split("T")[0]; }
function yearStart() { let d=new Date(); d.setMonth(0); d.setDate(1); return d.toISOString().split("T")[0]; }
function getRange(f, t){ return {from: document.getElementById(f).value, to: document.getElementById(t).value}; }
function setEl(id, v){ document.getElementById(id).value = v; }
function setPrintDates(id, from, to){ let el=document.getElementById(id); if(el) el.innerText=`Period: ${from}  →  ${to}`; }

function showToast(msg){
    let t=document.getElementById("toast"); t.innerText=msg; t.classList.add("show");
    clearTimeout(toastTimer); toastTimer=setTimeout(()=>t.classList.remove("show"),3000);
}

const REPORT_TAB_ORDER = ["sales","transactions","b2b","inventory","farm","spoilage","production","pl"];
const REPORT_TAB_PERMISSIONS = {
    sales: "tab_reports_sales",
    transactions: "tab_reports_transactions",
    b2b: "tab_reports_b2b",
    inventory: "tab_reports_inventory",
    farm: "tab_reports_farm",
    spoilage: "tab_reports_spoilage",
    production: "tab_reports_production",
    pl: "tab_reports_pl",
};

function ensureTabMetadata(){
    document.querySelectorAll(".tabs .tab").forEach((btn, index) => {
        if(!btn.dataset.tab && REPORT_TAB_ORDER[index]){
            btn.dataset.tab = REPORT_TAB_ORDER[index];
        }
    });
}

function getTabPermission(tab){
    return REPORT_TAB_PERMISSIONS[tab] || null;
}

function isTabAllowed(tab){
    const permission = getTabPermission(tab);
    return !permission || hasPermission(permission);
}

function getSectionElement(tab){
    return document.getElementById(`section-${tab}`);
}

function ensureSectionStatus(tab){
    const section = getSectionElement(tab);
    if(!section) return null;
    let status = section.querySelector(".report-status");
    if(!status){
        status = document.createElement("div");
        status.className = "report-status";
        status.style.cssText = "display:none;margin:0 0 14px;padding:12px 14px;border-radius:10px;border:1px solid var(--border2);font-size:13px;line-height:1.5;";
        const firstChild = section.firstElementChild;
        if(firstChild && firstChild.classList.contains("print-header")){
            firstChild.insertAdjacentElement("afterend", status);
        } else {
            section.prepend(status);
        }
    }
    return status;
}

function setSectionStatus(tab, kind, message){
    const status = ensureSectionStatus(tab);
    if(!status) return;
    if(!message){
        status.style.display = "none";
        status.textContent = "";
        return;
    }
    const palettes = {
        info: "background:rgba(77,159,255,.08);border-color:rgba(77,159,255,.25);color:var(--blue);",
        error: "background:rgba(255,77,109,.08);border-color:rgba(255,77,109,.25);color:var(--danger);",
        empty: "background:rgba(255,181,71,.08);border-color:rgba(255,181,71,.25);color:var(--warn);",
    };
    status.style.cssText = `display:block;margin:0 0 14px;padding:12px 14px;border-radius:10px;border:1px solid var(--border2);font-size:13px;line-height:1.5;${palettes[kind] || palettes.info}`;
    status.textContent = message;
}

async function fetchReportJson(url){
    const response = await fetch(url, { credentials: "same-origin" });
    const contentType = response.headers.get("content-type") || "";
    let payload = null;
    if(contentType.includes("application/json")){
        payload = await response.json().catch(() => null);
    } else {
        const text = await response.text().catch(() => "");
        if(!response.ok){
            throw new Error(text || `Request failed (${response.status})`);
        }
        throw new Error("Unexpected non-JSON response from reports endpoint.");
    }
    if(!response.ok){
        const detail = payload && (payload.detail || payload.message || payload.error);
        throw new Error(detail || `Request failed (${response.status})`);
    }
    return payload;
}

async function runReportLoader(tab, loader){
    if(!isTabAllowed(tab)){
        setSectionStatus(tab, "error", "You do not have permission to view this report.");
        return;
    }
    setSectionStatus(tab, "info", "Loading report...");
    try{
        await loader();
        setSectionStatus(tab, "", "");
    } catch(error){
        console.error(`Report load failed for ${tab}:`, error);
        setSectionStatus(tab, "error", error && error.message ? error.message : "Could not load this report.");
        showToast(`Could not load ${tab} report`);
    }
}

function getDownloadFilename(response, fallback){
    const disposition = response.headers.get("Content-Disposition") || "";
    const utf8Match = disposition.match(/filename\\*=UTF-8''([^;]+)/i);
    if(utf8Match) return decodeURIComponent(utf8Match[1]);
    const plainMatch = disposition.match(/filename="?([^\";]+)"?/i);
    return plainMatch ? plainMatch[1] : fallback;
}

async function exportSection(tab){
    const build = {
        sales:      ()=>{ let r=getRange("sales-from","sales-to"); return `/reports/export/sales?date_from=${r.from}&date_to=${r.to}`; },
        b2b:        ()=>{ let r=getRange("b2b-from","b2b-to");     return `/reports/export/b2b-statement?date_from=${r.from}&date_to=${r.to}`; },
        inventory:  ()=>{ let mode=document.getElementById("inv-mode")?.value || "snapshot"; let from=document.getElementById("inv-from")?.value || ""; let to=document.getElementById("inv-to")?.value || ""; return `/reports/export/inventory?mode=${mode}${mode==="movement"?`&date_from=${from}&date_to=${to}`:""}`; },
        farm:       ()=>{ let r=getRange("farm-from","farm-to");   return `/reports/export/farm-intake?date_from=${r.from}&date_to=${r.to}`; },
        spoilage:   ()=>{ let r=getRange("spl-from","spl-to");     return `/reports/export/spoilage?date_from=${r.from}&date_to=${r.to}`; },
        production: ()=>{ let r=getRange("prod-from","prod-to");   return `/reports/export/production?date_from=${r.from}&date_to=${r.to}`; },
        pl:           ()=>{ let r=getRange("pl-from","pl-to");   return `/reports/export/pl?date_from=${r.from}&date_to=${r.to}`; },
        transactions: ()=>{ let r=getRange("tx-from","tx-to"); let s=document.getElementById("tx-source").value; return `/reports/export/transactions?date_from=${r.from}&date_to=${r.to}${s?"&source="+s:""}`; },
    };
    const fallbackFilename = `${tab}_report.xlsx`;
    showToast("Preparing Excel...");
    try{
        const response = await fetch(build[tab](), { credentials: "same-origin" });
        if(!response.ok){
            const contentType = response.headers.get("content-type") || "";
            let message = `Excel export failed (${response.status}).`;
            if(contentType.includes("application/json")){
                const payload = await response.json().catch(()=>null);
                if(payload && payload.detail) message = payload.detail;
            } else if(response.status >= 500){
                message = "Excel export failed on the server. Please try again.";
            } else {
                const text = await response.text().catch(()=> "");
                if(text && !text.trim().startsWith("<!DOCTYPE") && !text.trim().startsWith("<html")){
                    message = text.trim().slice(0, 160);
                }
            }
            showToast(message);
            return;
        }
        const blob = await response.blob();
        const filename = getDownloadFilename(response, fallbackFilename);
        const blobUrl = URL.createObjectURL(blob);
        const link = document.createElement("a");
        link.href = blobUrl;
        link.download = filename;
        document.body.appendChild(link);
        link.click();
        link.remove();
        setTimeout(() => URL.revokeObjectURL(blobUrl), 1000);
        showToast("Excel downloaded");
    } catch(e){
        console.error("Export failed:", e);
        showToast("Excel export failed. Please check your connection and try again.");
    }
}

/* ── TRANSACTIONS ── */
async function loadTransactions(){
    let r      = getRange("tx-from","tx-to");
    let source = document.getElementById("tx-source").value;
    let url    = `/reports/api/transactions?date_from=${r.from}&date_to=${r.to}${source?"&source="+source:""}`;
    let data   = await fetchReportJson(url);
    const statsRow = document.querySelector("#section-transactions .stats-row");
    if (statsRow) {
        statsRow.innerHTML = `
            <div class="stat-card lime"><div class="stat-label">Rows</div><div class="stat-value lime" id="tx-count">${data.total_rows}</div></div>
            <div class="stat-card green"><div class="stat-label">Money In</div><div class="stat-value green" id="tx-money-in">${data.money_in.toFixed(2)}</div></div>
            <div class="stat-card warn"><div class="stat-label">Money Out</div><div class="stat-value warn" id="tx-money-out">${data.money_out.toFixed(2)}</div></div>
            <div class="stat-card blue"><div class="stat-label">Stock In</div><div class="stat-value blue" id="tx-stock-in">${data.stock_in.toFixed(2)}</div></div>
            <div class="stat-card sc-danger"><div class="stat-label">Stock Out</div><div class="stat-value sv-danger" id="tx-stock-out">${data.stock_out.toFixed(2)}</div></div>
        `;
    }
    const txTableHead = document.querySelector("#tx-body")?.closest("table")?.querySelector("thead tr");
    if (txTableHead) {
        txTableHead.innerHTML = "<th>Date</th><th>Reference</th><th>Type</th><th>Source</th><th>Counterparty Type</th><th>Counterparty</th><th>By</th><th>SKU</th><th>Product</th><th>Qty</th><th>Unit Price</th><th>Money Effect</th><th>Stock Effect</th><th>Direction</th><th>Payment</th><th>Status</th>";
    }
    setPrintDates("ph-tx-dates", r.from, r.to);
    document.getElementById("tx-body").innerHTML = data.rows.length
        ? data.rows.map(r => `<tr>
            <td class="mono" style="font-size:11px;white-space:nowrap">${r.date}</td>
            <td class="mono" style="font-size:11px;color:var(--blue)">${r.reference}</td>
            <td>${r.transaction_type}</td>
            <td>${r.source}</td>
            <td>${r.counterparty_type}</td>
            <td class="name">${r.counterparty_name}</td>
            <td style="font-size:12px;color:var(--muted)">${r.user_name}</td>
            <td class="mono" style="font-size:11px;color:var(--muted)">${r.sku}</td>
            <td>${r.product}${r.notes?`<br><span style="font-size:10px;color:var(--muted)">${r.notes}</span>`:""}</td>
            <td class="mono">${r.qty.toFixed(2)}</td>
            <td class="mono">${r.unit_price.toFixed(2)}</td>
            <td class="mono" style="color:${r.money_effect>=0?"var(--green)":"var(--danger)"};font-weight:700">${r.money_effect.toFixed(2)}</td>
            <td class="mono" style="color:${r.stock_effect>=0?"var(--green)":"var(--danger)"};font-weight:700">${r.stock_effect.toFixed(2)}</td>
            <td>${r.direction}</td>
            <td>${r.payment_method}</td>
            <td>${r.status}</td>
        </tr>`).join("")
        : `<tr><td colspan="16" style="text-align:center;color:var(--muted);padding:40px">No transactions in this period</td></tr>`;
    return;

    document.getElementById("tx-count").innerText   = data.total_rows;
    document.getElementById("tx-revenue").innerText = data.total_revenue.toFixed(2);
    document.getElementById("tx-qty").innerText     = data.total_qty.toFixed(2);
    document.getElementById("tx-discount").innerText= data.total_discount.toFixed(2);
    setPrintDates("ph-tx-dates", r.from, r.to);

    const payColor = (m) => {
        if(!m) return "var(--muted)";
        m = m.toLowerCase();
        if(m.includes("visa") || m.includes("card")) return "var(--blue)";
        if(m.includes("cash"))                        return "var(--green)";
        if(m.includes("consign"))                     return "var(--teal)";
        if(m.includes("transfer"))                    return "var(--purple)";
        if(m.includes("credit") || m.includes("exchange") || m.includes("refund")) return "var(--danger)";
        return "var(--sub)";
    };
    const statusColor = (s) => {
        if(s==="paid")      return "var(--green)";
        if(s==="unpaid")    return "var(--warn)";
        if(s==="partial")   return "var(--blue)";
        if(s==="consignment") return "var(--teal)";
        if(s==="refunded")  return "var(--danger)";
        if(s==="received")  return "var(--warn)";
        if(s==="posted")    return "var(--purple)";
        return "var(--muted)";
    };

    document.getElementById("tx-body").innerHTML = data.rows.length
        ? data.rows.map(r => {
            const isRef = r.row_type === "refund";
            const isReceipt = r.row_type === "receipt";
            const isExpense = r.row_type === "expense";
            const isOutflow = isRef || isReceipt || isExpense;
            const rowStyle = isRef
                ? 'style="background:rgba(255,77,109,.04);"'
                : isReceipt
                    ? 'style="background:rgba(255,181,71,.06);"'
                    : isExpense
                        ? 'style="background:rgba(168,85,247,.06);"'
                            : '';
            const numColor = isRef ? "var(--danger)" : isReceipt ? "var(--warn)" : isExpense ? "var(--purple)" : "var(--green)";
            const refBadge = isRef
                ? `<br><span style="font-size:9px;font-weight:800;letter-spacing:.5px;color:var(--danger);background:rgba(255,77,109,.15);padding:1px 5px;border-radius:4px">↩ REFUND</span>`
                : isReceipt
                    ? `<br><span style="font-size:9px;font-weight:800;letter-spacing:.5px;color:var(--warn);background:rgba(255,181,71,.16);padding:1px 5px;border-radius:4px">↑ RECEIVE</span>`
                    : isExpense
                        ? `<br><span style="font-size:9px;font-weight:800;letter-spacing:.5px;color:var(--purple);background:rgba(168,85,247,.16);padding:1px 5px;border-radius:4px">↓ EXPENSE</span>`
                            : "";
            return `<tr ${rowStyle}>
                <td class="mono" style="font-size:11px;white-space:nowrap">${r.date}</td>
                <td class="mono" style="font-size:11px;color:${isRef?"var(--danger)":isReceipt?"var(--warn)":isExpense?"var(--purple)":"var(--lime)"}">${r.invoice_number}${refBadge}</td>
                <td style="font-size:11px;color:${isRef?"var(--danger)":isReceipt?"var(--warn)":isExpense?"var(--purple)":"var(--sub)"}">${r.source}</td>
                <td class="name" style="white-space:nowrap">${r.customer}</td>
                <td style="font-size:12px;color:var(--muted);white-space:nowrap">${r.user_name}</td>
                <td class="mono" style="font-size:11px;color:var(--muted)">${r.sku}</td>
                <td style="font-weight:600;white-space:nowrap">${r.product}${isOutflow&&r.reason?`<br><span style="font-size:10px;color:var(--muted);font-weight:400">${r.reason}</span>`:""}</td>
                <td class="mono" style="color:${isRef?"var(--danger)":isReceipt?"var(--amber)":isExpense?"var(--purple)":"var(--blue)"};font-weight:700">${r.qty.toFixed(2)}</td>
                <td class="mono">${r.unit_price.toFixed(2)}</td>
                <td class="mono" style="color:${numColor};font-weight:700">${isOutflow?"−":""}${Math.abs(r.line_total).toFixed(2)}</td>
                <td class="mono" style="color:${r.discount>0?"var(--warn)":"var(--muted)"}">${r.discount>0?"-"+r.discount.toFixed(2):"—"}</td>
                <td class="mono" style="color:${r.discount_pct>0?"var(--warn)":"var(--muted)"}">${r.discount_pct>0?r.discount_pct.toFixed(1)+"%":"—"}</td>
                <td style="font-size:12px;font-weight:700;color:${payColor(r.payment_method)}">${r.payment_method}</td>
                <td class="mono" style="font-weight:700;color:${isOutflow?"var(--danger)":"inherit"}">${isOutflow?"−":""}${Math.abs(r.invoice_total).toFixed(2)}</td>
                <td><span style="font-size:11px;font-weight:700;padding:2px 8px;border-radius:20px;background:rgba(0,0,0,.2);color:${statusColor(r.status)}">${r.status}</span></td>
              </tr>`;
        }).join("")
        : `<tr><td colspan="15" style="text-align:center;color:var(--muted);padding:40px">No transactions in this period</td></tr>`;
}

/* ── SALES ── */
async function loadSales(){
    let r = getRange("sales-from","sales-to");
    let data = await fetchReportJson(`/reports/api/sales?date_from=${r.from}&date_to=${r.to}`);
    const statsRow = document.querySelector("#section-sales .stats-row");
    if (statsRow) {
        statsRow.innerHTML = `
            <div class="stat-card sc-blue"><div class="stat-label">Gross Sales</div><div class="stat-value sv-blue" id="s-gross">${data.gross_sales.toFixed(2)}</div></div>
            <div class="stat-card" style="border-color:rgba(255,77,109,.3);background:rgba(255,77,109,.04);position:relative;overflow:hidden;">
                <div style="position:absolute;top:0;left:0;right:0;height:2px;background:linear-gradient(90deg,#ff4d6d,transparent)"></div>
                <div class="stat-label" style="color:#ff4d6d">Refunds</div>
                <div class="stat-value" style="color:#ff4d6d;font-family:var(--mono)" id="s-refunds">${data.refunds > 0 ? "−" + data.refunds.toFixed(2) : "0.00"}</div>
                <div style="font-size:11px;color:rgba(255,77,109,.6);margin-top:4px" id="s-refund-count">${data.refund_count} refund${data.refund_count !== 1 ? "s" : ""}</div>
            </div>
            <div class="stat-card sc-green"><div class="stat-label">Net Sales</div><div class="stat-value sv-green" id="s-net">${data.net_sales.toFixed(2)}</div></div>
            <div class="stat-card sc-orange"><div class="stat-label">Cash Collected</div><div class="stat-value sv-orange" id="s-collected">${data.cash_collected.toFixed(2)}</div></div>
            <div class="stat-card sc-danger"><div class="stat-label">Outstanding</div><div class="stat-value sv-danger" id="s-outstanding">${data.outstanding.toFixed(2)}</div></div>`;
    }
    const salesHead = document.querySelector("#sales-daily")?.closest("table")?.querySelector("thead tr");
    if (salesHead) {
        salesHead.innerHTML = "<th>Date</th><th>Gross Sales</th><th style='color:#ff4d6d'>Refunds</th><th>Net Sales</th><th>Cash Collected</th>";
    }
    setPrintDates("ph-sales-dates", data.date_from, data.date_to);
    document.getElementById("sales-daily").innerHTML = data.daily.length
        ? data.daily.map(d=>`<tr>
            <td class="mono">${d.date}</td>
            <td class="mono" style="color:var(--blue)">${d.gross_sales.toFixed(2)}</td>
            <td class="mono" style="color:#ff4d6d;font-weight:${d.refunds>0?700:400}">${d.refunds>0?"−"+d.refunds.toFixed(2):"—"}</td>
            <td class="mono" style="color:var(--green);font-weight:700">${d.net_sales.toFixed(2)}</td>
            <td class="mono" style="color:var(--orange)">${d.cash_collected.toFixed(2)}</td>
          </tr>`).join("")
        : `<tr><td colspan="5" style="text-align:center;color:var(--muted);padding:30px">No sales in this period</td></tr>`;

    let legacyMaxR = data.top_products.length ? data.top_products[0].revenue : 1;
    document.getElementById("sales-top").innerHTML = data.top_products.length
        ? data.top_products.map(p=>`<div class="bar-row">
            <div class="bar-label">${p.name}</div>
            <div class="bar-track"><div class="bar-fill" style="width:${(p.revenue/legacyMaxR*100).toFixed(1)}%;background:linear-gradient(90deg,var(--green),var(--lime))"></div></div>
            <div class="bar-val" style="color:var(--green)">${p.revenue.toFixed(0)}</div>
          </div>`).join("")
        : `<div style="color:var(--muted);font-size:13px">No data</div>`;

    let posHtml = `
        <div class="table-title" style="margin-top:28px">POS Invoices — ${data.pos_records.length} transactions</div>
        <div class="table-wrap">
        <table><thead><tr><th>Invoice #</th><th>Date / Time</th><th>Customer</th><th>By</th><th>Payment</th><th>Items</th><th style="text-align:right">Gross</th><th style="text-align:right">Collected</th><th style="text-align:right">Outstanding</th></tr></thead><tbody>`;
    if(data.pos_records.length){
        posHtml += data.pos_records.map(inv=>`
            <tr>
                <td class="mono" style="font-size:11px;color:var(--blue)">${inv.invoice_number}</td>
                <td class="mono" style="font-size:12px;color:var(--muted)">${inv.datetime}</td>
                <td class="name" style="font-size:12px">${inv.customer}</td>
                <td style="font-size:12px;color:var(--muted);white-space:nowrap">${inv.user_name}</td>
                <td style="font-size:12px">${inv.payment}</td>
                <td style="font-size:12px;color:var(--sub)">
                    ${inv.items.map(it=>`<span style="display:inline-block;background:var(--card2);border:1px solid var(--border2);border-radius:5px;padding:1px 7px;margin:2px;white-space:nowrap">${it.qty%1===0?it.qty.toFixed(0):it.qty.toFixed(2)} × ${it.name} <span style="color:var(--muted)">${it.total.toFixed(2)}</span></span>`).join("")}
                </td>
                <td class="mono" style="text-align:right">${inv.total.toFixed(2)}</td>
                <td class="mono" style="text-align:right;color:var(--green);font-weight:700">${inv.cash_collected.toFixed(2)}</td>
                <td class="mono" style="text-align:right;color:${inv.outstanding>0?"var(--warn)":"var(--muted)"}">${inv.outstanding>0?inv.outstanding.toFixed(2):"—"}</td>
            </tr>`).join("");
    } else {
        posHtml += `<tr><td colspan="9" style="text-align:center;color:var(--muted);padding:24px">No POS invoices</td></tr>`;
    }
    posHtml += `</tbody></table></div>`;

    const typeLabel = {cash:"Cash", full_payment:"Full Payment", consignment:"Consignment"};
    let b2bHtml = `
        <div class="table-title" style="margin-top:22px">B2B Invoices — ${data.b2b_records.length} invoices</div>
        <div class="table-wrap">
        <table><thead><tr><th>Invoice #</th><th>Client</th><th>Date / Time</th><th>By</th><th>Type</th><th>Items</th><th style="text-align:right">Invoiced</th><th style="text-align:right">Paid</th><th style="text-align:right">Outstanding</th></tr></thead><tbody>`;
    if(data.b2b_records.length){
        b2bHtml += data.b2b_records.map(inv=>`
            <tr>
                <td class="mono" style="font-size:11px;color:var(--blue)">${inv.invoice_number}</td>
                <td class="name" style="font-size:13px">${inv.client}</td>
                <td class="mono" style="font-size:12px;color:var(--muted)">${inv.datetime}</td>
                <td style="font-size:12px;color:var(--muted);white-space:nowrap">${inv.user_name}</td>
                <td style="font-size:12px">${typeLabel[inv.invoice_type]||inv.invoice_type}</td>
                <td style="font-size:12px;color:var(--sub)">${inv.items.map(it=>`<span style="display:inline-block;background:var(--card2);border:1px solid var(--border2);border-radius:5px;padding:1px 7px;margin:2px;white-space:nowrap">${it.qty%1===0?it.qty.toFixed(0):it.qty.toFixed(2)} × ${it.name} <span style="color:var(--muted)">${it.total.toFixed(2)}</span></span>`).join("")}</td>
                <td class="mono" style="text-align:right">${inv.total.toFixed(2)}</td>
                <td class="mono" style="text-align:right;color:var(--green)">${inv.amount_paid.toFixed(2)}</td>
                <td class="mono" style="text-align:right;color:${inv.balance_due>0?"var(--warn)":"var(--muted)"};font-weight:${inv.balance_due>0?700:400}">${inv.balance_due>0?inv.balance_due.toFixed(2):"—"}</td>
            </tr>`).join("");
    } else {
        b2bHtml += `<tr><td colspan="9" style="text-align:center;color:var(--muted);padding:24px">No B2B invoices</td></tr>`;
    }
    b2bHtml += `</tbody></table></div>`;

    let b2bCollectionsHtml = `
        <div class="table-title" style="margin-top:22px">B2B Client Collections — ${data.b2b_payment_records.length} payment${data.b2b_payment_records.length!==1?"s":""}</div>
        <div class="table-wrap">
        <table><thead><tr><th>Reference</th><th>Client</th><th>Date / Time</th><th>By</th><th style="text-align:right">Amount</th><th>Notes</th></tr></thead><tbody>`;
    if(data.b2b_payment_records.length){
        b2bCollectionsHtml += data.b2b_payment_records.map(payment=>`
            <tr>
                <td class="mono" style="font-size:11px;color:var(--teal)">${payment.reference}</td>
                <td class="name" style="font-size:13px">${payment.client}</td>
                <td class="mono" style="font-size:12px;color:var(--muted)">${payment.datetime}</td>
                <td style="font-size:12px;color:var(--muted);white-space:nowrap">${payment.user_name}</td>
                <td class="mono" style="text-align:right;color:var(--green);font-weight:700">${payment.amount.toFixed(2)}</td>
                <td style="font-size:12px;color:var(--muted)">${payment.notes || "—"}</td>
            </tr>`).join("");
    } else {
        b2bCollectionsHtml += `<tr><td colspan="6" style="text-align:center;color:var(--muted);padding:24px">No B2B client payments in this period</td></tr>`;
    }
    b2bCollectionsHtml += `</tbody></table></div>`;

    let refHtml = "";
    if(data.refund_records && data.refund_records.length){
        refHtml = `
        <div style="margin-top:28px;display:flex;align-items:center;gap:14px;padding:14px 18px;background:rgba(255,77,109,.06);border:1px solid rgba(255,77,109,.2);border-radius:12px;">
            <div>
                <div style="font-size:13px;font-weight:700;color:#ff4d6d;letter-spacing:.3px">Refunds — ${data.refund_records.length} refund${data.refund_records.length!==1?"s":""}</div>
                <div style="font-size:11px;color:rgba(255,77,109,.6);margin-top:2px">Shown separately from sales and collections</div>
            </div>
            <div style="margin-left:auto;font-family:var(--mono);font-size:22px;font-weight:800;color:#ff4d6d">−${data.refunds.toFixed(2)}</div>
        </div>
        <div class="table-wrap" style="border-color:rgba(255,77,109,.18);">
        <table><thead style="background:rgba(255,77,109,.05)"><tr><th style="color:#ff4d6d">Ref #</th><th>Source</th><th>Counterparty</th><th>Date / Time</th><th>Processed By</th><th>Reason</th><th>Method</th><th style="text-align:right;color:#ff4d6d">Amount</th></tr></thead>
        <tbody>
            ${data.refund_records.map(row=>`<tr>
                <td class="mono" style="font-size:11px;color:#ff4d6d;font-weight:700">${row.refund_number}</td>
                <td>${row.source}</td>
                <td class="name">${row.counterparty}</td>
                <td class="mono" style="font-size:12px;color:var(--muted)">${row.datetime}</td>
                <td style="font-size:12px;color:var(--muted)">${row.processed_by}</td>
                <td style="font-size:12px;color:var(--sub)">${row.reason}</td>
                <td style="font-size:12px">${row.refund_method}</td>
                <td class="mono" style="text-align:right;font-weight:700;color:#ff4d6d">−${row.total.toFixed(2)}</td>
            </tr>`).join("")}
        </tbody></table></div>`;
    }
    document.getElementById("sales-records").innerHTML = posHtml + b2bHtml + b2bCollectionsHtml + refHtml;
    return;
}

/* ── B2B ── */
async function loadB2B(){
    let r = getRange("b2b-from","b2b-to");
    let data = await fetchReportJson(`/reports/api/b2b-statement?date_from=${r.from}&date_to=${r.to}`);
    document.getElementById("b-clients").innerText     = data.length;
    document.getElementById("b-invoiced").innerText    = data.reduce((s,c)=>s+c.total_invoiced,0).toFixed(2);
    document.getElementById("b-outstanding").innerText = data.reduce((s,c)=>s+c.outstanding,0).toFixed(2);
    setPrintDates("ph-b2b-dates", r.from, r.to);
    document.getElementById("b2b-body").innerHTML = data.length
        ? data.map(c=>`<tr>
            <td class="name">${c.name}</td>
            <td style="font-size:12px">${c.phone}</td>
            <td style="font-size:12px">${String(c.payment_terms || "-").replaceAll("_"," ")}</td>
            <td class="mono">${c.total_invoiced.toFixed(2)}</td>
            <td class="mono" style="color:var(--green)">${c.total_paid.toFixed(2)}</td>
            <td class="mono" style="color:${c.outstanding>0?"var(--warn)":"var(--muted)"};font-weight:${c.outstanding>0?700:400}">${c.outstanding>0?c.outstanding.toFixed(2):"—"}</td>
            <td class="mono" style="color:var(--muted)">${c.invoice_count}</td>
          </tr>`).join("")
        : `<tr><td colspan="7" style="text-align:center;color:var(--muted);padding:30px">No data for this period</td></tr>`;
}

/* ── INVENTORY ── */
async function loadInventory(){
    const mode = document.getElementById("inv-mode")?.value || "snapshot";
    const from = document.getElementById("inv-from")?.value || "";
    const to = document.getElementById("inv-to")?.value || "";
    const url = mode === "movement"
        ? `/reports/api/inventory?mode=movement&date_from=${from}&date_to=${to}`
        : "/reports/api/inventory?mode=snapshot";
    let data = await fetchReportJson(url);
    const filterBar = document.querySelector("#section-inventory .filter-bar");
    if (filterBar && !document.getElementById("inv-mode")) {
        filterBar.innerHTML = `
            <label>Mode</label>
            <select id="inv-mode">
                <option value="snapshot">Stock Snapshot</option>
                <option value="movement">Stock Movement</option>
            </select>
            <label>From</label><input type="date" id="inv-from">
            <label>To</label><input type="date" id="inv-to">
            <div class="filter-sep"></div>
            <button class="btn btn-lime" onclick="loadInventory()">Apply</button>
            <button class="btn btn-excel" onclick="exportSection('inventory')">â¬‡ Excel</button>
            <button class="btn btn-print" onclick="window.print()">ðŸ–¨ Print</button>`;
        document.getElementById("inv-mode").value = mode;
        document.getElementById("inv-from").value = from || monthStart();
        document.getElementById("inv-to").value = to || today();
    }
    if (data.mode === "movement") {
        document.getElementById("inv-count").innerText = data.total_products;
        document.getElementById("inv-value").innerText = data.summary.stock_in.toFixed(2);
        document.getElementById("inv-low").innerText = data.summary.stock_out.toFixed(2);
        document.getElementById("ph-inv-dates").innerText = `Movement period: ${data.date_from}  →  ${data.date_to}`;
        document.getElementById("inv-body").closest("table").querySelector("thead").innerHTML = "<tr><th>SKU</th><th>Product</th><th>Category</th><th>Unit</th><th>Stock In</th><th>Stock Out</th><th>Receipts</th><th>Sales/Usage</th><th>Spoilage</th><th>Transfers In</th><th>Transfers Out</th><th>Adjustments Net</th><th>Net Movement</th></tr>";
        document.getElementById("inv-body").innerHTML = data.products.length ? data.products.map(p=>`<tr>
            <td class="mono" style="font-size:11px;color:var(--muted)">${p.sku}</td>
            <td class="name">${p.name}</td>
            <td>${p.category}</td>
            <td>${p.unit}</td>
            <td class="mono" style="color:var(--green)">${p.stock_in.toFixed(2)}</td>
            <td class="mono" style="color:var(--danger)">${p.stock_out.toFixed(2)}</td>
            <td class="mono">${p.receipts.toFixed(2)}</td>
            <td class="mono">${p.sales_usage.toFixed(2)}</td>
            <td class="mono">${p.spoilage.toFixed(2)}</td>
            <td class="mono">${p.transfers_in.toFixed(2)}</td>
            <td class="mono">${p.transfers_out.toFixed(2)}</td>
            <td class="mono">${p.adjustments_net.toFixed(2)}</td>
            <td class="mono" style="color:${p.net_movement>=0?"var(--green)":"var(--danger)"}">${p.net_movement.toFixed(2)}</td>
          </tr>`).join("") : `<tr><td colspan="13" style="text-align:center;color:var(--muted);padding:30px">No movement in this period</td></tr>`;
    } else {
        document.getElementById("inv-count").innerText = data.total_products;
        document.getElementById("inv-value").innerText = data.total_value.toFixed(2);
        document.getElementById("inv-low").innerText = data.low_count;
        document.getElementById("ph-inv-dates").innerText = `Snapshot as of ${today()}`;
        document.getElementById("inv-body").closest("table").querySelector("thead").innerHTML = "<tr><th>SKU</th><th>Product</th><th>Category</th><th>Stock</th><th>Unit</th><th>Price</th><th>Stock Value</th><th>Threshold</th><th>Last Move</th><th>Status</th></tr>";
        document.getElementById("inv-body").innerHTML = data.products.map(p=>`<tr>
            <td class="mono" style="font-size:11px;color:var(--muted)">${p.sku}</td>
            <td class="name">${p.name}</td>
            <td>${p.category}</td>
            <td class="mono" style="color:${p.low_stock?"var(--danger)":"var(--text)"};font-weight:700">${p.stock.toFixed(2)}</td>
            <td style="font-size:12px;color:var(--muted)">${p.unit}</td>
            <td class="mono">${p.price.toFixed(2)}</td>
            <td class="mono" style="color:var(--blue)">${p.value.toFixed(2)}</td>
            <td class="mono">${p.threshold.toFixed(2)}</td>
            <td class="mono">${p.last_move_at}</td>
            <td><span class="badge ${p.low_stock?"badge-low":"badge-ok"}">${p.low_stock?"Low Stock":"OK"}${p.dead_stock?" · Dead Stock":""}</span></td>
          </tr>`).join("");
    }
    return;
}

/* ── FARM ── */
async function loadFarm(){
    let r = getRange("farm-from","farm-to");
    let data = await fetchReportJson(`/reports/api/farm-intake?date_from=${r.from}&date_to=${r.to}`);
    setPrintDates("ph-farm-dates", r.from, r.to);
    const summaryRows = data.summary.length
        ? data.summary.map(row=>`<tr>
            <td class="name">${row.farm}</td>
            <td class="mono">${row.delivery_count}</td>
            <td class="mono">${row.line_count}</td>
            <td class="mono" style="color:var(--green)">${row.total_qty.toFixed(2)}</td>
            <td>${row.top_product}</td>
          </tr>`).join("")
        : `<tr><td colspan="5" style="text-align:center;color:var(--muted);padding:24px">No farm intake summary in this period</td></tr>`;
    const detailRows = data.detail.length
        ? data.detail.map(row=>`<tr>
            <td class="name">${row.farm}</td>
            <td class="mono">${row.date}</td>
            <td class="mono">${row.delivery_number}</td>
            <td class="mono">${row.sku}</td>
            <td>${row.product}</td>
            <td class="mono">${row.qty.toFixed(2)}</td>
            <td>${row.unit}</td>
            <td>${row.received_by}</td>
            <td>${row.user_name}</td>
          </tr>`).join("")
        : `<tr><td colspan="9" style="text-align:center;color:var(--muted);padding:24px">No farm intake detail in this period</td></tr>`;
    document.getElementById("farm-content").innerHTML = `
        <div class="stats-row">
            <div class="stat-card sc-blue"><div class="stat-label">Farms</div><div class="stat-value sv-blue">${data.totals.farm_count}</div></div>
            <div class="stat-card sc-green"><div class="stat-label">Deliveries</div><div class="stat-value sv-green">${data.totals.delivery_count}</div></div>
            <div class="stat-card sc-orange"><div class="stat-label">Line Items</div><div class="stat-value sv-orange">${data.totals.line_count}</div></div>
            <div class="stat-card sc-green"><div class="stat-label">Total Qty</div><div class="stat-value sv-green">${data.totals.total_qty.toFixed(2)}</div></div>
        </div>
        <div class="table-wrap" style="margin-bottom:18px">
            <div class="table-title">Farm Intake Summary</div>
            <table><thead><tr><th>Farm</th><th>Deliveries</th><th>Line Items</th><th>Total Qty</th><th>Top Product</th></tr></thead><tbody>${summaryRows}</tbody></table>
        </div>
        <div class="table-wrap">
            <div class="table-title">Farm Intake Detail</div>
            <table><thead><tr><th>Farm</th><th>Date</th><th>Delivery #</th><th>SKU</th><th>Product</th><th>Qty</th><th>Unit</th><th>Received By</th><th>Performed By</th></tr></thead><tbody>${detailRows}</tbody></table>
        </div>`;
    return;
    let summaryHtml = data.farms.map((farm,fi)=>{
        let color = fi===0?"var(--lime)":"var(--teal)";
        let maxQty = farm.products.length ? farm.products[0].total_qty : 1;
        return `<div class="table-wrap" style="margin-bottom:12px">
            <div class="table-title">
                <span>${fi===0?"🌿":"♻️"} ${farm.name}</span>
                <span>${farm.delivery_count} deliveries — ${farm.total_qty.toFixed(1)} total</span>
            </div>
            <div style="padding:14px 16px">
                ${farm.products.length
                    ? farm.products.map(p=>`<div class="bar-row">
                        <div class="bar-label">${p.name}</div>
                        <div class="bar-track"><div class="bar-fill" style="width:${(p.total_qty/maxQty*100).toFixed(1)}%;background:linear-gradient(90deg,${color},var(--green))"></div></div>
                        <div class="bar-val" style="color:${color}">${p.total_qty.toFixed(1)} ${p.unit}</div>
                      </div>`).join("")
                    : `<div style="color:var(--muted);font-size:13px">No deliveries in this period.</div>`}
            </div>
        </div>`;
    }).join("");
    let deliveriesHtml = `<div class="table-wrap">
        <div class="table-title">Delivery Records</div>
        <table><thead><tr><th>Delivery #</th><th>Farm</th><th>Date</th><th>Received By</th><th>Qty</th><th>Items</th><th>By</th><th>Notes</th></tr></thead><tbody>`;
    if(data.deliveries.length){
        deliveriesHtml += data.deliveries.map(d=>`<tr>
            <td class="mono" style="font-size:11px;color:var(--lime)">${d.delivery_number}</td>
            <td class="name">${d.farm}</td>
            <td class="mono" style="font-size:12px;color:var(--muted)">${d.delivery_date}</td>
            <td style="font-size:12px">${d.received_by}</td>
            <td class="mono" style="color:var(--green)">${d.total_qty.toFixed(2)}</td>
            <td class="mono">${d.total_items}</td>
            <td style="font-size:12px;color:var(--muted);white-space:nowrap">${d.user_name}</td>
            <td style="font-size:12px;color:var(--muted)">${d.notes||"—"}</td>
        </tr>`).join("");
    } else {
        deliveriesHtml += `<tr><td colspan="8" style="text-align:center;color:var(--muted);padding:30px">No deliveries in this period</td></tr>`;
    }
    deliveriesHtml += `</tbody></table></div>`;
    document.getElementById("farm-content").innerHTML = summaryHtml + deliveriesHtml;
}


/* ── SPOILAGE ── */
async function loadSpoilage(){
    let r = getRange("spl-from","spl-to");
    let data = await fetchReportJson(`/reports/api/spoilage?date_from=${r.from}&date_to=${r.to}`);
    document.getElementById("spl-count").innerText = data.total_count;
    document.getElementById("spl-qty").innerText   = data.total_qty.toFixed(2);
    setPrintDates("ph-spl-dates", r.from, r.to);
    let maxP = data.by_product.length ? data.by_product[0].qty : 1;
    document.getElementById("spl-by-product").innerHTML = data.by_product.length
        ? data.by_product.map(p=>`<div class="bar-row">
            <div class="bar-label">${p.name}</div>
            <div class="bar-track"><div class="bar-fill" style="width:${(p.qty/maxP*100).toFixed(1)}%;background:linear-gradient(90deg,var(--danger),var(--orange))"></div></div>
            <div class="bar-val" style="color:var(--danger)">${p.qty.toFixed(1)}</div>
          </div>`).join("")
        : `<div style="color:var(--muted);font-size:13px">No data</div>`;
    let maxReasonQty = data.by_reason.length ? data.by_reason[0].qty : 1;
    document.getElementById("spl-by-reason").innerHTML = data.by_reason.length
        ? data.by_reason.map(r=>`<div class="bar-row">
            <div class="bar-label">${r.reason}</div>
            <div class="bar-track"><div class="bar-fill" style="width:${(r.qty/maxReasonQty*100).toFixed(1)}%;background:linear-gradient(90deg,var(--warn),var(--orange))"></div></div>
            <div class="bar-val" style="color:var(--warn)">${r.qty.toFixed(1)}</div>
          </div>`).join("")
        : `<div style="color:var(--muted);font-size:13px">No data</div>`;
    document.getElementById("spl-body").innerHTML = data.records.length
        ? data.records.map(r=>`<tr>
            <td class="mono" style="font-size:11px;color:var(--danger)">${r.ref}</td>
            <td class="name">${r.product}</td>
            <td class="mono" style="color:var(--danger)">-${r.qty.toFixed(2)} ${r.unit}</td>
            <td style="font-size:12px">${r.reason}</td>
            <td style="font-size:12px;color:var(--muted)">${r.farm}</td>
            <td class="mono" style="font-size:12px">${r.date}</td>
            <td style="font-size:12px;color:var(--muted);white-space:nowrap">${r.user_name}</td>
            <td style="font-size:12px;color:var(--muted)">${r.notes}</td>
          </tr>`).join("")
        : `<tr><td colspan="8" style="text-align:center;color:var(--muted);padding:30px">No spoilage in this period</td></tr>`;
}

/* ── PRODUCTION ── */
async function loadProduction(){
    let r = getRange("prod-from","prod-to");
    let data = await fetchReportJson(`/reports/api/production?date_from=${r.from}&date_to=${r.to}`);
    document.getElementById("prod-proc").innerText = data.total_processing;
    document.getElementById("prod-pkg").innerText  = data.total_packaging;
    document.getElementById("prod-loss").innerText = data.avg_loss_pct.toFixed(1)+"%";
    setPrintDates("ph-prod-dates", r.from, r.to);
    document.getElementById("prod-body").innerHTML = data.batches.length
        ? data.batches.map(b=>`<tr>
            <td class="mono" style="font-size:12px;color:${b.type==="Packaging"?"var(--teal)":"var(--orange)"}">${b.batch_number}</td>
            <td><span style="font-size:11px;font-weight:700;padding:2px 8px;border-radius:20px;background:${b.type==="Packaging"?"rgba(45,212,191,.1)":"rgba(251,146,60,.1)"};color:${b.type==="Packaging"?"var(--teal)":"var(--orange)"}">${b.type}</span></td>
            <td class="name" style="font-size:12px">${b.recipe}</td>
            <td style="font-size:11px;color:var(--sub);max-width:130px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">${b.inputs_str||"—"}</td>
            <td style="font-size:11px;color:var(--green);max-width:130px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">${b.outputs_str||"—"}</td>
            <td class="mono" style="color:${b.waste_pct<10?"var(--green)":b.waste_pct<25?"var(--warn)":"var(--danger)"}">${b.waste_pct.toFixed(1)}%</td>
            <td class="mono" style="font-size:12px;color:var(--muted)">${b.date}</td>
            <td style="font-size:12px;color:var(--muted);white-space:nowrap">${b.user_name}</td>
          </tr>`).join("")
        : `<tr><td colspan="8" style="text-align:center;color:var(--muted);padding:30px">No batches in this period</td></tr>`;
}


/* ── P&L ── */
async function loadPL(){
    let r = getRange("pl-from","pl-to");
    let data = await fetchReportJson(`/reports/api/pl?date_from=${r.from}&date_to=${r.to}`);
    setPrintDates("ph-pl-dates", data.date_from, data.date_to);
    let isProfit = data.net_profit >= 0;

    const refLabel = {b2b:"B2B Sale", b2b_payment:"B2B Payment", pos:"POS Sale", consignment:"Consignment", payroll:"Payroll", spoilage:"Spoilage", manual:"Manual Entry"};

    function renderEntries(entries, color){
        if(!entries.length) return "";
        return `<div style="background:var(--bg);border-radius:8px;margin:4px 0 10px 24px;overflow:hidden">
            <table style="width:100%;border-collapse:collapse;font-size:12px">
                <thead><tr style="background:var(--card2)">
                    <th style="padding:6px 12px;text-align:left;color:var(--muted);font-weight:700;letter-spacing:.5px">Date</th>
                    <th style="padding:6px 12px;text-align:left;color:var(--muted);font-weight:700;letter-spacing:.5px">Type</th>
                    <th style="padding:6px 12px;text-align:left;color:var(--muted);font-weight:700;letter-spacing:.5px">Description</th>
                    <th style="padding:6px 12px;text-align:right;color:var(--muted);font-weight:700;letter-spacing:.5px">Amount (EGP)</th>
                </tr></thead>
                <tbody>
                ${entries.map(e=>`<tr style="border-top:1px solid var(--border)">
                    <td style="padding:7px 12px;font-family:var(--mono);color:var(--muted)">${e.date}</td>
                    <td style="padding:7px 12px"><span style="background:var(--card2);border-radius:4px;padding:1px 7px;font-size:11px;color:var(--sub)">${refLabel[e.ref_type]||e.ref_type}</span></td>
                    <td style="padding:7px 12px;color:var(--sub)">${e.description}</td>
                    <td style="padding:7px 12px;text-align:right;font-family:var(--mono);font-weight:700;color:${color}">${e.amount.toFixed(2)}</td>
                </tr>`).join("")}
                </tbody>
            </table>
        </div>`;
    }

    function renderAccountLine(item, color, expanded=false){
        let id = "pl-"+item.code.replace(/\\W/g,"");
        return `
            <div class="pl-row" style="cursor:pointer;user-select:none" onclick="togglePLDetail('${id}')">
                <span style="color:var(--sub);display:flex;align-items:center;gap:8px">
                    <span id="${id}-icon" style="color:var(--muted);font-size:11px;transition:transform .2s">${item.entries.length?"▶":""}</span>
                    ${item.code} — ${item.name}
                    <span style="font-size:11px;color:var(--muted)">(${item.entries.length} entries)</span>
                </span>
                <span class="mono" style="color:${color}">${item.amount.toFixed(2)}</span>
            </div>
            <div id="${id}" style="display:none">${renderEntries(item.entries, color)}</div>`;
    }

    document.getElementById("pl-content").innerHTML = `
        <div class="stats-row">
            <div class="stat-card sc-green"><div class="stat-label">Total Revenue</div><div class="stat-value sv-green">${data.total_revenue.toFixed(2)}</div></div>
            <div class="stat-card sc-danger"><div class="stat-label">Total Expenses</div><div class="stat-value sv-danger">${data.total_expense.toFixed(2)}</div></div>
            <div class="stat-card ${isProfit?"sc-green":"sc-danger"}">
                <div class="stat-label">Net ${isProfit?"Profit":"Loss"}</div>
                <div class="stat-value ${isProfit?"sv-green":"sv-danger"}">${Math.abs(data.net_profit).toFixed(2)}</div>
            </div>
        </div>
        ${data.warning ? `<div style="font-size:12px;color:var(--warn);margin-bottom:12px">${data.warning}</div>` : ``}
        <div style="font-size:12px;color:var(--muted);margin-bottom:12px">💡 Click any account line to expand its journal entries</div>
        <div class="pl-section">
            <div class="pl-header">Revenue</div>
            ${data.revenue_lines.map(r=>renderAccountLine(r,"var(--green)")).join("") || `<div class="pl-row"><span style="color:var(--muted)">No revenue entries</span><span></span></div>`}
            <div class="pl-row pl-total"><span>Total Revenue</span><span class="mono" style="color:var(--green)">${data.total_revenue.toFixed(2)}</span></div>
        </div>
        <div class="pl-section">
            <div class="pl-header">Expenses</div>
            ${data.expense_lines.map(e=>renderAccountLine(e,"var(--danger)")).join("") || `<div class="pl-row"><span style="color:var(--muted)">No expense entries</span><span></span></div>`}
            <div class="pl-row pl-total"><span>Total Expenses</span><span class="mono" style="color:var(--danger)">${data.total_expense.toFixed(2)}</span></div>
        </div>
        <div class="pl-section">
            <div class="pl-row pl-net" style="background:${isProfit?"rgba(0,255,157,.06)":"rgba(255,77,109,.06)"};border-top:2px solid ${isProfit?"var(--green)":"var(--danger)"}">
                <span>Net ${isProfit?"Profit":"Loss"}</span>
                <span class="mono" style="color:${isProfit?"var(--green)":"var(--danger)"};font-size:20px">${Math.abs(data.net_profit).toFixed(2)}</span>
            </div>
        </div>`;
}

const __rawReportLoaders = {
    sales: loadSales,
    transactions: loadTransactions,
    b2b: loadB2B,
    inventory: loadInventory,
    farm: loadFarm,
    spoilage: loadSpoilage,
    production: loadProduction,
    pl: loadPL,
};

loadSales = () => runReportLoader("sales", __rawReportLoaders.sales);
loadTransactions = () => runReportLoader("transactions", __rawReportLoaders.transactions);
loadB2B = () => runReportLoader("b2b", __rawReportLoaders.b2b);
loadInventory = () => runReportLoader("inventory", __rawReportLoaders.inventory);
loadFarm = () => runReportLoader("farm", __rawReportLoaders.farm);
loadSpoilage = () => runReportLoader("spoilage", __rawReportLoaders.spoilage);
loadProduction = () => runReportLoader("production", __rawReportLoaders.production);
loadPL = () => runReportLoader("pl", __rawReportLoaders.pl);

function togglePLDetail(id){
    let el   = document.getElementById(id);
    let icon = document.getElementById(id+"-icon");
    if(!el) return;
    let open = el.style.display === "none" || el.style.display === "";
    el.style.display   = open ? "block" : "none";
    if(icon) icon.style.transform = open ? "rotate(90deg)" : "";
}

/* ── INIT: set default dates ── */
(function initDates(){
    let m = monthStart(), y = yearStart(), t = today();
    setEl("sales-from", m); setEl("sales-to", t);
    setEl("tx-from",    m); setEl("tx-to",    t);
    setEl("b2b-from",   m); setEl("b2b-to",   t);
    setEl("farm-from",  m); setEl("farm-to",   t);
    setEl("spl-from",   m); setEl("spl-to",    t);
    setEl("prod-from",  m); setEl("prod-to",   t);
    setEl("pl-from",    y); setEl("pl-to",     t);
    const invMode = document.getElementById("inv-mode");
    const invFrom = document.getElementById("inv-from");
    const invTo = document.getElementById("inv-to");
    if(invMode) invMode.value = "snapshot";
    if(invFrom) invFrom.value = m;
    if(invTo) invTo.value = t;
})();

ensureTabMetadata();
</script>
</body>
</html>"""