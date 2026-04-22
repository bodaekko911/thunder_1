from __future__ import annotations

import io
import json
import uuid
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

import openpyxl
from fastapi import HTTPException
from openpyxl.utils.datetime import from_excel
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.log import ActivityLog, record
from app.models.expense import ExpenseCategory
from app.models.farm import Farm
from app.models.user import User
from app.schemas.expense import ExpenseCategoryCreate, ExpenseCreate
from app.services.expense_service import (
    create_category,
    create_expense_entry,
    delete_expense_entry,
)


@dataclass
class ParsedExpenseRow:
    excel_row: int
    category_name: str
    amount: Decimal
    expense_date: date
    farm_id: int | None
    farm_name: str | None
    general_expense: bool
    notes: str | None


def _normalize_key(value: object) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("_", " ").replace("-", " ")
    return " ".join(text.split())


def _normalize_name(value: object) -> str:
    return " ".join(str(value or "").strip().split())


def _find_header_index(headers: list[str], aliases: list[str]) -> int | None:
    normalized = [_normalize_key(header) for header in headers]
    alias_set = {_normalize_key(alias) for alias in aliases}
    for idx, header in enumerate(normalized):
        if header in alias_set:
            return idx + 1
    return None


def _parse_amount(value: object) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, Decimal):
        return value.quantize(Decimal("0.01"))
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return Decimal(str(value)).quantize(Decimal("0.01"))

    text = str(value).strip()
    if not text:
        return None
    cleaned = text.replace(",", "").replace("$", "")
    if cleaned.startswith("(") and cleaned.endswith(")"):
        cleaned = f"-{cleaned[1:-1]}"
    try:
        return Decimal(cleaned).quantize(Decimal("0.01"))
    except InvalidOperation:
        return None


def _parse_date_value(value: object) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            parsed = from_excel(value)
        except Exception:
            parsed = None
        if isinstance(parsed, datetime):
            return parsed.date()
        if isinstance(parsed, date):
            return parsed

    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d", "%d-%m-%Y", "%m-%d-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    try:
        return date.fromisoformat(text[:10])
    except ValueError:
        return None


async def _load_category_map(db: AsyncSession) -> dict[str, ExpenseCategory]:
    result = await db.execute(
        select(ExpenseCategory).where(ExpenseCategory.is_active == "1")
    )
    categories = result.scalars().all()
    return {_normalize_name(category.name).lower(): category for category in categories}


async def _load_farm_map(db: AsyncSession) -> dict[str, Farm]:
    result = await db.execute(select(Farm).where(Farm.is_active == 1))
    farms = result.scalars().all()
    return {_normalize_name(farm.name).lower(): farm for farm in farms}


def _empty_row(values: list[object]) -> bool:
    return all(str(value or "").strip() == "" for value in values)


def _clean_notes(value: object) -> str | None:
    text = _normalize_name(value)
    return text or None


def _batch_payload_from_log(log: ActivityLog) -> dict:
    try:
        payload = json.loads(log.description or "{}")
    except json.JSONDecodeError:
        payload = {}
    if not isinstance(payload, dict):
        payload = {}
    return payload


async def list_expense_import_batches(db: AsyncSession) -> dict:
    result = await db.execute(
        select(ActivityLog)
        .where(
            ActivityLog.module == "Import",
            ActivityLog.action.in_(["expense_import_batch", "expense_import_revert"]),
        )
        .order_by(ActivityLog.created_at.desc())
    )
    logs = result.scalars().all()
    batches: dict[str, dict] = {}

    for log in logs:
        batch_id = log.ref_id
        if not batch_id:
            continue
        if log.action == "expense_import_batch":
            payload = _batch_payload_from_log(log)
            batches[batch_id] = {
                "batch_id": batch_id,
                "filename": payload.get("filename") or "expenses.xlsx",
                "ran_on": log.created_at.date().isoformat() if log.created_at else None,
                "rows_read": payload.get("rows_read", 0),
                "rows_imported": payload.get("rows_imported", 0),
                "rows_skipped": payload.get("rows_skipped", 0),
                "expense_records_created": payload.get("expense_records_created", 0),
                "notes_imported": payload.get("notes_imported", 0),
                "general_expense_rows": payload.get("general_expense_rows", 0),
                "reverted": False,
                "reverted_on": None,
            }
        elif log.action == "expense_import_revert" and batch_id in batches:
            batches[batch_id]["reverted"] = True
            batches[batch_id]["reverted_on"] = log.created_at.date().isoformat() if log.created_at else None

    ordered = sorted(
        batches.values(),
        key=lambda item: (item["ran_on"] or "", item["batch_id"]),
        reverse=True,
    )
    return {"batches": ordered}


async def revert_expense_import_batch(
    db: AsyncSession,
    batch_id: str,
    current_user: User,
) -> dict:
    summary_result = await db.execute(
        select(ActivityLog).where(
            ActivityLog.module == "Import",
            ActivityLog.action == "expense_import_batch",
            ActivityLog.ref_id == batch_id,
        )
    )
    summary_log = summary_result.scalar_one_or_none()
    if not summary_log:
        raise HTTPException(status_code=404, detail="Expense import batch not found")

    reverted_result = await db.execute(
        select(ActivityLog).where(
            ActivityLog.module == "Import",
            ActivityLog.action == "expense_import_revert",
            ActivityLog.ref_id == batch_id,
        )
    )
    if reverted_result.scalar_one_or_none():
        return {"ok": True, "batch_id": batch_id, "already_reverted": True, "deleted_expenses": 0}

    item_result = await db.execute(
        select(ActivityLog)
        .where(
            ActivityLog.module == "Import",
            ActivityLog.action == "expense_import_item",
            ActivityLog.ref_type == batch_id,
        )
        .order_by(ActivityLog.id.desc())
    )
    item_logs = item_result.scalars().all()

    deleted_expenses = 0
    skipped_missing = 0
    seen_expense_ids: set[int] = set()
    for log in item_logs:
        try:
            expense_id = int(log.ref_id)
        except (TypeError, ValueError):
            continue
        if expense_id in seen_expense_ids:
            continue
        seen_expense_ids.add(expense_id)
        try:
            await delete_expense_entry(db, expense_id, current_user)
            deleted_expenses += 1
        except HTTPException as exc:
            if exc.status_code == 404:
                skipped_missing += 1
                continue
            raise

    record(
        db,
        "Import",
        "expense_import_revert",
        json.dumps({"batch_id": batch_id, "deleted_expenses": deleted_expenses, "skipped_missing": skipped_missing}),
        user=current_user,
        ref_type="expense_import_batch",
        ref_id=batch_id,
    )
    await db.commit()
    return {
        "ok": True,
        "batch_id": batch_id,
        "deleted_expenses": deleted_expenses,
        "skipped_missing": skipped_missing,
    }


async def import_expenses(
    *,
    db: AsyncSession,
    workbook_bytes: bytes,
    filename: str,
    current_user,
    dry_run: bool = True,
) -> dict:
    workbook = openpyxl.load_workbook(io.BytesIO(workbook_bytes), data_only=True)
    sheet = workbook.active
    headers = [str(sheet.cell(1, col).value or "") for col in range(1, sheet.max_column + 1)]

    col_category = _find_header_index(headers, ["Category", "Expense Category"])
    col_amount = _find_header_index(headers, ["Amount", "Expense Amount", "Value"])
    col_farm = _find_header_index(headers, ["Farm", "Farm Name"])
    col_date = _find_header_index(headers, ["Date", "Expense Date"])
    col_notes = _find_header_index(headers, ["Notes", "Note", "Description", "Expense Notes"])

    missing = []
    if not col_category:
        missing.append("Category")
    if not col_amount:
        missing.append("Amount")
    if not col_date:
        missing.append("Date")
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"Missing required column(s): {', '.join(missing)}",
        )

    category_map = await _load_category_map(db)
    farm_map = await _load_farm_map(db)
    rows: list[ParsedExpenseRow] = []
    errors: list[dict] = []
    missing_categories: dict[str, str] = {}
    resolved_farm_rows = 0
    general_expense_rows = 0
    total_amount = Decimal("0.00")
    date_values: list[date] = []

    for row_idx in range(2, sheet.max_row + 1):
        row_values = [sheet.cell(row_idx, col).value for col in range(1, sheet.max_column + 1)]
        if _empty_row(row_values):
            continue

        raw_category = sheet.cell(row_idx, col_category).value if col_category else None
        raw_amount = sheet.cell(row_idx, col_amount).value if col_amount else None
        raw_farm = sheet.cell(row_idx, col_farm).value if col_farm else None
        raw_date = sheet.cell(row_idx, col_date).value if col_date else None
        raw_notes = sheet.cell(row_idx, col_notes).value if col_notes else None

        category_name = _normalize_name(raw_category)
        amount = _parse_amount(raw_amount)
        expense_date = _parse_date_value(raw_date)
        farm_name = _normalize_name(raw_farm)
        notes = _clean_notes(raw_notes)

        row_errors: list[str] = []
        if not category_name:
            row_errors.append("Category is required")
        if amount is None:
            row_errors.append("Amount is required and must be numeric")
        elif amount <= 0:
            row_errors.append("Amount must be greater than 0")
        if expense_date is None:
            row_errors.append("Date is required and must be a valid date")

        category_key = category_name.lower()
        category = category_map.get(category_key) if category_name else None
        if category_name and category is None and category_key not in missing_categories:
            missing_categories[category_key] = category_name

        farm = None
        general_expense = not farm_name
        if not general_expense:
            farm = farm_map.get(farm_name.lower())
            if farm is None:
                row_errors.append(f"Farm '{farm_name}' was not found")

        if row_errors:
            errors.append(
                {
                    "row": row_idx,
                    "category": category_name or "",
                    "amount": "" if raw_amount is None else str(raw_amount),
                    "farm": farm_name or "General Expense",
                    "date": "" if raw_date is None else str(raw_date),
                    "reason": "; ".join(row_errors),
                }
            )
            continue

        rows.append(
            ParsedExpenseRow(
                excel_row=row_idx,
                category_name=category_name,
                amount=amount,
                expense_date=expense_date,
                farm_id=farm.id if farm else None,
                farm_name=farm.name if farm else None,
                general_expense=general_expense,
                notes=notes,
            )
        )
        if general_expense:
            general_expense_rows += 1
        elif farm is not None:
            resolved_farm_rows += 1
        total_amount += amount
        date_values.append(expense_date)

    auto_created_categories: list[dict] = []
    batch_id = str(uuid.uuid4())
    if not dry_run and missing_categories:
        for category_name in missing_categories.values():
            created = await create_category(
                db,
                ExpenseCategoryCreate(name=category_name, description="Auto-created by Expenses import"),
            )
            auto_created_categories.append(created)
        category_map = await _load_category_map(db)

    expenses_created = 0
    notes_imported = 0
    if not dry_run:
        for row in rows:
            category = category_map.get(row.category_name.lower())
            if category is None:
                errors.append(
                    {
                        "row": row.excel_row,
                        "category": row.category_name,
                        "amount": str(row.amount),
                        "farm": row.farm_name or "General Expense",
                        "date": row.expense_date.isoformat(),
                        "reason": f"Category '{row.category_name}' could not be created",
                    }
                )
                continue
            payload = ExpenseCreate(
                category_id=category.id,
                expense_date=row.expense_date.isoformat(),
                amount=float(row.amount),
                payment_method="cash",
                description=row.notes,
                farm_id=row.farm_id,
            )
            created = await create_expense_entry(db, payload, current_user)
            expenses_created += 1
            if row.notes:
                notes_imported += 1
            record(
                db,
                "Import",
                "expense_import_item",
                json.dumps(
                    {
                        "batch_id": batch_id,
                        "ref_number": created.get("ref_number"),
                        "category": row.category_name,
                        "amount": float(row.amount),
                        "farm": row.farm_name or "General Expense",
                    }
                ),
                user=current_user,
                ref_type=batch_id,
                ref_id=created.get("id"),
            )
        record(
            db,
            "Import",
            "expense_import_batch",
            json.dumps(
                {
                    "batch_id": batch_id,
                    "filename": filename,
                    "rows_read": len(rows) + len(errors),
                    "rows_imported": expenses_created,
                    "rows_skipped": len(errors),
                    "expense_records_created": expenses_created,
                    "categories_auto_created": len(auto_created_categories),
                    "farms_resolved": resolved_farm_rows,
                    "general_expense_rows": general_expense_rows,
                    "notes_imported": notes_imported,
                }
            ),
            user=current_user,
            ref_type="expense_import_batch",
            ref_id=batch_id,
        )
        await db.commit()

    categories_auto_created = len(auto_created_categories) if not dry_run else len(missing_categories)
    valid_rows = len(rows)
    notes_present = sum(1 for row in rows if row.notes)
    summary = {
        "rows_read": valid_rows + len(errors),
        "rows_imported": valid_rows if dry_run else expenses_created,
        "rows_skipped": len(errors),
        "expenses_created": expenses_created,
        "expenses_would_create": valid_rows,
        "expense_records_created": expenses_created,
        "categories_auto_created": categories_auto_created,
        "farms_resolved": resolved_farm_rows,
        "general_expense_rows": general_expense_rows,
        "notes_imported": notes_present if dry_run else notes_imported,
        "earliest_date": min(date_values).isoformat() if date_values else None,
        "latest_date": max(date_values).isoformat() if date_values else None,
        "total_amount": float(total_amount),
    }
    warnings = []
    if general_expense_rows:
        warnings.append(
            f"{general_expense_rows} row(s) will be recorded as General Expense because Farm was blank."
        )

    return {
        "ok": len(errors) == 0,
        "dry_run": dry_run,
        "filename": filename,
        "summary": summary,
        "errors": errors,
        "warnings": warnings,
        "batch_id": None if dry_run or expenses_created == 0 else batch_id,
        "revert_available": (not dry_run and expenses_created > 0),
        "auto_created_categories": [
            {"name": category["name"], "account_code": category.get("account_code")}
            for category in auto_created_categories
        ] if not dry_run else [
            {"name": category_name, "account_code": None}
            for category_name in missing_categories.values()
        ],
    }
