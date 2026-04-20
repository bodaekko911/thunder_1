"""
b2b_sales_import_service.py

Back-fill historical B2B (wholesale) sales from an Excel sheet into
b2b_invoices / b2b_invoice_items (and consignments / consignment_items).

Operating modes
───────────────
history_only           Create B2BInvoice + B2BInvoiceItem rows and post
                       journals (cash 1000/4000; AR/deferred 1100/2200).
                       For consignment rows, also creates Consignment +
                       ConsignmentItem records.  Does NOT decrement stock.
with_stock_adjustment  Everything above PLUS decrements products.stock
                       and writes StockMove rows.

Idempotency
───────────
Every real (non-dry-run) import carries a UUID4 batch_id stamped on every
B2BInvoice.import_batch_id (and Consignment.import_batch_id).
Duplicate detection: per (client_id, date, payment_type) group, check for an
existing invoice with notes LIKE 'Imported from %'.  Override with force=True.
"""

import io
import uuid
from collections import Counter
from datetime import date, datetime, timezone
from decimal import Decimal
from typing import Optional

import openpyxl
from sqlalchemy import Date as SADate, cast, insert, select, text
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import defer
from sqlalchemy.sql import func as sa_func

from app.models.b2b import (
    B2BClient, B2BClientPrice, B2BInvoice, B2BInvoiceItem,
    Consignment, ConsignmentItem,
)
from app.models.inventory import StockMove
from app.models.product import Product
from app.services.b2b_shared import post_journal, seed_deferred_revenue
from app.services.barcode_service import normalize_barcode_value

MIN_IMPORT_DATE = date(2026, 1, 1)
VALID_MODES = frozenset({"history_only", "with_stock_adjustment"})
_SKU_MAX = 80
_NAME_MAX = 200

# ── Payment type normalisation ────────────────────────────────────────────────

_PAYMENT_MAP: dict[str, str] = {
    "cash": "cash", "paid": "cash", "full": "cash", "cod": "cash",
    "immediate": "cash", "pay now": "cash", "pay_now": "cash",
    "full_payment": "full_payment", "full payment": "full_payment",
    "credit": "full_payment", "net15": "full_payment", "net30": "full_payment",
    "net60": "full_payment", "on account": "full_payment", "account": "full_payment",
    "invoiced": "full_payment",
    "consignment": "consignment", "cons": "consignment",
    "on consignment": "consignment", "sale or return": "consignment", "sor": "consignment",
}


def _normalize_payment_type(val) -> Optional[str]:
    if val is None:
        return None
    raw = " ".join(str(val).strip().lower().split())
    return _PAYMENT_MAP.get(raw)


# ── Row-level helpers ─────────────────────────────────────────────────────────

def _parse_date(val) -> Optional[date]:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    if not s or s.lower() in ("none", ""):
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _safe_float(val) -> Optional[float]:
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _normalize_sku(val) -> str:
    if val is None:
        return ""
    if isinstance(val, (int, float)):
        try:
            return str(int(val))
        except (ValueError, TypeError):
            pass
    return str(val).strip()


def _build_imported_product_name(item_hint: str, sku: str) -> str:
    candidate = (item_hint or "").strip()
    if not candidate:
        candidate = f"Imported Product {sku}"
    return candidate[:_NAME_MAX]


def _find_col(headers: list[str], candidates: list[str]) -> Optional[int]:
    for name in candidates:
        for i, h in enumerate(headers):
            if h == name.lower().strip():
                return i + 1
    return None


async def _column_exists(db: AsyncSession, table_name: str, column_name: str) -> bool:
    result = await db.execute(
        text(
            """
            SELECT EXISTS (
                SELECT 1
                FROM information_schema.columns
                WHERE table_schema = current_schema()
                  AND table_name = :table_name
                  AND column_name = :column_name
            )
            """
        ),
        {"table_name": table_name, "column_name": column_name},
    )
    return bool(result.scalar())


async def _find_existing_product_by_norm_sku(db: AsyncSession, norm_sku: str) -> Product | None:
    if not norm_sku:
        return None

    result = await db.execute(
        select(Product).options(defer(Product.created_by_import_batch))
    )
    for product in result.scalars().all():
        if normalize_barcode_value(product.sku) == norm_sku:
            return product
    return None


async def _next_unique_reference(
    db: AsyncSession,
    model,
    field_name: str,
    prefix: str,
    start_number: int,
    width: int,
) -> tuple[str, int]:
    number = max(1, start_number)
    field = getattr(model, field_name)
    while True:
        candidate = f"{prefix}{str(number).zfill(width)}"
        existing = await db.execute(select(model.id).where(field == candidate))
        if existing.scalar_one_or_none() is None:
            return candidate, number
        number += 1


# ── Main entry point ──────────────────────────────────────────────────────────

async def import_b2b_sales(
    db: AsyncSession,
    workbook_bytes: bytes,
    filename: str,
    current_user_id: int,
    dry_run: bool = True,
    mode: str = "history_only",
    force: bool = False,
) -> dict:
    if mode not in VALID_MODES:
        mode = "history_only"

    product_import_batch_supported = await _column_exists(db, "products", "created_by_import_batch")
    invoice_import_batch_supported = await _column_exists(db, "b2b_invoices", "import_batch_id")
    consignment_import_batch_supported = await _column_exists(db, "consignments", "import_batch_id")

    wb = openpyxl.load_workbook(io.BytesIO(workbook_bytes), data_only=True)
    ws = wb.active

    # ── Detect columns ──────────────────────────────────────────────────────
    raw_headers = [
        str(ws.cell(1, c).value or "").strip().lower()
        for c in range(1, ws.max_column + 1)
    ]
    col_sku      = _find_col(raw_headers, ["sku", "code", "item code", "barcode"])
    col_item     = _find_col(raw_headers, ["item", "name", "product", "description"])
    col_qty      = _find_col(raw_headers, ["qty", "quantity", "amount"])
    col_price    = _find_col(raw_headers, ["price", "unit price", "unit_price", "sale price", "sales price"])
    col_discount = _find_col(raw_headers, ["discount", "discount %", "disc", "disc %", "discount pct", "discount_pct"])
    col_ptype    = _find_col(raw_headers, ["payment type", "payment_type", "payment", "type", "payment method"])
    col_client   = _find_col(raw_headers, ["client name", "client_name", "client", "customer", "company"])
    col_date     = _find_col(raw_headers, ["date", "sale date", "invoice date", "order date"])

    missing = [n for n, c in [
        ("SKU", col_sku), ("QTY", col_qty), ("Price", col_price),
        ("Payment type", col_ptype), ("Client name", col_client), ("Date", col_date),
    ] if not c]
    if missing:
        return {
            "dry_run": dry_run, "mode": mode, "file": filename,
            "error": f"Required column(s) not found: {', '.join(missing)}",
            "errors": [],
        }

    def _cell(row, col):
        return ws.cell(row, col).value if col else None

    # ── Parse & validate every data row ────────────────────────────────────
    rows_read = 0
    all_rows: list[dict] = []

    for rn in range(2, ws.max_row + 1):
        if all(_cell(rn, c) is None for c in range(1, ws.max_column + 1)):
            continue
        rows_read += 1

        raw_sku      = _cell(rn, col_sku)
        raw_qty      = _cell(rn, col_qty)
        raw_price    = _cell(rn, col_price)
        raw_disc     = _cell(rn, col_discount)
        raw_ptype    = _cell(rn, col_ptype)
        raw_client   = _cell(rn, col_client)
        raw_date     = _cell(rn, col_date)
        raw_item     = _cell(rn, col_item) if col_item else None

        sku          = _normalize_sku(raw_sku)
        qty          = _safe_float(raw_qty)
        price        = _safe_float(raw_price)
        disc_raw     = _safe_float(raw_disc)
        discount_pct = disc_raw if disc_raw is not None else 0.0
        payment_type = _normalize_payment_type(raw_ptype)
        client_name  = str(raw_client).strip() if raw_client is not None else ""
        sale_date    = _parse_date(raw_date)
        item_hint    = str(raw_item).strip() if raw_item else ""

        row_errors: list[str] = []
        if not sku:
            row_errors.append("SKU is required")
        if qty is None:
            row_errors.append("QTY is required and must be numeric")
        elif qty <= 0:
            row_errors.append("QTY must be > 0")
        if price is None:
            row_errors.append("Price is required and must be numeric")
        elif price < 0:
            row_errors.append("Price must be >= 0")
        if not (0 <= discount_pct <= 100):
            row_errors.append(f"Discount {discount_pct} is out of range (0–100)")
        if payment_type is None:
            row_errors.append(
                f"Payment type '{raw_ptype}' not recognised — "
                "accepted: cash, full_payment, consignment (or aliases)"
            )
        if not client_name:
            row_errors.append("Client name is required")
        if sale_date is None:
            row_errors.append(f"Date '{raw_date}' is invalid — use YYYY-MM-DD, DD/MM/YYYY, or MM/DD/YYYY")
        elif sale_date < MIN_IMPORT_DATE:
            row_errors.append(f"Date {sale_date} is before the minimum import date (2026-01-01)")

        all_rows.append({
            "row":          rn,
            "sku":          sku or str(raw_sku or ""),
            "item":         item_hint,
            "qty":          qty,
            "price":        price,
            "discount_pct": discount_pct,
            "payment_type": payment_type,
            "client":       client_name,
            "date":         sale_date,
            "errors":       row_errors,
        })

    # ── Group by (client_lower, date, payment_type) ─────────────────────────
    groups: dict[tuple, list[dict]] = {}
    for r in all_rows:
        key = (r["client"].lower(), r["date"], r["payment_type"])
        groups.setdefault(key, []).append(r)

    valid_groups:   dict[tuple, list[dict]] = {}
    invalid_groups: dict[tuple, list[dict]] = {}
    for key, rows in groups.items():
        if any(r["errors"] for r in rows):
            invalid_groups[key] = rows
        else:
            valid_groups[key] = rows

    rows_skipped = sum(len(rows) for rows in invalid_groups.values())

    errors_report: list[dict] = []
    for rows in invalid_groups.values():
        for r in rows:
            for msg in r["errors"]:
                errors_report.append({
                    "row": r["row"], "sku": r["sku"],
                    "client": r["client"],
                    "date": str(r["date"]) if r["date"] else "",
                    "reason": msg,
                })

    # ── Pre-compute summary stats ───────────────────────────────────────────
    valid_rows  = [r for rows in valid_groups.values() for r in rows]
    valid_dates = [r["date"] for r in valid_rows if r["date"]]
    earliest_date = min(valid_dates).isoformat() if valid_dates else None
    latest_date   = max(valid_dates).isoformat() if valid_dates else None

    total_subtotal = sum(
        r["qty"] * r["price"]
        for r in valid_rows if r["qty"] is not None and r["price"] is not None
    )
    total_invoiced = sum(
        r["qty"] * r["price"] * (1 - r["discount_pct"] / 100)
        for r in valid_rows if r["qty"] is not None and r["price"] is not None
    )
    total_discount = round(total_subtotal - total_invoiced, 2)
    total_subtotal = round(total_subtotal, 2)
    total_invoiced = round(total_invoiced, 2)

    # by_payment_type preview
    by_ptype_groups: dict[str, dict] = {}
    for (client_lower, d, ptype), rows in valid_groups.items():
        if ptype not in by_ptype_groups:
            by_ptype_groups[ptype] = {"invoices": 0, "total": 0.0}
        by_ptype_groups[ptype]["invoices"] += 1
        by_ptype_groups[ptype]["total"] += round(sum(
            r["qty"] * r["price"] * (1 - r["discount_pct"] / 100)
            for r in rows if r["qty"] and r["price"]
        ), 2)

    consignments_would_create = sum(
        1 for (_, _, pt), _ in valid_groups.items() if pt == "consignment"
    )

    # ── Discount suggestions (read-only, done even for dry_run) ────────────
    # Group valid rows by client name
    client_rows_map: dict[str, list[dict]] = {}
    for r in valid_rows:
        client_rows_map.setdefault(r["client"].lower(), []).append(r)

    # Compute mode discount per unique client name
    client_mode_discount: dict[str, float] = {}
    for cname_lower, crows in client_rows_map.items():
        counts = Counter(r["discount_pct"] for r in crows)
        client_mode_discount[cname_lower] = counts.most_common(1)[0][0]

    # Look up existing clients
    existing_client_map: dict[str, "B2BClient"] = {}
    for cname_lower in client_rows_map:
        _cr = await db.execute(
            select(B2BClient).where(sa_func.lower(B2BClient.name) == cname_lower)
        )
        obj = _cr.scalar_one_or_none()
        if obj:
            existing_client_map[cname_lower] = obj

    discount_pct_suggestions: list[dict] = []
    for cname_lower, mode_disc in client_mode_discount.items():
        if mode_disc == 0:
            continue
        existing = existing_client_map.get(cname_lower)
        if existing is None:
            continue  # new client — handled in auto_created_clients
        current_pct = float(existing.discount_pct or 0)
        if current_pct == 0:
            discount_pct_suggestions.append({
                "client": existing.name,
                "current": 0.0,
                "suggested": mode_disc,
                "applied": True,
            })
        else:
            discount_pct_suggestions.append({
                "client": existing.name,
                "current": current_pct,
                "suggested": mode_disc,
                "applied": False,
                "note": "Client has existing discount_pct — manual review needed",
            })

    # Auto-created clients preview (for dry_run)
    auto_created_preview: list[dict] = []
    for cname_lower, crows in client_rows_map.items():
        if cname_lower in existing_client_map:
            continue
        first_pt = crows[0]["payment_type"] or "cash"
        mode_disc = client_mode_discount.get(cname_lower, 0.0)
        auto_created_preview.append({
            "name": crows[0]["client"],
            "payment_terms": first_pt,
            "discount_pct": mode_disc,
        })

    _all_p = await db.execute(
        select(Product).options(defer(Product.created_by_import_batch))
    )
    all_products = _all_p.scalars().all()
    products_by_norm_sku = {normalize_barcode_value(p.sku): p for p in all_products}

    products_auto_created_preview = 0
    auto_created_products_preview: list[dict] = []
    product_warnings_preview: list[str] = []
    preview_created_skus: set[str] = set()
    preview_name_registry: dict[str, str] = {}
    for r in valid_rows:
        raw_sku_str = r["sku"]
        if len(raw_sku_str) > _SKU_MAX:
            raw_sku_str = raw_sku_str[:_SKU_MAX]
        norm_sku = normalize_barcode_value(raw_sku_str)
        if not norm_sku or norm_sku in products_by_norm_sku or norm_sku in preview_created_skus:
            continue

        item_name = _build_imported_product_name(r["item"], raw_sku_str)
        if norm_sku in preview_name_registry:
            first_name = preview_name_registry[norm_sku]
            if first_name.lower() != item_name.lower():
                warn_msg = (
                    f"SKU '{raw_sku_str}' had conflicting item names: "
                    f"'{first_name}' vs '{item_name}' — used the first."
                )
                if warn_msg not in product_warnings_preview:
                    product_warnings_preview.append(warn_msg)
            item_name = first_name
        else:
            preview_name_registry[norm_sku] = item_name

        preview_created_skus.add(norm_sku)
        products_auto_created_preview += 1
        auto_created_products_preview.append({
            "sku": raw_sku_str,
            "name": item_name,
        })

    if products_auto_created_preview > 0:
        product_warnings_preview.insert(
            0,
            f"{products_auto_created_preview} products would be auto-created with cost = 0. "
            "Review Products and set costs if needed."
        )

    # ── Dry-run: return preview without touching DB ─────────────────────────
    if dry_run:
        return {
            "dry_run": True,
            "mode": mode,
            "file": filename,
            "batch_id": None,
            "summary": {
                "rows_read":             rows_read,
                "invoices_created":      0,
                "invoices_would_create": len(valid_groups),
                "line_items":            len(valid_rows),
                "clients_auto_created":  len(auto_created_preview),
                "products_auto_created": products_auto_created_preview,
                "consignments_created":  0,
                "consignments_would_create": consignments_would_create,
                "rows_skipped":          rows_skipped,
                "earliest_date":         earliest_date,
                "latest_date":           latest_date,
                "total_subtotal":        total_subtotal,
                "total_discount":        total_discount,
                "total_invoiced":        total_invoiced,
                "by_payment_type":       {
                    k: {"invoices": v["invoices"], "total": round(v["total"], 2)}
                    for k, v in by_ptype_groups.items()
                },
                "client_prices_created": 0,
                "client_prices_updated": 0,
            },
            "discount_pct_suggestions": discount_pct_suggestions,
            "auto_created_clients":     auto_created_preview,
            "auto_created_products":    auto_created_products_preview[:50],
            "warnings":                 product_warnings_preview,
            "errors":                   errors_report,
        }

    # ── Real import ─────────────────────────────────────────────────────────
    await seed_deferred_revenue(db)

    batch_id                 = str(uuid.uuid4())
    invoices_created         = 0
    consignments_created     = 0
    clients_auto_created     = 0
    products_auto_created    = 0
    auto_created_client_names: list[dict] = []
    auto_created_products_list: list[dict] = []
    warnings_list: list[str] = []
    today_str                = date.today().isoformat()

    # Fetch all active products once
    _all_p = await db.execute(
        select(Product).options(defer(Product.created_by_import_batch))
    )
    all_products = _all_p.scalars().all()
    products_by_norm_sku = {normalize_barcode_value(p.sku): p for p in all_products}
    newly_created_by_norm_sku: dict[str, Product] = {}
    sku_name_registry: dict[str, str] = {}

    # Fetch starting MAX ids for bulk-safe invoice numbering
    _r = await db.execute(select(sa_func.max(B2BInvoice.id)))
    max_b2b_id = _r.scalar() or 0
    _r = await db.execute(select(sa_func.max(Consignment.id)))
    max_cons_id = _r.scalar() or 0
    b2b_counter  = 0
    cons_counter = 0

    # Track client_id → list of (product_id, qty, post_discount_unit_price)
    # for B2BClientPrice and discount propagation after all commits
    client_line_history: dict[int, list[tuple]] = {}

    for (client_key, sale_date, payment_type), rows in valid_groups.items():
        client_name = rows[0]["client"]

        # ── Resolve client ─────────────────────────────────────────────────
        _cr = await db.execute(
            select(B2BClient).where(sa_func.lower(B2BClient.name) == client_key)
        )
        client = _cr.scalar_one_or_none()
        if not client:
            mode_disc = client_mode_discount.get(client_key, 0.0)
            first_pt  = payment_type
            client = B2BClient(
                name=client_name,
                payment_terms=first_pt,
                discount_pct=mode_disc,
                credit_limit=0,
                outstanding=0,
                is_active=True,
            )
            db.add(client)
            await db.flush()
            clients_auto_created += 1
            auto_created_client_names.append({
                "name": client_name,
                "payment_terms": first_pt,
                "discount_pct": mode_disc,
            })

        # ── Duplicate detection ────────────────────────────────────────────
        if not force and sale_date is not None:
            _dup = await db.execute(
                select(B2BInvoice).where(
                    B2BInvoice.client_id == client.id,
                    cast(B2BInvoice.created_at, SADate) == sale_date,
                    B2BInvoice.invoice_type == payment_type,
                    B2BInvoice.notes.like("Imported from %"),
                )
            )
            if _dup.scalar_one_or_none() is not None:
                for r in rows:
                    errors_report.append({
                        "row": r["row"], "sku": r["sku"],
                        "client": client_name,
                        "date": str(sale_date),
                        "reason": "Duplicate: invoice already exists for this client+date+type from a prior import",
                    })
                rows_skipped += len(rows)
                continue

        # ── Resolve products ───────────────────────────────────────────────
        line_items: list[tuple] = []  # (product, qty, unit_price, discount_pct, line_total)
        line_errors: list[dict] = []
        group_created_norm_skus: list[str] = []
        group_created_products_count = 0

        for r in rows:
            raw_sku_str = r["sku"]
            if len(raw_sku_str) > _SKU_MAX:
                truncated_sku = raw_sku_str[:_SKU_MAX]
                line_errors.append({
                    "row": r["row"], "sku": raw_sku_str,
                    "client": client_name, "date": str(sale_date),
                    "reason": f"SKU truncated from {len(raw_sku_str)} to {_SKU_MAX} chars: '{truncated_sku}'",
                })
                raw_sku_str = truncated_sku

            norm_sku = normalize_barcode_value(raw_sku_str)
            if not norm_sku:
                line_errors.append({
                    "row": r["row"], "sku": r["sku"],
                    "client": client_name, "date": str(sale_date),
                    "reason": "SKU is blank or invalid after normalization",
                })
                continue

            product = products_by_norm_sku.get(norm_sku) or newly_created_by_norm_sku.get(norm_sku)
            if not product:
                item_name = _build_imported_product_name(r["item"], raw_sku_str)

                if norm_sku in sku_name_registry:
                    first_name = sku_name_registry[norm_sku]
                    if first_name.lower() != item_name.lower():
                        warn_msg = (
                            f"SKU '{raw_sku_str}' had conflicting item names: "
                            f"'{first_name}' vs '{item_name}' — used the first."
                        )
                        if warn_msg not in warnings_list:
                            warnings_list.append(warn_msg)
                    item_name = first_name
                else:
                    sku_name_registry[norm_sku] = item_name

                product_payload = {
                    "sku": raw_sku_str,
                    "name": item_name,
                    "price": round(float(r["price"]), 2),
                    "cost": 0,
                    "stock": 0,
                    "min_stock": 5,
                    "unit": "pcs",
                    "is_active": True,
                    "category": "Imported - Historical B2B",
                    "item_type": "finished",
                }

                existing_product = await _find_existing_product_by_norm_sku(db, norm_sku)
                if existing_product:
                    product = existing_product
                    if product.is_active is False:
                        product.is_active = True
                else:
                    try:
                        async with db.begin_nested():
                            if product_import_batch_supported:
                                product = Product(**product_payload, created_by_import_batch=batch_id)
                                db.add(product)
                                await db.flush()
                            else:
                                product_insert = await db.execute(
                                    insert(Product.__table__).values(**product_payload).returning(Product.id)
                                )
                                product_id = product_insert.scalar_one()
                                fetched_product = await db.execute(
                                    select(Product)
                                    .options(defer(Product.created_by_import_batch))
                                    .where(Product.id == product_id)
                                )
                                product = fetched_product.scalar_one()
                    except IntegrityError:
                        existing_product = await _find_existing_product_by_norm_sku(db, norm_sku)
                        if not existing_product:
                            raise
                        product = existing_product
                        if product.is_active is False:
                            product.is_active = True

                newly_created_by_norm_sku[norm_sku] = product
                if existing_product is None:
                    group_created_norm_skus.append(norm_sku)
                    group_created_products_count += 1
                    products_auto_created += 1
                    auto_created_products_list.append({
                        "sku": raw_sku_str,
                        "name": item_name,
                    })

            disc = r["discount_pct"]
            line_total = round(float(r["qty"]) * float(r["price"]) * (1 - disc / 100), 2)
            line_items.append((product, float(r["qty"]), float(r["price"]), disc, line_total))

        if line_errors:
            errors_report.extend(line_errors)
            rows_skipped += len(rows)
            if group_created_norm_skus:
                await db.rollback()
                for norm_sku in group_created_norm_skus:
                    newly_created_by_norm_sku.pop(norm_sku, None)
                products_auto_created -= group_created_products_count
                del auto_created_products_list[-group_created_products_count:]
            continue

        # ── Build Invoice ──────────────────────────────────────────────────
        subtotal_inv = round(sum(float(r["qty"]) * float(r["price"])
                                 for r in rows if r["qty"] and r["price"]), 2)
        total_inv    = round(sum(lt for _, _, _, _, lt in line_items), 2)
        discount_inv = round(subtotal_inv - total_inv, 2)

        sale_dt = datetime(sale_date.year, sale_date.month, sale_date.day, 12, 0, 0)

        status       = "paid"  if payment_type == "cash" else "unpaid"
        amount_paid  = total_inv if payment_type == "cash" else 0.0

        b2b_counter += 1
        invoice_number, next_b2b_number = await _next_unique_reference(
            db, B2BInvoice, "invoice_number", "HB2B-", max_b2b_id + b2b_counter, 5
        )
        b2b_counter = max(0, next_b2b_number - max_b2b_id)

        try:
            invoice_payload = {
                "invoice_number": invoice_number,
                "client_id": client.id,
                "user_id": current_user_id,
                "invoice_type": payment_type,
                "status": status,
                "payment_method": payment_type,
                "subtotal": subtotal_inv,
                "discount": discount_inv,
                "total": total_inv,
                "amount_paid": amount_paid,
                "notes": f"Imported from {filename} on {today_str}",
                "created_at": sale_dt,
            }
            if invoice_import_batch_supported:
                invoice = B2BInvoice(**invoice_payload, import_batch_id=batch_id)
                db.add(invoice)
                await db.flush()
                invoice_id = invoice.id
            else:
                invoice_insert = await db.execute(
                    insert(B2BInvoice.__table__).values(**invoice_payload).returning(B2BInvoice.id)
                )
                invoice_id = invoice_insert.scalar_one()

            for product, qty, unit_price, disc, line_total in line_items:
                db.add(B2BInvoiceItem(
                    invoice_id=invoice_id,
                    product_id=product.id,
                    qty=qty,
                    unit_price=unit_price,
                    total=line_total,
                ))

                if mode == "with_stock_adjustment":
                    before = float(product.stock)
                    after  = before - qty
                    product.stock = after
                    db.add(StockMove(
                        product_id=product.id, type="out", qty=-qty,
                        qty_before=before, qty_after=after,
                        ref_type="b2b", ref_id=invoice_id,
                        note=f"Historical B2B import — {invoice_number}",
                        user_id=current_user_id,
                    ))

                # Track for B2BClientPrice & discount propagation
                post_discount_unit = line_total / qty if qty else 0.0
                client_line_history.setdefault(client.id, []).append(
                    (product.id, qty, post_discount_unit)
                )

            # ── Journals ────────────────────────────────────────────────────
            if payment_type == "cash":
                await post_journal(
                    db, f"B2B Historical Cash Sale — {invoice_number}", "b2b",
                    [("1000", total_inv, 0), ("4000", 0, total_inv)],
                    user_id=current_user_id,
                    created_at=sale_dt,
                    ref_id=invoice_id,
                )
            else:  # full_payment or consignment
                await post_journal(
                    db,
                    f"B2B Historical {'Consignment' if payment_type == 'consignment' else 'Full Payment'} — {invoice_number}",
                    "b2b",
                    [("1100", total_inv, 0), ("2200", 0, total_inv)],
                    user_id=current_user_id,
                    created_at=sale_dt,
                    ref_id=invoice_id,
                )
                client.outstanding = Decimal(str(float(client.outstanding) + total_inv))

            # ── Consignment record ───────────────────────────────────────────
            if payment_type == "consignment":
                cons_counter += 1
                cons_ref, next_cons_number = await _next_unique_reference(
                    db, Consignment, "ref_number", "HCONS-", max_cons_id + cons_counter, 4
                )
                cons_counter = max(0, next_cons_number - max_cons_id)
                consignment_payload = {
                    "ref_number": cons_ref,
                    "client_id": client.id,
                    "invoice_id": invoice_id,
                    "user_id": current_user_id,
                    "status": "active",
                    "created_at": sale_dt,
                    "notes": f"Imported from {filename} on {today_str}",
                }
                if consignment_import_batch_supported:
                    consignment = Consignment(**consignment_payload, import_batch_id=batch_id)
                    db.add(consignment)
                    await db.flush()
                    consignment_id = consignment.id
                else:
                    consignment_insert = await db.execute(
                        insert(Consignment.__table__).values(**consignment_payload).returning(Consignment.id)
                    )
                    consignment_id = consignment_insert.scalar_one()
                for product, qty, unit_price, disc, line_total in line_items:
                    post_disc_unit = line_total / qty if qty else 0.0
                    db.add(ConsignmentItem(
                        consignment_id=consignment_id,
                        product_id=product.id,
                        qty_sent=qty,
                        qty_sold=0,
                        qty_returned=0,
                        unit_price=post_disc_unit,
                    ))
                consignments_created += 1

            await db.commit()
            invoices_created += 1

        except Exception as exc:
            await db.rollback()
            for norm_sku in group_created_norm_skus:
                newly_created_by_norm_sku.pop(norm_sku, None)
            products_auto_created -= group_created_products_count
            if group_created_products_count:
                del auto_created_products_list[-group_created_products_count:]
            errors_report.append({
                "row": rows[0]["row"], "sku": "",
                "client": client_name, "date": str(sale_date),
                "reason": f"DB error creating invoice: {exc}",
            })
            rows_skipped += len(rows)
            continue

        for norm_sku in group_created_norm_skus:
            products_by_norm_sku[norm_sku] = newly_created_by_norm_sku[norm_sku]

    # ── Discount propagation + B2BClientPrice (post-all-commits) ──────────
    client_prices_created = 0
    client_prices_updated = 0

    for client_id, lines in client_line_history.items():
        # Re-fetch client (it may have been flushed in a rolled-back session)
        _c = await db.execute(select(B2BClient).where(B2BClient.id == client_id))
        client = _c.scalar_one_or_none()
        if not client:
            continue

        # Discount propagation from already-set discount_pct on auto-created clients
        # For existing clients, apply the suggestion rule
        cname_lower = client.name.lower()
        mode_disc = client_mode_discount.get(cname_lower, 0.0)
        if mode_disc > 0 and float(client.discount_pct or 0) == 0:
            # Only apply if not already set (auto-created clients already have it set)
            current_pct = float(client.discount_pct or 0)
            if current_pct == 0:
                client.discount_pct = Decimal(str(mode_disc))

        # B2BClientPrice: group by product_id
        product_lines: dict[int, list[tuple]] = {}
        for (prod_id, qty, pdp) in lines:
            product_lines.setdefault(prod_id, []).append((qty, pdp))

        for prod_id, qty_pdp_list in product_lines.items():
            if len(qty_pdp_list) < 2:
                continue
            total_qty = sum(q for q, _ in qty_pdp_list)
            wavg = sum(q * p for q, p in qty_pdp_list) / total_qty if total_qty else 0.0
            wavg = round(wavg, 2)

            _cp = await db.execute(
                select(B2BClientPrice).where(
                    B2BClientPrice.client_id == client_id,
                    B2BClientPrice.product_id == prod_id,
                )
            )
            cp = _cp.scalar_one_or_none()
            if cp:
                cp.price = Decimal(str(wavg))
                client_prices_updated += 1
            else:
                db.add(B2BClientPrice(
                    client_id=client_id,
                    product_id=prod_id,
                    price=Decimal(str(wavg)),
                ))
                client_prices_created += 1

    try:
        await db.commit()
    except Exception:
        await db.rollback()

    if products_auto_created > 0:
        warnings_list.insert(
            0,
            f"{products_auto_created} products were auto-created with cost = 0. "
            "Review Products and set costs if needed."
        )

    return {
        "dry_run": False,
        "mode": mode,
        "file": filename,
        "batch_id": batch_id,
        "summary": {
            "rows_read":             rows_read,
            "invoices_created":      invoices_created,
            "line_items":            sum(len(rows) for rows in valid_groups.values()),
            "clients_auto_created":  clients_auto_created,
            "products_auto_created": products_auto_created,
            "consignments_created":  consignments_created,
            "rows_skipped":          rows_skipped,
            "earliest_date":         earliest_date,
            "latest_date":           latest_date,
            "total_subtotal":        total_subtotal,
            "total_discount":        total_discount,
            "total_invoiced":        total_invoiced,
            "by_payment_type":       {
                k: {"invoices": v["invoices"], "total": round(v["total"], 2)}
                for k, v in by_ptype_groups.items()
            },
            "client_prices_created": client_prices_created,
            "client_prices_updated": client_prices_updated,
        },
        "discount_pct_suggestions": discount_pct_suggestions,
        "auto_created_clients":     auto_created_client_names[:50],
        "auto_created_products":    auto_created_products_list[:50],
        "warnings":                 warnings_list,
        "errors":                   errors_report,
    }
