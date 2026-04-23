from __future__ import annotations

import io
import json
import uuid
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

import openpyxl
from fastapi import HTTPException
from openpyxl.utils.datetime import from_excel
from sqlalchemy import or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.log import ActivityLog, record
from app.models.inventory import StockMove
from app.models.product import Product
from app.models.receipt import ProductReceipt
from app.models.user import User
from app.services.receive_service import (
    PRODUCT_TYPE_PACKAGING,
    PRODUCT_TYPE_PRODUCTS,
    ReceiptCreate,
    create_receipt,
    delete_receipt,
)


PRODUCT_TYPE_ALIASES = {
    "products": PRODUCT_TYPE_PRODUCTS,
    "product": PRODUCT_TYPE_PRODUCTS,
    "finished": PRODUCT_TYPE_PRODUCTS,
    "packaging materials": PRODUCT_TYPE_PACKAGING,
    "packaging material": PRODUCT_TYPE_PACKAGING,
    "packaging": PRODUCT_TYPE_PACKAGING,
    "packaging_materials": PRODUCT_TYPE_PACKAGING,
}


def _normalize_key(value: object) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("_", " ").replace("-", " ")
    return " ".join(text.split())


def _find_header_index(headers: list[str], aliases: list[str]) -> int | None:
    normalized = [_normalize_key(header) for header in headers]
    alias_set = {_normalize_key(alias) for alias in aliases}
    for idx, header in enumerate(normalized):
        if header in alias_set:
            return idx + 1
    return None


def _normalize_sku(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        if "." in text:
            return str(int(float(text)))
    except (TypeError, ValueError):
        pass
    return text


def _parse_decimal(value: object, *, scale: str) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, Decimal):
        return value.quantize(Decimal(scale))
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return Decimal(str(value)).quantize(Decimal(scale))

    text = str(value).strip()
    if not text:
        return None
    cleaned = text.replace(",", "").replace("$", "")
    if cleaned.startswith("(") and cleaned.endswith(")"):
        cleaned = f"-{cleaned[1:-1]}"
    try:
        return Decimal(cleaned).quantize(Decimal(scale))
    except InvalidOperation:
        return None


def _parse_date_value(value: object) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            parsed = from_excel(value)
        except Exception:
            parsed = None
        if isinstance(parsed, datetime):
            return parsed.date()
        if isinstance(parsed, date):
            return parsed
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d", "%d-%m-%Y", "%m-%d-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    try:
        return date.fromisoformat(text[:10])
    except ValueError:
        return None


def _normalize_product_type(value: object) -> str | None:
    key = _normalize_key(value)
    if not key:
        return None
    return PRODUCT_TYPE_ALIASES.get(key)


def _batch_payload_from_log(log: ActivityLog) -> dict:
    try:
        payload = json.loads(log.description or "{}")
    except json.JSONDecodeError:
        payload = {}
    return payload if isinstance(payload, dict) else {}


async def _load_product_map(db: AsyncSession) -> dict[str, Product]:
    result = await db.execute(
        select(Product)
        .where(or_(Product.is_active.is_(True), Product.is_active.is_(None)))
    )
    products = result.scalars().all()
    mapping = {}
    for product in products:
        sku = _normalize_sku(product.sku)
        if sku:
            mapping[sku] = product
    return mapping


async def import_receive_products(
    *,
    db: AsyncSession,
    workbook_bytes: bytes,
    filename: str,
    current_user: User,
    dry_run: bool = True,
) -> dict:
    workbook = openpyxl.load_workbook(io.BytesIO(workbook_bytes), data_only=True)
    sheet = workbook.active
    headers = [str(sheet.cell(1, col).value or "") for col in range(1, sheet.max_column + 1)]

    col_sku = _find_header_index(headers, ["SKU", "Item Code", "Code"])
    col_product = _find_header_index(headers, ["Product", "Item", "Name", "Product Name"])
    col_qty = _find_header_index(headers, ["QTY", "Qty", "Quantity"])
    col_unit_price = _find_header_index(headers, ["Unit Price", "Unit Cost", "Cost", "Price"])
    col_product_type = _find_header_index(headers, ["Product Type", "Type"])
    col_date = _find_header_index(headers, ["Date", "Receive Date"])

    missing = []
    if not col_sku:
        missing.append("SKU")
    if not col_qty:
        missing.append("QTY")
    if not col_unit_price:
        missing.append("Unit Price")
    if not col_product_type:
        missing.append("Product Type")
    if not col_date:
        missing.append("Date")
    if missing:
        raise HTTPException(status_code=400, detail=f"Missing required column(s): {', '.join(missing)}")

    product_map = await _load_product_map(db)
    rows_to_import: list[dict] = []
    errors: list[dict] = []
    resolved_products = 0
    products_rows = 0
    packaging_rows = 0
    total_cost = Decimal("0.00")
    all_dates: list[date] = []

    for row_idx in range(2, sheet.max_row + 1):
        values = [sheet.cell(row_idx, col).value for col in range(1, sheet.max_column + 1)]
        if all(str(value or "").strip() == "" for value in values):
            continue

        raw_sku = sheet.cell(row_idx, col_sku).value if col_sku else None
        raw_product = sheet.cell(row_idx, col_product).value if col_product else None
        raw_qty = sheet.cell(row_idx, col_qty).value if col_qty else None
        raw_unit_price = sheet.cell(row_idx, col_unit_price).value if col_unit_price else None
        raw_product_type = sheet.cell(row_idx, col_product_type).value if col_product_type else None
        raw_date = sheet.cell(row_idx, col_date).value if col_date else None

        sku = _normalize_sku(raw_sku)
        product_name = str(raw_product or "").strip()
        qty = _parse_decimal(raw_qty, scale="0.001")
        unit_price = _parse_decimal(raw_unit_price, scale="0.01")
        product_type = _normalize_product_type(raw_product_type)
        receive_date = _parse_date_value(raw_date)

        row_errors: list[str] = []
        if not sku:
            row_errors.append("SKU is required")
        product = product_map.get(sku) if sku else None
        if sku and product is None:
            row_errors.append(f"Product with SKU '{sku}' was not found")
        if qty is None:
            row_errors.append("QTY is required and must be numeric")
        elif qty <= 0:
            row_errors.append("QTY must be greater than 0")
        if unit_price is None:
            row_errors.append("Unit Price is required and must be numeric")
        elif unit_price < 0:
            row_errors.append("Unit Price must be zero or greater")
        if product_type is None:
            row_errors.append("Product Type is required and must be Products or Packaging Materials")
        if receive_date is None:
            row_errors.append("Date is required and must be a valid date")

        if row_errors:
            errors.append(
                {
                    "row": row_idx,
                    "sku": sku or "",
                    "product": product_name,
                    "qty": "" if raw_qty is None else str(raw_qty),
                    "unit_price": "" if raw_unit_price is None else str(raw_unit_price),
                    "product_type": "" if raw_product_type is None else str(raw_product_type),
                    "date": "" if raw_date is None else str(raw_date),
                    "reason": "; ".join(row_errors),
                }
            )
            continue

        rows_to_import.append(
            {
                "row": row_idx,
                "product": product,
                "sku": sku,
                "product_name": product_name or (product.name if product else ""),
                "qty": qty,
                "unit_price": unit_price,
                "product_type": product_type,
                "receive_date": receive_date,
            }
        )
        resolved_products += 1
        if product_type == PRODUCT_TYPE_PRODUCTS:
            products_rows += 1
        else:
            packaging_rows += 1
        total_cost += (qty * unit_price).quantize(Decimal("0.01"))
        all_dates.append(receive_date)

    receipts_created = 0
    stock_moves_created = 0
    batch_id = str(uuid.uuid4())
    if not dry_run:
        for row in rows_to_import:
            payload = ReceiptCreate(
                product_id=row["product"].id,
                qty=float(row["qty"]),
                unit_cost=float(row["unit_price"]),
                product_type=row["product_type"],
                receive_date=row["receive_date"],
                supplier_ref=None,
                notes=None,
            )
            created = await create_receipt(db, payload, current_user)
            receipts_created += 1
            stock_moves_created += 1
            record(
                db,
                "Import",
                "receive_import_item",
                json.dumps(
                    {
                        "batch_id": batch_id,
                        "ref_number": created.get("ref_number"),
                        "sku": row["sku"],
                        "product_type": row["product_type"],
                        "qty": float(row["qty"]),
                        "unit_price": float(row["unit_price"]),
                    }
                ),
                user=current_user,
                ref_type=batch_id,
                ref_id=created.get("id"),
            )

        record(
            db,
            "Import",
            "receive_import_batch",
            json.dumps(
                {
                    "batch_id": batch_id,
                    "filename": filename,
                    "rows_read": len(rows_to_import) + len(errors),
                    "rows_imported": receipts_created,
                    "rows_skipped": len(errors),
                    "products_resolved": resolved_products,
                    "products_auto_created": 0,
                    "receive_records_created": receipts_created,
                    "stock_moves_created": stock_moves_created,
                    "products_rows_count": products_rows,
                    "packaging_rows_count": packaging_rows,
                }
            ),
            user=current_user,
            ref_type="receive_import_batch",
            ref_id=batch_id,
        )
        await db.commit()

    summary = {
        "rows_read": len(rows_to_import) + len(errors),
        "rows_imported": len(rows_to_import) if dry_run else receipts_created,
        "rows_skipped": len(errors),
        "products_resolved": resolved_products,
        "products_auto_created": 0,
        "receive_records_created": len(rows_to_import) if dry_run else receipts_created,
        "stock_moves_created": len(rows_to_import) if dry_run else stock_moves_created,
        "products_rows_count": products_rows,
        "packaging_rows_count": packaging_rows,
        "earliest_date": min(all_dates).isoformat() if all_dates else None,
        "latest_date": max(all_dates).isoformat() if all_dates else None,
        "total_cost": float(total_cost),
    }
    return {
        "ok": len(errors) == 0,
        "dry_run": dry_run,
        "filename": filename,
        "summary": summary,
        "errors": errors,
        "batch_id": None if dry_run or receipts_created == 0 else batch_id,
        "revert_available": (not dry_run and receipts_created > 0),
    }


async def list_receive_import_batches(db: AsyncSession) -> dict:
    result = await db.execute(
        select(ActivityLog)
        .where(
            ActivityLog.module == "Import",
            ActivityLog.action.in_(["receive_import_batch", "receive_import_revert"]),
        )
        .order_by(ActivityLog.created_at.desc())
    )
    logs = result.scalars().all()
    batches: dict[str, dict] = {}
    for log in logs:
        batch_id = log.ref_id
        if not batch_id:
            continue
        if log.action == "receive_import_batch":
            payload = _batch_payload_from_log(log)
            batches[batch_id] = {
                "batch_id": batch_id,
                "filename": payload.get("filename") or "receive_products.xlsx",
                "ran_on": log.created_at.date().isoformat() if log.created_at else None,
                "rows_read": payload.get("rows_read", 0),
                "rows_imported": payload.get("rows_imported", 0),
                "rows_skipped": payload.get("rows_skipped", 0),
                "receive_records_created": payload.get("receive_records_created", 0),
                "stock_moves_created": payload.get("stock_moves_created", 0),
                "products_rows_count": payload.get("products_rows_count", 0),
                "packaging_rows_count": payload.get("packaging_rows_count", 0),
                "reverted": False,
                "reverted_on": None,
            }
        elif log.action == "receive_import_revert" and batch_id in batches:
            batches[batch_id]["reverted"] = True
            batches[batch_id]["reverted_on"] = log.created_at.date().isoformat() if log.created_at else None
    ordered = sorted(batches.values(), key=lambda item: (item["ran_on"] or "", item["batch_id"]), reverse=True)
    return {"batches": ordered}


async def revert_receive_import_batch(
    db: AsyncSession,
    batch_id: str,
    current_user: User,
) -> dict:
    summary_result = await db.execute(
        select(ActivityLog).where(
            ActivityLog.module == "Import",
            ActivityLog.action == "receive_import_batch",
            ActivityLog.ref_id == batch_id,
        )
    )
    summary_log = summary_result.scalar_one_or_none()
    if not summary_log:
        raise HTTPException(status_code=404, detail="Receive Products import batch not found")

    reverted_result = await db.execute(
        select(ActivityLog).where(
            ActivityLog.module == "Import",
            ActivityLog.action == "receive_import_revert",
            ActivityLog.ref_id == batch_id,
        )
    )
    if reverted_result.scalar_one_or_none():
        return {"ok": True, "batch_id": batch_id, "already_reverted": True, "deleted_receipts": 0}

    item_result = await db.execute(
        select(ActivityLog)
        .where(
            ActivityLog.module == "Import",
            ActivityLog.action == "receive_import_item",
            ActivityLog.ref_type == batch_id,
        )
        .order_by(ActivityLog.id.desc())
    )
    item_logs = item_result.scalars().all()
    deleted_receipts = 0
    skipped_missing = 0
    seen_ids: set[int] = set()
    for log in item_logs:
        try:
            receipt_id = int(log.ref_id)
        except (TypeError, ValueError):
            continue
        if receipt_id in seen_ids:
            continue
        seen_ids.add(receipt_id)
        try:
            await delete_receipt(db, receipt_id, current_user)
            deleted_receipts += 1
        except HTTPException as exc:
            if exc.status_code == 404:
                skipped_missing += 1
                continue
            raise

    record(
        db,
        "Import",
        "receive_import_revert",
        json.dumps({"batch_id": batch_id, "deleted_receipts": deleted_receipts, "skipped_missing": skipped_missing}),
        user=current_user,
        ref_type="receive_import_batch",
        ref_id=batch_id,
    )
    await db.commit()
    return {
        "ok": True,
        "batch_id": batch_id,
        "deleted_receipts": deleted_receipts,
        "skipped_missing": skipped_missing,
    }
