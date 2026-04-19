"""
sales_import_service.py

Back-fill historical retail sales from an Excel sheet into invoices / invoice_items.

Operating modes
───────────────
history_only          Create Invoice + InvoiceItem rows only.  No stock, no journals.
with_journals         Also post the same 1000/4000 double-entry journal the POS uses,
                      date-stamped with the historical sale date.
with_stock_and_journals  Full live-sale behaviour: decrement stock, write StockMoves,
                      post journals.  Rarely appropriate for a backfill.

Idempotency
───────────
Every real (non-dry-run) import carries a UUID4 batch_id stamped on every
Invoice.import_batch_id.  Re-running the same file after a successful import is
detected via "notes LIKE 'Imported from %'" on the same (customer, date).  The
caller can override with force=True.

Product auto-creation
─────────────────────
When a row's SKU is not in the products table, a Product is created inline with
category="Imported - Historical" and created_by_import_batch=batch_id.
cost defaults to 0 unless the caller passes default_cost_ratio (cost = price × ratio).
"""

import io
import uuid
from datetime import date, datetime, timezone
from typing import Optional

import openpyxl
from sqlalchemy import Date as SADate, cast, or_, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.sql import func as sa_func

from app.models.customer import Customer
from app.models.inventory import StockMove
from app.models.invoice import Invoice, InvoiceItem
from app.models.product import Product
from app.services.barcode_service import normalize_barcode_value
from app.services.location_inventory_service import sync_product_stock_to_default_location
from app.services.pos_service import get_walk_in_customer_id, post_journal

MIN_IMPORT_DATE = date(2026, 1, 1)
VALID_MODES = frozenset({"history_only", "with_journals", "with_stock_and_journals"})
_SKU_MAX  = 80
_NAME_MAX = 200
_CUST_MAX = 150


# ── Row-level helpers ─────────────────────────────────────────────────────────

def _parse_date(val) -> Optional[date]:
    """Accept openpyxl datetime/date objects and common string formats."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    if not s or s.lower() in ("none", ""):
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _safe_float(val) -> Optional[float]:
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _normalize_sku(val) -> str:
    if val is None:
        return ""
    if isinstance(val, (int, float)):
        try:
            return str(int(val))
        except (ValueError, TypeError):
            pass
    return str(val).strip()


def _find_col(headers: list[str], candidates: list[str]) -> Optional[int]:
    """Return 1-based column index of the first matching header, or None."""
    for name in candidates:
        for i, h in enumerate(headers):
            if h == name.lower().strip():
                return i + 1
    return None


# ── Main service entry point ──────────────────────────────────────────────────

async def import_sales(
    db: AsyncSession,
    workbook_bytes: bytes,
    filename: str,
    current_user_id: int,
    dry_run: bool = True,
    mode: str = "history_only",
    force: bool = False,
    default_cost_ratio: Optional[float] = None,
) -> dict:
    """
    Parse an Excel workbook and import historical sales.

    Returns a result dict matching the documented response shape.
    In dry_run=True mode, simulates customer/product resolution without writing.
    """
    if mode not in VALID_MODES:
        mode = "history_only"

    wb = openpyxl.load_workbook(io.BytesIO(workbook_bytes), data_only=True)
    ws = wb.active

    # ── Detect columns ──────────────────────────────────────────────────────
    raw_headers = [
        str(ws.cell(1, c).value or "").strip().lower()
        for c in range(1, ws.max_column + 1)
    ]
    col_sku      = _find_col(raw_headers, ["sku", "code", "item code", "barcode"])
    col_item     = _find_col(raw_headers, ["item", "name", "product", "description"])
    col_qty      = _find_col(raw_headers, ["qty", "quantity", "amount"])
    col_price    = _find_col(raw_headers, ["price", "unit price", "sale price", "sales price"])
    col_customer = _find_col(raw_headers, ["customer", "customer name", "client"])
    col_date     = _find_col(raw_headers, ["date", "sale date", "invoice date"])

    missing = [n for n, c in [("SKU", col_sku), ("QTY", col_qty), ("Price", col_price), ("Date", col_date)] if not c]
    if missing:
        return {
            "dry_run": dry_run,
            "mode": mode,
            "file": filename,
            "error": f"Required column(s) not found in sheet: {', '.join(missing)}",
            "errors": [],
        }

    def _cell(row, col):
        if not col:
            return None
        return ws.cell(row, col).value

    # ── Parse & validate every data row ────────────────────────────────────
    rows_read = 0
    all_rows: list[dict] = []

    for rn in range(2, ws.max_row + 1):
        if all(_cell(rn, c) is None for c in range(1, ws.max_column + 1)):
            continue
        rows_read += 1

        raw_sku      = _cell(rn, col_sku)
        raw_qty      = _cell(rn, col_qty)
        raw_price    = _cell(rn, col_price)
        raw_customer = _cell(rn, col_customer)
        raw_date     = _cell(rn, col_date)
        raw_item     = _cell(rn, col_item) if col_item else None

        sku           = _normalize_sku(raw_sku)
        qty           = _safe_float(raw_qty)
        price         = _safe_float(raw_price)
        customer_name = str(raw_customer).strip() if raw_customer is not None else ""
        sale_date     = _parse_date(raw_date)
        item_hint     = str(raw_item).strip() if raw_item else ""

        row_errors: list[str] = []
        if not sku:
            row_errors.append("SKU is required")
        if qty is None:
            row_errors.append("QTY is required and must be numeric")
        elif qty <= 0:
            row_errors.append("QTY must be > 0")
        if price is None:
            row_errors.append("Price is required and must be numeric")
        elif price < 0:
            row_errors.append("Price must be >= 0")
        if sale_date is None:
            row_errors.append(f"Date '{raw_date}' is invalid — use YYYY-MM-DD, DD/MM/YYYY, or MM/DD/YYYY")
        elif sale_date < MIN_IMPORT_DATE:
            row_errors.append(f"Date {sale_date} is before the minimum import date (2026-01-01)")

        all_rows.append({
            "row":      rn,
            "sku":      sku or str(raw_sku or ""),
            "item":     item_hint,
            "qty":      qty,
            "price":    price,
            "customer": customer_name,
            "date":     sale_date,
            "errors":   row_errors,
        })

    # ── Group by (normalised customer name, date) ───────────────────────────
    groups: dict[tuple, list[dict]] = {}
    for r in all_rows:
        key = (r["customer"].lower(), r["date"])
        groups.setdefault(key, []).append(r)

    valid_groups:   dict[tuple, list[dict]] = {}
    invalid_groups: dict[tuple, list[dict]] = {}
    for key, rows in groups.items():
        if any(r["errors"] for r in rows):
            invalid_groups[key] = rows
        else:
            valid_groups[key] = rows

    rows_skipped = sum(len(rows) for rows in invalid_groups.values())

    errors_report: list[dict] = []
    for rows in invalid_groups.values():
        for r in rows:
            for msg in r["errors"]:
                errors_report.append({
                    "row":      r["row"],
                    "sku":      r["sku"],
                    "customer": r["customer"],
                    "date":     str(r["date"]) if r["date"] else "",
                    "reason":   msg,
                })

    valid_rows = [r for rows in valid_groups.values() for r in rows]
    all_valid_dates = [r["date"] for r in valid_rows if r["date"]]
    total_value = sum(
        r["qty"] * r["price"]
        for r in valid_rows
        if r["qty"] is not None and r["price"] is not None
    )
    earliest_date = min(all_valid_dates).isoformat() if all_valid_dates else None
    latest_date   = max(all_valid_dates).isoformat() if all_valid_dates else None

    # ── Load existing products once (used in both dry-run and real) ─────────
    _all_p = await db.execute(
        select(Product).where(or_(Product.is_active.is_(True), Product.is_active.is_(None)))
    )
    all_products = _all_p.scalars().all()
    products_by_norm_sku: dict[str, Product] = {
        normalize_barcode_value(p.sku): p for p in all_products
    }

    # ── Per-batch state ─────────────────────────────────────────────────────
    batch_id = str(uuid.uuid4()) if not dry_run else None
    today_str = date.today().isoformat()

    invoices_created       = 0
    line_items_created     = 0
    customers_auto_created = 0
    products_auto_created  = 0
    auto_created_customer_names: list[str]  = []
    auto_created_products_list:  list[dict] = []
    warnings_list: list[str] = []

    # In-memory registries to avoid double-creating within the same import run
    newly_created_by_norm_sku: dict[str, Product]  = {}
    new_customers_by_name_lower: dict[str, Customer] = {}
    sku_name_registry: dict[str, str] = {}  # norm_sku → first-seen item name

    # ── Process each invoice group ──────────────────────────────────────────
    for (customer_key, sale_date), rows in valid_groups.items():
        customer_name = rows[0]["customer"]

        # ── Resolve / simulate customer ────────────────────────────────────
        if not customer_name:
            customer_id = (await get_walk_in_customer_id(db)) if not dry_run else -1
        else:
            # Check in-session registry first to avoid double-create
            customer = new_customers_by_name_lower.get(customer_name.lower())
            if not customer:
                _cr = await db.execute(
                    select(Customer).where(
                        sa_func.lower(Customer.name) == customer_name.lower()
                    )
                )
                customer = _cr.scalar_one_or_none()

            if not customer:
                # Truncate customer name if needed
                stored_name = customer_name[:_CUST_MAX]
                if len(customer_name) > _CUST_MAX:
                    warnings_list.append(
                        f"Customer name truncated to {_CUST_MAX} chars: '{stored_name}'"
                    )
                if not dry_run:
                    customer = Customer(
                        name=stored_name,
                        created_by_import_batch=batch_id,
                    )
                    db.add(customer)
                    await db.flush()
                    new_customers_by_name_lower[customer_name.lower()] = customer
                else:
                    customer = Customer(name=stored_name)
                    new_customers_by_name_lower[customer_name.lower()] = customer
                customers_auto_created += 1
                auto_created_customer_names.append(customer_name)
                customer_id = customer.id if not dry_run else -1
            else:
                new_customers_by_name_lower[customer_name.lower()] = customer
                customer_id = customer.id

        # ── Duplicate detection (real-run only) ────────────────────────────
        if not dry_run and not force and sale_date is not None:
            _dup = await db.execute(
                select(Invoice).where(
                    Invoice.customer_id == customer_id,
                    cast(Invoice.created_at, SADate) == sale_date,
                    Invoice.notes.like("Imported from %"),
                )
            )
            if _dup.scalar_one_or_none() is not None:
                for r in rows:
                    errors_report.append({
                        "row":      r["row"],
                        "sku":      r["sku"],
                        "customer": customer_name or "Walk-in",
                        "date":     str(sale_date),
                        "reason":   "Duplicate: invoice already exists for this customer and date from a prior import",
                    })
                rows_skipped += len(rows)
                continue

        # ── Resolve / auto-create products for each line item ──────────────
        line_items: list[tuple] = []  # (product, qty, unit_price, line_total)

        for r in rows:
            raw_sku_str = r["sku"]

            # Truncate SKU if too long
            if len(raw_sku_str) > _SKU_MAX:
                truncated_sku = raw_sku_str[:_SKU_MAX]
                errors_report.append({
                    "row":      r["row"],
                    "sku":      raw_sku_str,
                    "customer": customer_name or "Walk-in",
                    "date":     str(sale_date),
                    "reason":   f"SKU truncated from {len(raw_sku_str)} to {_SKU_MAX} chars: '{truncated_sku}'",
                })
                raw_sku_str = truncated_sku

            norm_sku = normalize_barcode_value(raw_sku_str)

            product = products_by_norm_sku.get(norm_sku) or newly_created_by_norm_sku.get(norm_sku)
            if not product:
                # Determine item name (first-seen wins for duplicate SKUs)
                raw_item_name = r["item"] or f"Imported {raw_sku_str}"
                item_name = raw_item_name[:_NAME_MAX]

                if norm_sku in sku_name_registry:
                    first_name = sku_name_registry[norm_sku]
                    if first_name.lower() != item_name.lower():
                        warn_msg = (
                            f"SKU '{raw_sku_str}' had conflicting item names: "
                            f"'{first_name}' vs '{item_name}' — used the first."
                        )
                        if warn_msg not in warnings_list:
                            warnings_list.append(warn_msg)
                    item_name = first_name
                else:
                    sku_name_registry[norm_sku] = item_name

                price_val = float(r["price"])
                cost_val  = (
                    round(price_val * default_cost_ratio, 4)
                    if default_cost_ratio is not None
                    else 0.0
                )

                new_product = Product(
                    sku=raw_sku_str,
                    name=item_name,
                    price=price_val,
                    cost=cost_val,
                    stock=0,
                    min_stock=5,
                    unit="pcs",
                    is_active=True,
                    category="Imported - Historical",
                    created_by_import_batch=batch_id,
                )

                if not dry_run:
                    db.add(new_product)
                    await db.flush()

                newly_created_by_norm_sku[norm_sku] = new_product
                products_auto_created += 1
                auto_created_products_list.append({
                    "sku":   raw_sku_str,
                    "name":  item_name,
                    "price": price_val,
                    "cost":  cost_val,
                })
                product = new_product

            line_total = float(r["qty"]) * float(r["price"])
            line_items.append((product, float(r["qty"]), float(r["price"]), line_total))

        # In dry-run, count the group and move on
        if dry_run:
            invoices_created += 1
            line_items_created += len(line_items)
            continue

        # ── Build Invoice ──────────────────────────────────────────────────
        subtotal = sum(lt for _, _, _, lt in line_items)
        sale_dt  = datetime(
            sale_date.year, sale_date.month, sale_date.day, 12, 0, 0
        ) if sale_date else datetime.now(timezone.utc)

        try:
            invoice = Invoice(
                customer_id=customer_id,
                user_id=current_user_id,
                payment_method="historical_import",
                subtotal=round(subtotal, 2),
                discount=0,
                total=round(subtotal, 2),
                status="paid",
                notes=f"Imported from {filename} on {today_str}",
                created_at=sale_dt,
                import_batch_id=batch_id,
            )
            db.add(invoice)
            await db.flush()
            invoice.invoice_number = f"HIST-{str(invoice.id).zfill(5)}"

            for product, qty, unit_price, line_total in line_items:
                db.add(InvoiceItem(
                    invoice_id=invoice.id,
                    product_id=product.id,
                    sku=product.sku,
                    name=product.name,
                    qty=qty,
                    unit_price=unit_price,
                    total=round(line_total, 2),
                ))

                if mode == "with_stock_and_journals":
                    before = float(product.stock)
                    after  = before - qty
                    _, location_stock = await sync_product_stock_to_default_location(
                        db, product=product
                    )
                    loc_before = float(location_stock.qty)
                    location_stock.qty = loc_before - qty
                    product.stock = after
                    db.add(StockMove(
                        product_id=product.id,
                        type="out",
                        qty=-qty,
                        qty_before=before,
                        qty_after=after,
                        ref_type="invoice",
                        ref_id=invoice.id,
                        note=f"Historical import — {invoice.invoice_number}",
                        user_id=current_user_id,
                    ))

            if mode in ("with_journals", "with_stock_and_journals"):
                await post_journal(
                    db,
                    f"Historical Sale — {invoice.invoice_number}",
                    [
                        ("1000", round(subtotal, 2), 0),
                        ("4000", 0, round(subtotal, 2)),
                    ],
                    user_id=current_user_id,
                    created_at=sale_dt,
                )

            await db.commit()
            invoices_created  += 1
            line_items_created += len(line_items)

        except Exception as exc:
            await db.rollback()
            errors_report.append({
                "row":      rows[0]["row"],
                "sku":      "",
                "customer": customer_name or "Walk-in",
                "date":     str(sale_date),
                "reason":   f"DB error creating invoice: {exc}",
            })
            rows_skipped += len(rows)

    # ── Build warnings ──────────────────────────────────────────────────────
    if products_auto_created > 0 and default_cost_ratio is None:
        warnings_list.insert(0,
            f"{products_auto_created} products were auto-created with cost = 0. "
            "Gross margin reports may be overstated. "
            "Set costs in Products → filter 'Imported - Historical' category."
        )

    # ── Return result ───────────────────────────────────────────────────────
    if dry_run:
        return {
            "dry_run":  True,
            "mode":     mode,
            "file":     filename,
            "batch_id": None,
            "summary": {
                "rows_read":              rows_read,
                "invoices_created":       0,
                "invoices_would_create":  invoices_created,
                "line_items":             line_items_created,
                "customers_auto_created": customers_auto_created,
                "products_auto_created":  products_auto_created,
                "rows_skipped":           rows_skipped,
                "earliest_date":          earliest_date,
                "latest_date":            latest_date,
                "total_value":            round(total_value, 2),
            },
            "errors":                  errors_report,
            "auto_created_customers":  auto_created_customer_names[:50],
            "auto_created_products":   auto_created_products_list[:50],
            "warnings":                warnings_list,
        }

    return {
        "dry_run":  False,
        "mode":     mode,
        "file":     filename,
        "batch_id": batch_id,
        "summary": {
            "rows_read":              rows_read,
            "invoices_created":       invoices_created,
            "line_items":             line_items_created,
            "customers_auto_created": customers_auto_created,
            "products_auto_created":  products_auto_created,
            "rows_skipped":           rows_skipped,
            "earliest_date":          earliest_date,
            "latest_date":            latest_date,
            "total_value":            round(total_value, 2),
        },
        "errors":                  errors_report,
        "auto_created_customers":  auto_created_customer_names[:50],
        "auto_created_products":   auto_created_products_list[:50],
        "warnings":                warnings_list,
    }
