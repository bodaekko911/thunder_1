from __future__ import annotations

import io
import uuid
from collections import defaultdict
from datetime import date, datetime, timezone
from typing import Optional

import openpyxl
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.farm import Farm, FarmDelivery
from app.models.inventory import StockMove
from app.models.product import Product
from app.services.barcode_service import normalize_barcode_value
from app.services.farm_intake_service import create_farm_delivery

_SKU_MAX = 80
_FARM_MAX = 150
_ITEM_MAX = 200


def _parse_date(val) -> Optional[date]:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    raw = str(val).strip()
    if not raw or raw.lower() == "none":
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


def _safe_float(val) -> Optional[float]:
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _normalize_sku(val) -> str:
    if val is None:
        return ""
    if isinstance(val, (int, float)):
        try:
            return str(int(val)).strip()
        except (ValueError, TypeError):
            pass
    return str(val).strip()


def _find_col(headers: list[str], candidates: list[str]) -> Optional[int]:
    for name in candidates:
        for i, h in enumerate(headers):
            if h == name.lower().strip():
                return i + 1
    return None


def _normalize_farm_name(value) -> str:
    return " ".join(str(value or "").strip().split())


def _extract_batch_id(notes: str | None) -> str | None:
    if not notes or "[Import Batch:" not in notes:
        return None
    try:
        return notes.split("[Import Batch:", 1)[1].split("]", 1)[0].strip()
    except (IndexError, AttributeError):
        return None


def build_import_note(filename: str, batch_id: str, record_stock_movement: bool) -> str:
    stock_label = "recorded" if record_stock_movement else "disabled"
    timestamp = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    return f"Imported from {filename} on {timestamp} [Import Batch: {batch_id}] [Stock Moves: {stock_label}]"


async def import_farm_intake(
    db: AsyncSession,
    workbook_bytes: bytes,
    filename: str,
    current_user_id: int,
    dry_run: bool = True,
    record_stock_movement: bool = True,
) -> dict:
    wb = openpyxl.load_workbook(io.BytesIO(workbook_bytes), data_only=True)
    ws = wb.active

    raw_headers = [str(ws.cell(1, c).value or "").strip().lower() for c in range(1, ws.max_column + 1)]
    col_sku = _find_col(raw_headers, ["sku", "code", "item code", "barcode"])
    col_product = _find_col(raw_headers, ["product", "item", "name", "product name", "description"])
    col_qty = _find_col(raw_headers, ["qty", "quantity", "amount"])
    col_farm = _find_col(raw_headers, ["farm", "farm name"])
    col_date = _find_col(raw_headers, ["date", "delivery date", "intake date"])

    missing = [n for n, c in [("SKU", col_sku), ("QTY", col_qty), ("Farm", col_farm), ("Date", col_date)] if not c]
    if missing:
        return {
            "dry_run": dry_run,
            "record_stock_movement": record_stock_movement,
            "file": filename,
            "error": f"Required column(s) not found in sheet: {', '.join(missing)}",
            "errors": [],
        }

    def _cell(row, col):
        if not col:
            return None
        return ws.cell(row, col).value

    rows_read = 0
    errors: list[dict] = []
    parsed_rows: list[dict] = []

    for rn in range(2, ws.max_row + 1):
        if all(_cell(rn, c) is None for c in range(1, ws.max_column + 1)):
            continue
        rows_read += 1

        raw_sku = _cell(rn, col_sku)
        raw_product = _cell(rn, col_product)
        raw_qty = _cell(rn, col_qty)
        raw_farm = _cell(rn, col_farm)
        raw_date = _cell(rn, col_date)

        sku = _normalize_sku(raw_sku)[:_SKU_MAX]
        item_name = str(raw_product or "").strip()[:_ITEM_MAX]
        farm_name = _normalize_farm_name(raw_farm)[:_FARM_MAX]
        qty = _safe_float(raw_qty)
        parsed_date = _parse_date(raw_date)

        row_errors = []
        if not sku:
            row_errors.append("SKU is required")
        elif not normalize_barcode_value(sku):
            row_errors.append("SKU is invalid")
        if qty is None:
            row_errors.append("QTY must be numeric")
        elif qty <= 0:
            row_errors.append("QTY must be greater than 0")
        if not farm_name:
            row_errors.append("Farm is required")
        if not parsed_date:
            row_errors.append(f"Date '{raw_date}' is invalid — use YYYY-MM-DD, DD/MM/YYYY, or MM/DD/YYYY")

        if row_errors:
            errors.append(
                {
                    "row": rn,
                    "sku": sku,
                    "product": item_name,
                    "farm": farm_name,
                    "date": str(raw_date or ""),
                    "reason": "; ".join(row_errors),
                }
            )
            continue

        parsed_rows.append(
            {
                "row": rn,
                "sku": sku,
                "norm_sku": normalize_barcode_value(sku),
                "product": item_name,
                "qty": qty,
                "farm": farm_name,
                "farm_key": farm_name.lower(),
                "date": parsed_date,
            }
        )

    products_by_norm_sku: dict[str, Product] = {}
    farms_by_name: dict[str, Farm] = {}

    product_result = await db.execute(select(Product))
    for product in product_result.scalars().all():
        norm_sku = normalize_barcode_value(product.sku)
        if norm_sku:
            products_by_norm_sku[norm_sku] = product

    farm_result = await db.execute(select(Farm).where(Farm.is_active == 1))
    for farm in farm_result.scalars().all():
        farms_by_name[_normalize_farm_name(farm.name).lower()] = farm

    farms_auto_created: list[dict] = []
    rows_ready: list[dict] = []
    for row in parsed_rows:
        product = products_by_norm_sku.get(row["norm_sku"])
        if not product:
            label = row["product"] or row["sku"]
            errors.append(
                {
                    "row": row["row"],
                    "sku": row["sku"],
                    "product": label,
                    "farm": row["farm"],
                    "date": row["date"].isoformat(),
                    "reason": f"Product SKU '{row['sku']}' not found",
                }
            )
            continue

        farm = farms_by_name.get(row["farm_key"])
        if farm is None:
            farm = Farm(name=row["farm"], is_active=1)
            farms_by_name[row["farm_key"]] = farm
            farms_auto_created.append({"name": farm.name})
            if not dry_run:
                db.add(farm)
                await db.flush()

        rows_ready.append(
            {
                **row,
                "product_id": product.id,
                "product_name": product.name,
                "farm_id": farm.id if getattr(farm, "id", None) is not None else None,
                "farm_model": farm,
            }
        )

    grouped_rows: dict[tuple[str, date], list[dict]] = defaultdict(list)
    for row in rows_ready:
        grouped_rows[(row["farm_key"], row["date"])].append(row)

    batch_id = str(uuid.uuid4()) if not dry_run and rows_ready else None
    note = build_import_note(filename, batch_id, record_stock_movement) if batch_id else None

    deliveries_created = 0
    stock_moves_created = 0

    if not dry_run:
        for (_farm_key, delivery_date), items in grouped_rows.items():
            farm = items[0]["farm_model"]
            delivery, created_moves = await create_farm_delivery(
                db,
                farm=farm,
                delivery_date=delivery_date,
                user_id=current_user_id,
                items=[{"product_id": item["product_id"], "qty": item["qty"], "notes": None} for item in items],
                notes=note,
                record_stock_movement=record_stock_movement,
                activity_user=None,
            )
            deliveries_created += 1
            stock_moves_created += created_moves
        await db.commit()
    else:
        deliveries_created = len(grouped_rows)
        if record_stock_movement:
            stock_moves_created = len(rows_ready)
        await db.rollback()

    valid_dates = [row["date"] for row in rows_ready]
    summary = {
        "rows_read": rows_read,
        "rows_imported": len(rows_ready),
        "rows_skipped": len(errors),
        "farm_deliveries_created": deliveries_created,
        "farms_auto_created": len(farms_auto_created),
        "products_auto_created": 0,
        "stock_movement_recorded": bool(record_stock_movement),
        "stock_moves_created": stock_moves_created,
        "earliest_date": min(valid_dates).isoformat() if valid_dates else None,
        "latest_date": max(valid_dates).isoformat() if valid_dates else None,
    }

    warnings = []
    if record_stock_movement:
        warnings.append("Stock movement recording was enabled, so imported intake updated product stock and wrote inventory history.")
    else:
        warnings.append("Stock movement recording was disabled, so intake records were imported without changing product stock or writing inventory history.")
    if farms_auto_created:
        warnings.append(f"{len(farms_auto_created)} farm(s) were auto-created from the sheet.")

    return {
        "dry_run": dry_run,
        "file": filename,
        "batch_id": batch_id,
        "record_stock_movement": record_stock_movement,
        "summary": summary,
        "errors": errors,
        "warnings": warnings,
        "auto_created_farms": farms_auto_created,
    }


async def list_farm_intake_import_batches(db: AsyncSession) -> dict:
    result = await db.execute(select(FarmDelivery).options(selectinload(FarmDelivery.items)))
    deliveries = result.scalars().all()

    buckets: dict[str, dict] = {}
    for delivery in deliveries:
        batch_id = _extract_batch_id(delivery.notes)
        if not batch_id:
            continue
        bucket = buckets.setdefault(
            batch_id,
            {
                "batch_id": batch_id,
                "filename": "",
                "ran_on": None,
                "delivery_count": 0,
                "row_count": 0,
                "stock_recorded": False,
            },
        )
        notes = delivery.notes or ""
        if "Imported from " in notes and not bucket["filename"]:
            try:
                bucket["filename"] = notes.split("Imported from ", 1)[1].split(" on ", 1)[0]
            except (IndexError, AttributeError):
                bucket["filename"] = ""
        if "[Stock Moves: recorded]" in notes:
            bucket["stock_recorded"] = True
        created_at = delivery.created_at.date().isoformat() if delivery.created_at else None
        if created_at and (bucket["ran_on"] is None or created_at < bucket["ran_on"]):
            bucket["ran_on"] = created_at
        bucket["delivery_count"] += 1
        bucket["row_count"] += len(delivery.items or [])

    batches = sorted(buckets.values(), key=lambda row: row["ran_on"] or "", reverse=True)
    return {"batches": batches}


async def revert_farm_intake_import_batch(db: AsyncSession, batch_id: str) -> dict:
    result = await db.execute(select(FarmDelivery).options(selectinload(FarmDelivery.items)))
    deliveries = [delivery for delivery in result.scalars().all() if _extract_batch_id(delivery.notes) == batch_id]
    if not deliveries:
        return {"ok": True, "deleted_deliveries": 0, "deleted_stock_moves": 0}

    deleted_stock_moves = 0
    for delivery in deliveries:
        move_result = await db.execute(
            select(StockMove).where(
                StockMove.ref_type == "farm_intake",
                StockMove.ref_id == delivery.id,
            )
        )
        move_rows = move_result.scalars().all()
        moves_by_product: dict[int, list[StockMove]] = {}
        for move in move_rows:
            moves_by_product.setdefault(move.product_id, []).append(move)

        for item in delivery.items:
            product_result = await db.execute(select(Product).where(Product.id == item.product_id))
            product = product_result.scalar_one_or_none()
            product_moves = moves_by_product.get(item.product_id, [])
            if product and product_moves:
                product.stock = float(product.stock or 0) - float(item.qty)
                deleted_stock_moves += len(product_moves)
                for move in product_moves:
                    await db.delete(move)
        await db.delete(delivery)

    await db.commit()
    return {"ok": True, "deleted_deliveries": len(deliveries), "deleted_stock_moves": deleted_stock_moves}
