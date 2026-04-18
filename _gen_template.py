"""Generate sales_import_template.xlsx into app/static/ (and repo root as fallback)."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Sales"

headers = ["SKU", "Item", "QTY", "Price", "Customer", "Date"]
hdr_font = Font(bold=True, color="00FF9D")
hdr_fill = PatternFill("solid", fgColor="0F1424")
for col, h in enumerate(headers, 1):
    c = ws.cell(1, col, h)
    c.font = hdr_font
    c.fill = hdr_fill
    c.alignment = Alignment(horizontal="center")

for col, w in enumerate([14, 28, 8, 10, 22, 14], 1):
    ws.column_dimensions[get_column_letter(col)].width = w

ws.append(["SKU-001", "Olive Oil 500ml",  3,  15.50, "Ahmed Al-Rashid", "2026-01-15"])
ws.append(["SKU-002", "Tahini 250g",      10,  8.00, "Ahmed Al-Rashid", "2026-01-15"])
ws.append(["SKU-001", "Olive Oil 500ml",  1,  15.50, "Sara Khalil",     "2026-01-20"])

readme = wb.create_sheet("README")
readme.column_dimensions["A"].width = 20
readme.column_dimensions["B"].width = 70
readme.append(["Column", "Rules"])
readme["A1"].font = Font(bold=True)
readme["B1"].font = Font(bold=True)

rules = [
    ("SKU",      "Required. Must match a product SKU in the ERP (whitespace stripped). "
                 "Numeric-looking SKUs are normalised (e.g. 12345.0 → 12345)."),
    ("Item",     "Optional. Product description — used only in error messages when the SKU "
                 "is not found. Not stored on the invoice."),
    ("QTY",      "Required. Numeric, must be > 0. Decimals accepted."),
    ("Price",    "Required. Unit price at the time of sale. May differ from the current "
                 "product price. Must be >= 0."),
    ("Customer", "Optional. Customer name. Leave blank for Walk-in Customer. If the name "
                 "does not already exist, a new Customer record is created automatically."),
    ("Date",     "Required. Must be >= 2026-01-01. Accepted formats: YYYY-MM-DD, "
                 "DD/MM/YYYY, MM/DD/YYYY. Excel date cells are also accepted."),
    ("", ""),
    ("Grouping", "Multiple rows with the same Customer + Date are combined into a single "
                 "invoice. Leave Customer blank and they go to the Walk-in invoice for "
                 "that date."),
    ("Dry run",  "Always preview with Dry run checked first. A summary shows what would "
                 "be created without writing anything to the DB. Uncheck Dry run only for "
                 "the final confirmed import."),
]
for key, val in rules:
    readme.append([key, val])
    if key:
        readme.cell(readme.max_row, 1).font = Font(bold=True)

import os, pathlib
_here = pathlib.Path(__file__).parent
for dest in [_here / "app" / "static" / "sales_import_template.xlsx",
             _here / "sales_import_template.xlsx"]:
    wb.save(dest)
    print(f"Saved → {dest}")
